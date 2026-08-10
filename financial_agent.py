# -*- coding: utf-8 -*-
"""
财务 ERP 系统 - 八模块架构 + 用户登录
================================

模块一：离线记账（免费，不需要 API）
模块二：离线报表（免费，不需要 API）
模块三：AI 智能问答（需要 API Key）
模块四：库存管理（免费，自动生成凭证）
模块五：多平台商品管理（免费）
模块六：电商CRM会员（免费）
模块七：全渠道订单OMS（免费）
模块八：BI数据报表（免费）

运行方法：
  pip install streamlit openai pandas openpyxl plotly
  streamlit run financial_agent.py
"""

import json
import os
import re
import hashlib
import sqlite3
from datetime import datetime
import streamlit as st
import pandas as pd
from openai import OpenAI

# set_page_config 必须是第一个 Streamlit 命令
st.set_page_config(page_title="财务 ERP 系统", page_icon="💰", layout="wide")

# ============================================================
# 数据目录：固定路径，确保数据不会因工作目录变化而丢失
# ============================================================
import pathlib
_DATA_DIR = pathlib.Path.home() / ".finance_erp_data"
_DATA_DIR.mkdir(parents=True, exist_ok=True)

# ============================================================
# 数据库兼容层：SQLite / PostgreSQL 自动切换
# ----------------------------------------------------------
# - 本地沙盒（无 SUPABASE_DB_URL）→ 使用 SQLite（原有逻辑）
# - Streamlit Cloud（有 SUPABASE_DB_URL）→ 使用 Supabase PostgreSQL
#   数据永久保存，不受应用重启影响
# ============================================================

# 尝试获取 Supabase 连接字符串
def _get_supabase_url():
    """从环境变量或 Streamlit Secrets 获取 Supabase 连接字符串"""
    url = os.environ.get("SUPABASE_DB_URL", "")
    if not url:
        try:
            url = st.secrets.get("SUPABASE_DB_URL", "")
        except Exception:
            pass
    return url

_SUPABASE_URL = _get_supabase_url()
_USE_POSTGRES = bool(_SUPABASE_URL)

# 表主键映射（用于 INSERT OR REPLACE → ON CONFLICT 转换）
_TABLE_PK = {
    'opening_balances': 'account_code',
    'custom_accounts': 'full_code',
    'products': 'product_code',
    'products_standard': 'spu_code',
    'product_skus': 'sku_code',
    'platform_shops': 'shop_id',
    'bom_headers': 'bom_code',
    'product_categories': 'category_code',
    'member_levels': 'level_code',
    'customer_tags': 'tag_name',
    'orders': 'order_id',
    'logistics_companies': 'company_code',
}


def _db_path_to_schema(db_path):
    """将数据库文件路径转换为 PostgreSQL schema 名称"""
    name = db_path.replace(".db", "")
    name = re.sub(r'[^a-zA-Z0-9_]', '_', name)
    return name


def _adapt_sql(sql):
    """将 SQLite SQL 语法转换为 PostgreSQL 兼容语法"""
    # ? 占位符 → %s
    sql = sql.replace("?", "%s")
    # AUTOINCREMENT → SERIAL
    sql = sql.replace("INTEGER PRIMARY KEY AUTOINCREMENT", "SERIAL PRIMARY KEY")
    # datetime('now') → CURRENT_TIMESTAMP
    sql = sql.replace("datetime('now')", "CURRENT_TIMESTAMP")
    # INSERT OR IGNORE → INSERT ... ON CONFLICT DO NOTHING
    sql = _convert_insert_or_ignore(sql)
    # INSERT OR REPLACE → INSERT ... ON CONFLICT ... DO UPDATE
    sql = _convert_insert_or_replace(sql)
    return sql


def _convert_insert_or_ignore(sql):
    """将 SQLite 的 INSERT OR IGNORE 转换为 PostgreSQL 的 ON CONFLICT DO NOTHING"""
    pattern = r'INSERT\s+OR\s+IGNORE\s+INTO\s+(\w+)'
    match = re.search(pattern, sql, re.IGNORECASE)
    if not match:
        return sql
    table = match.group(1)
    pk = _TABLE_PK.get(table)
    # 先把 INSERT OR IGNORE 替换为普通 INSERT
    sql = re.sub(r'INSERT\s+OR\s+IGNORE\s+INTO', 'INSERT INTO', sql, flags=re.IGNORECASE)
    if pk:
        sql = sql.rstrip().rstrip(';')
        sql += f"\n        ON CONFLICT ({pk}) DO NOTHING"
    return sql


def _convert_insert_or_replace(sql):
    """将 SQLite 的 INSERT OR REPLACE 转换为 PostgreSQL 的 ON CONFLICT"""
    pattern = r'INSERT\s+OR\s+REPLACE\s+INTO\s+(\w+)\s*\(([^)]+)\)\s*VALUES'
    match = re.search(pattern, sql, re.IGNORECASE | re.DOTALL)

    if not match:
        return sql

    table = match.group(1)
    cols_str = match.group(2)
    cols = [c.strip() for c in cols_str.split(',')]

    pk = _TABLE_PK.get(table)
    if not pk:
        return sql.replace("INSERT OR REPLACE", "INSERT")

    non_pk_cols = [c for c in cols if c != pk]

    old_part = match.group(0)
    new_part = f"INSERT INTO {table} ({cols_str}) VALUES"
    sql = sql.replace(old_part, new_part, 1)

    if non_pk_cols:
        set_clause = ", ".join(f"{c} = EXCLUDED.{c}" for c in non_pk_cols)
        sql = sql.rstrip().rstrip(';')
        sql += f"\n        ON CONFLICT ({pk}) DO UPDATE SET {set_clause}"
    else:
        sql = sql.rstrip().rstrip(';')
        sql += f"\n        ON CONFLICT ({pk}) DO NOTHING"

    return sql


if _USE_POSTGRES:
    try:
        import psycopg2
        from psycopg2.extras import RealDictCursor
    except ImportError:
        # psycopg2 未安装时自动回退到 SQLite
        _USE_POSTGRES = False

if _USE_POSTGRES:
    class _PgCursor:
        """PostgreSQL 游标包装器，自动转换 SQLite SQL"""

        def __init__(self, cursor):
            self._cursor = cursor

        def execute(self, sql, params=None):
            pg_sql = _adapt_sql(sql)
            if params is not None:
                self._cursor.execute(pg_sql, params)
            else:
                self._cursor.execute(pg_sql)

        def executemany(self, sql, params_seq):
            pg_sql = _adapt_sql(sql)
            self._cursor.executemany(pg_sql, params_seq)

        def fetchall(self):
            return self._cursor.fetchall()

        def fetchone(self):
            return self._cursor.fetchone()

        def fetchmany(self, size=None):
            return self._cursor.fetchmany(size) if size else self._cursor.fetchmany()

        @property
        def lastrowid(self):
            return None

        @property
        def rowcount(self):
            return self._cursor.rowcount

        @property
        def description(self):
            return self._cursor.description

        def close(self):
            self._cursor.close()

    class _PgConnection:
        """PostgreSQL 连接包装器，模拟 sqlite3.Connection 接口"""

        def __init__(self, db_path):
            self._schema = _db_path_to_schema(db_path)
            self._conn = psycopg2.connect(_SUPABASE_URL)
            self._conn.autocommit = True
            self._row_factory = None

            cur = self._conn.cursor()
            cur.execute(f"CREATE SCHEMA IF NOT EXISTS {self._schema}")
            cur.execute(f"SET search_path TO {self._schema}")
            cur.close()

        @property
        def row_factory(self):
            return self._row_factory

        @row_factory.setter
        def row_factory(self, value):
            self._row_factory = value

        def cursor(self):
            if self._row_factory is not None:
                return _PgCursor(self._conn.cursor(cursor_factory=RealDictCursor))
            return _PgCursor(self._conn.cursor())

        def commit(self):
            self._conn.commit()

        def rollback(self):
            self._conn.rollback()

        def close(self):
            self._conn.close()

        def execute(self, sql, params=None):
            cur = self.cursor()
            cur.execute(sql, params)
            return cur

    class _SQLiteCompat:
        """模拟 sqlite3 模块，底层使用 PostgreSQL"""
        Row = dict

        def connect(self, db_path, **kwargs):
            return _PgConnection(db_path)

    sqlite3 = _SQLiteCompat()
else:
    import sqlite3


# ============================================================
# 全局辅助：会计格式化函数
# ============================================================
def fmt_money(v):
    """将数值格式化为会计专用格式：千分位逗号 + 两位小数，如 300,000.00"""
    try:
        return f"{float(v):,.2f}"
    except (TypeError, ValueError):
        return "0.00"


def fmt_money_df(df, columns):
    """对 DataFrame 的指定列应用会计格式化（返回新 DataFrame）"""
    result = df.copy()
    for col in columns:
        if col in result.columns:
            result[col] = result[col].apply(lambda x: fmt_money(x) if pd.notna(x) else x)
    return result


def _format_money_cb(key):
    """on_change 回调：将金额输入框的值自动格式化为千分位逗号"""
    val = st.session_state.get(key, "")
    if val and str(val).strip():
        cleaned = str(val).replace(",", "").replace("，", "").replace(" ", "").strip()
        try:
            num = float(cleaned)
            st.session_state[key] = f"{num:,.2f}"
        except ValueError:
            pass  # 非数字，保持原样


def money_input(label, default_value=0.0, key=None, min_value=0.0,
                label_visibility="visible", placeholder="0.00",
                on_change_cb=None, on_change_args=None):
    """
    会计专用金额输入框（替代 st.number_input）。
    - 初始值显示千分位逗号，如 300,000.00
    - 用户可以输入带逗号或不带逗号的数字
    - 通过 on_change 回调自动格式化千分位逗号
    - 返回 float 类型的数值
    - on_change_cb: 自定义 on_change 回调（替代默认的 _format_money_cb）
    - on_change_args: 传给回调的参数元组
    """
    init_flag = f"_money_init_{key}"
    if init_flag not in st.session_state:
        if default_value and float(default_value) > 0:
            st.session_state[key] = f"{float(default_value):,.2f}"
        else:
            st.session_state[key] = ""
        st.session_state[init_flag] = True

    _cb = on_change_cb if on_change_cb else _format_money_cb
    _cb_args = tuple(on_change_args) if on_change_args else (key,)

    try:
        raw = st.text_input(
            label, key=key, label_visibility=label_visibility,
            placeholder=placeholder, on_change=_cb, args=_cb_args
        )
    except Exception as e:
        # 表单内不允许 on_change，退回普通 text_input
        st.session_state[f"_dbg_except_{key}"] = str(e)
        raw = st.text_input(
            label, key=key, label_visibility=label_visibility,
            placeholder=placeholder,
        )

    if raw is None or str(raw).strip() == "":
        return 0.0

    cleaned = str(raw).replace(",", "").replace("，", "").replace(" ", "").strip()
    try:
        num = float(cleaned)
        if min_value is not None and num < min_value:
            return float(min_value)
        return num
    except ValueError:
        return 0.0


# ============================================================
# 第 0.5 部分：用户认证系统
# ============================================================

AUTH_DB = str(_DATA_DIR / "erp_users.db")


def init_auth_db():
    """创建用户认证数据库"""
    conn = sqlite3.connect(AUTH_DB)
    c = conn.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS users (
            username TEXT PRIMARY KEY,
            password_hash TEXT NOT NULL,
            display_name TEXT,
            company TEXT,
            created_at TEXT
        )
    """)
    # 多公司管理表：一个用户下可以管理多家公司
    c.execute("""
        CREATE TABLE IF NOT EXISTS user_companies (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL,
            company_code TEXT NOT NULL,
            company_name TEXT NOT NULL,
            tax_id TEXT,
            address TEXT,
            phone TEXT,
            is_default INTEGER DEFAULT 0,
            created_at TEXT,
            UNIQUE(username, company_code)
        )
    """)
    conn.commit()
    conn.close()


def hash_password(password):
    """密码哈希"""
    return hashlib.sha256(password.encode()).hexdigest()


def register_user(username, password, display_name, company):
    """注册新用户"""
    conn = sqlite3.connect(AUTH_DB)
    c = conn.cursor()
    c.execute("SELECT username FROM users WHERE username = ?", (username,))
    if c.fetchone():
        conn.close()
        return False, "用户名已存在"
    from datetime import datetime
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    c.execute("""
        INSERT INTO users (username, password_hash, display_name, company, created_at)
        VALUES (?, ?, ?, ?, ?)
    """, (username, hash_password(password), display_name, company, now_str))
    # 注册时同时创建第一家公司
    if company:
        company_code = "CO001"
    else:
        company = "默认公司"
        company_code = "CO001"
    c.execute("""
        INSERT INTO user_companies (username, company_code, company_name, is_default, created_at)
        VALUES (?, ?, ?, 1, ?)
    """, (username, company_code, company, now_str))
    conn.commit()
    conn.close()
    return True, "注册成功"


def verify_user(username, password):
    """验证用户登录"""
    conn = sqlite3.connect(AUTH_DB)
    c = conn.cursor()
    c.execute("SELECT username, password_hash, display_name, company FROM users WHERE username = ?",
              (username,))
    row = c.fetchone()
    if not row:
        conn.close()
        return False, "用户名不存在"
    if row[1] != hash_password(password):
        conn.close()
        return False, "密码错误"
    # 获取该用户的所有公司
    c.execute("""
        SELECT company_code, company_name, is_default
        FROM user_companies WHERE username = ?
        ORDER BY is_default DESC, id ASC
    """, (username,))
    companies = [{"code": r[0], "name": r[1], "is_default": bool(r[2])} for r in c.fetchall()]
    conn.close()
    return True, {"username": row[0], "display_name": row[2], "company": row[3], "companies": companies}


def create_company(username, company_code, company_name, tax_id="", address="", phone=""):
    """为用户新建一家公司"""
    conn = sqlite3.connect(AUTH_DB)
    c = conn.cursor()
    from datetime import datetime
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    try:
        c.execute("""
            INSERT INTO user_companies (username, company_code, company_name, tax_id, address, phone, is_default, created_at)
            VALUES (?, ?, ?, ?, ?, ?, 0, ?)
        """, (username, company_code, company_name, tax_id, address, phone, now_str))
        conn.commit()
        conn.close()
        return True, "公司创建成功"
    except Exception as e:
        conn.close()
        if "UNIQUE" in str(e):
            return False, "公司编码已存在，请换一个"
        return False, str(e)


def get_user_companies(username):
    """获取用户的所有公司列表"""
    conn = sqlite3.connect(AUTH_DB)
    c = conn.cursor()
    c.execute("""
        SELECT company_code, company_name, tax_id, address, phone, is_default, created_at
        FROM user_companies WHERE username = ?
        ORDER BY is_default DESC, id ASC
    """, (username,))
    rows = c.fetchall()
    conn.close()
    return [
        {"code": r[0], "name": r[1], "tax_id": r[2] or "", "address": r[3] or "",
         "phone": r[4] or "", "is_default": bool(r[5]), "created_at": r[6] or ""}
        for r in rows
    ]


def delete_company(username, company_code):
    """删除一家公司（不能删除默认公司或最后一家公司）"""
    conn = sqlite3.connect(AUTH_DB)
    c = conn.cursor()
    # 检查是否是最后一家公司
    c.execute("SELECT COUNT(*) FROM user_companies WHERE username = ?", (username,))
    count = c.fetchone()[0]
    if count <= 1:
        conn.close()
        return False, "至少保留一家公司，无法删除"
    # 检查是否是默认公司
    c.execute("SELECT is_default FROM user_companies WHERE username = ? AND company_code = ?",
              (username, company_code))
    row = c.fetchone()
    if not row:
        conn.close()
        return False, "公司不存在"
    if row[0]:
        conn.close()
        return False, "不能删除默认公司，请先切换到其他公司"
    c.execute("DELETE FROM user_companies WHERE username = ? AND company_code = ?",
              (username, company_code))
    conn.commit()
    conn.close()
    return True, "公司已删除"


def update_company(username, company_code, company_name=None, tax_id=None, address=None, phone=None):
    """更新公司信息"""
    conn = sqlite3.connect(AUTH_DB)
    c = conn.cursor()
    updates = []
    params = []
    if company_name is not None:
        updates.append("company_name = ?")
        params.append(company_name)
    if tax_id is not None:
        updates.append("tax_id = ?")
        params.append(tax_id)
    if address is not None:
        updates.append("address = ?")
        params.append(address)
    if phone is not None:
        updates.append("phone = ?")
        params.append(phone)
    if not updates:
        conn.close()
        return False, "没有需要更新的字段"
    params.extend([username, company_code])
    c.execute(f"""
        UPDATE user_companies SET {', '.join(updates)}
        WHERE username = ? AND company_code = ?
    """, params)
    conn.commit()
    conn.close()
    return True, "公司信息已更新"


def get_user_db_path(username, company_code=None):
    """获取用户专属数据库路径
    如果提供了 company_code，则每家公司有独立数据库
    否则回退到旧版单数据库路径（兼容）
    """
    if company_code:
        return str(_DATA_DIR / f"erp_data_{username}_{company_code}.db")
    return str(_DATA_DIR / f"erp_data_{username}.db")


def login_page():
    """登录/注册页面"""
    st.markdown("## 💰 财务 ERP 系统")
    st.caption("离线记账 · 自动报表 · AI 智能问答 · 库存管理 · 多平台商品 · CRM · 订单OMS · BI报表")

    tab_login, tab_register = st.tabs(["🔐 登录", "📝 注册"])

    with tab_login:
        with st.form("login_form"):
            username = st.text_input("用户名", key="login_user")
            password = st.text_input("密码", type="password", key="login_pwd")
            submitted = st.form_submit_button("登录")

            if submitted:
                if not username or not password:
                    st.error("请输入用户名和密码")
                else:
                    success, result = verify_user(username, password)
                    if success:
                        st.session_state["logged_in"] = True
                        st.session_state["username"] = username
                        st.session_state["display_name"] = result["display_name"]
                        st.session_state["company"] = result["company"]
                        st.session_state["user_companies"] = result.get("companies", [])
                        # 自动选择默认公司
                        if result.get("companies"):
                            default_co = result["companies"][0]
                            st.session_state["current_company_code"] = default_co["code"]
                            st.session_state["current_company_name"] = default_co["name"]
                        st.rerun()
                    else:
                        st.error(result)

    with tab_register:
        with st.form("register_form"):
            reg_user = st.text_input("设置用户名 *", key="reg_user", placeholder="字母或数字，如：admin")
            reg_pwd = st.text_input("设置密码 *", type="password", key="reg_pwd")
            reg_pwd2 = st.text_input("确认密码 *", type="password", key="reg_pwd2")
            reg_name = st.text_input("显示名称", key="reg_name", placeholder="如：张三")
            reg_company = st.text_input("公司名称", key="reg_company", placeholder="如：XX电子商务有限公司")
            reg_submitted = st.form_submit_button("注册")

            if reg_submitted:
                if not reg_user or not reg_pwd:
                    st.error("用户名和密码不能为空")
                elif reg_pwd != reg_pwd2:
                    st.error("两次输入的密码不一致")
                elif len(reg_pwd) < 4:
                    st.error("密码至少4位")
                else:
                    success, msg = register_user(reg_user, reg_pwd, reg_name or reg_user, reg_company)
                    if success:
                        st.success("注册成功！请切换到「登录」标签登录。")
                    else:
                        st.error(msg)

    st.markdown("---")
    st.caption("一个账号可管理多家公司，每家公司拥有独立的财务数据库。")


# ============================================================
# 第 0.8 部分：登录状态检查
# ============================================================

# 自动迁移旧数据库文件到固定目录
def _migrate_old_dbs():
    """将工作目录下已有的 .db 文件迁移到固定数据目录"""
    import shutil
    cwd = pathlib.Path.cwd()
    for db_file in cwd.glob("erp_*.db"):
        target = _DATA_DIR / db_file.name
        if not target.exists():
            try:
                shutil.copy2(str(db_file), str(target))
            except Exception:
                pass

_migrate_old_dbs()

init_auth_db()

if "logged_in" not in st.session_state:
    st.session_state["logged_in"] = False

if not st.session_state["logged_in"]:
    login_page()
    st.stop()

# 已登录 → 设置用户专属数据库
_username = st.session_state.get("username", "default")

# === 多公司管理：旧用户迁移 ===
# 如果用户登录后没有公司信息（旧用户首次登录），自动创建默认公司
if not st.session_state.get("user_companies"):
    _existing_companies = get_user_companies(_username)
    if _existing_companies:
        # 数据库中已有公司记录但 session_state 中没有
        st.session_state["user_companies"] = _existing_companies
        _default_co = _existing_companies[0]
        st.session_state["current_company_code"] = _default_co["code"]
        st.session_state["current_company_name"] = _default_co["name"]
    else:
        # 旧用户没有公司记录，自动创建一家默认公司
        _old_company_name = st.session_state.get("company", "") or "默认公司"
        create_company(_username, "CO001", _old_company_name)
        _existing_companies = get_user_companies(_username)
        st.session_state["user_companies"] = _existing_companies
        if _existing_companies:
            _default_co = _existing_companies[0]
            st.session_state["current_company_code"] = _default_co["code"]
            st.session_state["current_company_name"] = _default_co["name"]

# 根据当前选中的公司设置 DB_PATH
_current_co_code = st.session_state.get("current_company_code")
_current_co_name = st.session_state.get("current_company_name", "")

# 如果没有当前公司但有旧数据库文件，使用旧路径（兼容）
if _current_co_code:
    DB_PATH = get_user_db_path(_username, _current_co_code)
    # 检查是否有旧数据库需要迁移
    _old_db = get_user_db_path(_username)
    import os as _os_mod
    if _os_mod.path.exists(_old_db) and not _os_mod.path.exists(DB_PATH):
        import shutil as _shutil_mod
        try:
            _shutil_mod.copy2(_old_db, DB_PATH)
        except Exception:
            pass
else:
    DB_PATH = get_user_db_path(_username)


# ============================================================
# 第 1 部分：会计科目表
# ----------------------------------------------------------
# 这是整个系统的基础。
# 每个科目都有：
#   - code: 科目编码（国家标准编码）
#   - name: 科目名称
#   - category: 大类（资产/负债/权益/成本/损益-收入/损益-费用）
#   - direction: 余额方向（借/贷）
#     · 资产、成本类科目正常余额在借方
#     · 负债、权益类科目正常余额在贷方
#     · 收入类科目余额方向是贷方
#     · 费用类科目余额方向是借方
# ============================================================

ACCOUNT_CHART = [
    # ================================================================
    # 一、资产类（余额在借方，抵减科目余额在贷方）
    # ================================================================
    # --- 流动资产 ---
    {"code": "1001", "name": "库存现金",               "category": "资产", "direction": "借"},
    {"code": "1002", "name": "银行存款",               "category": "资产", "direction": "借"},
    {"code": "1012", "name": "其他货币资金",           "category": "资产", "direction": "借"},
    {"code": "1101", "name": "交易性金融资产",         "category": "资产", "direction": "借"},
    {"code": "1102", "name": "交易性金融资产公允价值变动", "category": "资产", "direction": "借"},
    {"code": "1121", "name": "应收票据",               "category": "资产", "direction": "借"},
    {"code": "1122", "name": "应收账款",               "category": "资产", "direction": "借"},
    {"code": "1123", "name": "预付账款",               "category": "资产", "direction": "借"},
    {"code": "1131", "name": "应收股利",               "category": "资产", "direction": "借"},
    {"code": "1132", "name": "应收利息",               "category": "资产", "direction": "借"},
    {"code": "1221", "name": "其他应收款",             "category": "资产", "direction": "借"},
    {"code": "1231", "name": "坏账准备",               "category": "资产", "direction": "贷"},   # 抵减
    {"code": "1241", "name": "应收代位追偿款",         "category": "资产", "direction": "借"},
    {"code": "1251", "name": "应收分保账款",           "category": "资产", "direction": "借"},
    {"code": "1252", "name": "应收分保合同准备金",     "category": "资产", "direction": "借"},
    {"code": "1401", "name": "在途物资",               "category": "资产", "direction": "借"},
    {"code": "1403", "name": "原材料",                 "category": "资产", "direction": "借"},
    {"code": "1405", "name": "库存商品",               "category": "资产", "direction": "借"},
    {"code": "1406", "name": "发出商品",               "category": "资产", "direction": "借"},
    {"code": "1408", "name": "委托加工物资",           "category": "资产", "direction": "借"},
    {"code": "1411", "name": "周转材料",               "category": "资产", "direction": "借"},
    {"code": "1421", "name": "消耗性生物资产",         "category": "资产", "direction": "借"},
    {"code": "1471", "name": "存货跌价准备",           "category": "资产", "direction": "贷"},   # 抵减
    {"code": "1481", "name": "持有待售资产",           "category": "资产", "direction": "借"},
    {"code": "1482", "name": "持有待售资产减值准备",   "category": "资产", "direction": "贷"},   # 抵减
    {"code": "1491", "name": "合同资产",               "category": "资产", "direction": "借"},
    {"code": "1492", "name": "合同资产减值准备",       "category": "资产", "direction": "贷"},   # 抵减
    # --- 非流动资产 ---
    {"code": "1501", "name": "债权投资",               "category": "资产", "direction": "借"},
    {"code": "1502", "name": "债权投资减值准备",       "category": "资产", "direction": "贷"},   # 抵减
    {"code": "1503", "name": "其他债权投资",           "category": "资产", "direction": "借"},
    {"code": "1504", "name": "其他权益工具投资",       "category": "资产", "direction": "借"},
    {"code": "1511", "name": "长期股权投资",           "category": "资产", "direction": "借"},
    {"code": "1512", "name": "长期股权投资减值准备",   "category": "资产", "direction": "贷"},   # 抵减
    {"code": "1521", "name": "投资性房地产",           "category": "资产", "direction": "借"},
    {"code": "1531", "name": "长期应收款",             "category": "资产", "direction": "借"},
    {"code": "1541", "name": "未实现融资收益",         "category": "资产", "direction": "贷"},
    {"code": "1601", "name": "固定资产",               "category": "资产", "direction": "借"},
    {"code": "1602", "name": "累计折旧",               "category": "资产", "direction": "贷"},   # 抵减
    {"code": "1603", "name": "固定资产减值准备",       "category": "资产", "direction": "贷"},   # 抵减
    {"code": "1604", "name": "在建工程",               "category": "资产", "direction": "借"},
    {"code": "1605", "name": "工程物资",               "category": "资产", "direction": "借"},
    {"code": "1606", "name": "固定资产清理",           "category": "资产", "direction": "借"},
    {"code": "1621", "name": "生产性生物资产",         "category": "资产", "direction": "借"},
    {"code": "1622", "name": "生产性生物资产累计折旧", "category": "资产", "direction": "贷"},   # 抵减
    {"code": "1623", "name": "公益性生物资产",         "category": "资产", "direction": "借"},
    {"code": "1641", "name": "使用权资产",             "category": "资产", "direction": "借"},
    {"code": "1642", "name": "使用权资产累计折旧",     "category": "资产", "direction": "贷"},   # 抵减
    {"code": "1651", "name": "油气资产",               "category": "资产", "direction": "借"},
    {"code": "1652", "name": "油气资产折耗",           "category": "资产", "direction": "贷"},   # 抵减
    {"code": "1701", "name": "无形资产",               "category": "资产", "direction": "借"},
    {"code": "1702", "name": "累计摊销",               "category": "资产", "direction": "贷"},   # 抵减
    {"code": "1703", "name": "无形资产减值准备",       "category": "资产", "direction": "贷"},   # 抵减
    {"code": "1801", "name": "长期待摊费用",           "category": "资产", "direction": "借"},
    {"code": "1811", "name": "递延所得税资产",         "category": "资产", "direction": "借"},
    {"code": "1901", "name": "待处理财产损溢",         "category": "资产", "direction": "借"},
    # --- 金融行业专用 ---
    {"code": "1021", "name": "结算备付金",             "category": "资产", "direction": "借"},
    {"code": "1031", "name": "存出保证金",             "category": "资产", "direction": "借"},
    {"code": "1111", "name": "买入返售金融资产",       "category": "资产", "direction": "借"},
    {"code": "1302", "name": "贷款损失准备",           "category": "资产", "direction": "贷"},   # 抵减
    {"code": "1303", "name": "贴现资产",               "category": "资产", "direction": "借"},
    {"code": "1304", "name": "拆出资金",               "category": "资产", "direction": "借"},
    {"code": "1431", "name": "贵金属",                 "category": "资产", "direction": "借"},
    {"code": "1441", "name": "抵债资产",               "category": "资产", "direction": "借"},
    {"code": "1451", "name": "损余物资",               "category": "资产", "direction": "借"},
    {"code": "1461", "name": "融资租赁资产",           "category": "资产", "direction": "借"},
    {"code": "1661", "name": "勘探费用",               "category": "资产", "direction": "借"},
    {"code": "1671", "name": "独立账户资产",           "category": "资产", "direction": "借"},

    # ================================================================
    # 二、负债类（余额在贷方）
    # ================================================================
    {"code": "2001", "name": "短期借款",               "category": "负债", "direction": "贷"},
    {"code": "2201", "name": "应付票据",               "category": "负债", "direction": "贷"},
    {"code": "2202", "name": "应付账款",               "category": "负债", "direction": "贷"},
    {"code": "2203", "name": "合同负债",               "category": "负债", "direction": "贷"},
    {"code": "2211", "name": "应付职工薪酬",           "category": "负债", "direction": "贷"},
    {"code": "2221", "name": "应交税费",               "category": "负债", "direction": "贷"},
    {"code": "2231", "name": "应付利息",               "category": "负债", "direction": "贷"},
    {"code": "2232", "name": "应付股利",               "category": "负债", "direction": "贷"},
    {"code": "2241", "name": "其他应付款",             "category": "负债", "direction": "贷"},
    {"code": "2501", "name": "长期借款",               "category": "负债", "direction": "贷"},
    {"code": "2502", "name": "应付债券",               "category": "负债", "direction": "贷"},
    {"code": "2701", "name": "长期应付款",             "category": "负债", "direction": "贷"},
    {"code": "2711", "name": "递延收益",               "category": "负债", "direction": "贷"},
    {"code": "2801", "name": "预计负债",               "category": "负债", "direction": "贷"},

    # ================================================================
    # 三、所有者权益类（余额在贷方，库存股为借方抵减）
    # ================================================================
    {"code": "4001", "name": "实收资本",               "category": "权益", "direction": "贷"},
    {"code": "4002", "name": "资本公积",               "category": "权益", "direction": "贷"},
    {"code": "4101", "name": "盈余公积",               "category": "权益", "direction": "贷"},
    {"code": "4103", "name": "本年利润",               "category": "权益", "direction": "贷"},
    {"code": "4104", "name": "利润分配",               "category": "权益", "direction": "贷"},
    {"code": "4201", "name": "库存股",                 "category": "权益", "direction": "借"},   # 抵减
    {"code": "4301", "name": "其他综合收益",           "category": "权益", "direction": "贷"},

    # ================================================================
    # 四、成本类（余额在借方）
    # ================================================================
    {"code": "5001", "name": "生产成本",               "category": "成本", "direction": "借"},
    {"code": "5101", "name": "制造费用",               "category": "成本", "direction": "借"},
    {"code": "5201", "name": "劳务成本",               "category": "成本", "direction": "借"},
    {"code": "5301", "name": "研发支出",               "category": "成本", "direction": "借"},
    {"code": "5401", "name": "合同履约",               "category": "成本", "direction": "借"},
    {"code": "5501", "name": "生产成本辅助核算",       "category": "成本", "direction": "借"},

    # ================================================================
    # 五、损益类 - 收入（余额在贷方）
    # ================================================================
    {"code": "6001", "name": "主营业务收入",           "category": "损益-收入", "direction": "贷"},
    {"code": "6021", "name": "手续费及佣金收入",       "category": "损益-收入", "direction": "贷"},
    {"code": "6041", "name": "租赁收入",               "category": "损益-收入", "direction": "贷"},
    {"code": "6051", "name": "其他业务收入",           "category": "损益-收入", "direction": "贷"},
    {"code": "6101", "name": "公允价值变动损益",       "category": "损益-收入", "direction": "贷"},
    {"code": "6102", "name": "套期损益",               "category": "损益-收入", "direction": "贷"},
    {"code": "6103", "name": "净敞口套期损益",         "category": "损益-收入", "direction": "贷"},
    {"code": "6111", "name": "投资收益",               "category": "损益-收入", "direction": "贷"},
    {"code": "6115", "name": "资产处置损益",           "category": "损益-收入", "direction": "贷"},
    {"code": "6117", "name": "其他收益",               "category": "损益-收入", "direction": "贷"},
    {"code": "6301", "name": "营业外收入",             "category": "损益-收入", "direction": "贷"},
    {"code": "6391", "name": "报废收入（保险）",       "category": "损益-收入", "direction": "贷"},

    # ================================================================
    # 五、损益类 - 费用（余额在借方）
    # ================================================================
    {"code": "6401", "name": "主营业务成本",           "category": "损益-费用", "direction": "借"},
    {"code": "6402", "name": "其他业务成本",           "category": "损益-费用", "direction": "借"},
    {"code": "6403", "name": "税金及附加",             "category": "损益-费用", "direction": "借"},
    {"code": "6405", "name": "研发费用",               "category": "损益-费用", "direction": "借"},
    {"code": "6601", "name": "销售费用",               "category": "损益-费用", "direction": "借"},
    {"code": "6602", "name": "管理费用",               "category": "损益-费用", "direction": "借"},
    {"code": "6603", "name": "财务费用",               "category": "损益-费用", "direction": "借"},
    {"code": "6604", "name": "勘探费用",               "category": "损益-费用", "direction": "借"},
    {"code": "6605", "name": "租赁费用",               "category": "损益-费用", "direction": "借"},
    {"code": "6606", "name": "汇兑损益",               "category": "损益-费用", "direction": "借"},
    {"code": "6607", "name": "退保金（保险）",         "category": "损益-费用", "direction": "借"},
    {"code": "6608", "name": "赔付支出（保险）",       "category": "损益-费用", "direction": "借"},
    {"code": "6611", "name": "保单红利支出（保险）",   "category": "损益-费用", "direction": "借"},
    {"code": "6621", "name": "分出保费（保险）",       "category": "损益-费用", "direction": "借"},
    {"code": "6622", "name": "分保费用（保险）",       "category": "损益-费用", "direction": "借"},
    {"code": "6631", "name": "手续及佣金支出（金融）", "category": "损益-费用", "direction": "借"},
    {"code": "6641", "name": "信用减值损失",           "category": "损益-费用", "direction": "借"},
    {"code": "6642", "name": "资产减值损失",           "category": "损益-费用", "direction": "借"},
    {"code": "6701", "name": "营业外支出",             "category": "损益-费用", "direction": "借"},
    {"code": "6711", "name": "所得税费用",             "category": "损益-费用", "direction": "借"},
    {"code": "6801", "name": "以前年度损益调整",       "category": "损益-费用", "direction": "借"},
    {"code": "6901", "name": "其他资产损失",           "category": "损益-费用", "direction": "借"},
    {"code": "6902", "name": "财务担保合同",           "category": "损益-费用", "direction": "借"},
]

# 科目编码 → 科目信息 的映射表，方便快速查找
ACCOUNT_MAP = {a["code"]: a for a in ACCOUNT_CHART}

# 科目名称列表，用于下拉选择框
ACCOUNT_NAMES = [f"{a['code']} {a['name']}" for a in ACCOUNT_CHART]




# ============================================================
# 第 2 部分：数据库初始化
# ============================================================
# DB_PATH 已在第 0.8 部分根据登录用户动态设置


def init_database():
    """创建数据库表（如果不存在的话）"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()

    # 科目表：存科目编码、名称、分类、余额方向
    c.execute("""
        CREATE TABLE IF NOT EXISTS accounts (
            code TEXT PRIMARY KEY,
            name TEXT,
            category TEXT,
            direction TEXT
        )
    """)

    # 期初余额表：记录每个科目的初始余额
    c.execute("""
        CREATE TABLE IF NOT EXISTS opening_balances (
            account_code TEXT PRIMARY KEY,
            opening_debit REAL DEFAULT 0,
            opening_credit REAL DEFAULT 0
        )
    """)

    # 凭证表：每一行是一条分录（借方或贷方）
    c.execute("""
        CREATE TABLE IF NOT EXISTS vouchers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            voucher_number TEXT,
            voucher_date TEXT,
            summary TEXT,
            account_code TEXT,
            account_name TEXT,
            debit_amount REAL DEFAULT 0,
            credit_amount REAL DEFAULT 0
        )
    """)

    # 产品档案表：产成品主数据
    c.execute("""
        CREATE TABLE IF NOT EXISTS products (
            product_code TEXT PRIMARY KEY,
            product_name TEXT NOT NULL,
            specification TEXT,
            unit TEXT,
            cost_price REAL DEFAULT 0,
            selling_price REAL DEFAULT 0,
            account_code TEXT DEFAULT '1405',
            account_name TEXT DEFAULT '库存商品',
            created_at TEXT
        )
    """)

    # 库存出入库记录表：每一次入库或出库都是一条记录
    c.execute("""
        CREATE TABLE IF NOT EXISTS inventory_movements (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            movement_date TEXT,
            movement_type TEXT,
            product_code TEXT,
            product_name TEXT,
            specification TEXT,
            unit TEXT,
            quantity REAL,
            unit_price REAL,
            total_amount REAL,
            summary TEXT,
            counterparty_account TEXT,
            counterparty_account_name TEXT,
            voucher_number TEXT,
            created_at TEXT
        )
    """)

    # ============================================================
    # 多平台商品管理模块表（参照用友U8存货档案设计）
    # ============================================================

    # 商品分类表
    c.execute("""
        CREATE TABLE IF NOT EXISTS product_categories (
            category_code TEXT PRIMARY KEY,
            category_name TEXT NOT NULL,
            parent_code TEXT,
            level INTEGER DEFAULT 1,
            is_active INTEGER DEFAULT 1,
            created_at TEXT
        )
    """)

    # 标准商品库（SPU表）—— ERP内部统一商品主数据
    # 参照用友U8存货档案三维架构：基础属性+业务属性+核算属性
    c.execute("""
        CREATE TABLE IF NOT EXISTS products_standard (
            spu_code TEXT PRIMARY KEY,
            spu_name TEXT NOT NULL,
            category_code TEXT,
            category_name TEXT,
            brand TEXT,
            specification TEXT,
            main_unit TEXT,
            barcode TEXT,
            product_type TEXT DEFAULT '外购',
            cost_price REAL DEFAULT 0,
            selling_price REAL DEFAULT 0,
            tax_rate REAL DEFAULT 0.13,
            supplier TEXT,
            default_warehouse TEXT,
            account_code TEXT DEFAULT '1405',
            account_name TEXT DEFAULT '库存商品',
            description TEXT,
            main_image TEXT,
            is_active INTEGER DEFAULT 1,
            created_at TEXT
        )
    """)

    # SKU规格表 —— 一个SPU下多个规格组合
    c.execute("""
        CREATE TABLE IF NOT EXISTS product_skus (
            sku_code TEXT PRIMARY KEY,
            spu_code TEXT NOT NULL,
            sku_name TEXT NOT NULL,
            spec_attrs TEXT,
            barcode TEXT,
            cost_price REAL DEFAULT 0,
            selling_price REAL DEFAULT 0,
            weight REAL DEFAULT 0,
            weight_unit TEXT DEFAULT '克',
            is_active INTEGER DEFAULT 1,
            created_at TEXT
        )
    """)

    # 平台店铺表 —— 各电商平台店铺信息（含API预留字段）
    c.execute("""
        CREATE TABLE IF NOT EXISTS platform_shops (
            shop_id TEXT PRIMARY KEY,
            platform TEXT NOT NULL,
            shop_name TEXT NOT NULL,
            shop_url TEXT,
            api_app_key TEXT,
            api_app_secret TEXT,
            api_access_token TEXT,
            api_token_expire TEXT,
            api_status TEXT DEFAULT '未对接',
            is_active INTEGER DEFAULT 1,
            created_at TEXT
        )
    """)

    # 平台商品绑定表（一品多商）—— 一个ERP商品对应多个平台链接
    c.execute("""
        CREATE TABLE IF NOT EXISTS platform_product_links (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            spu_code TEXT NOT NULL,
            sku_code TEXT,
            platform TEXT NOT NULL,
            shop_id TEXT,
            platform_product_id TEXT,
            platform_sku_id TEXT,
            platform_title TEXT,
            platform_item_url TEXT,
            platform_price REAL DEFAULT 0,
            platform_stock INTEGER DEFAULT 0,
            listing_status TEXT DEFAULT '未上架',
            last_sync_time TEXT,
            sync_status TEXT DEFAULT '未同步',
            created_at TEXT
        )
    """)

    # BOM表头（母件）
    c.execute("""
        CREATE TABLE IF NOT EXISTS bom_headers (
            bom_code TEXT PRIMARY KEY,
            parent_sku_code TEXT NOT NULL,
            parent_sku_name TEXT NOT NULL,
            bom_type TEXT DEFAULT '生产',
            version TEXT DEFAULT 'V1.0',
            total_cost REAL DEFAULT 0,
            status TEXT DEFAULT '启用',
            description TEXT,
            created_at TEXT
        )
    """)

    # BOM明细（子件）
    c.execute("""
        CREATE TABLE IF NOT EXISTS bom_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            bom_code TEXT NOT NULL,
            seq INTEGER DEFAULT 1,
            child_sku_code TEXT NOT NULL,
            child_sku_name TEXT NOT NULL,
            quantity REAL NOT NULL,
            unit TEXT,
            unit_cost REAL DEFAULT 0,
            total_cost REAL DEFAULT 0,
            loss_rate REAL DEFAULT 0,
            is_active INTEGER DEFAULT 1
        )
    """)

    # 商品标签表
    c.execute("""
        CREATE TABLE IF NOT EXISTS product_tags (
            tag_id INTEGER PRIMARY KEY AUTOINCREMENT,
            tag_name TEXT UNIQUE NOT NULL,
            tag_color TEXT DEFAULT '#1890ff',
            created_at TEXT
        )
    """)

    # 商品-标签关联表
    c.execute("""
        CREATE TABLE IF NOT EXISTS product_tag_links (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            spu_code TEXT NOT NULL,
            tag_id INTEGER NOT NULL
        )
    """)

    # ============================================================
    # CRM 会员模块表
    # ============================================================

    # 会员等级定义表
    c.execute("""
        CREATE TABLE IF NOT EXISTS member_levels (
            level_code TEXT PRIMARY KEY,
            level_name TEXT NOT NULL,
            min_points INTEGER DEFAULT 0,
            min_consumption REAL DEFAULT 0,
            discount_rate REAL DEFAULT 1.0,
            description TEXT,
            sort_order INTEGER DEFAULT 0,
            created_at TEXT
        )
    """)

    # 客户主表 —— 全平台客户数据聚合
    c.execute("""
        CREATE TABLE IF NOT EXISTS customers (
            customer_id TEXT PRIMARY KEY,
            customer_name TEXT NOT NULL,
            phone TEXT,
            email TEXT,
            gender TEXT,
            birthday TEXT,
            province TEXT,
            city TEXT,
            register_date TEXT,
            level_code TEXT DEFAULT 'L1',
            level_name TEXT DEFAULT '普通会员',
            total_points INTEGER DEFAULT 0,
            total_consumption REAL DEFAULT 0,
            total_orders INTEGER DEFAULT 0,
            avg_order_value REAL DEFAULT 0,
            last_order_date TEXT,
            max_order_value REAL DEFAULT 0,
            status TEXT DEFAULT '正常',
            remark TEXT,
            created_at TEXT
        )
    """)

    # 客户平台账号表 —— 一个客户可绑定多个平台账号
    c.execute("""
        CREATE TABLE IF NOT EXISTS customer_platform_accounts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            customer_id TEXT NOT NULL,
            platform TEXT NOT NULL,
            platform_user_id TEXT,
            platform_nick TEXT,
            shop_id TEXT,
            bind_date TEXT,
            created_at TEXT
        )
    """)

    # 客户消费记录表 —— 对接订单数据
    c.execute("""
        CREATE TABLE IF NOT EXISTS customer_consumption_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            customer_id TEXT NOT NULL,
            order_date TEXT,
            platform TEXT,
            order_id TEXT,
            order_amount REAL,
            product_summary TEXT,
            points_earned INTEGER DEFAULT 0,
            points_used INTEGER DEFAULT 0,
            created_at TEXT
        )
    """)

    # 客户积分变动日志表
    c.execute("""
        CREATE TABLE IF NOT EXISTS customer_points_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            customer_id TEXT NOT NULL,
            change_type TEXT,
            points_change INTEGER,
            balance_after INTEGER,
            description TEXT,
            related_order TEXT,
            created_at TEXT
        )
    """)

    # 客户标签定义表
    c.execute("""
        CREATE TABLE IF NOT EXISTS customer_tags (
            tag_id INTEGER PRIMARY KEY AUTOINCREMENT,
            tag_name TEXT UNIQUE NOT NULL,
            tag_type TEXT DEFAULT '自定义',
            tag_color TEXT DEFAULT '#1890ff',
            auto_rule TEXT,
            created_at TEXT
        )
    """)

    # 客户-标签关联表
    c.execute("""
        CREATE TABLE IF NOT EXISTS customer_tag_links (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            customer_id TEXT NOT NULL,
            tag_id INTEGER NOT NULL,
            is_auto INTEGER DEFAULT 0,
            created_at TEXT
        )
    """)

    # 客户黑名单表
    c.execute("""
        CREATE TABLE IF NOT EXISTS customer_blacklist (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            customer_id TEXT NOT NULL,
            customer_name TEXT,
            reason TEXT,
            block_type TEXT DEFAULT '临时',
            start_date TEXT,
            end_date TEXT,
            operator TEXT,
            status TEXT DEFAULT '生效中',
            created_at TEXT
        )
    """)

    # ============================================================
    # OMS 全渠道订单模块表
    # ============================================================

    # 订单主表
    c.execute("""
        CREATE TABLE IF NOT EXISTS orders (
            order_id TEXT PRIMARY KEY,
            platform TEXT NOT NULL,
            shop_id TEXT,
            shop_name TEXT,
            customer_id TEXT,
            customer_name TEXT,
            receiver_name TEXT,
            receiver_phone TEXT,
            receiver_province TEXT,
            receiver_city TEXT,
            receiver_district TEXT,
            receiver_address TEXT,
            order_type TEXT DEFAULT '现货',
            order_status TEXT DEFAULT '待付款',
            total_amount REAL DEFAULT 0,
            discount_amount REAL DEFAULT 0,
            shipping_fee REAL DEFAULT 0,
            actual_amount REAL DEFAULT 0,
            payment_method TEXT,
            payment_time TEXT,
            order_time TEXT,
            ship_time TEXT,
            complete_time TEXT,
            logistics_company TEXT,
            logistics_number TEXT,
            logistics_status TEXT,
            seller_remark TEXT,
            buyer_remark TEXT,
            invoice_required INTEGER DEFAULT 0,
            invoice_title TEXT,
            is_abnormal INTEGER DEFAULT 0,
            abnormal_reason TEXT,
            parent_order_id TEXT,
            split_from TEXT,
            created_at TEXT,
            updated_at TEXT
        )
    """)

    # 订单明细表
    c.execute("""
        CREATE TABLE IF NOT EXISTS order_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            order_id TEXT NOT NULL,
            spu_code TEXT,
            spu_name TEXT,
            sku_code TEXT,
            sku_name TEXT,
            spec_attrs TEXT,
            quantity INTEGER DEFAULT 1,
            unit_price REAL DEFAULT 0,
            total_price REAL DEFAULT 0,
            discount_price REAL DEFAULT 0,
            is_gift INTEGER DEFAULT 0
        )
    """)

    # 订单状态变更日志表
    c.execute("""
        CREATE TABLE IF NOT EXISTS order_status_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            order_id TEXT NOT NULL,
            from_status TEXT,
            to_status TEXT,
            operator TEXT,
            remark TEXT,
            created_at TEXT
        )
    """)

    # 异常订单拦截记录表
    c.execute("""
        CREATE TABLE IF NOT EXISTS abnormal_orders (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            order_id TEXT NOT NULL,
            customer_name TEXT,
            platform TEXT,
            intercept_reason TEXT,
            intercept_type TEXT DEFAULT '自动',
            risk_level TEXT DEFAULT '中',
            handle_status TEXT DEFAULT '待处理',
            handle_result TEXT,
            operator TEXT,
            created_at TEXT,
            handled_at TEXT
        )
    """)

    # 订单合并/拆分记录表
    c.execute("""
        CREATE TABLE IF NOT EXISTS order_operations (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            operation_type TEXT NOT NULL,
            source_order_ids TEXT,
            target_order_id TEXT,
            reason TEXT,
            operator TEXT,
            created_at TEXT
        )
    """)

    # 物流公司信息表
    c.execute("""
        CREATE TABLE IF NOT EXISTS logistics_companies (
            company_code TEXT PRIMARY KEY,
            company_name TEXT NOT NULL,
            api_code TEXT,
            is_active INTEGER DEFAULT 1,
            sort_order INTEGER DEFAULT 0
        )
    """)

    # 售后单表
    c.execute("""
        CREATE TABLE IF NOT EXISTS after_sales_orders (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            after_sales_id TEXT UNIQUE,
            order_id TEXT NOT NULL,
            platform TEXT,
            shop_name TEXT,
            customer_id TEXT,
            customer_name TEXT,
            type TEXT DEFAULT '退货',
            reason TEXT,
            spu_code TEXT,
            spu_name TEXT,
            sku_code TEXT,
            sku_name TEXT,
            quantity INTEGER DEFAULT 1,
            refund_amount REAL DEFAULT 0,
            status TEXT DEFAULT '待处理',
            handle_result TEXT,
            operator TEXT,
            apply_date TEXT,
            handle_date TEXT,
            created_at TEXT
        )
    """)

    # 自定义二级科目表
    c.execute("""
        CREATE TABLE IF NOT EXISTS custom_accounts (
            full_code TEXT PRIMARY KEY,
            parent_code TEXT NOT NULL,
            parent_name TEXT NOT NULL,
            sub_code TEXT NOT NULL,
            sub_name TEXT NOT NULL,
            full_name TEXT NOT NULL,
            category TEXT NOT NULL,
            direction TEXT NOT NULL,
            tax_flag TEXT DEFAULT '',
            created_at TEXT
        )
    """)

    # 如果科目表是空的，把预置科目写进去
    c.execute("SELECT COUNT(*) FROM accounts")
    if c.fetchone()[0] == 0:
        for a in ACCOUNT_CHART:
            c.execute(
                "INSERT INTO accounts VALUES (?, ?, ?, ?)",
                (a["code"], a["name"], a["category"], a["direction"]),
            )

    # ============================================================
    # 预置常用二级科目（参照企业会计准则应用指南 + 常见财务软件模板）
    # 如果 custom_accounts 表为空，自动创建以下二级科目：
    #   一、资产类    1001/1012/1221/1601/1602/1701/1702/1801
    #   二、负债类    2211/2221/2241
    #   三、所有者权益  4001/4002/4101/4104
    #   四、成本类    5001/5101
    #   五、损益类    6601/6602/6603/6301/6701
    # 注：1002银行存款、1122应收账款、1123预付账款、2202应付账款等
    #     通常启用辅助核算而非预置二级科目，此处不预置。
    # ============================================================
    c.execute("SELECT COUNT(*) FROM custom_accounts")
    if c.fetchone()[0] == 0:
        _preset_subs = [
            # (parent_code, parent_name, sub_name, category, direction, tax_flag)

            # ===== 一、资产类 =====
            # 1001 库存现金
            ("1001", "库存现金", "人民币",           "资产", "借", ""),
            ("1001", "库存现金", "外币",             "资产", "借", ""),
            # 1012 其他货币资金
            ("1012", "其他货币资金", "外埠存款",       "资产", "借", ""),
            ("1012", "其他货币资金", "银行汇票存款",   "资产", "借", ""),
            ("1012", "其他货币资金", "银行本票存款",   "资产", "借", ""),
            ("1012", "其他货币资金", "信用证保证金",   "资产", "借", ""),
            ("1012", "其他货币资金", "存出投资款",     "资产", "借", ""),
            # 1221 其他应收款
            ("1221", "其他应收款", "员工借款",         "资产", "借", ""),
            ("1221", "其他应收款", "备用金",           "资产", "借", ""),
            ("1221", "其他应收款", "押金保证金",       "资产", "借", ""),
            ("1221", "其他应收款", "其他应收",         "资产", "借", ""),
            # 1601 固定资产
            ("1601", "固定资产", "房屋建筑物",         "资产", "借", ""),
            ("1601", "固定资产", "机器设备",           "资产", "借", ""),
            ("1601", "固定资产", "运输设备",           "资产", "借", ""),
            ("1601", "固定资产", "电子设备",           "资产", "借", ""),
            ("1601", "固定资产", "办公设备",           "资产", "借", ""),
            # 1602 累计折旧（与固定资产一一对应，贷方抵减）
            ("1602", "累计折旧", "房屋建筑物折旧",     "资产", "贷", ""),
            ("1602", "累计折旧", "机器设备折旧",       "资产", "贷", ""),
            ("1602", "累计折旧", "运输设备折旧",       "资产", "贷", ""),
            ("1602", "累计折旧", "电子设备折旧",       "资产", "贷", ""),
            ("1602", "累计折旧", "办公设备折旧",       "资产", "贷", ""),
            # 1701 无形资产
            ("1701", "无形资产", "土地使用权",         "资产", "借", ""),
            ("1701", "无形资产", "软件及专利权",       "资产", "借", ""),
            # 1702 累计摊销（与无形资产对应，贷方抵减）
            ("1702", "累计摊销", "土地使用权摊销",     "资产", "贷", ""),
            ("1702", "累计摊销", "软件及专利权摊销",   "资产", "贷", ""),
            # 1801 长期待摊费用
            ("1801", "长期待摊费用", "开办费",                   "资产", "借", ""),
            ("1801", "长期待摊费用", "租入固定资产改良支出",     "资产", "借", ""),

            # ===== 二、负债类 =====
            # 2211 应付职工薪酬（几乎所有软件标准预置）
            ("2211", "应付职工薪酬", "工资薪金",         "负债", "贷", "工资薪金"),
            ("2211", "应付职工薪酬", "职工福利费",       "负债", "贷", "职工福利费"),
            ("2211", "应付职工薪酬", "社会保险费",       "负债", "贷", "社保公积金"),
            ("2211", "应付职工薪酬", "住房公积金",       "负债", "贷", "社保公积金"),
            ("2211", "应付职工薪酬", "工会经费",         "负债", "贷", "工会经费"),
            ("2211", "应付职工薪酬", "职工教育经费",     "负债", "贷", "职工教育经费"),
            # 2221 应交税费（强制标准二级，全部软件预置）
            #   应交增值税的三级明细（进项/销项等）展开为二级科目
            ("2221", "应交税费", "应交增值税-进项税额",         "负债", "贷", ""),
            ("2221", "应交税费", "应交增值税-销项税额",         "负债", "贷", ""),
            ("2221", "应交税费", "应交增值税-进项税额转出",     "负债", "贷", ""),
            ("2221", "应交税费", "应交增值税-已交税金",         "负债", "贷", ""),
            ("2221", "应交税费", "未交增值税",                   "负债", "贷", ""),
            ("2221", "应交税费", "城建税",                       "负债", "贷", ""),
            ("2221", "应交税费", "教育费附加",                   "负债", "贷", ""),
            ("2221", "应交税费", "地方教育附加",                 "负债", "贷", ""),
            ("2221", "应交税费", "企业所得税",                   "负债", "贷", ""),
            ("2221", "应交税费", "个人所得税",                   "负债", "贷", ""),
            ("2221", "应交税费", "印花税",                       "负债", "贷", ""),
            ("2221", "应交税费", "房产税",                       "负债", "贷", ""),
            ("2221", "应交税费", "土地使用税",                   "负债", "贷", ""),
            # 2241 其他应付款
            ("2241", "其他应付款", "押金保证金",       "负债", "贷", ""),
            ("2241", "其他应付款", "代扣款项",         "负债", "贷", ""),
            ("2241", "其他应付款", "其他",             "负债", "贷", ""),

            # ===== 三、所有者权益类 =====
            # 4001 实收资本
            ("4001", "实收资本", "法人资本",           "权益", "贷", ""),
            ("4001", "实收资本", "个人资本",           "权益", "贷", ""),
            # 4002 资本公积
            ("4002", "资本公积", "资本溢价",           "权益", "贷", ""),
            # 4101 盈余公积
            ("4101", "盈余公积", "法定盈余公积",       "权益", "贷", ""),
            ("4101", "盈余公积", "任意盈余公积",       "权益", "贷", ""),
            # 4104 利润分配
            ("4104", "利润分配", "提取法定盈余公积",   "权益", "贷", ""),
            ("4104", "利润分配", "应付利润",           "权益", "贷", ""),
            ("4104", "利润分配", "未分配利润",         "权益", "贷", ""),

            # ===== 四、成本类（生产型企业模板预置）=====
            # 5001 生产成本
            ("5001", "生产成本", "直接材料",           "成本", "借", ""),
            ("5001", "生产成本", "直接人工",           "成本", "借", ""),
            ("5001", "生产成本", "制造费用转入",       "成本", "借", ""),
            # 5101 制造费用
            ("5101", "制造费用", "车间人工",           "成本", "借", ""),
            ("5101", "制造费用", "车间折旧",           "成本", "借", ""),
            ("5101", "制造费用", "车间水电费",         "成本", "借", ""),
            ("5101", "制造费用", "机物料消耗",         "成本", "借", ""),

            # ===== 五、损益类 =====
            # 6601 销售费用
            ("6601", "销售费用", "工资",               "损益-费用", "借", "工资薪金"),
            ("6601", "销售费用", "广告费及业务宣传费", "损益-费用", "借", "广告费"),
            ("6601", "销售费用", "运输费",             "损益-费用", "借", ""),
            ("6601", "销售费用", "差旅费",             "损益-费用", "借", ""),
            ("6601", "销售费用", "业务招待费",         "损益-费用", "借", "业务招待费"),
            ("6601", "销售费用", "折旧费",             "损益-费用", "借", ""),
            ("6601", "销售费用", "职工福利费",         "损益-费用", "借", "职工福利费"),
            ("6601", "销售费用", "销售佣金",           "损益-费用", "借", ""),
            # 6602 管理费用（通用模板标配，预置最完整）
            ("6602", "管理费用", "工资薪金",           "损益-费用", "借", "工资薪金"),
            ("6602", "管理费用", "办公费",             "损益-费用", "借", ""),
            ("6602", "管理费用", "差旅费",             "损益-费用", "借", ""),
            ("6602", "管理费用", "业务招待费",         "损益-费用", "借", "业务招待费"),
            ("6602", "管理费用", "折旧费",             "损益-费用", "借", ""),
            ("6602", "管理费用", "水电费",             "损益-费用", "借", ""),
            ("6602", "管理费用", "租赁费",             "损益-费用", "借", ""),
            ("6602", "管理费用", "社保公积金",         "损益-费用", "借", "社保公积金"),
            ("6602", "管理费用", "咨询审计费",         "损益-费用", "借", ""),
            ("6602", "管理费用", "印花税及税费",       "损益-费用", "借", ""),
            ("6602", "管理费用", "职工福利费",         "损益-费用", "借", "职工福利费"),
            ("6602", "管理费用", "研发费用",           "损益-费用", "借", "研发费用"),
            ("6602", "管理费用", "工会经费",           "损益-费用", "借", "工会经费"),
            ("6602", "管理费用", "职工教育经费",       "损益-费用", "借", "职工教育经费"),
            ("6602", "管理费用", "邮电通讯费",         "损益-费用", "借", ""),
            ("6602", "管理费用", "车辆使用费",         "损益-费用", "借", ""),
            ("6602", "管理费用", "会议费",             "损益-费用", "借", ""),
            # 6603 财务费用
            ("6603", "财务费用", "利息支出",           "损益-费用", "借", ""),
            ("6603", "财务费用", "利息收入",           "损益-费用", "贷", ""),
            ("6603", "财务费用", "银行手续费",         "损益-费用", "借", ""),
            ("6603", "财务费用", "汇兑损益",           "损益-费用", "借", ""),
            # 6301 营业外收入
            ("6301", "营业外收入", "政府补助",         "损益-收入", "贷", ""),
            ("6301", "营业外收入", "罚款收入",         "损益-收入", "贷", ""),
            ("6301", "营业外收入", "其他",             "损益-收入", "贷", ""),
            # 6701 营业外支出
            ("6701", "营业外支出", "罚款支出",         "损益-费用", "借", ""),
            ("6701", "营业外支出", "公益性捐赠",       "损益-费用", "借", "公益性捐赠"),
            ("6701", "营业外支出", "资产盘亏损失",     "损益-费用", "借", ""),
        ]
        # 按父科目分别编号（01、02、03...），避免跨科目编号混乱
        _sub_counters = {}
        for p_code, p_name, s_name, cat, direction, tax_flag in _preset_subs:
            _sub_counters[p_code] = _sub_counters.get(p_code, 0) + 1
            sub_code = f"{_sub_counters[p_code]:02d}"
            full_code = f"{p_code}{sub_code}"
            full_name = f"{p_name}-{s_name}"
            c.execute("""
                INSERT OR IGNORE INTO custom_accounts
                (full_code, parent_code, parent_name, sub_code, sub_name, full_name,
                 category, direction, tax_flag, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (full_code, p_code, p_name, sub_code, s_name, full_name,
                  cat, direction, tax_flag,
                  datetime.now().strftime("%Y-%m-%d %H:%M:%S")))

    conn.commit()
    conn.close()


def get_next_voucher_number():
    """获取下一个凭证编号，如：记字第001号"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT COUNT(DISTINCT voucher_number) FROM vouchers")
    count = c.fetchone()[0]
    conn.close()
    return f"记字第{count + 1:03d}号"


def save_voucher(voucher_number, voucher_date, summary, lines):
    """
    保存一张凭证到数据库
    lines 是一个列表，每个元素：{account_code, account_name, debit, credit}
    """
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    for line in lines:
        c.execute("""
            INSERT INTO vouchers
            (voucher_number, voucher_date, summary, account_code, account_name, debit_amount, credit_amount)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (
            voucher_number, voucher_date, summary,
            line["account_code"], line["account_name"],
            float(line["debit"]), float(line["credit"]),
        ))
    conn.commit()
    conn.close()
    get_all_vouchers.clear()


def delete_voucher(voucher_number):
    """删除一张凭证（所有分录行）"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("DELETE FROM vouchers WHERE voucher_number = ?", (voucher_number,))
    conn.commit()
    conn.close()
    get_all_vouchers.clear()
    get_all_opening_balances.clear()
    calc_account_balance.clear()


def update_voucher(voucher_number, new_date, new_summary, new_lines):
    """
    修改一张凭证：先删除旧分录，再写入新分录。
    new_lines: [{account_code, account_name, debit, credit}, ...]
    """
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    # 删除旧分录
    c.execute("DELETE FROM vouchers WHERE voucher_number = ?", (voucher_number,))
    # 写入新分录
    for line in new_lines:
        c.execute("""
            INSERT INTO vouchers
            (voucher_number, voucher_date, summary, account_code, account_name, debit_amount, credit_amount)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (
            voucher_number,
            new_date,
            new_summary,
            line["account_code"],
            line["account_name"],
            float(line["debit"]),
            float(line["credit"]),
        ))
    conn.commit()
    conn.close()
    get_all_vouchers.clear()
    get_all_opening_balances.clear()
    calc_account_balance.clear()


def get_voucher_by_number(voucher_number):
    """获取一张凭证的所有分录，返回列表 [{account_code, account_name, debit, credit}, ...]"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("""
        SELECT account_code, account_name, debit_amount, credit_amount
        FROM vouchers WHERE voucher_number = ?
        ORDER BY id
    """, (voucher_number,))
    rows = c.fetchall()
    conn.close()
    return [
        {"account_code": r[0], "account_name": r[1], "debit": r[2] or 0, "credit": r[3] or 0}
        for r in rows
    ]


def get_voucher_info(voucher_number):
    """获取凭证的基本信息（日期、摘要）"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("""
        SELECT voucher_date, summary FROM vouchers
        WHERE voucher_number = ? LIMIT 1
    """, (voucher_number,))
    row = c.fetchone()
    conn.close()
    if row:
        return {"date": row[0], "summary": row[1]}
    return None


@st.cache_data(ttl=60, show_spinner=False)
def get_all_vouchers():
    """获取所有凭证，返回 DataFrame，缓存 1 分钟"""
    conn = sqlite3.connect(DB_PATH)
    df = pd.read_sql_query("SELECT * FROM vouchers ORDER BY id", conn)
    conn.close()
    return df


@st.cache_data(ttl=30, show_spinner=False)
def get_all_opening_balances():
    """一次查询获取所有期初余额，返回 {account_code: (debit, credit)} 字典"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT account_code, opening_debit, opening_credit FROM opening_balances")
    rows = c.fetchall()
    conn.close()
    return {row[0]: (row[1] or 0, row[2] or 0) for row in rows}


def get_opening_balance(account_code):
    """获取某个科目的期初余额（借方和贷方）"""
    _all = get_all_opening_balances()
    if account_code in _all:
        return _all[account_code]
    return 0, 0


def save_opening_balance(account_code, debit, credit):
    """保存某个科目的期初余额"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("""
        INSERT OR REPLACE INTO opening_balances
        (account_code, opening_debit, opening_credit)
        VALUES (?, ?, ?)
    """, (account_code, float(debit), float(credit)))
    conn.commit()
    conn.close()
    get_all_opening_balances.clear()


# ============================================================
# 第 2.2 部分：自定义二级科目管理
# ============================================================

# 税务特殊科目标记
TAX_SPECIAL_FLAGS = {
    "": "无特殊税务标记",
    "业务招待费": "业务招待费：按实际发生额60%扣除，且不超过当年销售收入的5‰",
    "广告费": "广告费：不超过当年销售（营业）收入15%的部分准予扣除",
    "公益性捐赠": "公益性捐赠：不超过年度利润总额12%的部分准予扣除",
    "职工福利费": "职工福利费：不超过工资薪金总额14%的部分准予扣除",
    "工会经费": "工会经费：不超过工资薪金总额2%的部分准予扣除",
    "职工教育经费": "职工教育经费：不超过工资薪金总额8%的部分准予扣除",
    "研发费用": "研发费用：按规定加计扣除（一般企业75%，特定领域100%）",
}


def get_custom_accounts(parent_code=None):
    """获取自定义二级科目列表，可按一级科目筛选"""
    conn = sqlite3.connect(DB_PATH)
    if parent_code:
        df = pd.read_sql_query(
            "SELECT * FROM custom_accounts WHERE parent_code = ? ORDER BY full_code",
            conn, params=(parent_code,)
        )
    else:
        df = pd.read_sql_query(
            "SELECT * FROM custom_accounts ORDER BY parent_code, sub_code", conn
        )
    conn.close()
    return df


def get_next_sub_code(parent_code):
    """获取下一个二级科目序号，如 '01', '02'..."""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute(
        "SELECT sub_code FROM custom_accounts WHERE parent_code = ? ORDER BY sub_code DESC LIMIT 1",
        (parent_code,)
    )
    row = c.fetchone()
    conn.close()
    if row:
        next_num = int(row[0]) + 1
    else:
        next_num = 1
    return f"{next_num:02d}"


def add_custom_account(parent_code, parent_name, sub_name, category, direction, tax_flag=""):
    """添加一个自定义二级科目"""
    sub_code = get_next_sub_code(parent_code)
    full_code = f"{parent_code}{sub_code}"
    full_name = f"{parent_name}-{sub_name}"
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("""
        INSERT INTO custom_accounts
        (full_code, parent_code, parent_name, sub_code, sub_name, full_name, category, direction, tax_flag, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (full_code, parent_code, parent_name, sub_code, sub_name,
          full_name, category, direction, tax_flag,
          datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
    conn.commit()
    conn.close()
    return full_code, full_name


def delete_custom_account(full_code):
    """删除一个自定义二级科目（同时清理期初余额）"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("DELETE FROM custom_accounts WHERE full_code = ?", (full_code,))
    c.execute("DELETE FROM opening_balances WHERE account_code = ?", (full_code,))
    conn.commit()
    conn.close()


@st.cache_data(ttl=300, show_spinner=False)
def get_all_accounts_for_display():
    """
    合并系统科目 + 自定义二级科目，返回统一格式的列表。
    缓存 5 分钟，减少数据库查询次数。
    每个元素: {"code", "name", "display_name", "category", "direction",
               "is_custom", "parent_code", "tax_flag", "level"}
    level: 1=一级科目, 2=二级科目
    """
    result = []
    # 系统一级科目
    for a in ACCOUNT_CHART:
        result.append({
            "code": a["code"],
            "name": a["name"],
            "display_name": f"{a['code']} {a['name']}",
            "category": a["category"],
            "direction": a["direction"],
            "is_custom": False,
            "parent_code": None,
            "tax_flag": "",
            "level": 1,
        })
    # 自定义二级科目
    df = get_custom_accounts()
    for _, row in df.iterrows():
        result.append({
            "code": row["full_code"],
            "name": row["full_name"],
            "display_name": f"{row['full_code']} {row['full_name']}  ← {row['parent_name']}",
            "category": row["category"],
            "direction": row["direction"],
            "is_custom": True,
            "parent_code": row["parent_code"],
            "tax_flag": row.get("tax_flag", ""),
            "level": 2,
        })
    return result


@st.cache_data(ttl=300, show_spinner=False)
def get_account_names_dynamic():
    """获取动态科目名称列表（用于下拉选择框），缓存 5 分钟"""
    accounts = get_all_accounts_for_display()
    return [a["display_name"] for a in accounts]


def get_account_info_dynamic(code):
    """根据科目编码获取科目信息（先查系统科目，再查自定义科目）"""
    # 先查系统一级科目
    acc = ACCOUNT_MAP.get(code)
    if acc:
        return {
            "code": acc["code"],
            "name": acc["name"],
            "display_name": f"{acc['code']} {acc['name']}",
            "category": acc["category"],
            "direction": acc["direction"],
            "is_custom": False,
            "parent_code": None,
            "tax_flag": "",
            "level": 1,
        }
    # 再查自定义二级科目
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT * FROM custom_accounts WHERE full_code = ?", (code,))
    row = c.fetchone()
    conn.close()
    if row:
        return {
            "code": row[0],
            "name": row[5],
            "display_name": f"{row[0]} {row[5]}",
            "category": row[6],
            "direction": row[7],
            "is_custom": True,
            "parent_code": row[1],
            "tax_flag": row[8] or "",
            "level": 2,
        }
    return None


@st.cache_data(ttl=300, show_spinner=False)
def get_sub_account_codes(parent_code):
    """获取某个一级科目下所有二级科目的编码列表"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute(
        "SELECT full_code FROM custom_accounts WHERE parent_code = ?",
        (parent_code,)
    )
    codes = [row[0] for row in c.fetchall()]
    conn.close()
    return codes

def _now():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


@st.cache_data(ttl=60, show_spinner=False)
def get_all_products():
    """获取所有产品档案，返回 DataFrame"""
    conn = sqlite3.connect(DB_PATH)
    df = pd.read_sql_query("SELECT * FROM products ORDER BY product_code", conn)
    conn.close()
    return df


def add_product(product_code, product_name, specification, unit,
                cost_price, selling_price, account_code, account_name):
    """新增一个产品档案"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("""
        INSERT OR REPLACE INTO products
        (product_code, product_name, specification, unit,
         cost_price, selling_price, account_code, account_name, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (product_code, product_name, specification, unit,
          float(cost_price), float(selling_price),
          account_code, account_name, _now()))
    conn.commit()
    conn.close()


def delete_product(product_code):
    """删除一个产品档案"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("DELETE FROM products WHERE product_code = ?", (product_code,))
    conn.commit()
    conn.close()


def get_product_stock(product_code):
    """
    计算某个产品的当前库存数量。
    入库为正，出库为负，累加即为当前库存。
    """
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("""
        SELECT
            COALESCE(SUM(CASE WHEN movement_type = '入库' THEN quantity ELSE 0 END), 0)
            - COALESCE(SUM(CASE WHEN movement_type = '出库' THEN quantity ELSE 0 END), 0)
        FROM inventory_movements
        WHERE product_code = ?
    """, (product_code,))
    stock = c.fetchone()[0]
    conn.close()
    return stock or 0


@st.cache_data(ttl=60, show_spinner=False)
def get_all_stock():
    """获取所有产品的当前库存汇总，返回 DataFrame"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("""
        SELECT
            p.product_code,
            p.product_name,
            p.specification,
            p.unit,
            p.cost_price,
            p.selling_price,
            COALESCE(SUM(CASE WHEN m.movement_type = '入库' THEN m.quantity ELSE 0 END), 0)
              - COALESCE(SUM(CASE WHEN m.movement_type = '出库' THEN m.quantity ELSE 0 END), 0)
              AS current_stock,
            COALESCE(SUM(CASE WHEN m.movement_type = '入库' THEN m.total_amount ELSE 0 END), 0)
              - COALESCE(SUM(CASE WHEN m.movement_type = '出库' THEN m.total_amount ELSE 0 END), 0)
              AS stock_value
        FROM products p
        LEFT JOIN inventory_movements m ON p.product_code = m.product_code
        GROUP BY p.product_code
        ORDER BY p.product_code
    """)
    rows = c.fetchall()
    conn.close()
    df = pd.DataFrame(rows, columns=[
        "产品编码", "产品名称", "规格", "单位",
        "成本单价", "销售单价", "当前库存", "库存金额"
    ])
    return df


def add_inventory_movement(movement_date, movement_type, product_code,
                           product_name, specification, unit,
                           quantity, unit_price, summary,
                           counterparty_account, counterparty_account_name):
    """
    记录一条出入库记录，并自动生成对应的会计凭证。
    返回生成的凭证编号。
    """
    total_amount = float(quantity) * float(unit_price)

    # --- 1. 保存库存记录 ---
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()

    # 获取下一个凭证编号
    c.execute("SELECT COUNT(DISTINCT voucher_number) FROM vouchers")
    count = c.fetchone()[0]
    voucher_number = f"库字第{count + 1:03d}号"

    c.execute("""
        INSERT INTO inventory_movements
        (movement_date, movement_type, product_code, product_name,
         specification, unit, quantity, unit_price, total_amount,
         summary, counterparty_account, counterparty_account_name,
         voucher_number, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (movement_date, movement_type, product_code, product_name,
          specification, unit, float(quantity), float(unit_price),
          total_amount, summary,
          counterparty_account, counterparty_account_name,
          voucher_number, _now()))
    conn.commit()
    conn.close()

    # --- 2. 自动生成会计凭证 ---
    # 产品关联的库存科目（默认 1405 库存商品）
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT account_code, account_name FROM products WHERE product_code = ?",
              (product_code,))
    row = c.fetchone()
    conn.close()

    if row:
        stock_code, stock_name = row
    else:
        stock_code, stock_name = "1405", "库存商品"

    if movement_type == "入库":
        # 入库：借 库存商品，贷 对方科目（如银行存款/应付账款）
        lines = [
            {"account_code": stock_code, "account_name": stock_name,
             "debit": total_amount, "credit": 0},
            {"account_code": counterparty_account,
             "account_name": counterparty_account_name,
             "debit": 0, "credit": total_amount},
        ]
    else:
        # 出库：借 对方科目（如主营业务成本），贷 库存商品
        lines = [
            {"account_code": counterparty_account,
             "account_name": counterparty_account_name,
             "debit": total_amount, "credit": 0},
            {"account_code": stock_code, "account_name": stock_name,
             "debit": 0, "credit": total_amount},
        ]

    save_voucher(voucher_number, movement_date, summary, lines)

    return voucher_number


def get_inventory_movements(product_code=None):
    """获取库存变动记录，可选按产品筛选。返回 DataFrame"""
    conn = sqlite3.connect(DB_PATH)
    if product_code:
        df = pd.read_sql_query(
            "SELECT * FROM inventory_movements WHERE product_code = ? ORDER BY movement_date, id",
            conn, params=(product_code,))
    else:
        df = pd.read_sql_query(
            "SELECT * FROM inventory_movements ORDER BY movement_date, id",
            conn)
    conn.close()
    return df


# ============================================================
# 第 2.5 部分：多平台商品管理 —— 数据库操作函数
# ----------------------------------------------------------
# 参照用友U8存货档案 + 多平台电商 ERP 设计：
#   SPU（标准商品库） -> SKU（规格组合） -> 平台店铺绑定（一品多商）
#   BOM（物料清单）  -> 分类树 / 标签体系 -> 批量导入导出
# ============================================================


# ---------- 1. SPU（标准商品库）管理 ----------

@st.cache_data(ttl=60, show_spinner=False)
def get_all_spus():
    """获取所有标准商品（SPU），返回 DataFrame"""
    conn = sqlite3.connect(DB_PATH)
    df = pd.read_sql_query("SELECT * FROM products_standard ORDER BY spu_code", conn)
    conn.close()
    return df


def get_spu_by_code(spu_code):
    """根据 SPU 编码获取单个 SPU 信息，返回 dict 或 None"""
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    c.execute("SELECT * FROM products_standard WHERE spu_code = ?", (spu_code,))
    row = c.fetchone()
    conn.close()
    if not row:
        return None
    return dict(row)


def add_spu(spu_code, spu_name, category_code, category_name, brand, specification,
            main_unit, barcode, product_type, cost_price, selling_price, tax_rate,
            supplier, default_warehouse, account_code, account_name, description):
    """新增或更新一个标准商品（SPU）。使用 INSERT OR REPLACE 实现覆盖。"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("""
        INSERT OR REPLACE INTO products_standard
        (spu_code, spu_name, category_code, category_name, brand, specification,
         main_unit, barcode, product_type, cost_price, selling_price, tax_rate,
         supplier, default_warehouse, account_code, account_name, description,
         is_active, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1, ?)
    """, (spu_code, spu_name, category_code, category_name, brand, specification,
          main_unit, barcode, product_type, float(cost_price), float(selling_price),
          float(tax_rate), supplier, default_warehouse, account_code, account_name,
          description, _now()))
    conn.commit()
    conn.close()


def delete_spu(spu_code):
    """删除一个 SPU，并级联清理其下的 SKU、平台绑定、标签关联与相关 BOM"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    # 收集该 SPU 下的 SKU 编码，用于清理 BOM
    c.execute("SELECT sku_code FROM product_skus WHERE spu_code = ?", (spu_code,))
    sku_codes = [r[0] for r in c.fetchall()]
    c.execute("DELETE FROM product_skus WHERE spu_code = ?", (spu_code,))
    c.execute("DELETE FROM platform_product_links WHERE spu_code = ?", (spu_code,))
    c.execute("DELETE FROM product_tag_links WHERE spu_code = ?", (spu_code,))
    if sku_codes:
        placeholders = ",".join("?" for _ in sku_codes)
        c.execute(f"SELECT bom_code FROM bom_headers WHERE parent_sku_code IN ({placeholders})",
                  sku_codes)
        bom_codes = [r[0] for r in c.fetchall()]
        if bom_codes:
            bph = ",".join("?" for _ in bom_codes)
            c.execute(f"DELETE FROM bom_items WHERE bom_code IN ({bph})", bom_codes)
        c.execute(f"DELETE FROM bom_headers WHERE parent_sku_code IN ({placeholders})", sku_codes)
    c.execute("DELETE FROM products_standard WHERE spu_code = ?", (spu_code,))
    conn.commit()
    conn.close()


# ---------- 2. SKU（规格组合）管理 ----------

def get_all_skus():
    """获取所有 SKU（含所属 SPU 名称），返回 DataFrame"""
    conn = sqlite3.connect(DB_PATH)
    df = pd.read_sql_query(
        "SELECT s.*, p.spu_name FROM product_skus s "
        "LEFT JOIN products_standard p ON s.spu_code = p.spu_code "
        "ORDER BY s.spu_code, s.sku_code", conn)
    conn.close()
    return df


def get_skus_by_spu(spu_code):
    """获取某个 SPU 下的所有 SKU，返回 DataFrame"""
    conn = sqlite3.connect(DB_PATH)
    df = pd.read_sql_query(
        "SELECT * FROM product_skus WHERE spu_code = ? ORDER BY sku_code",
        conn, params=(spu_code,))
    conn.close()
    return df


def add_sku(sku_code, spu_code, sku_name, spec_attrs, barcode,
            cost_price, selling_price, weight, weight_unit):
    """新增或更新一个 SKU"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("""
        INSERT OR REPLACE INTO product_skus
        (sku_code, spu_code, sku_name, spec_attrs, barcode, cost_price,
         selling_price, weight, weight_unit, is_active, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 1, ?)
    """, (sku_code, spu_code, sku_name, spec_attrs, barcode,
          float(cost_price), float(selling_price), float(weight), weight_unit, _now()))
    conn.commit()
    conn.close()


def delete_sku(sku_code):
    """删除一个 SKU，并清理平台绑定、BOM（作为母件或子件）"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("DELETE FROM platform_product_links WHERE sku_code = ?", (sku_code,))
    # 作为母件的 BOM
    c.execute("SELECT bom_code FROM bom_headers WHERE parent_sku_code = ?", (sku_code,))
    bom_codes = [r[0] for r in c.fetchall()]
    if bom_codes:
        bph = ",".join("?" for _ in bom_codes)
        c.execute(f"DELETE FROM bom_items WHERE bom_code IN ({bph})", bom_codes)
    c.execute("DELETE FROM bom_headers WHERE parent_sku_code = ?", (sku_code,))
    # 作为子件的明细
    c.execute("DELETE FROM bom_items WHERE child_sku_code = ?", (sku_code,))
    c.execute("DELETE FROM product_skus WHERE sku_code = ?", (sku_code,))
    conn.commit()
    conn.close()


# ---------- 3. 平台店铺管理 ----------

@st.cache_data(ttl=60, show_spinner=False)
def get_all_shops():
    """获取所有平台店铺，返回 DataFrame"""
    conn = sqlite3.connect(DB_PATH)
    df = pd.read_sql_query(
        "SELECT * FROM platform_shops ORDER BY platform, shop_id", conn)
    conn.close()
    return df


def add_shop(shop_id, platform, shop_name, shop_url, app_key, app_secret,
             access_token, token_expire, api_status):
    """新增或更新一个平台店铺（含 API 预留字段）"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("""
        INSERT OR REPLACE INTO platform_shops
        (shop_id, platform, shop_name, shop_url, api_app_key, api_app_secret,
         api_access_token, api_token_expire, api_status, is_active, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 1, ?)
    """, (shop_id, platform, shop_name, shop_url, app_key, app_secret,
          access_token, token_expire, api_status, _now()))
    conn.commit()
    conn.close()


def delete_shop(shop_id):
    """删除一个平台店铺（保留其历史绑定记录）"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("DELETE FROM platform_shops WHERE shop_id = ?", (shop_id,))
    conn.commit()
    conn.close()


# ---------- 4. 平台商品绑定（一品多商）管理 ----------

def get_platform_links(platform=None, shop_id=None):
    """获取平台商品绑定记录，可选按平台/店铺筛选。返回 DataFrame"""
    conn = sqlite3.connect(DB_PATH)
    query = (
        "SELECT l.*, p.spu_name, s.sku_name, sh.shop_name "
        "FROM platform_product_links l "
        "LEFT JOIN products_standard p ON l.spu_code = p.spu_code "
        "LEFT JOIN product_skus s ON l.sku_code = s.sku_code "
        "LEFT JOIN platform_shops sh ON l.shop_id = sh.shop_id "
        "WHERE 1=1"
    )
    params = []
    if platform:
        query += " AND l.platform = ?"
        params.append(platform)
    if shop_id:
        query += " AND l.shop_id = ?"
        params.append(shop_id)
    query += " ORDER BY l.platform, l.spu_code, l.id"
    df = pd.read_sql_query(query, conn, params=params)
    conn.close()
    return df


def get_links_by_spu(spu_code):
    """获取某个 SPU 的所有平台绑定，返回 DataFrame"""
    conn = sqlite3.connect(DB_PATH)
    df = pd.read_sql_query(
        "SELECT l.*, sh.shop_name FROM platform_product_links l "
        "LEFT JOIN platform_shops sh ON l.shop_id = sh.shop_id "
        "WHERE l.spu_code = ? ORDER BY l.platform, l.id",
        conn, params=(spu_code,))
    conn.close()
    return df


def add_platform_link(spu_code, sku_code, platform, shop_id, platform_product_id,
                      platform_sku_id, platform_title, platform_item_url,
                      platform_price, platform_stock, listing_status):
    """新增一条平台商品绑定记录"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("""
        INSERT INTO platform_product_links
        (spu_code, sku_code, platform, shop_id, platform_product_id, platform_sku_id,
         platform_title, platform_item_url, platform_price, platform_stock,
         listing_status, last_sync_time, sync_status, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, '未同步', ?)
    """, (spu_code, sku_code, platform, shop_id, platform_product_id, platform_sku_id,
          platform_title, platform_item_url, float(platform_price), int(platform_stock),
          listing_status, _now(), _now()))
    conn.commit()
    conn.close()


def update_listing_status(link_id, listing_status):
    """更新某条平台绑定的上下架状态"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("""
        UPDATE platform_product_links
        SET listing_status = ?, last_sync_time = ?, sync_status = '已同步'
        WHERE id = ?
    """, (listing_status, _now(), link_id))
    conn.commit()
    conn.close()


# ---------- 5. BOM（物料清单）管理 ----------

def get_all_boms():
    """获取所有 BOM 表头，返回 DataFrame"""
    conn = sqlite3.connect(DB_PATH)
    df = pd.read_sql_query("SELECT * FROM bom_headers ORDER BY bom_code", conn)
    conn.close()
    return df


def add_bom(bom_code, parent_sku_code, parent_sku_name, bom_type, version, description):
    """新增或更新一个 BOM 表头"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("""
        INSERT OR REPLACE INTO bom_headers
        (bom_code, parent_sku_code, parent_sku_name, bom_type, version,
         total_cost, status, description, created_at)
        VALUES (?, ?, ?, ?, ?, 0, '启用', ?, ?)
    """, (bom_code, parent_sku_code, parent_sku_name, bom_type, version,
          description, _now()))
    conn.commit()
    conn.close()


def get_bom_items(bom_code):
    """获取某个 BOM 的子件明细，返回 DataFrame"""
    conn = sqlite3.connect(DB_PATH)
    df = pd.read_sql_query(
        "SELECT * FROM bom_items WHERE bom_code = ? ORDER BY seq, id",
        conn, params=(bom_code,))
    conn.close()
    return df


def add_bom_item(bom_code, seq, child_sku_code, child_sku_name, quantity,
                 unit, unit_cost, loss_rate):
    """新增一条 BOM 子件明细，并重算 BOM 总成本。
    单行成本 = 数量 * 单价 * (1 + 损耗率)
    """
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    quantity = float(quantity)
    unit_cost = float(unit_cost)
    loss_rate = float(loss_rate)
    line_total = quantity * unit_cost * (1 + loss_rate)
    c.execute("""
        INSERT INTO bom_items
        (bom_code, seq, child_sku_code, child_sku_name, quantity, unit,
         unit_cost, total_cost, loss_rate, is_active)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 1)
    """, (bom_code, int(seq), child_sku_code, child_sku_name, quantity, unit,
          unit_cost, line_total, loss_rate))
    c.execute("SELECT COALESCE(SUM(total_cost), 0) FROM bom_items WHERE bom_code = ?", (bom_code,))
    new_total = c.fetchone()[0]
    c.execute("UPDATE bom_headers SET total_cost = ? WHERE bom_code = ?", (new_total, bom_code))
    conn.commit()
    conn.close()


def delete_bom(bom_code):
    """删除一个 BOM（表头 + 明细）"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("DELETE FROM bom_items WHERE bom_code = ?", (bom_code,))
    c.execute("DELETE FROM bom_headers WHERE bom_code = ?", (bom_code,))
    conn.commit()
    conn.close()


def calc_bom_cost(bom_code):
    """计算 BOM 总成本 = SUM(子件数量 * 子件单价 * (1 + 损耗率))"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("""
        SELECT COALESCE(SUM(quantity * unit_cost * (1 + loss_rate)), 0)
        FROM bom_items WHERE bom_code = ?
    """, (bom_code,))
    total = c.fetchone()[0]
    conn.close()
    return total


# ---------- 6. 商品分类管理 ----------

def get_all_categories():
    """获取所有商品分类，返回 DataFrame"""
    conn = sqlite3.connect(DB_PATH)
    df = pd.read_sql_query(
        "SELECT * FROM product_categories ORDER BY level, category_code", conn)
    conn.close()
    return df


def add_category(category_code, category_name, parent_code, level):
    """新增或更新一个商品分类"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("""
        INSERT OR REPLACE INTO product_categories
        (category_code, category_name, parent_code, level, is_active, created_at)
        VALUES (?, ?, ?, ?, 1, ?)
    """, (category_code, category_name, parent_code, int(level), _now()))
    conn.commit()
    conn.close()


# ---------- 7. 商品标签管理 ----------

def get_all_tags():
    """获取所有商品标签，返回 DataFrame"""
    conn = sqlite3.connect(DB_PATH)
    df = pd.read_sql_query("SELECT * FROM product_tags ORDER BY tag_id", conn)
    conn.close()
    return df


def add_tag(tag_name, tag_color):
    """新增一个商品标签（名称重复则忽略）"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    try:
        c.execute("INSERT INTO product_tags (tag_name, tag_color, created_at) VALUES (?, ?, ?)",
                  (tag_name, tag_color, _now()))
        conn.commit()
    except sqlite3.IntegrityError:
        pass  # tag_name 唯一约束冲突，忽略
    conn.close()


def get_tags_by_spu(spu_code):
    """获取某个 SPU 已绑定的所有标签，返回 DataFrame"""
    conn = sqlite3.connect(DB_PATH)
    df = pd.read_sql_query(
        "SELECT t.* FROM product_tags t "
        "INNER JOIN product_tag_links l ON t.tag_id = l.tag_id "
        "WHERE l.spu_code = ? ORDER BY t.tag_id",
        conn, params=(spu_code,))
    conn.close()
    return df


def add_product_tag(spu_code, tag_id):
    """给 SPU 添加一个标签（已存在则忽略）"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT id FROM product_tag_links WHERE spu_code = ? AND tag_id = ?",
              (spu_code, tag_id))
    if not c.fetchone():
        c.execute("INSERT INTO product_tag_links (spu_code, tag_id) VALUES (?, ?)",
                  (spu_code, tag_id))
        conn.commit()
    conn.close()


# ---------- 8. 批量导入导出 ----------

def _safe_str(val):
    """Excel 导入时的字符串安全转换，空/nan -> 空串"""
    if val is None:
        return ""
    s = str(val).strip()
    if s.lower() == "nan":
        return ""
    return s


def _safe_float(val, default=0.0):
    """Excel 导入时的浮点安全转换"""
    try:
        if val is None:
            return default
        s = str(val).strip().lower()
        if s in ("", "nan", "none"):
            return default
        return float(val)
    except (ValueError, TypeError):
        return default


def export_products_excel():
    """导出 SPU 与 SKU 数据为 Excel（两个工作表），返回 BytesIO"""
    from io import BytesIO
    spus = get_all_spus()
    skus = get_all_skus()
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        spus.to_excel(writer, index=False, sheet_name="SPU标准商品")
        skus.to_excel(writer, index=False, sheet_name="SKU规格")
    output.seek(0)
    return output


def import_products_from_excel(file_bytes):
    """从 Excel 导入 SPU 与 SKU 数据。
    返回 (spu_count, sku_count)。
    工作表名：SPU标准商品 / SKU规格
    """
    from io import BytesIO
    xls = pd.ExcelFile(BytesIO(file_bytes), engine="openpyxl")
    spu_count = 0
    sku_count = 0
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()

    if "SPU标准商品" in xls.sheet_names:
        spu_df = xls.parse("SPU标准商品")
        for _, row in spu_df.iterrows():
            spu_code = _safe_str(row.get("spu_code"))
            spu_name = _safe_str(row.get("spu_name"))
            if not spu_code or not spu_name:
                continue
            c.execute("""
                INSERT OR REPLACE INTO products_standard
                (spu_code, spu_name, category_code, category_name, brand, specification,
                 main_unit, barcode, product_type, cost_price, selling_price, tax_rate,
                 supplier, default_warehouse, account_code, account_name, description,
                 is_active, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1, ?)
            """, (
                spu_code, spu_name,
                _safe_str(row.get("category_code")), _safe_str(row.get("category_name")),
                _safe_str(row.get("brand")), _safe_str(row.get("specification")),
                _safe_str(row.get("main_unit")), _safe_str(row.get("barcode")),
                _safe_str(row.get("product_type")) or "外购",
                _safe_float(row.get("cost_price", 0)),
                _safe_float(row.get("selling_price", 0)),
                _safe_float(row.get("tax_rate", 0.13), 0.13),
                _safe_str(row.get("supplier")), _safe_str(row.get("default_warehouse")),
                _safe_str(row.get("account_code")) or "1405",
                _safe_str(row.get("account_name")) or "库存商品",
                _safe_str(row.get("description")), _now(),
            ))
            spu_count += 1

    if "SKU规格" in xls.sheet_names:
        sku_df = xls.parse("SKU规格")
        for _, row in sku_df.iterrows():
            sku_code = _safe_str(row.get("sku_code"))
            spu_code = _safe_str(row.get("spu_code"))
            sku_name = _safe_str(row.get("sku_name"))
            if not sku_code or not spu_code or not sku_name:
                continue
            c.execute("""
                INSERT OR REPLACE INTO product_skus
                (sku_code, spu_code, sku_name, spec_attrs, barcode, cost_price,
                 selling_price, weight, weight_unit, is_active, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 1, ?)
            """, (
                sku_code, spu_code, sku_name,
                _safe_str(row.get("spec_attrs")), _safe_str(row.get("barcode")),
                _safe_float(row.get("cost_price", 0)),
                _safe_float(row.get("selling_price", 0)),
                _safe_float(row.get("weight", 0)),
                _safe_str(row.get("weight_unit")) or "克", _now(),
            ))
            sku_count += 1

    conn.commit()
    conn.close()
    return spu_count, sku_count


# ============================================================
# 第 2.9 部分：CRM 会员模块数据库函数
# ============================================================

# ---------- 1. 会员等级管理 ----------

@st.cache_data(ttl=60, show_spinner=False)
def get_all_levels():
    """查询所有会员等级，按 sort_order 排序，返回 DataFrame"""
    conn = sqlite3.connect(DB_PATH)
    df = pd.read_sql_query(
        "SELECT * FROM member_levels ORDER BY sort_order", conn)
    conn.close()
    return df


def add_level(level_code, level_name, min_points, min_consumption,
              discount_rate, description, sort_order):
    """新增或更新一个会员等级"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("""
        INSERT OR REPLACE INTO member_levels
        (level_code, level_name, min_points, min_consumption, discount_rate,
         description, sort_order, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    """, (level_code, level_name, int(min_points), float(min_consumption),
          float(discount_rate), description, int(sort_order), _now()))
    conn.commit()
    conn.close()


def delete_level(level_code):
    """删除一个会员等级"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("DELETE FROM member_levels WHERE level_code = ?", (level_code,))
    conn.commit()
    conn.close()


def check_and_update_level(customer_id):
    """根据客户积分和消费额自动升级等级。
    遍历所有等级（按 sort_order 升序），找同时满足 min_points 与 min_consumption
    的最高等级；若与当前等级不同则更新。
    """
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute(
        "SELECT total_points, total_consumption FROM customers WHERE customer_id = ?",
        (customer_id,))
    row = c.fetchone()
    if not row:
        conn.close()
        return
    total_points = row[0] or 0
    total_consumption = row[1] or 0

    levels = c.execute(
        "SELECT level_code, level_name, min_points, min_consumption "
        "FROM member_levels ORDER BY sort_order").fetchall()
    if not levels:
        conn.close()
        return

    target_code = None
    target_name = None
    for level_code, level_name, min_points, min_consumption in levels:
        if total_points >= (min_points or 0) and total_consumption >= (min_consumption or 0):
            target_code = level_code
            target_name = level_name

    if target_code:
        c.execute(
            "SELECT level_code FROM customers WHERE customer_id = ?", (customer_id,))
        cur = c.fetchone()
        cur_code = cur[0] if cur else None
        if cur_code != target_code:
            c.execute(
                "UPDATE customers SET level_code = ?, level_name = ? WHERE customer_id = ?",
                (target_code, target_name, customer_id))
            conn.commit()
    conn.close()


def init_preset_levels():
    """如果会员等级表为空，预置 4 个默认等级"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT COUNT(*) FROM member_levels")
    if c.fetchone()[0] == 0:
        presets = [
            ("L1", "普通会员", 0, 0, 1.0, "注册即享", 1),
            ("L2", "银卡会员", 500, 2000, 0.95, "积分满500且消费满2000", 2),
            ("L3", "金卡会员", 2000, 8000, 0.9, "积分满2000且消费满8000", 3),
            ("L4", "钻石会员", 5000, 20000, 0.85, "积分满5000且消费满20000", 4),
        ]
        for lvl in presets:
            c.execute("""
                INSERT INTO member_levels
                (level_code, level_name, min_points, min_consumption, discount_rate,
                 description, sort_order, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (lvl[0], lvl[1], lvl[2], lvl[3], lvl[4], lvl[5], lvl[6], _now()))
        conn.commit()
    conn.close()


# ---------- 2. 客户管理 ----------

@st.cache_data(ttl=60, show_spinner=False)
def get_all_customers():
    """查询所有客户，返回 DataFrame"""
    conn = sqlite3.connect(DB_PATH)
    df = pd.read_sql_query("SELECT * FROM customers ORDER BY customer_id", conn)
    conn.close()
    return df


def get_customer_by_id(customer_id):
    """查询单个客户，返回字典或 None"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT * FROM customers WHERE customer_id = ?", (customer_id,))
    row = c.fetchone()
    result = None
    if row:
        cols = [d[0] for d in c.description]
        result = dict(zip(cols, row))
    conn.close()
    return result


def add_customer(customer_id, customer_name, phone, email, gender, birthday,
                 province, city, register_date, remark):
    """新增或更新一个客户（已存在时仅更新档案字段，保留消费统计）"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT customer_id FROM customers WHERE customer_id = ?", (customer_id,))
    exists = c.fetchone() is not None
    if exists:
        c.execute("""
            UPDATE customers SET customer_name=?, phone=?, email=?, gender=?,
                birthday=?, province=?, city=?, register_date=?, remark=?
            WHERE customer_id=?
        """, (customer_name, phone, email, gender, birthday, province, city,
              register_date, remark, customer_id))
    else:
        c.execute("""
            INSERT INTO customers
            (customer_id, customer_name, phone, email, gender, birthday, province,
             city, register_date, level_code, level_name, total_points, total_consumption,
             total_orders, avg_order_value, last_order_date, max_order_value, status,
             remark, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'L1', '普通会员', 0, 0, 0, 0, NULL, 0, '正常', ?, ?)
        """, (customer_id, customer_name, phone, email, gender, birthday, province,
              city, register_date, remark, _now()))
    conn.commit()
    conn.close()


def update_customer_stats(customer_id):
    """更新客户的消费统计（总额、订单数、客单价、最大订单、最后消费日期）"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("""
        SELECT COUNT(*), COALESCE(SUM(order_amount), 0), COALESCE(MAX(order_amount), 0),
               MAX(order_date)
        FROM customer_consumption_log WHERE customer_id = ?
    """, (customer_id,))
    row = c.fetchone()
    total_orders = row[0] or 0
    total_consumption = row[1] or 0
    max_order_value = row[2] or 0
    last_order_date = row[3]
    avg_order_value = total_consumption / total_orders if total_orders > 0 else 0
    c.execute("""
        UPDATE customers SET total_consumption=?, total_orders=?, avg_order_value=?,
            max_order_value=?, last_order_date=?
        WHERE customer_id=?
    """, (total_consumption, total_orders, avg_order_value, max_order_value,
          last_order_date, customer_id))
    conn.commit()
    conn.close()


def delete_customer(customer_id):
    """级联删除客户及其所有关联数据（平台账号、消费记录、积分日志、标签关联、黑名单）"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("DELETE FROM customer_platform_accounts WHERE customer_id = ?", (customer_id,))
    c.execute("DELETE FROM customer_consumption_log WHERE customer_id = ?", (customer_id,))
    c.execute("DELETE FROM customer_points_log WHERE customer_id = ?", (customer_id,))
    c.execute("DELETE FROM customer_tag_links WHERE customer_id = ?", (customer_id,))
    c.execute("DELETE FROM customer_blacklist WHERE customer_id = ?", (customer_id,))
    c.execute("DELETE FROM customers WHERE customer_id = ?", (customer_id,))
    conn.commit()
    conn.close()


def add_consumption(customer_id, order_date, platform, order_id, order_amount,
                    product_summary, points_rate=1):
    """记录一笔消费，自动计算积分、更新统计、记录积分日志、检查等级升级。
    返回本次获得的积分。
    """
    points_earned = int(float(order_amount) * float(points_rate))

    # 1. 记录消费
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("""
        INSERT INTO customer_consumption_log
        (customer_id, order_date, platform, order_id, order_amount,
         product_summary, points_earned, points_used, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, 0, ?)
    """, (customer_id, order_date, platform, order_id, float(order_amount),
          product_summary, points_earned, _now()))
    conn.commit()
    conn.close()

    # 2. 更新客户统计
    update_customer_stats(customer_id)

    # 3. 记录积分变动日志并更新 total_points
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT total_points FROM customers WHERE customer_id = ?", (customer_id,))
    row = c.fetchone()
    current_points = (row[0] if row else 0) or 0
    balance_after = current_points + points_earned
    c.execute("""
        INSERT INTO customer_points_log
        (customer_id, change_type, points_change, balance_after, description,
         related_order, created_at)
        VALUES (?, '消费获得', ?, ?, ?, ?, ?)
    """, (customer_id, points_earned, balance_after,
          f"消费 {order_amount} 元获得积分", order_id, _now()))
    c.execute("UPDATE customers SET total_points = ? WHERE customer_id = ?",
              (balance_after, customer_id))
    conn.commit()
    conn.close()

    # 4. 检查等级升级
    check_and_update_level(customer_id)
    return points_earned


def get_consumption_log(customer_id):
    """获取客户的消费记录，返回 DataFrame"""
    conn = sqlite3.connect(DB_PATH)
    df = pd.read_sql_query(
        "SELECT * FROM customer_consumption_log WHERE customer_id = ? ORDER BY id DESC",
        conn, params=(customer_id,))
    conn.close()
    return df


# ---------- 3. 平台账号聚合 ----------

def get_customer_platform_accounts(customer_id):
    """获取客户绑定的平台账号，返回 DataFrame"""
    conn = sqlite3.connect(DB_PATH)
    df = pd.read_sql_query(
        "SELECT * FROM customer_platform_accounts WHERE customer_id = ? ORDER BY id",
        conn, params=(customer_id,))
    conn.close()
    return df


def add_platform_account(customer_id, platform, platform_user_id, platform_nick, shop_id):
    """新增一个平台账号绑定"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("""
        INSERT INTO customer_platform_accounts
        (customer_id, platform, platform_user_id, platform_nick, shop_id, bind_date, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    """, (customer_id, platform, platform_user_id, platform_nick, shop_id,
          datetime.now().strftime("%Y-%m-%d"), _now()))
    conn.commit()
    conn.close()


def delete_platform_account(account_id):
    """删除一个平台账号绑定"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("DELETE FROM customer_platform_accounts WHERE id = ?", (account_id,))
    conn.commit()
    conn.close()


# ---------- 4. 积分管理 ----------

def get_points_log(customer_id):
    """获取客户积分变动日志，返回 DataFrame"""
    conn = sqlite3.connect(DB_PATH)
    df = pd.read_sql_query(
        "SELECT * FROM customer_points_log WHERE customer_id = ? ORDER BY id DESC",
        conn, params=(customer_id,))
    conn.close()
    return df


def adjust_points(customer_id, points_change, change_type, description, related_order):
    """调整客户积分（正数增加，负数扣减），记录日志、更新 total_points、检查等级升级。"""
    points_change = int(points_change)
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT total_points FROM customers WHERE customer_id = ?", (customer_id,))
    row = c.fetchone()
    current_points = (row[0] if row else 0) or 0
    balance_after = current_points + points_change
    c.execute("""
        INSERT INTO customer_points_log
        (customer_id, change_type, points_change, balance_after, description,
         related_order, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    """, (customer_id, change_type, points_change, balance_after, description,
          related_order, _now()))
    c.execute("UPDATE customers SET total_points = ? WHERE customer_id = ?",
              (balance_after, customer_id))
    conn.commit()
    conn.close()
    check_and_update_level(customer_id)


# ---------- 5. 客户标签管理 ----------

def get_all_customer_tags():
    """查询所有客户标签，返回 DataFrame"""
    conn = sqlite3.connect(DB_PATH)
    df = pd.read_sql_query("SELECT * FROM customer_tags ORDER BY tag_id", conn)
    conn.close()
    return df


def add_customer_tag(tag_name, tag_type, tag_color, auto_rule):
    """新增一个客户标签"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("""
        INSERT OR REPLACE INTO customer_tags
        (tag_name, tag_type, tag_color, auto_rule, created_at)
        VALUES (?, ?, ?, ?, ?)
    """, (tag_name, tag_type, tag_color, auto_rule, _now()))
    conn.commit()
    conn.close()


def get_customer_tags(customer_id):
    """联表查询客户已绑定的标签，返回 DataFrame"""
    conn = sqlite3.connect(DB_PATH)
    df = pd.read_sql_query("""
        SELECT t.tag_id, t.tag_name, t.tag_type, t.tag_color, t.auto_rule,
               l.is_auto, l.created_at AS bind_time
        FROM customer_tag_links l
        JOIN customer_tags t ON l.tag_id = t.tag_id
        WHERE l.customer_id = ?
        ORDER BY l.id
    """, conn, params=(customer_id,))
    conn.close()
    return df


def bind_customer_tag(customer_id, tag_id, is_auto=0):
    """为客户绑定一个标签（先清理同一对绑定避免重复）"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute(
        "DELETE FROM customer_tag_links WHERE customer_id = ? AND tag_id = ?",
        (customer_id, tag_id))
    c.execute("""
        INSERT INTO customer_tag_links (customer_id, tag_id, is_auto, created_at)
        VALUES (?, ?, ?, ?)
    """, (customer_id, tag_id, is_auto, _now()))
    conn.commit()
    conn.close()


def unbind_customer_tag(customer_id, tag_id):
    """解除客户的某个标签绑定"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute(
        "DELETE FROM customer_tag_links WHERE customer_id = ? AND tag_id = ?",
        (customer_id, tag_id))
    conn.commit()
    conn.close()


# ---------- 6. 黑名单管理 ----------

def get_blacklist():
    """获取黑名单列表，返回 DataFrame"""
    conn = sqlite3.connect(DB_PATH)
    df = pd.read_sql_query(
        "SELECT * FROM customer_blacklist ORDER BY id DESC", conn)
    conn.close()
    return df


def add_to_blacklist(customer_id, customer_name, reason, block_type,
                     start_date, end_date, operator):
    """添加黑名单记录，同时更新客户 status 为'黑名单'"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("""
        INSERT INTO customer_blacklist
        (customer_id, customer_name, reason, block_type, start_date, end_date,
         operator, status, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, '生效中', ?)
    """, (customer_id, customer_name, reason, block_type, start_date, end_date,
          operator, _now()))
    c.execute(
        "UPDATE customers SET status = '黑名单' WHERE customer_id = ?",
        (customer_id,))
    conn.commit()
    conn.close()


def remove_from_blacklist(blacklist_id):
    """移除黑名单记录，并恢复客户 status 为'正常'"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute(
        "SELECT customer_id FROM customer_blacklist WHERE id = ?", (blacklist_id,))
    row = c.fetchone()
    customer_id = row[0] if row else None
    c.execute("DELETE FROM customer_blacklist WHERE id = ?", (blacklist_id,))
    if customer_id:
        c.execute(
            "UPDATE customers SET status = '正常' WHERE customer_id = ?",
            (customer_id,))
    conn.commit()
    conn.close()


# ============================================================
# 第 2.95 部分：CRM 分析辅助函数
# ============================================================

def get_all_platform_accounts_for_analysis():
    """获取所有客户平台账号数据，用于分析平台来源分布"""
    conn = sqlite3.connect(DB_PATH)
    df = pd.read_sql_query("SELECT * FROM customer_platform_accounts ORDER BY platform", conn)
    conn.close()
    return df


def get_all_consumption_for_analysis():
    """获取所有消费记录，用于RFM分析"""
    conn = sqlite3.connect(DB_PATH)
    df = pd.read_sql_query(
        "SELECT * FROM customer_consumption_log ORDER BY order_date DESC",
        conn,
    )
    conn.close()
    return df


# ============================================================
# 第 2.7 部分：OMS 全渠道订单管理函数（模块七）
# ----------------------------------------------------------
# 订单状态流转：
#   待付款 -> 待审核 -> 待发货 -> 待打单 -> 待揽收 -> 已发货 -> 已签收 -> 已完成
#   异常分支：任意状态 -> 已取消 / 已退款
# ============================================================

# OMS 常量定义
OMS_PLATFORMS = ["淘宝", "拼多多", "抖音", "视频号", "京东", "快手"]
OMS_ORDER_TYPES = ["现货", "预售"]
OMS_STATUSES = ["待付款", "待审核", "待发货", "待打单", "待揽收",
                "已发货", "已签收", "已完成", "已取消", "已退款"]
OMS_RISK_LEVELS = ["高", "中", "低"]
OMS_HANDLE_STATUSES = ["待处理", "已处理"]
OMS_PAYMENT_METHODS = ["在线支付", "货到付款", "银行转账", "微信支付", "支付宝"]


# ---------- 1. 订单基础管理 ----------

@st.cache_data(ttl=30, show_spinner=False)
def get_all_orders(status_filter=None, platform_filter=None):
    """查询订单列表，支持按状态/平台筛选，返回 DataFrame"""
    conn = sqlite3.connect(DB_PATH)
    query = "SELECT * FROM orders WHERE 1=1"
    params = []
    if status_filter:
        query += " AND order_status = ?"
        params.append(status_filter)
    if platform_filter:
        query += " AND platform = ?"
        params.append(platform_filter)
    query += " ORDER BY created_at DESC"
    df = pd.read_sql_query(query, conn, params=params)
    conn.close()
    return df


def get_order_by_id(order_id):
    """查询单个订单详情，返回字典或 None"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT * FROM orders WHERE order_id = ?", (order_id,))
    row = c.fetchone()
    result = None
    if row:
        cols = [d[0] for d in c.description]
        result = dict(zip(cols, row))
    conn.close()
    return result


def get_order_items(order_id):
    """获取订单明细，返回 DataFrame"""
    conn = sqlite3.connect(DB_PATH)
    df = pd.read_sql_query(
        "SELECT * FROM order_items WHERE order_id = ? ORDER BY id",
        conn, params=(order_id,))
    conn.close()
    return df


def get_order_status_log(order_id):
    """获取订单状态变更日志，返回 DataFrame"""
    conn = sqlite3.connect(DB_PATH)
    df = pd.read_sql_query(
        "SELECT * FROM order_status_log WHERE order_id = ? ORDER BY id DESC",
        conn, params=(order_id,))
    conn.close()
    return df


def add_order(order_id, platform, shop_id, shop_name, customer_id, customer_name,
              receiver_name, receiver_phone, receiver_province, receiver_city,
              receiver_district, receiver_address, order_type, total_amount,
              discount_amount, shipping_fee, actual_amount, payment_method,
              order_time, seller_remark, buyer_remark, invoice_required,
              invoice_title):
    """新增或覆盖一个订单（初始状态为'待付款'）"""
    now = _now()
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("""
        INSERT OR REPLACE INTO orders
        (order_id, platform, shop_id, shop_name, customer_id, customer_name,
         receiver_name, receiver_phone, receiver_province, receiver_city,
         receiver_district, receiver_address, order_type, order_status,
         total_amount, discount_amount, shipping_fee, actual_amount,
         payment_method, order_time, seller_remark, buyer_remark,
         invoice_required, invoice_title, is_abnormal, created_at, updated_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, '待付款',
                ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 0, ?, ?)
    """, (order_id, platform, shop_id, shop_name, customer_id, customer_name,
          receiver_name, receiver_phone, receiver_province, receiver_city,
          receiver_district, receiver_address, order_type,
          float(total_amount or 0), float(discount_amount or 0),
          float(shipping_fee or 0), float(actual_amount or 0),
          payment_method, order_time, seller_remark, buyer_remark,
          int(bool(invoice_required)), invoice_title, now, now))
    # 记录初始状态日志
    c.execute("""
        INSERT INTO order_status_log
        (order_id, from_status, to_status, operator, remark, created_at)
        VALUES (?, NULL, '待付款', '系统', '订单创建', ?)
    """, (order_id, now))
    conn.commit()
    conn.close()


def add_order_item(order_id, spu_code, spu_name, sku_code, sku_name, spec_attrs,
                   quantity, unit_price, total_price, is_gift):
    """添加一条订单明细"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("""
        INSERT INTO order_items
        (order_id, spu_code, spu_name, sku_code, sku_name, spec_attrs,
         quantity, unit_price, total_price, is_gift)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (order_id, spu_code, spu_name, sku_code, sku_name, spec_attrs,
          int(quantity or 0), float(unit_price or 0),
          float(total_price or 0), int(bool(is_gift))))
    conn.commit()
    conn.close()


def update_order_status(order_id, new_status, operator, remark=""):
    """更新订单状态，记录日志，并自动更新对应时间字段"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT order_status FROM orders WHERE order_id = ?", (order_id,))
    row = c.fetchone()
    if not row:
        conn.close()
        return False
    old_status = row[0]
    now = _now()
    # 自动更新对应时间字段
    if new_status == "待审核":
        c.execute("UPDATE orders SET order_status=?, payment_time=?, updated_at=? WHERE order_id=?",
                  (new_status, now, now, order_id))
    elif new_status == "已发货":
        c.execute("UPDATE orders SET order_status=?, ship_time=?, updated_at=? WHERE order_id=?",
                  (new_status, now, now, order_id))
    elif new_status == "已完成":
        c.execute("UPDATE orders SET order_status=?, complete_time=?, updated_at=? WHERE order_id=?",
                  (new_status, now, now, order_id))
    else:
        c.execute("UPDATE orders SET order_status=?, updated_at=? WHERE order_id=?",
                  (new_status, now, order_id))
    c.execute("""
        INSERT INTO order_status_log
        (order_id, from_status, to_status, operator, remark, created_at)
        VALUES (?, ?, ?, ?, ?, ?)
    """, (order_id, old_status, new_status, operator, remark, now))
    conn.commit()
    conn.close()
    return True


def delete_order(order_id):
    """级联删除订单及其明细、日志、异常记录"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("DELETE FROM order_items WHERE order_id = ?", (order_id,))
    c.execute("DELETE FROM order_status_log WHERE order_id = ?", (order_id,))
    c.execute("DELETE FROM abnormal_orders WHERE order_id = ?", (order_id,))
    c.execute("DELETE FROM orders WHERE order_id = ?", (order_id,))
    conn.commit()
    conn.close()


# ---------- 2. 智能审核 ----------

def auto_review_order(order_id):
    """自动审核订单，通过则状态变为'待发货'，否则标记异常"""
    order = get_order_by_id(order_id)
    if not order:
        return {"passed": False, "reason": "订单不存在"}
    if order.get("order_status") != "待审核":
        return {"passed": False, "reason": f"订单不在待审核状态（当前：{order.get('order_status')}）"}

    reasons = []
    # 收货地址完整性
    addr_fields = ["receiver_name", "receiver_phone", "receiver_province",
                   "receiver_city", "receiver_district", "receiver_address"]
    for f in addr_fields:
        if not order.get(f):
            reasons.append("收货地址不完整")
            break
    # 金额校验
    if float(order.get("actual_amount") or 0) <= 0:
        reasons.append("实付金额必须大于0")
    # 明细非空
    items = get_order_items(order_id)
    if items.empty:
        reasons.append("订单明细为空")
    # 黑名单校验
    bl_df = get_blacklist()
    if not bl_df.empty and order.get("customer_id"):
        active_bl = bl_df[(bl_df["customer_id"] == order.get("customer_id")) &
                          (bl_df["status"] == "生效中")]
        if not active_bl.empty:
            reasons.append("客户在黑名单中")

    if reasons:
        reason_str = "；".join(reasons)
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        now = _now()
        c.execute("UPDATE orders SET is_abnormal=1, abnormal_reason=?, updated_at=? WHERE order_id=?",
                  (reason_str, now, order_id))
        c.execute("""
            INSERT INTO abnormal_orders
            (order_id, customer_name, platform, intercept_reason,
             intercept_type, risk_level, handle_status, created_at)
            VALUES (?, ?, ?, ?, '自动', '中', '待处理', ?)
        """, (order_id, order.get("customer_name"), order.get("platform"),
              reason_str, now))
        conn.commit()
        conn.close()
        return {"passed": False, "reason": reason_str}
    else:
        update_order_status(order_id, "待发货", "系统自动审核", "自动审核通过")
        return {"passed": True, "reason": "审核通过"}


def batch_auto_review():
    """批量审核所有'待审核'订单，返回汇总结果"""
    df = get_all_orders(status_filter="待审核")
    results = {"total": len(df), "passed": 0, "failed": 0, "details": []}
    for _, row in df.iterrows():
        r = auto_review_order(row["order_id"])
        if r["passed"]:
            results["passed"] += 1
        else:
            results["failed"] += 1
        results["details"].append({
            "order_id": row["order_id"],
            "passed": r["passed"],
            "reason": r["reason"],
        })
    return results


# ---------- 3. 订单合并 / 拆分 ----------

def merge_orders(order_ids, operator, reason):
    """将多个订单合并为一个新订单，明细合并，原订单标记 parent_order_id"""
    if not order_ids or len(order_ids) < 2:
        return None
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    placeholders = ",".join("?" for _ in order_ids)
    c.execute(f"SELECT * FROM orders WHERE order_id IN ({placeholders})", order_ids)
    rows = c.fetchall()
    cols = [d[0] for d in c.description]
    orders = [dict(zip(cols, r)) for r in rows]
    if not orders:
        conn.close()
        return None

    base = orders[0]
    now = _now()
    new_order_id = "MR" + datetime.now().strftime("%Y%m%d%H%M%S%f")
    total_amount = sum(float(o.get("total_amount") or 0) for o in orders)
    discount_amount = sum(float(o.get("discount_amount") or 0) for o in orders)
    shipping_fee = float(base.get("shipping_fee") or 0)
    actual_amount = total_amount - discount_amount + shipping_fee

    c.execute("""
        INSERT INTO orders
        (order_id, platform, shop_id, shop_name, customer_id, customer_name,
         receiver_name, receiver_phone, receiver_province, receiver_city,
         receiver_district, receiver_address, order_type, order_status,
         total_amount, discount_amount, shipping_fee, actual_amount,
         payment_method, order_time, seller_remark, buyer_remark,
         created_at, updated_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, '待发货',
                ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (new_order_id, base.get("platform"), base.get("shop_id"),
          base.get("shop_name"), base.get("customer_id"), base.get("customer_name"),
          base.get("receiver_name"), base.get("receiver_phone"),
          base.get("receiver_province"), base.get("receiver_city"),
          base.get("receiver_district"), base.get("receiver_address"),
          base.get("order_type") or "现货", total_amount, discount_amount,
          shipping_fee, actual_amount, base.get("payment_method"), now,
          f"合并自：{','.join(order_ids)}", "", now, now))

    # 合并明细
    c.execute(f"SELECT * FROM order_items WHERE order_id IN ({placeholders})",
              order_ids)
    item_rows = c.fetchall()
    item_cols = [d[0] for d in c.description]
    for ir in item_rows:
        it = dict(zip(item_cols, ir))
        c.execute("""
            INSERT INTO order_items
            (order_id, spu_code, spu_name, sku_code, sku_name, spec_attrs,
             quantity, unit_price, total_price, is_gift)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (new_order_id, it.get("spu_code"), it.get("spu_name"),
              it.get("sku_code"), it.get("sku_name"), it.get("spec_attrs"),
              it.get("quantity"), it.get("unit_price"),
              it.get("total_price"), it.get("is_gift")))

    # 标记原订单
    for oid in order_ids:
        c.execute("UPDATE orders SET parent_order_id=?, updated_at=? WHERE order_id=?",
                  (new_order_id, now, oid))

    c.execute("""
        INSERT INTO order_operations
        (operation_type, source_order_ids, target_order_id, reason, operator, created_at)
        VALUES ('合并', ?, ?, ?, ?, ?)
    """, (",".join(order_ids), new_order_id, reason, operator, now))
    conn.commit()
    conn.close()
    return new_order_id


def split_order(order_id, split_items, operator, reason):
    """将一个订单的部分明细拆分为新订单。split_items: [{sku_code, quantity}]"""
    order = get_order_by_id(order_id)
    if not order:
        return None
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    now = _now()
    new_order_id = order_id + "-S" + datetime.now().strftime("%H%M%S")

    c.execute("""
        INSERT INTO orders
        (order_id, platform, shop_id, shop_name, customer_id, customer_name,
         receiver_name, receiver_phone, receiver_province, receiver_city,
         receiver_district, receiver_address, order_type, order_status,
         payment_method, order_time, split_from, seller_remark, buyer_remark,
         created_at, updated_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, '待发货', ?, ?, ?, ?, ?, ?, ?)
    """, (new_order_id, order.get("platform"), order.get("shop_id"),
          order.get("shop_name"), order.get("customer_id"), order.get("customer_name"),
          order.get("receiver_name"), order.get("receiver_phone"),
          order.get("receiver_province"), order.get("receiver_city"),
          order.get("receiver_district"), order.get("receiver_address"),
          order.get("order_type") or "现货", order.get("payment_method"),
          order.get("order_time"), order_id, f"拆分自：{order_id}", "", now, now))

    split_total = 0.0
    for si in split_items:
        sku_code = si.get("sku_code")
        qty = int(si.get("quantity") or 0)
        if not sku_code or qty <= 0:
            continue
        c.execute("SELECT * FROM order_items WHERE order_id=? AND sku_code=?",
                  (order_id, sku_code))
        row = c.fetchone()
        if not row:
            continue
        cols = [d[0] for d in c.description]
        it = dict(zip(cols, row))
        orig_qty = int(it.get("quantity") or 0)
        unit_price = float(it.get("unit_price") or 0)
        if qty > orig_qty:
            qty = orig_qty
        new_qty = orig_qty - qty
        if new_qty <= 0:
            c.execute("DELETE FROM order_items WHERE id=?", (it.get("id"),))
        else:
            c.execute("UPDATE order_items SET quantity=?, total_price=? WHERE id=?",
                      (new_qty, new_qty * unit_price, it.get("id")))
        c.execute("""
            INSERT INTO order_items
            (order_id, spu_code, spu_name, sku_code, sku_name, spec_attrs,
             quantity, unit_price, total_price, is_gift)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (new_order_id, it.get("spu_code"), it.get("spu_name"),
              it.get("sku_code"), it.get("sku_name"), it.get("spec_attrs"),
              qty, unit_price, qty * unit_price, it.get("is_gift")))
        split_total += qty * unit_price

    # 更新新订单金额
    c.execute("UPDATE orders SET total_amount=?, discount_amount=0, shipping_fee=0, actual_amount=? WHERE order_id=?",
              (split_total, split_total, new_order_id))
    # 减少原订单金额
    c.execute("SELECT total_amount, actual_amount FROM orders WHERE order_id=?", (order_id,))
    orig_row = c.fetchone()
    orig_total = float(orig_row[0] or 0) if orig_row else 0
    orig_actual = float(orig_row[1] or 0) if orig_row else 0
    c.execute("UPDATE orders SET total_amount=?, actual_amount=?, updated_at=? WHERE order_id=?",
              (orig_total - split_total, orig_actual - split_total, now, order_id))

    c.execute("""
        INSERT INTO order_operations
        (operation_type, source_order_ids, target_order_id, reason, operator, created_at)
        VALUES ('拆分', ?, ?, ?, ?, ?)
    """, (order_id, new_order_id, reason, operator, now))
    conn.commit()
    conn.close()
    return new_order_id


# ---------- 4. 异常订单拦截 ----------

def check_abnormal(order_id):
    """检查订单异常条件，标记异常并记录到 abnormal_orders"""
    order = get_order_by_id(order_id)
    if not order:
        return {"abnormal": False, "reason": "", "risk": "低"}
    reasons = []
    risk = "低"
    actual = float(order.get("actual_amount") or 0)
    # 金额异常
    if actual <= 0:
        reasons.append("金额异常：实付金额为0")
        risk = "高"
    elif actual > 1000000:
        reasons.append("金额异常：金额过大（超过100万）")
        risk = "高"
    # 地址不全
    if not (order.get("receiver_province") and order.get("receiver_city") and
            order.get("receiver_address")):
        reasons.append("收货地址不完整")
        if risk != "高":
            risk = "中"
    # 黑名单客户
    bl_df = get_blacklist()
    if not bl_df.empty and order.get("customer_id"):
        active_bl = bl_df[(bl_df["customer_id"] == order.get("customer_id")) &
                          (bl_df["status"] == "生效中")]
        if not active_bl.empty:
            reasons.append("黑名单客户")
            risk = "高"
    # 库存不足
    items = get_order_items(order_id)
    conn_chk = sqlite3.connect(DB_PATH)
    c_chk = conn_chk.cursor()
    for _, it in items.iterrows():
        sku_code = it.get("sku_code")
        qty = int(it.get("quantity") or 0)
        if not sku_code or qty <= 0:
            continue
        c_chk.execute("SELECT COUNT(*) FROM inventory_movements WHERE product_code=?",
                      (sku_code,))
        has_records = c_chk.fetchone()[0] > 0
        if has_records:
            stock = get_product_stock(sku_code)
            if stock < qty:
                name = it.get("sku_name") or sku_code
                reasons.append(f"库存不足：{name}（库存{stock}，需{qty}）")
                if risk != "高":
                    risk = "中"
    conn_chk.close()

    if reasons:
        reason_str = "；".join(reasons)
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        now = _now()
        c.execute("UPDATE orders SET is_abnormal=1, abnormal_reason=?, updated_at=? WHERE order_id=?",
                  (reason_str, now, order_id))
        # 避免重复记录
        c.execute("SELECT id FROM abnormal_orders WHERE order_id=? AND handle_status='待处理'",
                  (order_id,))
        if not c.fetchone():
            c.execute("""
                INSERT INTO abnormal_orders
                (order_id, customer_name, platform, intercept_reason,
                 intercept_type, risk_level, handle_status, created_at)
                VALUES (?, ?, ?, ?, '自动', ?, '待处理', ?)
            """, (order_id, order.get("customer_name"), order.get("platform"),
                  reason_str, risk, now))
        conn.commit()
        conn.close()
        return {"abnormal": True, "reason": reason_str, "risk": risk}
    return {"abnormal": False, "reason": "", "risk": "低"}


def get_abnormal_orders():
    """获取异常订单记录列表，返回 DataFrame"""
    conn = sqlite3.connect(DB_PATH)
    df = pd.read_sql_query(
        "SELECT * FROM abnormal_orders ORDER BY id DESC", conn)
    conn.close()
    return df


def handle_abnormal(abnormal_id, handle_result, operator):
    """标记异常订单处理结果，handle_status 改为'已处理'"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    now = _now()
    c.execute("""
        UPDATE abnormal_orders
        SET handle_status='已处理', handle_result=?, operator=?, handled_at=?
        WHERE id=?
    """, (handle_result, operator, now, abnormal_id))
    conn.commit()
    conn.close()


def manual_intercept_order(order_id, reason, risk_level, operator):
    """手动拦截订单：标记异常并记录到 abnormal_orders（手动）"""
    order = get_order_by_id(order_id)
    if not order:
        return
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    now = _now()
    c.execute("UPDATE orders SET is_abnormal=1, abnormal_reason=?, updated_at=? WHERE order_id=?",
              (reason, now, order_id))
    c.execute("SELECT id FROM abnormal_orders WHERE order_id=? AND handle_status='待处理'",
              (order_id,))
    if not c.fetchone():
        c.execute("""
            INSERT INTO abnormal_orders
            (order_id, customer_name, platform, intercept_reason,
             intercept_type, risk_level, handle_status, created_at)
            VALUES (?, ?, ?, ?, '手动', ?, '待处理', ?)
        """, (order_id, order.get("customer_name"), order.get("platform"),
              reason, risk_level, now))
    conn.commit()
    conn.close()


# ---------- 5. 物流同步 ----------

def get_logistics_companies():
    """获取启用的物流公司列表，返回 DataFrame"""
    conn = sqlite3.connect(DB_PATH)
    df = pd.read_sql_query(
        "SELECT * FROM logistics_companies WHERE is_active=1 ORDER BY sort_order",
        conn)
    conn.close()
    return df


def add_logistics_company(company_code, company_name, api_code):
    """新增或更新一个物流公司"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("""
        INSERT OR REPLACE INTO logistics_companies
        (company_code, company_name, api_code, is_active, sort_order)
        VALUES (?, ?, ?, 1, 0)
    """, (company_code, company_name, api_code))
    conn.commit()
    conn.close()


def delete_logistics_company(company_code):
    """删除一个物流公司"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("DELETE FROM logistics_companies WHERE company_code=?", (company_code,))
    conn.commit()
    conn.close()


def ship_order(order_id, logistics_company, logistics_number, operator):
    """填写物流信息，状态从'待揽收'变为'已发货'，记录 ship_time"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT order_status FROM orders WHERE order_id=?", (order_id,))
    row = c.fetchone()
    old_status = row[0] if row else ""
    now = _now()
    c.execute("""
        UPDATE orders
        SET order_status='已发货', logistics_company=?, logistics_number=?,
            ship_time=?, updated_at=?
        WHERE order_id=?
    """, (logistics_company, logistics_number, now, now, order_id))
    c.execute("""
        INSERT INTO order_status_log
        (order_id, from_status, to_status, operator, remark, created_at)
        VALUES (?, ?, '已发货', ?, ?, ?)
    """, (order_id, old_status, operator,
          f"物流公司：{logistics_company}，单号：{logistics_number}", now))
    conn.commit()
    conn.close()


def batch_ship(orders_data, operator):
    """批量发货，orders_data: [{order_id, logistics_company, logistics_number}]"""
    count = 0
    for od in orders_data:
        if od.get("order_id") and od.get("logistics_number"):
            ship_order(od["order_id"], od.get("logistics_company", ""),
                       od.get("logistics_number"), operator)
            count += 1
    return count


# ---------- 6. 统计看板 ----------

@st.cache_data(ttl=30, show_spinner=False)
def get_order_stats():
    """订单统计看板数据，返回 dict"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    # 各状态订单数量
    c.execute("SELECT order_status, COUNT(*) FROM orders GROUP BY order_status")
    status_counts = {r[0]: r[1] for r in c.fetchall()}
    # 各平台订单数量
    c.execute("SELECT platform, COUNT(*) FROM orders GROUP BY platform")
    platform_counts = {r[0]: r[1] for r in c.fetchall()}
    # 今日新增
    today = datetime.now().strftime("%Y-%m-%d")
    c.execute("SELECT COUNT(*) FROM orders WHERE created_at LIKE ?", (today + "%",))
    today_new = c.fetchone()[0]
    # 待处理（待审核+待发货+待打单+待揽收）
    pending_statuses = ("待审核", "待发货", "待打单", "待揽收")
    ph = ",".join("?" for _ in pending_statuses)
    c.execute(f"SELECT COUNT(*) FROM orders WHERE order_status IN ({ph})",
              pending_statuses)
    pending = c.fetchone()[0]
    # 异常数
    c.execute("SELECT COUNT(*) FROM orders WHERE is_abnormal=1")
    abnormal = c.fetchone()[0]
    conn.close()
    return {
        "status_counts": status_counts,
        "platform_counts": platform_counts,
        "today_new": today_new,
        "pending": pending,
        "abnormal": abnormal,
    }


# ---------- 7. 预置数据 ----------

def init_preset_logistics():
    """如果物流公司表为空，插入预置物流公司"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT COUNT(*) FROM logistics_companies")
    if c.fetchone()[0] == 0:
        presets = [
            ("SF", "顺丰速运", "shunfeng", 1),
            ("ZTO", "中通快递", "zhongtong", 2),
            ("YTO", "圆通速递", "yuantong", 3),
            ("STO", "申通快递", "shentong", 4),
            ("YD", "韵达快递", "yunda", 5),
            ("JD", "京东物流", "jd", 6),
            ("EMS", "EMS", "ems", 7),
            ("DBL", "德邦快递", "debang", 8),
        ]
        for code, name, api_code, sort_order in presets:
            c.execute("""
                INSERT OR REPLACE INTO logistics_companies
                (company_code, company_name, api_code, is_active, sort_order)
                VALUES (?, ?, ?, 1, ?)
            """, (code, name, api_code, sort_order))
        conn.commit()
    conn.close()


# ---------- 8. 售后单管理 ----------

def get_all_after_sales(status_filter=None):
    """查询售后单列表，支持按状态筛选，返回 DataFrame"""
    conn = sqlite3.connect(DB_PATH)
    query = "SELECT * FROM after_sales_orders WHERE 1=1"
    params = []
    if status_filter:
        query += " AND status = ?"
        params.append(status_filter)
    query += " ORDER BY created_at DESC"
    df = pd.read_sql_query(query, conn, params=params)
    conn.close()
    return df


def add_after_sales(after_sales_id, order_id, platform, shop_name, customer_id,
                    customer_name, type, reason, spu_code, spu_name, sku_code,
                    sku_name, quantity, refund_amount, apply_date):
    """新增一条售后单"""
    now = _now()
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("""
        INSERT INTO after_sales_orders
        (after_sales_id, order_id, platform, shop_name, customer_id, customer_name,
         type, reason, spu_code, spu_name, sku_code, sku_name, quantity,
         refund_amount, status, apply_date, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, '待处理', ?, ?)
    """, (after_sales_id, order_id, platform, shop_name, customer_id, customer_name,
          type, reason, spu_code, spu_name, sku_code, sku_name,
          int(quantity or 1), float(refund_amount or 0), apply_date, now))
    conn.commit()
    conn.close()


def update_after_sales_status(after_sales_id, status, handle_result, operator):
    """更新售后单处理状态"""
    now = _now()
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("""
        UPDATE after_sales_orders
        SET status=?, handle_result=?, operator=?, handle_date=?
        WHERE after_sales_id=?
    """, (status, handle_result, operator, now, after_sales_id))
    conn.commit()
    conn.close()


def delete_after_sales(after_sales_id):
    """删除一条售后单"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("DELETE FROM after_sales_orders WHERE after_sales_id=?", (after_sales_id,))
    conn.commit()
    conn.close()


# ============================================================
# 第 2.5 部分：BI 数据分析函数（模块八）
# ----------------------------------------------------------
# 所有函数支持日期范围筛选 date_from, date_to（均为字符串）
# 订单按 order_time 过滤；售后按 apply_date 过滤；
# 库存按 movement_date 过滤；新客按 register_date 过滤。
# ============================================================

def _bi_date_where(column, date_from, date_to):
    """构造日期范围 WHERE 片段与参数列表"""
    clauses = []
    params = []
    if date_from:
        clauses.append(f"date({column}) >= date(?)")
        params.append(date_from)
    if date_to:
        clauses.append(f"date({column}) <= date(?)")
        params.append(date_to)
    return " AND ".join(clauses), params


def bi_overview(date_from=None, date_to=None):
    """经营总览统计，返回 dict"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    where, params = _bi_date_where("order_time", date_from, date_to)
    date_cond = f" AND {where}" if where else ""

    # 总订单数
    c.execute(f"SELECT COUNT(*) FROM orders WHERE 1=1{date_cond}", params)
    total_orders = c.fetchone()[0]

    # 已完成订单数
    c.execute(
        f"SELECT COUNT(*) FROM orders WHERE order_status IN ('已完成','已签收'){date_cond}",
        params)
    completed_orders = c.fetchone()[0]

    # 总营收（已完成/已签收订单的实付金额）
    c.execute(
        f"SELECT COALESCE(SUM(actual_amount),0) FROM orders WHERE order_status IN ('已完成','已签收'){date_cond}",
        params)
    total_revenue = c.fetchone()[0] or 0

    avg_order_value = total_revenue / completed_orders if completed_orders else 0

    # 退款总额（售后单中已退款/已完成的退款金额）
    a_where, a_params = _bi_date_where("apply_date", date_from, date_to)
    a_cond = f" AND {a_where}" if a_where else ""
    c.execute(
        f"SELECT COALESCE(SUM(refund_amount),0) FROM after_sales_orders WHERE status IN ('已退款','已完成'){a_cond}",
        a_params)
    total_refund = c.fetchone()[0] or 0

    refund_rate = total_refund / total_revenue if total_revenue else 0

    # 客户统计
    c.execute("SELECT COUNT(*) FROM customers")
    total_customers = c.fetchone()[0]
    c_where, c_params = _bi_date_where("register_date", date_from, date_to)
    if c_where:
        c.execute(f"SELECT COUNT(*) FROM customers WHERE {c_where}", c_params)
        new_customers = c.fetchone()[0]
    else:
        new_customers = total_customers

    # 商品销售件数
    c.execute(
        f"""SELECT COALESCE(SUM(oi.quantity),0)
            FROM order_items oi JOIN orders o ON oi.order_id = o.order_id
            WHERE 1=1{date_cond}""",
        params)
    total_products_sold = c.fetchone()[0] or 0

    # 异常订单数
    c.execute(f"SELECT COUNT(*) FROM orders WHERE is_abnormal=1{date_cond}", params)
    abnormal_count = c.fetchone()[0]

    # 售后单数
    c.execute(
        f"SELECT COUNT(*) FROM after_sales_orders WHERE 1=1{a_cond}", a_params)
    after_sales_count = c.fetchone()[0]
    after_sales_rate = after_sales_count / total_orders if total_orders else 0

    conn.close()
    return {
        "total_orders": total_orders,
        "completed_orders": completed_orders,
        "total_revenue": total_revenue,
        "avg_order_value": avg_order_value,
        "total_refund": total_refund,
        "refund_rate": refund_rate,
        "total_customers": total_customers,
        "new_customers": new_customers,
        "total_products_sold": total_products_sold,
        "abnormal_count": abnormal_count,
        "after_sales_count": after_sales_count,
        "after_sales_rate": after_sales_rate,
    }


def bi_shop_analysis(date_from=None, date_to=None):
    """店铺经营分析，返回 DataFrame"""
    conn = sqlite3.connect(DB_PATH)
    where, params = _bi_date_where("order_time", date_from, date_to)
    date_cond = f" AND {where}" if where else ""
    query = f"""
        SELECT
            o.shop_name AS shop_name,
            o.platform AS platform,
            COUNT(*) AS order_count,
            COALESCE(SUM(o.actual_amount),0) AS revenue,
            COALESCE(AVG(o.actual_amount),0) AS avg_order_value
        FROM orders o
        WHERE o.shop_name IS NOT NULL AND o.shop_name != ''{date_cond}
        GROUP BY o.shop_name, o.platform
        ORDER BY revenue DESC
    """
    df = pd.read_sql_query(query, conn, params=params)
    # 退款统计
    a_where, a_params = _bi_date_where("apply_date", date_from, date_to)
    a_cond = f" AND {a_where}" if a_where else ""
    refund_df = pd.read_sql_query(
        f"""SELECT shop_name,
                   COUNT(*) AS refund_count,
                   COALESCE(SUM(refund_amount),0) AS refund_amount
            FROM after_sales_orders
            WHERE shop_name IS NOT NULL AND shop_name != ''{a_cond}
            GROUP BY shop_name""",
        conn, params=a_params)
    # 完成率
    comp_df = pd.read_sql_query(
        f"""SELECT shop_name,
                   SUM(CASE WHEN order_status IN ('已完成','已签收') THEN 1 ELSE 0 END) AS completed,
                   COUNT(*) AS total
            FROM orders
            WHERE shop_name IS NOT NULL AND shop_name != ''{date_cond}
            GROUP BY shop_name""",
        conn, params=params)
    conn.close()

    if not df.empty:
        df = df.merge(refund_df, on="shop_name", how="left")
        df = df.merge(comp_df, on="shop_name", how="left")
        df["refund_count"] = df["refund_count"].fillna(0).astype(int)
        df["refund_amount"] = df["refund_amount"].fillna(0)
        df["completion_rate"] = df.apply(
            lambda r: round(r["completed"] / r["total"], 4) if r["total"] else 0, axis=1)
        df = df[["shop_name", "platform", "order_count", "revenue",
                 "avg_order_value", "refund_count", "refund_amount", "completion_rate"]]
    return df


def bi_product_sales(date_from=None, date_to=None):
    """商品销售分析，返回 DataFrame"""
    conn = sqlite3.connect(DB_PATH)
    where, params = _bi_date_where("o.order_time", date_from, date_to)
    date_cond = f" AND {where}" if where else ""
    query = f"""
        SELECT
            oi.spu_code AS spu_code,
            oi.spu_name AS spu_name,
            COALESCE(SUM(oi.quantity),0) AS total_qty,
            COALESCE(SUM(oi.total_price),0) AS total_revenue,
            COALESCE(AVG(oi.unit_price),0) AS avg_price,
            COUNT(DISTINCT oi.order_id) AS order_count
        FROM order_items oi
        JOIN orders o ON oi.order_id = o.order_id
        WHERE oi.spu_code IS NOT NULL AND oi.spu_code != ''{date_cond}
        GROUP BY oi.spu_code, oi.spu_name
        ORDER BY total_revenue DESC
    """
    df = pd.read_sql_query(query, conn, params=params)
    # 退款件数
    a_where, a_params = _bi_date_where("apply_date", date_from, date_to)
    a_cond = f" AND {a_where}" if a_where else ""
    refund_df = pd.read_sql_query(
        f"""SELECT spu_code, COALESCE(SUM(quantity),0) AS refund_qty
            FROM after_sales_orders
            WHERE spu_code IS NOT NULL AND spu_code != ''{a_cond}
            GROUP BY spu_code""",
        conn, params=a_params)
    conn.close()

    if not df.empty:
        df = df.merge(refund_df, on="spu_code", how="left")
        df["refund_qty"] = df["refund_qty"].fillna(0).astype(int)
        df = df[["spu_code", "spu_name", "total_qty", "total_revenue",
                 "avg_price", "order_count", "refund_qty"]]
    return df


def bi_inventory_turnover(date_from=None, date_to=None):
    """库存周转分析，返回 DataFrame"""
    conn = sqlite3.connect(DB_PATH)
    where, params = _bi_date_where("m.movement_date", date_from, date_to)
    date_cond = f" AND {where}" if where else ""
    query = f"""
        SELECT
            p.product_code AS product_code,
            p.product_name AS product_name,
            COALESCE(SUM(CASE WHEN m.movement_type='入库' THEN m.quantity ELSE 0 END),0) AS inbound_qty,
            COALESCE(SUM(CASE WHEN m.movement_type='出库' THEN m.quantity ELSE 0 END),0) AS outbound_qty,
            COALESCE(SUM(CASE WHEN m.movement_type='入库' THEN m.total_amount ELSE 0 END),0)
              - COALESCE(SUM(CASE WHEN m.movement_type='出库' THEN m.total_amount ELSE 0 END),0) AS stock_value
        FROM products p
        LEFT JOIN inventory_movements m ON p.product_code = m.product_code
        WHERE 1=1{date_cond}
        GROUP BY p.product_code, p.product_name
        ORDER BY p.product_code
    """
    df = pd.read_sql_query(query, conn, params=params)
    conn.close()

    # 当前库存（全量，不受日期范围影响）
    conn = sqlite3.connect(DB_PATH)
    stock_df = pd.read_sql_query(
        """SELECT product_code,
                  COALESCE(SUM(CASE WHEN movement_type='入库' THEN quantity ELSE 0 END),0)
                    - COALESCE(SUM(CASE WHEN movement_type='出库' THEN quantity ELSE 0 END),0) AS current_stock
           FROM inventory_movements
           GROUP BY product_code""", conn)
    conn.close()

    if not df.empty:
        df = df.merge(stock_df, on="product_code", how="left")
        df["current_stock"] = df["current_stock"].fillna(0)
        df["turnover_rate"] = df.apply(
            lambda r: round(r["outbound_qty"] / r["current_stock"], 4)
            if r["current_stock"] else 0, axis=1)
        df = df[["product_code", "product_name", "inbound_qty", "outbound_qty",
                 "current_stock", "stock_value", "turnover_rate"]]
    return df


def bi_after_sales_analysis(date_from=None, date_to=None):
    """售后分析，返回 dict"""
    conn = sqlite3.connect(DB_PATH)
    where, params = _bi_date_where("apply_date", date_from, date_to)
    date_cond = f" AND {where}" if where else ""

    c = conn.cursor()
    c.execute(f"SELECT COUNT(*) FROM after_sales_orders WHERE 1=1{date_cond}", params)
    total_count = c.fetchone()[0]
    c.execute(f"SELECT COUNT(*) FROM after_sales_orders WHERE type='退货'{date_cond}", params)
    refund_count = c.fetchone()[0]
    c.execute(f"SELECT COUNT(*) FROM after_sales_orders WHERE type='换货'{date_cond}", params)
    exchange_count = c.fetchone()[0]
    c.execute(
        f"SELECT COALESCE(SUM(refund_amount),0) FROM after_sales_orders WHERE 1=1{date_cond}",
        params)
    total_refund_amount = c.fetchone()[0] or 0

    # 退款率（售后单数 / 订单数）
    o_where, o_params = _bi_date_where("order_time", date_from, date_to)
    o_cond = f" AND {o_where}" if o_where else ""
    c.execute(f"SELECT COUNT(*) FROM orders WHERE 1=1{o_cond}", o_params)
    order_count = c.fetchone()[0]
    refund_rate = total_count / order_count if order_count else 0

    # 售后原因 TOP
    top_reasons = pd.read_sql_query(
        f"""SELECT reason, COUNT(*) AS count
            FROM after_sales_orders
            WHERE reason IS NOT NULL AND reason != ''{date_cond}
            GROUP BY reason ORDER BY count DESC LIMIT 10""",
        conn, params=params)

    # 平台分布
    platform_distribution = pd.read_sql_query(
        f"""SELECT platform, COUNT(*) AS count
            FROM after_sales_orders
            WHERE platform IS NOT NULL AND platform != ''{date_cond}
            GROUP BY platform ORDER BY count DESC""",
        conn, params=params)
    conn.close()

    return {
        "total_count": total_count,
        "refund_count": refund_count,
        "exchange_count": exchange_count,
        "total_refund_amount": total_refund_amount,
        "refund_rate": refund_rate,
        "top_reasons": top_reasons,
        "platform_distribution": platform_distribution,
    }


def bi_platform_distribution(date_from=None, date_to=None):
    """平台销售分布，返回 DataFrame"""
    conn = sqlite3.connect(DB_PATH)
    where, params = _bi_date_where("order_time", date_from, date_to)
    date_cond = f" AND {where}" if where else ""
    df = pd.read_sql_query(
        f"""SELECT platform,
                   COUNT(*) AS order_count,
                   COALESCE(SUM(actual_amount),0) AS revenue
            FROM orders
            WHERE platform IS NOT NULL AND platform != ''{date_cond}
            GROUP BY platform ORDER BY revenue DESC""",
        conn, params=params)
    conn.close()
    if not df.empty:
        total_rev = df["revenue"].sum()
        df["percentage"] = df["revenue"].apply(
            lambda x: round(x / total_rev, 4) if total_rev else 0)
        df = df[["platform", "order_count", "revenue", "percentage"]]
    return df


def bi_daily_trend(date_from=None, date_to=None):
    """每日销售趋势，返回 DataFrame"""
    conn = sqlite3.connect(DB_PATH)
    where, params = _bi_date_where("o.order_time", date_from, date_to)
    date_cond = f" AND {where}" if where else ""
    df = pd.read_sql_query(
        f"""SELECT substr(o.order_time,1,10) AS date,
                   COUNT(*) AS order_count,
                   COALESCE(SUM(o.actual_amount),0) AS revenue
            FROM orders o
            WHERE o.order_time IS NOT NULL{date_cond}
            GROUP BY date ORDER BY date""",
        conn, params=params)
    # 每日退款
    a_where, a_params = _bi_date_where("apply_date", date_from, date_to)
    a_cond = f" AND {a_where}" if a_where else ""
    refund_df = pd.read_sql_query(
        f"""SELECT substr(apply_date,1,10) AS date,
                   COALESCE(SUM(refund_amount),0) AS refund_amount
            FROM after_sales_orders
            WHERE apply_date IS NOT NULL{a_cond}
            GROUP BY date""",
        conn, params=a_params)
    conn.close()
    if not df.empty:
        df = df.merge(refund_df, on="date", how="left")
        df["refund_amount"] = df["refund_amount"].fillna(0)
        df = df[["date", "order_count", "revenue", "refund_amount"]]
    return df


def bi_hourly_distribution(date_from=None, date_to=None):
    """下单时段分布，返回 DataFrame"""
    conn = sqlite3.connect(DB_PATH)
    where, params = _bi_date_where("order_time", date_from, date_to)
    date_cond = f" AND {where}" if where else ""
    df = pd.read_sql_query(
        f"""SELECT CAST(substr(order_time,12,2) AS INTEGER) AS hour,
                   COUNT(*) AS order_count,
                   COALESCE(SUM(actual_amount),0) AS revenue
            FROM orders
            WHERE order_time IS NOT NULL AND length(order_time) >= 13{date_cond}
            GROUP BY hour ORDER BY hour""",
        conn, params=params)
    conn.close()
    # 补全 0-23 小时
    full_hours = pd.DataFrame({"hour": list(range(24))})
    if not df.empty:
        df["hour"] = df["hour"].astype(int)
        df = full_hours.merge(df, on="hour", how="left")
        df["order_count"] = df["order_count"].fillna(0).astype(int)
        df["revenue"] = df["revenue"].fillna(0)
    else:
        df = full_hours
        df["order_count"] = 0
        df["revenue"] = 0.0
    return df


# ============================================================
# 第 3 部分：核心会计计算函数
# ============================================================

@st.cache_data(ttl=30, show_spinner=False)
def calc_account_balance(account_code):
    """
    计算某个科目的期末余额
    返回：(期初借方, 期初贷方, 本期借方发生额, 本期贷方发生额, 期末余额)

    计算规则：
      借方科目（资产/成本）：期末 = 期初借方 + 本期借方 - 本期贷方
      贷方科目（负债/权益）：期末 = 期初贷方 + 本期贷方 - 本期借方

    如果是一级科目，自动汇总其名下所有二级科目的余额。
    """
    # 先判断是系统科目还是自定义二级科目
    account = ACCOUNT_MAP.get(account_code)
    if account:
        direction = account["direction"]
        is_primary = True
    else:
        # 查自定义二级科目
        info = get_account_info_dynamic(account_code)
        if info:
            direction = info["direction"]
            is_primary = False
        else:
            return 0, 0, 0, 0, 0

    # 本科目自身的期初余额
    opening_debit, opening_credit = get_opening_balance(account_code)

    # 本科目自身的本期发生额
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("""
        SELECT COALESCE(SUM(debit_amount), 0), COALESCE(SUM(credit_amount), 0)
        FROM vouchers WHERE account_code = ?
    """, (account_code,))
    period_debit, period_credit = c.fetchone()

    # 如果是一级科目，汇总所有二级科目的余额
    if is_primary:
        sub_codes = get_sub_account_codes(account_code)
        for sub_code in sub_codes:
            sub_od, sub_oc = get_opening_balance(sub_code)
            opening_debit += sub_od
            opening_credit += sub_oc
            c.execute("""
                SELECT COALESCE(SUM(debit_amount), 0), COALESCE(SUM(credit_amount), 0)
                FROM vouchers WHERE account_code = ?
            """, (sub_code,))
            sub_pd, sub_pc = c.fetchone()
            period_debit += sub_pd or 0
            period_credit += sub_pc or 0

    conn.close()

    period_debit = period_debit or 0
    period_credit = period_credit or 0

    # 计算期末余额
    if direction == "借":
        ending_balance = opening_debit + period_debit - period_credit
    else:
        ending_balance = opening_credit + period_credit - period_debit

    return opening_debit, opening_credit, period_debit, period_credit, ending_balance


# ============================================================
# 期末自动结转功能
# ----------------------------------------------------------
# 月末：将所有损益类科目余额结转至"本年利润"（4103）
#   收入类（贷方余额）→ 借记收入科目，贷记 4103
#   费用类（借方余额）→ 借记 4103，贷记费用科目
# 年末：再将 4103 余额转入"利润分配-未分配利润"（410403）
#   并可选提取法定盈余公积（净利润的 10%）
# ============================================================

# 需要结转的损益类科目（按类别分组）
_INCOME_ACCOUNT_CODES = [
    "6001", "6021", "6041", "6051",       # 主营/其他业务收入、租赁收入
    "6101", "6102", "6103",               # 公允价值变动损益等
    "6111", "6115", "6117",               # 投资收益、资产处置损益、其他收益
    "6301",                                # 营业外收入
]

_EXPENSE_ACCOUNT_CODES = [
    "6401", "6402", "6403",               # 主营/其他业务成本、税金及附加
    "6405",                                # 研发费用
    "6601", "6602", "6603",               # 销售/管理/财务费用
    "6604", "6605", "6606",               # 勘探/租赁费用、汇兑损益
    "6641", "6642",                        # 信用/资产减值损失
    "6701",                                # 营业外支出
    "6711",                                # 所得税费用
]

# 本年利润科目
_PROFIT_ACCOUNT_CODE = "4103"
# 利润分配-未分配利润科目编码（自定义二级科目）
_UNDISTRIBUTED_PROFIT_CODE = "410403"
# 盈余公积-法定盈余公积
_LEGAL_RESERVE_CODE = "410101"
# 利润分配-提取法定盈余公积
_EXTRACT_LEGAL_RESERVE_CODE = "410401"


def get_carryforward_preview():
    """
    计算期末结转预览数据。
    返回 dict:
      - income_items: [{code, name, balance, direction}]  收入类科目（贷方余额）
      - expense_items: [{code, name, balance, direction}] 费用类科目（借方余额）
      - total_income: 收入合计
      - total_expense: 费用合计
      - net_profit: 净利润（收入 - 费用）
      - profit_balance: 本年利润科目当前余额
    """
    # 清除缓存以确保数据最新
    calc_account_balance.clear()
    get_all_opening_balances.clear()
    get_all_vouchers.clear()

    income_items = []
    expense_items = []

    # 收入类科目（正常余额在贷方）
    for code in _INCOME_ACCOUNT_CODES:
        acc = ACCOUNT_MAP.get(code)
        if not acc:
            continue
        od, oc, pd, pc, ending = calc_account_balance(code)
        # 收入类科目贷方余额 = 期初贷方 + 本期贷方 - 本期借方
        balance = ending
        if abs(balance) > 0.01:
            income_items.append({
                "code": code,
                "name": acc["name"],
                "balance": balance,
                "direction": "贷" if balance > 0 else "借",
            })

    # 费用类科目（正常余额在借方）
    for code in _EXPENSE_ACCOUNT_CODES:
        acc = ACCOUNT_MAP.get(code)
        if not acc:
            continue
        od, oc, pd, pc, ending = calc_account_balance(code)
        balance = ending
        if abs(balance) > 0.01:
            expense_items.append({
                "code": code,
                "name": acc["name"],
                "balance": balance,
                "direction": "借" if balance > 0 else "贷",
            })

    total_income = sum(item["balance"] for item in income_items if item["balance"] > 0)
    total_expense = sum(item["balance"] for item in expense_items if item["balance"] > 0)
    # 负余额的收入类科目视为费用
    total_income -= sum(abs(item["balance"]) for item in income_items if item["balance"] < 0)
    # 负余额的费用类科目视为收入
    total_expense -= sum(abs(item["balance"]) for item in expense_items if item["balance"] < 0)

    # 本年利润当前余额
    _, _, _, _, profit_balance = calc_account_balance(_PROFIT_ACCOUNT_CODE)

    return {
        "income_items": income_items,
        "expense_items": expense_items,
        "total_income": total_income,
        "total_expense": total_expense,
        "net_profit": total_income - total_expense,
        "profit_balance": profit_balance,
    }


def execute_monthly_carryforward(carryforward_date):
    """
    执行月末损益结转：将所有损益类科目余额结转至"本年利润"。
    生成两张凭证：
      1. 结转收入：借各收入科目，贷 4103
      2. 结转费用：借 4103，贷各费用科目
    返回：生成的凭证编号列表
    """
    preview = get_carryforward_preview()
    date_str = carryforward_date.strftime("%Y-%m-%d")
    voucher_numbers = []

    # —— 凭证 1：结转本期收入 ——
    income_lines = []
    for item in preview["income_items"]:
        balance = item["balance"]
        if balance > 0.01:
            # 贷方余额 → 借记结转
            income_lines.append({
                "account_code": item["code"],
                "account_name": item["name"],
                "debit": balance,
                "credit": 0,
            })
        elif balance < -0.01:
            # 借方余额（净损失）→ 贷记结转
            income_lines.append({
                "account_code": item["code"],
                "account_name": item["name"],
                "debit": 0,
                "credit": abs(balance),
            })

    total_income_credit = sum(l["debit"] for l in income_lines)
    total_income_debit = sum(l["credit"] for l in income_lines)
    income_net = total_income_credit - total_income_debit

    if income_lines:
        # 添加本年利润抵消行
        if income_net > 0:
            income_lines.append({
                "account_code": _PROFIT_ACCOUNT_CODE,
                "account_name": ACCOUNT_MAP[_PROFIT_ACCOUNT_CODE]["name"],
                "debit": 0,
                "credit": income_net,
            })
        elif income_net < 0:
            income_lines.append({
                "account_code": _PROFIT_ACCOUNT_CODE,
                "account_name": ACCOUNT_MAP[_PROFIT_ACCOUNT_CODE]["name"],
                "debit": abs(income_net),
                "credit": 0,
            })

        v_num = get_next_voucher_number()
        save_voucher(v_num, date_str, "月末结转：结转本期收入至本年利润", income_lines)
        voucher_numbers.append(v_num)

    # —— 凭证 2：结转本期费用 ——
    expense_lines = []
    for item in preview["expense_items"]:
        balance = item["balance"]
        if balance > 0.01:
            # 借方余额 → 贷记结转
            expense_lines.append({
                "account_code": item["code"],
                "account_name": item["name"],
                "debit": 0,
                "credit": balance,
            })
        elif balance < -0.01:
            # 贷方余额 → 借记结转
            expense_lines.append({
                "account_code": item["code"],
                "account_name": item["name"],
                "debit": abs(balance),
                "credit": 0,
            })

    total_expense_debit = sum(l["debit"] for l in expense_lines)
    total_expense_credit = sum(l["credit"] for l in expense_lines)
    expense_net = total_expense_credit - total_expense_debit

    if expense_lines:
        if expense_net > 0:
            expense_lines.insert(0, {
                "account_code": _PROFIT_ACCOUNT_CODE,
                "account_name": ACCOUNT_MAP[_PROFIT_ACCOUNT_CODE]["name"],
                "debit": expense_net,
                "credit": 0,
            })
        elif expense_net < 0:
            expense_lines.insert(0, {
                "account_code": _PROFIT_ACCOUNT_CODE,
                "account_name": ACCOUNT_MAP[_PROFIT_ACCOUNT_CODE]["name"],
                "debit": 0,
                "credit": abs(expense_net),
            })

        v_num = get_next_voucher_number()
        save_voucher(v_num, date_str, "月末结转：结转本期费用至本年利润", expense_lines)
        voucher_numbers.append(v_num)

    # 清除缓存
    calc_account_balance.clear()
    get_all_vouchers.clear()
    get_all_opening_balances.clear()

    return voucher_numbers


def execute_yearly_carryforward(carryforward_date, extract_legal_reserve=True, reserve_rate=0.10):
    """
    执行年末结转：
      1. 先执行月末损益结转（确保所有损益科目清零）
      2. 将"本年利润"余额转入"利润分配-未分配利润"
      3. （可选）提取法定盈余公积
    返回：{voucher_numbers, net_profit, reserve_amount}
    """
    date_str = carryforward_date.strftime("%Y-%m-%d")
    all_voucher_numbers = []

    # 步骤 1：先执行月末损益结转
    monthly_vouchers = execute_monthly_carryforward(carryforward_date)
    all_voucher_numbers.extend(monthly_vouchers)

    # 步骤 2：本年利润 → 利润分配-未分配利润
    _, _, _, _, profit_balance = calc_account_balance(_PROFIT_ACCOUNT_CODE)
    voucher_numbers = []

    if abs(profit_balance) > 0.01:
        lines = []
        if profit_balance > 0:
            # 净利润：借本年利润，贷未分配利润
            lines.append({
                "account_code": _PROFIT_ACCOUNT_CODE,
                "account_name": ACCOUNT_MAP[_PROFIT_ACCOUNT_CODE]["name"],
                "debit": profit_balance,
                "credit": 0,
            })
            lines.append({
                "account_code": _UNDISTRIBUTED_PROFIT_CODE,
                "account_name": "利润分配-未分配利润",
                "debit": 0,
                "credit": profit_balance,
            })
        else:
            # 净亏损：借未分配利润，贷本年利润
            abs_balance = abs(profit_balance)
            lines.append({
                "account_code": _UNDISTRIBUTED_PROFIT_CODE,
                "account_name": "利润分配-未分配利润",
                "debit": abs_balance,
                "credit": 0,
            })
            lines.append({
                "account_code": _PROFIT_ACCOUNT_CODE,
                "account_name": ACCOUNT_MAP[_PROFIT_ACCOUNT_CODE]["name"],
                "debit": 0,
                "credit": abs_balance,
            })

        v_num = get_next_voucher_number()
        save_voucher(v_num, date_str, "年末结转：本年利润转入利润分配-未分配利润", lines)
        all_voucher_numbers.append(v_num)

    # 步骤 3：提取法定盈余公积
    reserve_amount = 0
    if extract_legal_reserve and profit_balance > 0:
        reserve_amount = round(profit_balance * reserve_rate, 2)
        if reserve_amount > 0.01:
            lines = [
                {
                    "account_code": _EXTRACT_LEGAL_RESERVE_CODE,
                    "account_name": "利润分配-提取法定盈余公积",
                    "debit": reserve_amount,
                    "credit": 0,
                },
                {
                    "account_code": _LEGAL_RESERVE_CODE,
                    "account_name": "盈余公积-法定盈余公积",
                    "debit": 0,
                    "credit": reserve_amount,
                },
            ]
            v_num = get_next_voucher_number()
            save_voucher(v_num, date_str, "年末结转：提取法定盈余公积", lines)
            all_voucher_numbers.append(v_num)

    # 清除缓存
    calc_account_balance.clear()
    get_all_vouchers.clear()
    get_all_opening_balances.clear()

    return {
        "voucher_numbers": all_voucher_numbers,
        "net_profit": profit_balance,
        "reserve_amount": reserve_amount,
    }


def get_account_ledger(account_code):
    """
    获取某个科目的明细账（所有凭证流水 + 运行余额）。
    返回 dict：
      - account_code, account_name, direction
      - opening_balance: 期初余额
      - entries: [{date, voucher_number, summary, debit, credit, balance, direction_label}]
      - total_debit, total_credit: 本期借贷合计
      - ending_balance: 期末余额
    """
    account = ACCOUNT_MAP.get(account_code)
    if account:
        direction = account["direction"]
        acc_name = account["name"]
    else:
        info = get_account_info_dynamic(account_code)
        if info:
            direction = info["direction"]
            acc_name = info["name"]
        else:
            return None

    opening_debit, opening_credit = get_opening_balance(account_code)
    if direction == "借":
        running = (opening_debit or 0) - (opening_credit or 0)
    else:
        running = (opening_credit or 0) - (opening_debit or 0)

    conn = sqlite3.connect(DB_PATH)
    df = pd.read_sql_query(
        "SELECT * FROM vouchers WHERE account_code = ? ORDER BY voucher_date, id",
        conn, params=(account_code,))
    conn.close()

    entries = []
    total_debit = 0.0
    total_credit = 0.0

    for _, row in df.iterrows():
        debit = float(row["debit_amount"] or 0)
        credit = float(row["credit_amount"] or 0)
        total_debit += debit
        total_credit += credit

        if direction == "借":
            running += debit - credit
        else:
            running += credit - debit

        direction_label = "借" if running >= 0 else "贷"
        entries.append({
            "date": row["voucher_date"],
            "voucher_number": row["voucher_number"],
            "summary": row["summary"],
            "debit": debit,
            "credit": credit,
            "balance": running,
            "direction_label": direction_label,
        })

    return {
        "account_code": account_code,
        "account_name": acc_name,
        "direction": direction,
        "opening_balance": running - (total_debit - total_credit) if direction == "借"
                         else running - (total_credit - total_debit),
        "entries": entries,
        "total_debit": total_debit,
        "total_credit": total_credit,
        "ending_balance": running,
    }


# ============================================================
# 第 4 部分：主界面 —— 八个 Tab
# ============================================================

# 只在首次加载时初始化数据库（避免每次 rerun 都创建 31 张表）
if not st.session_state.get("_db_initialized"):
    try:
        init_database()
        st.session_state["_db_initialized"] = True
    except Exception as _init_err:
        st.error(f"⚠️ 数据库初始化失败: {_init_err}")
        st.info(f"数据库模式: {'PostgreSQL' if _USE_POSTGRES else 'SQLite'} | DB_PATH: {DB_PATH}")
        st.stop()

# --- 侧边栏：系统导航 ---
with st.sidebar:
    st.markdown("## 💰 财务 ERP 系统")
    _disp = st.session_state.get("display_name", "")
    _comp_name = st.session_state.get("current_company_name", "")
    if _disp:
        st.markdown(f"👤 **{_disp}**")

    st.markdown("---")

    # 公司切换
    _user_companies = st.session_state.get("user_companies", [])
    _current_co_code = st.session_state.get("current_company_code", "")
    if _user_companies:
        _co_options = [f"{c['code']} | {c['name']}" for c in _user_companies]
        _current_idx = 0
        for i, c in enumerate(_user_companies):
            if c["code"] == _current_co_code:
                _current_idx = i
                break
        _selected_co = st.selectbox(
            "🏢 当前公司",
            _co_options,
            index=_current_idx,
            key="company_switcher",
        )
        _selected_co_code = _selected_co.split(" | ")[0] if " | " in _selected_co else _selected_co
        if _selected_co_code != _current_co_code:
            st.session_state["current_company_code"] = _selected_co_code
            st.session_state["current_company_name"] = _selected_co.split(" | ", 1)[1] if " | " in _selected_co else ""
            st.session_state["_db_initialized"] = False
            st.rerun()

    # 公司管理 & 退出
    _sb_col1, _sb_col2 = st.columns(2)
    with _sb_col1:
        if st.button("🏢 公司管理", key="company_mgmt_btn", use_container_width=True):
            st.session_state["_show_company_mgmt"] = not st.session_state.get("_show_company_mgmt", False)
            st.rerun()
    with _sb_col2:
        if st.button("🚪 退出登录", key="logout_btn", use_container_width=True):
            for k in list(st.session_state.keys()):
                del st.session_state[k]
            st.rerun()

    st.markdown("---")

    # 功能导航树（参照用友/金蝶业务域分组）
    st.markdown("### 📋 功能导航")
    st.markdown("""
    **💰 财务会计**
    > 📝 记账 · 📊 报表 · 🤖 AI问答

    **📦 供应链管理**
    > 📦 库存 · 🏪 商品 · 📋 订单

    **👥 客户关系**
    > 电商CRM会员管理

    **📈 数据分析**
    > BI智能数据报表
    """)
    st.caption("💡 点击上方对应标签页切换业务域")

# --- 主界面标题 ---
st.title("💰 财务 ERP 系统")
st.caption("离线记账 · 自动报表 · AI 智能问答 · 库存 · 商品 · CRM · 订单 · BI")

# --- 公司管理弹窗 ---
if st.session_state.get("_show_company_mgmt"):
    _mg_username = st.session_state.get("username", "default")
    st.markdown("### 🏢 公司管理")

    _mg_tab1, _mg_tab2, _mg_tab3 = st.tabs(["📋 公司列表", "➕ 新建公司", "✏️ 编辑公司"])

    with _mg_tab1:
        _companies_list = get_user_companies(_mg_username)
        if _companies_list:
            _list_data = []
            for c in _companies_list:
                _list_data.append({
                    "公司编码": c["code"],
                    "公司名称": c["name"],
                    "税号": c["tax_id"],
                    "联系电话": c["phone"],
                    "是否默认": "✅ 是" if c["is_default"] else "—",
                    "创建时间": c["created_at"],
                })
            st.dataframe(_list_data, use_container_width=True, hide_index=True)

            # 删除公司
            st.markdown("---")
            st.markdown("#### 🗑️ 删除公司")
            _del_options = [f"{c['code']} | {c['name']}" for c in _companies_list]
            _del_selected = st.selectbox("选择要删除的公司", _del_options, key="del_company_sel")
            if st.button("确认删除", key="del_company_btn", type="secondary"):
                _del_code = _del_selected.split(" | ")[0] if " | " in _del_selected else _del_selected
                _ok, _msg = delete_company(_mg_username, _del_code)
                if _ok:
                    st.success(_msg)
                    st.session_state["user_companies"] = get_user_companies(_mg_username)
                    if _del_code == st.session_state.get("current_company_code"):
                        _remaining = st.session_state["user_companies"]
                        if _remaining:
                            st.session_state["current_company_code"] = _remaining[0]["code"]
                            st.session_state["current_company_name"] = _remaining[0]["name"]
                        st.session_state["_db_initialized"] = False
                    st.rerun()
                else:
                    st.error(_msg)
        else:
            st.info("暂无公司记录")

    with _mg_tab2:
        with st.form("create_company_form"):
            _new_code = st.text_input("公司编码 *", placeholder="如：CO002", key="new_co_code")
            _new_name = st.text_input("公司名称 *", placeholder="如：XX科技有限公司", key="new_co_name")
            _new_tax = st.text_input("统一社会信用代码（税号）", placeholder="选填", key="new_co_tax")
            _new_addr = st.text_input("公司地址", placeholder="选填", key="new_co_addr")
            _new_phone = st.text_input("联系电话", placeholder="选填", key="new_co_phone")
            _new_submit = st.form_submit_button("➕ 创建公司")

            if _new_submit:
                if not _new_code or not _new_name:
                    st.error("公司编码和名称不能为空")
                else:
                    _ok, _msg = create_company(_mg_username, _new_code.strip(), _new_name.strip(),
                                               _new_tax.strip(), _new_addr.strip(), _new_phone.strip())
                    if _ok:
                        st.success(_msg)
                        st.session_state["user_companies"] = get_user_companies(_mg_username)
                        st.rerun()
                    else:
                        st.error(_msg)

    with _mg_tab3:
        _edit_companies = get_user_companies(_mg_username)
        if _edit_companies:
            _edit_options = [f"{c['code']} | {c['name']}" for c in _edit_companies]
            _edit_selected = st.selectbox("选择公司", _edit_options, key="edit_company_sel")
            _edit_code = _edit_selected.split(" | ")[0] if " | " in _edit_selected else _edit_selected

            _edit_co = None
            for c in _edit_companies:
                if c["code"] == _edit_code:
                    _edit_co = c
                    break

            if _edit_co:
                with st.form("edit_company_form"):
                    _ed_name = st.text_input("公司名称", value=_edit_co["name"], key="ed_co_name")
                    _ed_tax = st.text_input("税号", value=_edit_co["tax_id"], key="ed_co_tax")
                    _ed_addr = st.text_input("公司地址", value=_edit_co["address"], key="ed_co_addr")
                    _ed_phone = st.text_input("联系电话", value=_edit_co["phone"], key="ed_co_phone")
                    _ed_submit = st.form_submit_button("💾 保存修改")

                    if _ed_submit:
                        _ok, _msg = update_company(_mg_username, _edit_code,
                                                   company_name=_ed_name.strip(),
                                                   tax_id=_ed_tax.strip(),
                                                   address=_ed_addr.strip(),
                                                   phone=_ed_phone.strip())
                        if _ok:
                            st.success(_msg)
                            if _edit_code == st.session_state.get("current_company_code"):
                                st.session_state["current_company_name"] = _ed_name.strip()
                            st.session_state["user_companies"] = get_user_companies(_mg_username)
                            st.rerun()
                        else:
                            st.error(_msg)
        else:
            st.info("暂无可编辑的公司")

    st.markdown("---")

# === 两级导航：业务域分组（对标用友U8/金蝶K3模块分区）===
_grp_finance, _grp_supply, _grp_crm, _grp_bi = st.tabs([
    "💰 财务会计", "📦 供应链管理", "👥 客户关系", "📈 数据分析"
])

# --- 财务会计域：记账 + 报表 + AI问答 ---
with _grp_finance:
    st.markdown("##### 💰 财务会计 ｜ 总账核算 · 财务报表 · AI 辅助")
    tab1, tab2, tab3 = st.tabs(["📝 记账", "📊 报表", "🤖 AI 问答"])

# --- 供应链管理域：库存 + 商品 + 订单 ---
with _grp_supply:
    st.markdown("##### 📦 供应链管理 ｜ 库存核算 · 商品管理 · 订单履约")
    tab4, tab5, tab7 = st.tabs(["📦 库存管理", "🏪 多平台商品", "📋 订单 OMS"])

# --- 客户关系域：CRM会员 ---
with _grp_crm:
    st.markdown("##### 👥 客户关系管理 ｜ 会员体系 · 积分营销")
    tab6, = st.tabs(["👥 电商 CRM 会员"])

# --- 数据分析域：BI报表 ---
with _grp_bi:
    st.markdown("##### 📈 数据分析 ｜ 商业智能 · 经营洞察")
    tab8, = st.tabs(["📈 BI 数据报表"])


# ============================================================
# 【💰 财务会计】模块一：记账（总账核算）
# ============================================================
with tab1:
    st.header("📝 总账记账")
    st.caption("期初余额 · 录入凭证 · 凭证查询 · 明细账 · 科目管理 · 期末结转 ｜ 完全离线，无需 API")

    sub1, sub2, sub3, sub4, sub5, sub6 = st.tabs(["期初余额", "录入凭证", "凭证查询", "明细账", "科目管理", "🔄 期末结转"])

    # --- 期初余额 ---
    with sub1:
        st.subheader("期初余额录入")

        # 动态获取科目列表（系统科目 + 自定义二级科目）
        _dynamic_names = get_account_names_dynamic()

        # 录入区域：科目 + 金额
        cols_ob_in = st.columns([3, 2, 1])
        with cols_ob_in[0]:
            selected_ob = st.selectbox(
                "科目", _dynamic_names,
                key="ob_input_acc", label_visibility="collapsed",
                placeholder="搜索科目名称或编号...",
            )
        with cols_ob_in[1]:
            # 保存后清空输入框的标志位（必须在 widget 创建前处理）
            if st.session_state.pop("_clear_ob_amt", False):
                st.session_state["ob_input_amt"] = ""
                st.session_state["_money_init_ob_input_amt"] = True
            ob_amt_input = money_input(
                "金额", key="ob_input_amt", min_value=0.0,
                label_visibility="collapsed", placeholder="0.00",
            )
        with cols_ob_in[2]:
            save_clicked = st.button("💾 保存", key="ob_save_btn", use_container_width=True)

        code_ob = selected_ob.split(" ")[0]
        # 从显示名中提取科目名称（处理自定义科目带 ← 的情况）
        _name_part = selected_ob.split(" ", 1)[1] if " " in selected_ob else selected_ob
        name_ob = _name_part.split("  ←")[0].strip() if "  ←" in _name_part else _name_part.strip()
        acc_info_ob = get_account_info_dynamic(code_ob) or {}
        direction_ob = acc_info_ob.get("direction", "借")
        color_ob = "#1976d2" if direction_ob == "借" else "#e53935"
        st.caption(f"已选：**{code_ob} {name_ob}** ｜ 方向：<span style='color:{color_ob};font-weight:600'>{direction_ob}方</span>", unsafe_allow_html=True)

        if save_clicked:
            if ob_amt_input > 0:
                if direction_ob == "借":
                    save_opening_balance(code_ob, ob_amt_input, 0)
                else:
                    save_opening_balance(code_ob, 0, ob_amt_input)
                st.success(f"✅ 已保存：{name_ob} = {ob_amt_input:,.2f} 元（{direction_ob}方）")
                # 设置清空标志，在下一次 rerun 时清空输入框
                st.session_state["_clear_ob_amt"] = True
                st.rerun()
            else:
                st.warning("请输入金额")

        # 已保存的期初余额列表（包含一级科目和二级科目）
        all_opening = []
        # 系统一级科目
        for a in ACCOUNT_CHART:
            od_v, oc_v = get_opening_balance(a["code"])
            if od_v or oc_v:
                amt_v = od_v if a["direction"] == "借" else oc_v
                all_opening.append({
                    "code": a["code"],
                    "name": a["name"],
                    "direction": a["direction"],
                    "amount": amt_v,
                    "raw_debit": od_v,
                    "raw_credit": oc_v,
                    "level": 1,
                })
        # 自定义二级科目
        df_custom_ob = get_custom_accounts()
        for _, row in df_custom_ob.iterrows():
            od_v, oc_v = get_opening_balance(row["full_code"])
            if od_v or oc_v:
                amt_v = od_v if row["direction"] == "借" else oc_v
                all_opening.append({
                    "code": row["full_code"],
                    "name": row["full_name"],
                    "direction": row["direction"],
                    "amount": amt_v,
                    "raw_debit": od_v,
                    "raw_credit": oc_v,
                    "level": 2,
                })

        st.markdown("---")

        if all_opening:
            # 表头
            st.markdown(
                '<div style="display:flex; font-weight:600; font-size:13px; '
                'padding:6px 0; border-bottom:2px solid #333; margin-bottom:4px;">'
                '<span style="flex:1; text-align:center;">科目</span>'
                '<span style="flex:0.5; text-align:center;">方向</span>'
                '<span style="flex:1; text-align:center;">期初余额</span>'
                '<span style="flex:0.5; text-align:center;">删除</span>'
                '</div>',
                unsafe_allow_html=True,
            )
            for o in all_opening:
                cols_ob = st.columns([1, 0.5, 1, 0.5])
                with cols_ob[0]:
                    st.markdown(
                        f'<div style="padding-top:5px; font-size:14px;">'
                        f'{o["code"]} {o["name"]}</div>',
                        unsafe_allow_html=True,
                    )
                with cols_ob[1]:
                    c = "#1976d2" if o["direction"] == "借" else "#e53935"
                    st.markdown(
                        f'<div style="text-align:center; padding-top:5px; '
                        f'font-size:13px; color:{c}; font-weight:600;">{o["direction"]}方</div>',
                        unsafe_allow_html=True,
                    )
                with cols_ob[2]:
                    st.markdown(
                        f'<div style="text-align:center; padding-top:5px; '
                        f'font-size:14px; font-weight:600;">{o["amount"]:,.2f}</div>',
                        unsafe_allow_html=True,
                    )
                with cols_ob[3]:
                    if st.button("🗑", key=f"ob_del_{o['code']}", help=f"删除{o['name']}的期初余额"):
                        save_opening_balance(o["code"], 0, 0)
                        st.rerun()

            # 合计行
            total_debit_ob = sum(o["amount"] for o in all_opening if o["direction"] == "借")
            total_credit_ob = sum(o["amount"] for o in all_opening if o["direction"] == "贷")
            st.markdown(
                f'<div style="display:flex; font-weight:600; font-size:14px; '
                f'padding:8px 0; border-top:2px solid #333; border-bottom:2px solid #333;">'
                f'<span style="flex:1;">合计</span>'
                f'<span style="flex:0.5;"></span>'
                f'<span style="flex:1; text-align:center;">'
                f'借: {total_debit_ob:,.2f} ｜ 贷: {total_credit_ob:,.2f}</span>'
                f'<span style="flex:0.5;"></span>'
                f'</div>',
                unsafe_allow_html=True,
            )
            if abs(total_debit_ob - total_credit_ob) > 0.01:
                st.warning(f"⚠️ 借贷不平衡！差额：{abs(total_debit_ob - total_credit_ob):,.2f} 元")
            else:
                st.success("✅ 借贷平衡")
        else:
            st.info("💡 尚未录入任何期初余额。在上方选择科目并输入金额后点击「保存」。")

    # --- 录入凭证 ---
    with sub2:
        st.subheader("录入记账凭证")

        # 凭证金额专用回调：Enter 格式化 + 最后一行自动新增行
        def _voucher_money_cb(key):
            val = st.session_state.get(key, "")
            if not val or not str(val).strip():
                return
            cleaned = str(val).replace(",", "").replace("，", "").replace(" ", "").strip()
            try:
                num = float(cleaned)
                formatted = f"{num:,.2f}"
                current = str(val).strip()
                if current != formatted:
                    # Enter 格式化金额
                    st.session_state[key] = formatted
                    # 如果是最后一行的金额，标记需要新增行
                    parts = key.split("_")
                    if len(parts) >= 3:
                        _row_idx = int(parts[-1])
                        if _row_idx == st.session_state.get("voucher_row_count", 2) - 1:
                            st.session_state["_vch_add_row"] = True
            except ValueError:
                pass

        # 中文大写金额转换函数
        def to_chinese_amount(amount):
            """把数字金额转成中文大写，如：壹仟贰佰元整"""
            if abs(amount) < 0.005:
                return "零元整"
            digits = "零壹贰叁肆伍陆柒捌玖"
            units = ["", "拾", "佰", "仟", "万", "拾", "佰", "仟", "亿", "拾", "佰", "仟"]
            # 分离整数和小数部分
            integer_part = int(abs(amount))
            decimal_part = round(abs(amount) - integer_part, 2)

            # 整数部分转换
            if integer_part == 0:
                int_str = ""
            else:
                int_str = ""
                s = str(integer_part)
                length = len(s)
                for i, ch in enumerate(s):
                    n = int(ch)
                    pos = length - 1 - i
                    if n == 0:
                        if int_str and not int_str.endswith("零") and not int_str.endswith("亿") and not int_str.endswith("万"):
                            int_str += "零"
                    else:
                        int_str += digits[n] + units[pos]
                # 清理末尾多余的零
                int_str = int_str.rstrip("零")
                int_str += "元"

            # 小数部分
            if decimal_part == 0:
                dec_str = "整"
            else:
                jiao = int(decimal_part * 10)
                fen = int(round((decimal_part * 100) % 10))
                dec_str = ""
                if jiao > 0:
                    dec_str += digits[jiao] + "角"
                if fen > 0:
                    dec_str += digits[fen] + "分"
                if not dec_str:
                    dec_str = "整"

            result = int_str + dec_str
            if amount < 0:
                result = "负" + result
            return result

        # === 凭证头部 ===
        st.markdown(
            '<div style="text-align:center; padding:8px 0; font-size:20px; font-weight:600; '
            'border-bottom: 2px solid #333; margin-bottom:16px;">记 账 凭 证</div>',
            unsafe_allow_html=True,
        )

        # 凭证字、号、日期、附单据
        VOUCHER_TYPES = {"记": "记账凭证", "收": "收款凭证", "付": "付款凭证", "转": "转账凭证"}

        col_vtype, col_vnum, col_vdate, col_attach = st.columns([1, 1, 1.5, 1])
        with col_vtype:
            voucher_type = st.selectbox("凭证字", list(VOUCHER_TYPES.keys()), key="vtype")
        with col_vnum:
            next_num = get_next_voucher_number()
            # 提取数字部分
            base_num = int(next_num.replace("记字第", "").replace("号", "")) if "记字第" in next_num else 1
            vnum = st.number_input("字第 ___ 号", min_value=1, value=base_num, step=1, key="vnum")
        with col_vdate:
            voucher_date = st.date_input("凭证日期")
        with col_attach:
            attach_count = st.number_input("附单据", min_value=0, value=0, step=1, key="attach")

        full_voucher_number = f"{voucher_type}字第{vnum:03d}号"

        # === 分录明细表 ===
        st.markdown("**分录明细**（每行填一条，至少一行借、一行贷）")

        # 动态行数：默认2行，可增减
        if "voucher_row_count" not in st.session_state:
            st.session_state.voucher_row_count = 2

        # 增减行按钮
        col_add, col_del, col_info = st.columns([1, 1, 3])
        with col_add:
            if st.button("➕ 添加一行", key="add_row"):
                st.session_state.voucher_row_count += 1
                st.rerun()
        with col_del:
            if st.session_state.voucher_row_count > 2:
                if st.button("➖ 删除最后一行", key="del_row"):
                    # 清空最后一行的所有数据
                    last = st.session_state.voucher_row_count - 1
                    for suffix in ["sum", "acc", "debit", "credit"]:
                        k = f"row_{suffix}_{last}"
                        if k in st.session_state:
                            del st.session_state[k]
                        # 清除 money_input 初始化标记
                        ik = f"_money_init_{k}"
                        if ik in st.session_state:
                            del st.session_state[ik]
                    st.session_state.voucher_row_count -= 1
                    st.rerun()
        with col_info:
            st.caption(f"当前 {st.session_state.voucher_row_count} 行（填入金额按 Enter 格式化，最后一行自动新增）")

        row_count = st.session_state.voucher_row_count
        lines = []

        # 用 HTML 画表头
        st.markdown(
            '<div style="display:flex; font-weight:600; font-size:13px; '
            'padding:6px 0; border-bottom:2px solid #333; margin-bottom:4px;">'
            '<span style="width:40px; text-align:center;">行次</span>'
            '<span style="flex:1; text-align:center;">摘要</span>'
            '<span style="flex:1.2; text-align:center;">会计科目</span>'
            '<span style="width:120px; text-align:center;">借方金额</span>'
            '<span style="width:120px; text-align:center;">贷方金额</span>'
            '</div>',
            unsafe_allow_html=True,
        )

        for i in range(row_count):
            cols = st.columns([0.5, 2, 2.5, 1.5, 1.5])
            with cols[0]:
                st.markdown(
                    f'<div style="text-align:center; padding-top:8px; font-size:14px; '
                    f'color:#666;">{i+1}</div>',
                    unsafe_allow_html=True,
                )
            with cols[1]:
                row_summary = st.text_input(
                    f"摘要{i+1}", label_visibility="collapsed",
                    key=f"row_sum_{i}", placeholder="摘要"
                )
            with cols[2]:
                selected = st.selectbox(
                    f"科目{i+1}", get_account_names_dynamic(),
                    key=f"row_acc_{i}", label_visibility="collapsed"
                )
            with cols[3]:
                debit = money_input(
                    f"借{i+1}", key=f"row_debit_{i}", min_value=0.0,
                    label_visibility="collapsed",
                    on_change_cb=_voucher_money_cb, on_change_args=(f"row_debit_{i}",),
                )
            with cols[4]:
                credit = money_input(
                    f"贷{i+1}", key=f"row_credit_{i}", min_value=0.0,
                    label_visibility="collapsed",
                    on_change_cb=_voucher_money_cb, on_change_args=(f"row_credit_{i}",),
                )

            if debit > 0 or credit > 0:
                code = selected.split(" ")[0]
                # 从显示名中提取科目名称（处理自定义科目带 ← 的情况）
                _n_part = selected.split(" ", 1)[1] if " " in selected else selected
                name = _n_part.split("  ←")[0].strip() if "  ←" in _n_part else _n_part.strip()
                lines.append({
                    "account_code": code,
                    "account_name": name,
                    "debit": debit,
                    "credit": credit,
                    "summary": row_summary if row_summary else "",
                })

        # === 最后一行金额格式化后自动新增行 ===
        if st.session_state.pop("_vch_add_row", False):
            st.session_state.voucher_row_count += 1
            st.rerun()

        # === 合计行 ===
        total_debit = sum(l["debit"] for l in lines)
        total_credit = sum(l["credit"] for l in lines)
        is_balanced = abs(total_debit - total_credit) < 0.01
        chinese_total = to_chinese_amount(max(total_debit, total_credit))

        st.markdown("---")
        st.markdown(
            f'<div style="display:flex; font-weight:600; font-size:14px; '
            f'padding:8px 0; border-top:2px solid #333; border-bottom:2px solid #333;">'
            f'<span style="width:40px;"></span>'
            f'<span style="flex:1;">合计：{chinese_total}</span>'
            f'<span style="flex:1.2;"></span>'
            f'<span style="width:120px; text-align:center;">{total_debit:,.2f}</span>'
            f'<span style="width:120px; text-align:center;">{total_credit:,.2f}</span>'
            f'</div>',
            unsafe_allow_html=True,
        )

        # 借贷平衡校验
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("借方合计", f"{total_debit:,.2f} 元")
        with col2:
            st.metric("贷方合计", f"{total_credit:,.2f} 元")
        with col3:
            if is_balanced:
                st.metric("校验结果", "✅ 平衡")
            else:
                st.metric("校验结果", "❌ 不平衡")

        if not is_balanced and (total_debit > 0 or total_credit > 0):
            st.error(f"借贷不平衡！差额：{abs(total_debit - total_credit):,.2f} 元。请调整金额使借贷相等。")

        # 保存 + 清空按钮
        col_save, col_clear = st.columns([1, 1])
        with col_save:
            if st.button("💾 保存凭证", type="primary", use_container_width=True):
                if not lines:
                    st.warning("请至少填写一行分录")
                elif not is_balanced:
                    st.warning("借贷不平衡，无法保存。请调整金额。")
                elif len(lines) < 2:
                    st.warning("至少需要两行分录（一行借、一行贷）")
                else:
                    # 用第一行的摘要作为主摘要
                    main_summary = lines[0]["summary"] if lines[0]["summary"] else "综合凭证"
                    save_voucher(
                        full_voucher_number,
                        voucher_date.strftime("%Y-%m-%d"),
                        main_summary,
                        lines,
                    )
                    st.success(f"✅ 凭证已保存！编号：{full_voucher_number}")
                    # 重置行数为2，清除所有行数据
                    _max_rows = st.session_state.voucher_row_count
                    for _r in range(_max_rows):
                        for _suf in ["sum", "acc", "debit", "credit"]:
                            _rk = f"row_{_suf}_{_r}"
                            if _rk in st.session_state:
                                del st.session_state[_rk]
                            _ik = f"_money_init_{_rk}"
                            if _ik in st.session_state:
                                del st.session_state[_ik]
                    st.session_state.voucher_row_count = 2
                    st.rerun()
        with col_clear:
            if st.button("🔄 清空重填", use_container_width=True):
                _max_rows = st.session_state.voucher_row_count
                for _r in range(_max_rows):
                    for _suf in ["sum", "acc", "debit", "credit"]:
                        _rk = f"row_{_suf}_{_r}"
                        if _rk in st.session_state:
                            del st.session_state[_rk]
                        _ik = f"_money_init_{_rk}"
                        if _ik in st.session_state:
                            del st.session_state[_ik]
                st.session_state.voucher_row_count = 2
                st.rerun()

    # --- 凭证查询 ---
    with sub3:
        st.subheader("凭证查询")

        df = get_all_vouchers()
        if df.empty:
            st.info("暂无凭证记录。请先在「录入凭证」中添加。")
        else:
            # 按凭证编号分组显示
            voucher_numbers = df["voucher_number"].unique()

            st.write(f"共 {len(voucher_numbers)} 张凭证")

            # --- 导出凭证为格式化 Excel ---
            def _write_voucher_ws(ws, vdf_single, vnum, vdate, vsummary):
                """将一张记账凭证写入指定 worksheet（带专业格式）"""
                from openpyxl.styles import Font, Alignment, Border, Side
                import math

                _thin = Side(style='thin')
                _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
                _title_font = Font(name='Arial', size=18, bold=True)
                _hdr_font = Font(name='Arial', size=10, bold=True)
                _data_font = Font(name='Arial', size=10)
                _center = Alignment(horizontal='center', vertical='center')
                _left = Alignment(horizontal='left', vertical='center')
                _right = Alignment(horizontal='right', vertical='center')

                for col, w in [('A', 25), ('B', 30), ('C', 18), ('D', 18)]:
                    ws.column_dimensions[col].width = w

                # 第1行：标题
                ws.merge_cells('A1:D1')
                ws['A1'] = '记 账 凭 证'
                ws['A1'].font = _title_font
                ws['A1'].alignment = _center
                ws.row_dimensions[1].height = 30

                # 第2行：日期 / 凭证编号 / 附件
                ws['A2'] = f'日期：{vdate}'
                ws['A2'].font = _data_font
                ws['A2'].alignment = _left
                ws.merge_cells('B2:C2')
                ws['B2'] = f'凭证编号：{vnum}'
                ws['B2'].font = _data_font
                ws['B2'].alignment = _center
                ws['D2'] = '附件：    张'
                ws['D2'].font = _data_font
                ws['D2'].alignment = _right
                ws.row_dimensions[2].height = 20

                # 第3行：摘要
                ws.merge_cells('A3:D3')
                ws['A3'] = f'摘要：{vsummary}'
                ws['A3'].font = _data_font
                ws['A3'].alignment = _left
                ws.row_dimensions[3].height = 20

                # 第4行：表头
                for i, h in enumerate(['摘要', '科目', '借方金额', '贷方金额'], 1):
                    c = ws.cell(row=4, column=i, value=h)
                    c.font = _hdr_font
                    c.alignment = _center
                    c.border = _border
                ws.row_dimensions[4].height = 20

                # 数据行
                d_total = 0.0
                c_total = 0.0
                r = 5
                for _, row in vdf_single.iterrows():
                    ws.row_dimensions[r].height = 22
                    debit = float(row["debit_amount"] or 0)
                    credit = float(row["credit_amount"] or 0)
                    d_total += debit
                    c_total += credit
                    vals = [
                        vsummary,
                        row["account_name"],
                        f"{debit:,.2f}" if debit else "",
                        f"{credit:,.2f}" if credit else "",
                    ]
                    for ci, val in enumerate(vals, 1):
                        c = ws.cell(row=r, column=ci, value=val if val else None)
                        c.font = _data_font
                        c.border = _border
                        c.number_format = '@'
                        c.alignment = _left if ci <= 2 else _right
                    r += 1

                # 合计行
                ws.row_dimensions[r].height = 22
                for ci, val in enumerate(['合计', '', f"{d_total:,.2f}", f"{c_total:,.2f}"], 1):
                    c = ws.cell(row=r, column=ci, value=val if val else None)
                    c.font = _hdr_font
                    c.border = _border
                    c.number_format = '@'
                    c.alignment = _center if ci <= 2 else _right

                # 大写金额行
                r += 2
                ws.merge_cells(f'A{r}:D{r}')

                def _to_chinese(amount):
                    if abs(amount) < 0.005:
                        return "零元整"
                    digits = "零壹贰叁肆伍陆柒捌玖"
                    units = ["", "拾", "佰", "仟", "万", "拾", "佰", "仟", "亿", "拾", "佰", "仟"]
                    integer_part = int(abs(amount))
                    decimal_part = round(abs(amount) - integer_part, 2)
                    if integer_part == 0:
                        int_str = ""
                    else:
                        int_str = ""
                        s = str(integer_part)
                        length = len(s)
                        for i, ch in enumerate(s):
                            n = int(ch)
                            pos = length - 1 - i
                            if n == 0:
                                if int_str and not int_str.endswith("零") and not int_str.endswith("亿") and not int_str.endswith("万"):
                                    int_str += "零"
                            else:
                                int_str += digits[n] + units[pos]
                        int_str = int_str.rstrip("零") + "元"
                    if decimal_part == 0:
                        dec_str = "整"
                    else:
                        jiao = int(decimal_part * 10)
                        fen = int(round((decimal_part * 100) % 10))
                        dec_str = ""
                        if jiao > 0:
                            dec_str += digits[jiao] + "角"
                        if fen > 0:
                            dec_str += digits[fen] + "分"
                        if not dec_str:
                            dec_str = "整"
                    result = int_str + dec_str
                    return ("负" + result) if amount < 0 else result

                ws.cell(row=r, column=1).value = f'合计金额（大写）：{_to_chinese(d_total)}'
                ws.cell(row=r, column=1).font = _data_font
                ws.cell(row=r, column=1).alignment = _left

                # 制单 / 审核
                r += 2
                ws.cell(row=r, column=1).value = '制单：'
                ws.cell(row=r, column=1).font = _data_font
                ws.cell(row=r, column=3).value = '审核：'
                ws.cell(row=r, column=3).font = _data_font

            def export_single_voucher_excel(vdf_single, vnum, vdate, vsummary):
                """导出单张凭证为 Excel"""
                from openpyxl import Workbook
                import io
                wb = Workbook()
                ws = wb.active
                ws.title = vnum[:31]
                _write_voucher_ws(ws, vdf_single, vnum, vdate, vsummary)
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                return output

            def export_all_vouchers_excel():
                """导出所有凭证到一个 Excel（每张一个 sheet）"""
                from openpyxl import Workbook
                import io
                wb = Workbook()
                wb.remove(wb.active)
                for vnum in voucher_numbers:
                    vdf_single = df[df["voucher_number"] == vnum]
                    vdate = vdf_single.iloc[0]["voucher_date"]
                    vsummary = vdf_single.iloc[0]["summary"]
                    ws = wb.create_sheet(title=vnum[:31])
                    _write_voucher_ws(ws, vdf_single, vnum, vdate, vsummary)
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                return output

            st.download_button(
                label="📥 下载所有凭证（Excel）",
                data=export_all_vouchers_excel(),
                file_name="所有凭证.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

            for vnum in voucher_numbers:
                vdf = df[df["voucher_number"] == vnum]
                vdate = vdf.iloc[0]["voucher_date"]
                vsummary = vdf.iloc[0]["summary"]

                d_total = vdf["debit_amount"].sum()
                c_total = vdf["credit_amount"].sum()
                balanced = "✅" if abs(d_total - c_total) < 0.01 else "❌"

                with st.expander(f"{vnum} | {vdate} | {vsummary} | 借 {d_total:,.2f} = 贷 {c_total:,.2f} {balanced}"):
                    display_df = vdf[["account_name", "debit_amount", "credit_amount"]].copy()
                    display_df.columns = ["科目", "借方金额", "贷方金额"]
                    # 加合计行
                    display_df.loc[len(display_df)] = ["合计", d_total, c_total]
                    display_df = fmt_money_df(display_df, ["借方金额", "贷方金额"])
                    st.dataframe(display_df, use_container_width=True, hide_index=True)

                    st.download_button(
                        label=f"📥 下载此凭证（{vnum}）",
                        data=export_single_voucher_excel(vdf, vnum, vdate, vsummary),
                        file_name=f"凭证_{vnum}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"dl_{vnum}",
                    )

                    # --- 删除凭证 ---
                    st.markdown("---")
                    del_cols = st.columns([1, 3])
                    with del_cols[0]:
                        if st.button("🗑️ 删除此凭证", key=f"del_{vnum}",
                                     type="secondary", use_container_width=True):
                            st.session_state[f"_confirm_del_{vnum}"] = True
                    with del_cols[1]:
                        if st.session_state.get(f"_confirm_del_{vnum}"):
                            st.warning("⚠️ 确认要删除这张凭证吗？此操作不可撤销！")
                            confirm_cols = st.columns(2)
                            with confirm_cols[0]:
                                if st.button("✅ 确认删除", key=f"confirm_del_{vnum}",
                                             type="primary", use_container_width=True):
                                    delete_voucher(vnum)
                                    st.session_state.pop(f"_confirm_del_{vnum}", None)
                                    st.success(f"凭证 {vnum} 已删除！")
                                    st.rerun()
                            with confirm_cols[1]:
                                if st.button("❌ 取消", key=f"cancel_del_{vnum}",
                                             use_container_width=True):
                                    st.session_state.pop(f"_confirm_del_{vnum}", None)
                                    st.rerun()

                    # --- 修改凭证 ---
                    st.markdown("---")
                    edit_cols = st.columns([1, 3])
                    with edit_cols[0]:
                        if st.button("✏️ 修改此凭证", key=f"edit_{vnum}",
                                     use_container_width=True):
                            # 加载凭证数据到编辑模式
                            lines = get_voucher_by_number(vnum)
                            info = get_voucher_info(vnum)
                            st.session_state[f"_edit_voucher_{vnum}"] = {
                                "lines": lines,
                                "date": info["date"] if info else vdate,
                                "summary": info["summary"] if info else vsummary,
                            }

                    # 编辑界面
                    edit_data = st.session_state.get(f"_edit_voucher_{vnum}")
                    if edit_data:
                        st.markdown("#### ✏️ 修改凭证")
                        _dynamic_names_edit = get_account_names_dynamic()

                        # 修改日期和摘要
                        edit_c1, edit_c2 = st.columns([1, 2])
                        with edit_c1:
                            try:
                                _edit_date_obj = datetime.strptime(edit_data["date"], "%Y-%m-%d")
                            except (ValueError, TypeError):
                                _edit_date_obj = datetime.now()
                            new_date = st.date_input(
                                "凭证日期", value=_edit_date_obj,
                                key=f"edit_date_{vnum}",
                            )
                        with edit_c2:
                            new_summary = st.text_input(
                                "摘要", value=edit_data["summary"],
                                key=f"edit_summary_{vnum}",
                            )

                        # 修改分录行
                        st.markdown("##### 分录明细")
                        edit_lines = edit_data["lines"]

                        # 动态行数控制
                        if f"_edit_row_count_{vnum}" not in st.session_state:
                            st.session_state[f"_edit_row_count_{vnum}"] = len(edit_lines)

                        row_count = st.session_state[f"_edit_row_count_{vnum}"]
                        rc1, rc2 = st.columns([3, 1])
                        with rc2:
                            add_rows = st.number_input(
                                "行数", min_value=2, max_value=20,
                                value=max(row_count, 2), step=1,
                                key=f"edit_rowcnt_{vnum}",
                                label_visibility="collapsed",
                            )
                            if add_rows != row_count:
                                st.session_state[f"_edit_row_count_{vnum}"] = int(add_rows)
                                # 扩展或截断行
                                while len(edit_lines) < int(add_rows):
                                    edit_lines.append({
                                        "account_code": "",
                                        "account_name": "",
                                        "debit": 0,
                                        "credit": 0,
                                })
                                edit_lines = edit_lines[:int(add_rows)]
                                edit_data["lines"] = edit_lines
                                st.rerun()

                        # 渲染分录编辑表格
                        new_lines_data = []
                        total_debit = 0
                        total_credit = 0
                        for i in range(len(edit_lines)):
                            ec1, ec2, ec3, ec4, ec5 = st.columns([3, 2, 2, 1, 1])
                            with ec1:
                                # 构建当前选中值
                                cur_val = None
                                if edit_lines[i]["account_code"]:
                                    cur_val = f'{edit_lines[i]["account_code"]} {edit_lines[i]["account_name"]}'
                                idx_default = 0
                                if cur_val and cur_val in _dynamic_names_edit:
                                    idx_default = _dynamic_names_edit.index(cur_val)
                                sel = st.selectbox(
                                    f"科目{i+1}", _dynamic_names_edit,
                                    index=idx_default, key=f"edit_acc_{vnum}_{i}",
                                    label_visibility="collapsed",
                                )
                                if sel:
                                    parts = sel.split(" ", 1)
                                    acc_code = parts[0]
                                    acc_name = parts[1] if len(parts) > 1 else ""
                                else:
                                    acc_code = ""
                                    acc_name = ""
                            with ec2:
                                d_val = st.text_input(
                                    "借方", value=f"{edit_lines[i]['debit']:,.2f}" if edit_lines[i]['debit'] else "",
                                    key=f"edit_d_{vnum}_{i}",
                                    placeholder="0.00",
                                    label_visibility="collapsed",
                                )
                                d_num = float(str(d_val).replace(",", "").strip() or 0)
                            with ec3:
                                c_val = st.text_input(
                                    "贷方", value=f"{edit_lines[i]['credit']:,.2f}" if edit_lines[i]['credit'] else "",
                                    key=f"edit_c_{vnum}_{i}",
                                    placeholder="0.00",
                                    label_visibility="collapsed",
                                )
                                c_num = float(str(c_val).replace(",", "").strip() or 0)
                            with ec4:
                                if st.button("🗑", key=f"edit_del_row_{vnum}_{i}",
                                             help="删除此行"):
                                    if len(edit_lines) > 2:
                                        edit_lines.pop(i)
                                        st.session_state[f"_edit_row_count_{vnum}"] = len(edit_lines)
                                        st.rerun()
                            with ec5:
                                st.write("")

                            new_lines_data.append({
                                "account_code": acc_code,
                                "account_name": acc_name,
                                "debit": d_num,
                                "credit": c_num,
                            })
                            total_debit += d_num
                            total_credit += c_num

                        # 借贷平衡校验
                        balance_ok = abs(total_debit - total_credit) < 0.01
                        bal_color = "#2e7d32" if balance_ok else "#c62828"
                        st.markdown(
                            f'<div style="text-align:right; font-size:14px; '
                            f'color:{bal_color}; font-weight:600; padding:4px 0;">'
                            f'借方合计：{fmt_money(total_debit)} &nbsp;|&nbsp; '
                            f'贷方合计：{fmt_money(total_credit)} &nbsp;|&nbsp; '
                            f'{"✅ 平衡" if balance_ok else "❌ 不平衡"}'
                            f'</div>',
                            unsafe_allow_html=True,
                        )

                        # 保存 / 取消
                        save_cols = st.columns(2)
                        with save_cols[0]:
                            if st.button("💾 保存修改", key=f"save_edit_{vnum}",
                                         type="primary", use_container_width=True,
                                         disabled=not balance_ok):
                                new_date_str = new_date.strftime("%Y-%m-%d") if new_date else edit_data["date"]
                                # 过滤掉空行
                                filtered_lines = [l for l in new_lines_data if l["account_code"]]
                                if len(filtered_lines) >= 2 and abs(
                                    sum(l["debit"] for l in filtered_lines) -
                                    sum(l["credit"] for l in filtered_lines)
                                ) < 0.01:
                                    update_voucher(vnum, new_date_str, new_summary, filtered_lines)
                                    st.session_state.pop(f"_edit_voucher_{vnum}", None)
                                    st.session_state.pop(f"_edit_row_count_{vnum}", None)
                                    st.success(f"✅ 凭证 {vnum} 已修改成功！")
                                    st.rerun()
                                else:
                                    st.error("保存失败：请确保至少 2 行分录且借贷平衡。")
                        with save_cols[1]:
                            if st.button("❌ 取消修改", key=f"cancel_edit_{vnum}",
                                         use_container_width=True):
                                st.session_state.pop(f"_edit_voucher_{vnum}", None)
                                st.session_state.pop(f"_edit_row_count_{vnum}", None)
                                st.rerun()

    # --- 明细账 ---
    with sub4:
        st.subheader("科目明细账")
        st.caption("每个科目就像一个「碗」，凭证就是「瓢」，碗里水位随每张凭证的借贷变动而变化。选择一个科目，查看它的全部流水。")

        # 科目选择
        selected_account = st.selectbox(
            "选择科目",
            options=get_account_names_dynamic(),
            index=0,
            key="ledger_account_select",
        )

        if selected_account:
            # 解析科目编码
            sel_code = selected_account.split(" ")[0]
            ledger = get_account_ledger(sel_code)

            if not ledger:
                st.warning("未找到该科目信息")
            else:
                acc_name = f"{ledger['account_code']} {ledger['account_name']}"
                direction_text = "借方" if ledger["direction"] == "借" else "贷方"
                st.markdown(f"### {acc_name}（{direction_text}科目）")

                # 期初余额
                opening = ledger["opening_balance"]
                opening_dir = "借" if opening >= 0 else "贷"
                st.info(
                    f"期初余额：**{abs(opening):,.2f}** 元（{opening_dir}方）"
                )

                # 构建明细表
                if not ledger["entries"]:
                    st.warning("该科目暂无凭证记录。")
                else:
                    rows = []
                    for e in ledger["entries"]:
                        rows.append({
                            "日期": e["date"],
                            "凭证编号": e["voucher_number"],
                            "摘要": e["summary"],
                            "借方": f"{e['debit']:,.2f}" if e["debit"] else "",
                            "贷方": f"{e['credit']:,.2f}" if e["credit"] else "",
                            "方向": e["direction_label"],
                            "余额": f"{abs(e['balance']):,.2f}",
                        })

                    ledger_df = pd.DataFrame(rows)
                    st.dataframe(ledger_df, use_container_width=True, hide_index=True)

                    # 本期合计
                    td = ledger["total_debit"]
                    tc = ledger["total_credit"]
                    ending = ledger["ending_balance"]
                    ending_dir = "借" if ending >= 0 else "贷"

                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.metric("本期借方合计", f"{td:,.2f} 元")
                    with col2:
                        st.metric("本期贷方合计", f"{tc:,.2f} 元")
                    with col3:
                        st.metric("期末余额", f"{abs(ending):,.2f} 元（{ending_dir}方）")

                    # --- 导出 Excel ---
                    def export_ledger_excel(ledger_data):
                        """导出明细账为格式化 Excel"""
                        from openpyxl import Workbook
                        from openpyxl.styles import Font, Alignment, Border, Side
                        import io

                        wb = Workbook()
                        ws = wb.active
                        ws.title = ledger_data["account_name"][:31]

                        _thin = Side(style='thin')
                        _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
                        _title_font = Font(name='Arial', size=14, bold=True)
                        _hdr_font = Font(name='Arial', size=10, bold=True)
                        _data_font = Font(name='Arial', size=10)
                        _center = Alignment(horizontal='center', vertical='center')
                        _left = Alignment(horizontal='left', vertical='center')
                        _right = Alignment(horizontal='right', vertical='center')

                        for col, w in [('A', 12), ('B', 15), ('C', 25),
                                       ('D', 14), ('E', 14), ('F', 6), ('G', 16)]:
                            ws.column_dimensions[col].width = w

                        # 标题
                        ws.merge_cells('A1:G1')
                        ws['A1'] = f"{ledger_data['account_code']} {ledger_data['account_name']} 明细账"
                        ws['A1'].font = _title_font
                        ws['A1'].alignment = _center
                        ws.row_dimensions[1].height = 28

                        # 期初余额行
                        op = ledger_data["opening_balance"]
                        op_dir = "借" if op >= 0 else "贷"
                        ws.merge_cells('A2:G2')
                        ws['A2'] = f"期初余额：{abs(op):,.2f} 元（{op_dir}方）"
                        ws['A2'].font = _data_font
                        ws['A2'].alignment = _left
                        ws.row_dimensions[2].height = 20

                        # 表头
                        hdrs = ['日期', '凭证编号', '摘要', '借方金额', '贷方金额', '方向', '余额']
                        for i, h in enumerate(hdrs, 1):
                            c = ws.cell(row=3, column=i, value=h)
                            c.font = _hdr_font
                            c.alignment = _center
                            c.border = _border
                        ws.row_dimensions[3].height = 20

                        # 数据行
                        r = 4
                        for e in ledger_data["entries"]:
                            ws.row_dimensions[r].height = 20
                            vals = [
                                e["date"],
                                e["voucher_number"],
                                e["summary"],
                                f"{e['debit']:,.2f}" if e["debit"] else "",
                                f"{e['credit']:,.2f}" if e["credit"] else "",
                                e["direction_label"],
                                f"{abs(e['balance']):,.2f}",
                            ]
                            for ci, val in enumerate(vals, 1):
                                c = ws.cell(row=r, column=ci, value=val if val else None)
                                c.font = _data_font
                                c.border = _border
                                c.number_format = '@'
                                if ci == 3:
                                    c.alignment = _left
                                elif ci in (1, 2, 6):
                                    c.alignment = _center
                                else:
                                    c.alignment = _right
                            r += 1

                        # 合计行
                        ws.row_dimensions[r].height = 22
                        end_dir = "借" if ledger_data["ending_balance"] >= 0 else "贷"
                        totals = [
                            '本期合计及余额',
                            '',
                            '',
                            f"{ledger_data['total_debit']:,.2f}",
                            f"{ledger_data['total_credit']:,.2f}",
                            end_dir,
                            f"{abs(ledger_data['ending_balance']):,.2f}",
                        ]
                        for ci, val in enumerate(totals, 1):
                            c = ws.cell(row=r, column=ci, value=val if val else None)
                            c.font = _hdr_font
                            c.border = _border
                            c.number_format = '@'
                            if ci == 3:
                                c.alignment = _left
                            elif ci in (1, 2, 6):
                                c.alignment = _center
                            else:
                                c.alignment = _right

                        output = io.BytesIO()
                        wb.save(output)
                        output.seek(0)
                        return output

                    st.download_button(
                        label="📥 下载明细账（Excel）",
                        data=export_ledger_excel(ledger),
                        file_name=f"明细账_{ledger['account_code']}_{ledger['account_name']}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )

    # --- 科目管理 ---
    with sub5:
        st.subheader("科目管理")
        st.caption("自定义二级科目，例如：银行存款-工商银行、销售费用-业务招待费等。二级科目的编码自动在一级科目后追加两位。")

        # --- 添加二级科目 ---
        st.markdown("#### 添加二级科目")
        cols_add = st.columns([3, 3, 2, 2])

        with cols_add[0]:
            parent_options = [f"{a['code']} {a['name']}" for a in ACCOUNT_CHART]
            selected_parent = st.selectbox(
                "选择一级科目", parent_options,
                key="custom_acc_parent", label_visibility="collapsed",
                placeholder="搜索一级科目...",
            )

        parent_code = selected_parent.split(" ")[0]
        parent_name = selected_parent.split(" ", 1)[1]
        parent_info = ACCOUNT_MAP.get(parent_code, {})
        parent_direction = parent_info.get("direction", "借")
        parent_category = parent_info.get("category", "")
        next_sc = get_next_sub_code(parent_code)
        preview_full_code = f"{parent_code}{next_sc}"

        with cols_add[1]:
            sub_name_input = st.text_input(
                "二级科目名称", key="custom_acc_subname",
                label_visibility="collapsed", placeholder="如：工商银行、业务招待费",
            )

        with cols_add[2]:
            tax_options = list(TAX_SPECIAL_FLAGS.keys())
            tax_labels = [TAX_SPECIAL_FLAGS[k] for k in tax_options]
            selected_tax_idx = st.selectbox(
                "税务标记", range(len(tax_options)),
                format_func=lambda i: tax_labels[i],
                key="custom_acc_tax", label_visibility="collapsed",
            )
            selected_tax_flag = tax_options[selected_tax_idx]

        with cols_add[3]:
            add_clicked = st.button("➕ 添加", key="custom_acc_add_btn", use_container_width=True)

        # 预览信息
        color_dir = "#1976d2" if parent_direction == "借" else "#e53935"
        st.caption(
            f"预览：编码 **{preview_full_code}** ｜ "
            f"名称：{parent_name}-{sub_name_input or '（待输入）'} ｜ "
            f"方向：<span style='color:{color_dir};font-weight:600'>{parent_direction}方</span> ｜ "
            f"分类：{parent_category}"
            + (f" ｜ <span style='color:#e65100;font-weight:600'>⚠️ {selected_tax_flag}</span>" if selected_tax_flag else ""),
            unsafe_allow_html=True,
        )

        if add_clicked:
            if sub_name_input.strip():
                full_code, full_name = add_custom_account(
                    parent_code, parent_name, sub_name_input.strip(),
                    parent_category, parent_direction, selected_tax_flag,
                )
                st.success(f"✅ 已添加：{full_code} {full_name}")
                st.rerun()
            else:
                st.warning("请输入二级科目名称")

        st.markdown("---")

        # --- 已添加的二级科目列表 ---
        st.markdown("#### 已有二级科目")
        df_custom = get_custom_accounts()

        if df_custom.empty:
            st.info("💡 尚未添加任何自定义二级科目。在上方选择一级科目并输入名称后点击「添加」。")
        else:
            # 按一级科目分组显示
            for p_code in df_custom["parent_code"].unique():
                sub_df = df_custom[df_custom["parent_code"] == p_code].sort_values("sub_code")
                p_name = sub_df.iloc[0]["parent_name"]
                direction = sub_df.iloc[0]["direction"]
                color = "#1976d2" if direction == "借" else "#e53935"

                st.markdown(
                    f'<div style="font-weight:600; font-size:14px; margin-top:12px; margin-bottom:4px;">'
                    f'{p_code} {p_name}'
                    f' <span style="color:{color}; font-size:12px;">（{direction}方）</span>'
                    f'</div>',
                    unsafe_allow_html=True,
                )

                for _, row in sub_df.iterrows():
                    cols_row = st.columns([1, 2, 2, 1])
                    with cols_row[0]:
                        st.code(row["full_code"])
                    with cols_row[1]:
                        st.write(row["full_name"])
                    with cols_row[2]:
                        if row.get("tax_flag"):
                            st.markdown(
                                f'<span style="color:#e65100; font-size:12px;">⚠️ {row["tax_flag"]}</span>',
                                unsafe_allow_html=True,
                            )
                        else:
                            st.write("—")
                    with cols_row[3]:
                        if st.button("🗑", key=f"del_custom_{row['full_code']}",
                                     help=f"删除 {row['full_name']}"):
                            delete_custom_account(row["full_code"])
                            st.rerun()

        # --- 税务说明 ---
        st.markdown("---")
        st.markdown("#### 税务扣除规则说明")
        for flag, desc in TAX_SPECIAL_FLAGS.items():
            if flag:
                st.markdown(f"- **{flag}**：{desc}")

    # --- 期末结转 ---
    with sub6:
        st.subheader("🔄 期末自动结转")
        st.caption(
            "月末将所有损益类科目（收入、费用）余额结转至「本年利润」，结转后损益类科目余额为零。\n\n"
            "年末额外将「本年利润」余额转入「利润分配-未分配利润」，并可提取法定盈余公积。"
        )

        # 结转日期 & 类型选择
        cf_cols = st.columns([2, 2, 2])
        with cf_cols[0]:
            cf_date = st.date_input(
                "结转日期",
                value=datetime.now(),
                help="月末结转通常选当月最后一天，年末结转选12月31日",
            )
        with cf_cols[1]:
            cf_type = st.radio(
                "结转类型",
                ["月末结转（损益→本年利润）", "年末结转（含利润分配）"],
                label_visibility="collapsed",
            )
        with cf_cols[2]:
            is_year_end = "年末" in cf_type
            extract_reserve = False
            reserve_rate = 0.10
            if is_year_end:
                extract_reserve = st.checkbox(
                    "提取法定盈余公积（10%）",
                    value=True,
                    help="按净利润的 10% 提取，累计达注册资本 50% 后可不再提取",
                )
                if extract_reserve:
                    reserve_rate = st.number_input(
                        "计提比例", min_value=0.0, max_value=1.0,
                        value=0.10, step=0.01, format="%.2f",
                    )

        st.markdown("---")

        # 预览按钮
        if st.button("🔍 生成结转预览", type="secondary", use_container_width=True):
            st.session_state["_cf_preview_data"] = get_carryforward_preview()

        # 显示预览数据
        preview_data = st.session_state.get("_cf_preview_data")
        if preview_data:
            st.markdown("### 📋 结转预览")

            # 收入类科目
            if preview_data["income_items"]:
                st.markdown("#### 一、收入类科目（结转至本年利润·贷方）")
                inc_rows = []
                for item in preview_data["income_items"]:
                    inc_rows.append({
                        "科目编码": item["code"],
                        "科目名称": item["name"],
                        "期末余额": fmt_money(abs(item["balance"])),
                        "余额方向": item["direction"],
                        "结转方向": "借方" if item["balance"] > 0 else "贷方",
                        "结转金额": fmt_money(abs(item["balance"])),
                    })
                st.dataframe(pd.DataFrame(inc_rows), use_container_width=True, hide_index=True)
            else:
                st.info("收入类科目余额均为零，无需结转。")

            # 费用类科目
            if preview_data["expense_items"]:
                st.markdown("#### 二、费用类科目（结转至本年利润·借方）")
                exp_rows = []
                for item in preview_data["expense_items"]:
                    exp_rows.append({
                        "科目编码": item["code"],
                        "科目名称": item["name"],
                        "期末余额": fmt_money(abs(item["balance"])),
                        "余额方向": item["direction"],
                        "结转方向": "贷方" if item["balance"] > 0 else "借方",
                        "结转金额": fmt_money(abs(item["balance"])),
                    })
                st.dataframe(pd.DataFrame(exp_rows), use_container_width=True, hide_index=True)
            else:
                st.info("费用类科目余额均为零，无需结转。")

            # 汇总
            st.markdown("---")
            sum_cols = st.columns(3)
            with sum_cols[0]:
                st.metric("收入合计", fmt_money(preview_data["total_income"]))
            with sum_cols[1]:
                st.metric("费用合计", fmt_money(preview_data["total_expense"]))
            with sum_cols[2]:
                net = preview_data["net_profit"]
                st.metric(
                    "本期净利润（估）",
                    fmt_money(abs(net)),
                    delta=f"{'盈利' if net > 0 else '亏损'}" if net != 0 else "持平",
                )

            # 本年利润当前余额
            pb = preview_data["profit_balance"]
            if abs(pb) > 0.01:
                st.info(
                    f"📌 「本年利润（4103）」当前余额：{fmt_money(abs(pb))} 元"
                    f"（{'贷方' if pb > 0 else '借方'}）"
                )

            # 年末结转额外信息
            if is_year_end and net > 0 and extract_reserve:
                reserve_amt = round(net * reserve_rate, 2)
                st.warning(
                    f"📦 年末将额外执行：\n"
                    f"- 将本年利润余额转入「利润分配-未分配利润」\n"
                    f"- 提取法定盈余公积：{fmt_money(reserve_amt)} 元"
                    f"（净利润 {fmt_money(net)} × {reserve_rate:.0%}）"
                )

            # 执行按钮
            st.markdown("---")
            btn_label = "年末结转（含利润分配）" if is_year_end else "执行月末结转"
            if st.button(f"✅ 确认执行{btn_label}", type="primary", use_container_width=True):
                with st.spinner("正在执行结转，请稍候..."):
                    if is_year_end:
                        result = execute_yearly_carryforward(
                            cf_date,
                            extract_legal_reserve=extract_reserve,
                            reserve_rate=reserve_rate,
                        )
                        v_nums = result["voucher_numbers"]
                        net_profit = result["net_profit"]
                        reserve = result["reserve_amount"]
                    else:
                        v_nums = execute_monthly_carryforward(cf_date)
                        net_profit = preview_data["net_profit"]
                        reserve = 0

                    st.session_state["_cf_result"] = {
                        "vouchers": v_nums,
                        "net_profit": net_profit,
                        "reserve": reserve,
                        "is_year_end": is_year_end,
                    }
                    # 清除预览
                    st.session_state.pop("_cf_preview_data", None)
                    st.rerun()

        # 显示执行结果
        cf_result = st.session_state.get("_cf_result")
        if cf_result:
            st.success("✅ 期末结转执行成功！")
            st.markdown("### 📝 生成的结转凭证")

            for vn in cf_result["vouchers"]:
                v_df = get_all_vouchers()
                v_rows = v_df[v_df["voucher_number"] == vn]
                if not v_rows.empty:
                    st.markdown(f"**凭证编号：{vn}** — {v_rows.iloc[0]['summary']}")
                    display_df = v_rows[[
                        "account_code", "account_name", "debit_amount", "credit_amount"
                    ]].copy()
                    display_df.columns = ["科目编码", "科目名称", "借方金额", "贷方金额"]
                    display_df["借方金额"] = display_df["借方金额"].apply(fmt_money)
                    display_df["贷方金额"] = display_df["贷方金额"].apply(fmt_money)
                    st.dataframe(display_df, use_container_width=True, hide_index=True)

            # 净利润汇总
            net = cf_result["net_profit"]
            st.markdown("---")
            res_cols = st.columns(2)
            with res_cols[0]:
                st.metric("结转后净利润", fmt_money(abs(net)),
                          delta=f"{'盈利' if net > 0 else '亏损'}")
            with res_cols[1]:
                if cf_result["is_year_end"] and cf_result["reserve"] > 0:
                    st.metric("提取法定盈余公积", fmt_money(cf_result["reserve"]))
                else:
                    st.metric("损益类科目余额", "0.00 元", delta="已全部清零")

            st.info("💡 结转后可在「凭证查询」中查看结转凭证，在「报表」中查看更新后的利润表和资产负债表。")

            if st.button("清除结果"):
                st.session_state.pop("_cf_result", None)
                st.rerun()
        elif not preview_data:
            st.markdown("---")
            st.info("👆 点击「生成结转预览」查看需要结转的科目及金额。")

            with st.expander("📖 期末结转说明"):
                st.markdown("""
#### 月末结转流程
1. **结转收入**：将主营业务收入、其他业务收入、投资收益等收入类科目余额结转至「本年利润」
2. **结转费用**：将主营业务成本、销售费用、管理费用、财务费用等费用类科目余额结转至「本年利润」
3. 结转后所有损益类科目余额为零，「本年利润」反映当期净利润/亏损

#### 年末结转流程（在月末结转基础上）
4. **结转本年利润**：将「本年利润」全年累计余额转入「利润分配-未分配利润」
5. **提取法定盈余公积**：按净利润的 10% 提取（累计达注册资本 50% 后可不再提取）

#### 注意事项
- 结转会自动生成正式凭证，可在「凭证查询」中查看
- 月末结转每月执行一次，年末结转仅在 12 月末执行
- 执行前请确保当月所有日常凭证已录入完毕
""")


st.divider()


# ============================================================
# 【💰 财务会计】模块二：报表（财务报表）
# ============================================================
with tab2:
    st.header("📊 财务报表")
    st.caption("资产负债表 · 利润表 · 现金流量表 · 财务可视化分析 ｜ 自动生成，标准格式")

    # --- 公司信息输入 ---
    col_c1, col_c2, col_c3 = st.columns(3)
    with col_c1:
        company_name = st.text_input("编制单位", value=st.session_state.get("current_company_name", ""), placeholder="填写公司名称")
    with col_c2:
        report_date = st.date_input("报表出具日期", value=None)
    with col_c3:
        currency_unit = st.text_input("货币单位", value="元")

    report_date_str = report_date.strftime("%Y-%m-%d") if report_date else ""

    # --- 报表选择 ---
    report_type = st.radio("选择报表", ["资产负债表", "利润表", "现金流量表", "📊 财务可视化分析"], horizontal=True)

    # ================================================================
    # 资产负债表
    # ================================================================
    if report_type == "资产负债表":
        st.markdown("---")

        # --- 第1行：标题 ---
        st.markdown(
            '<div style="text-align:center; font-size:22px; font-weight:700; '
            'padding:12px 0;">资产负债表</div>',
            unsafe_allow_html=True,
        )

        # --- 第2行：公司名称 / 报表日期 / 货币单位 ---
        st.markdown(
            f'<div style="display:flex; justify-content:space-between; '
            f'font-size:13px; color:#555; padding:4px 12px 8px 12px;">'
            f'<span>编制单位：{company_name or "（未填写）"}</span>'
            f'<span>报表日期：{report_date_str or "（未选择）"}</span>'
            f'<span>货币单位：{currency_unit}</span>'
            f'</div>',
            unsafe_allow_html=True,
        )

        # --- 表格主体 ---
        # 获取所有科目余额
        asset_accounts = [a for a in ACCOUNT_CHART if a["category"] == "资产"]
        liab_accounts = [a for a in ACCOUNT_CHART if a["category"] == "负债"]
        equity_accounts = [a for a in ACCOUNT_CHART if a["category"] == "权益"]

        # --- 标准会计格式：左方资产，右方负债+所有者权益，底部合计相等 ---

        def fmt_amt(v):
            """格式化金额：0 显示空"""
            if abs(v) < 0.005:
                return ""
            return f"{v:,.2f}"

        def _build_bs_rows(accounts):
            """构建资产负债表数据行"""
            rows = []
            for idx, acc in enumerate(accounts, 1):
                od, oc, pd, pc, ending = calc_account_balance(acc["code"])
                opening = od if acc["direction"] == "借" else oc
                rows.append({
                    "code": acc["code"],
                    "name": acc["name"],
                    "ending": ending,
                    "opening": opening,
                })
            return rows

        asset_rows = _build_bs_rows(asset_accounts)
        liab_rows = _build_bs_rows(liab_accounts)
        equity_rows = _build_bs_rows(equity_accounts)

        asset_ending_total = sum(r["ending"] for r in asset_rows)
        asset_opening_total = sum(r["opening"] for r in asset_rows)
        liab_ending_total = sum(r["ending"] for r in liab_rows)
        liab_opening_total = sum(r["opening"] for r in liab_rows)
        equity_ending_total = sum(r["ending"] for r in equity_rows)
        equity_opening_total = sum(r["opening"] for r in equity_rows)

        # 右方 = 负债 + 所有者权益
        right_ending_total = liab_ending_total + equity_ending_total
        right_opening_total = liab_opening_total + equity_opening_total

        # 构建左方行列表（资产 + 合计行）
        left_rows = []
        for r in asset_rows:
            left_rows.append({"code": r["code"], "name": r["name"],
                              "ending": r["ending"], "opening": r["opening"],
                              "is_total": False})
        left_rows.append({"code": "", "name": "资产总计",
                          "ending": asset_ending_total, "opening": asset_opening_total,
                          "is_total": True})

        # 构建右方行列表（负债 + 负债合计 + 权益 + 权益合计 + 总合计）
        right_rows = []
        for r in liab_rows:
            right_rows.append({"code": r["code"], "name": r["name"],
                               "ending": r["ending"], "opening": r["opening"],
                               "is_total": False, "is_section": False})
        right_rows.append({"code": "", "name": "负债合计",
                           "ending": liab_ending_total, "opening": liab_opening_total,
                           "is_total": False, "is_section": True})
        for r in equity_rows:
            right_rows.append({"code": r["code"], "name": r["name"],
                               "ending": r["ending"], "opening": r["opening"],
                               "is_total": False, "is_section": False})
        right_rows.append({"code": "", "name": "所有者权益合计",
                           "ending": equity_ending_total, "opening": equity_opening_total,
                           "is_total": False, "is_section": True})
        right_rows.append({"code": "", "name": "负债和所有者权益总计",
                           "ending": right_ending_total, "opening": right_opening_total,
                           "is_total": True, "is_section": False})

        # 补齐行数（左右对齐）
        max_rows = max(len(left_rows), len(right_rows))
        while len(left_rows) < max_rows:
            left_rows.append({"code": "", "name": "", "ending": 0, "opening": 0, "is_total": False})
        while len(right_rows) < max_rows:
            right_rows.append({"code": "", "name": "", "ending": 0, "opening": 0,
                               "is_total": False, "is_section": False})

        # --- 渲染 HTML 表格（标准左右对照格式） ---
        html_rows = ""
        for i in range(max_rows):
            lr = left_rows[i]
            rr = right_rows[i]

            # 左方行样式
            l_style = 'font-weight:700; border-top:2px solid #333;' if lr.get("is_total") else ''
            l_ending = fmt_amt(lr["ending"]) if lr["ending"] else ""
            l_opening = fmt_amt(lr["opening"]) if lr["opening"] else ""

            # 右方行样式
            r_style = ""
            if rr.get("is_total"):
                r_style = 'font-weight:700; border-top:2px solid #333;'
            elif rr.get("is_section"):
                r_style = 'font-weight:600; border-top:1px solid #999;'
            r_ending = fmt_amt(rr["ending"]) if rr["ending"] else ""
            r_opening = fmt_amt(rr["opening"]) if rr["opening"] else ""

            html_rows += f"""
            <tr>
                <td style="text-align:center; {l_style}">{lr['code']}</td>
                <td style="text-align:left; {l_style}">{lr['name']}</td>
                <td style="text-align:right; {l_style}">{l_ending}</td>
                <td style="text-align:right; {l_style}">{l_opening}</td>
                <td style="text-align:center; {r_style}">{rr['code']}</td>
                <td style="text-align:left; {r_style}">{rr['name']}</td>
                <td style="text-align:right; {r_style}">{r_ending}</td>
                <td style="text-align:right; {r_style}">{r_opening}</td>
            </tr>"""

        st.markdown(f"""
        <table style="width:100%; border-collapse:collapse; font-size:13px; margin-top:8px;">
            <thead>
                <tr style="background:#f0f4f8;">
                    <th colspan="4" style="text-align:center; padding:8px; border:1px solid #ccc; font-size:15px;">资 产</th>
                    <th colspan="4" style="text-align:center; padding:8px; border:1px solid #ccc; font-size:15px;">负债和所有者权益</th>
                </tr>
                <tr style="background:#e8eef5;">
                    <th style="padding:4px 6px; border:1px solid #ccc; width:6%;">编码</th>
                    <th style="padding:4px 6px; border:1px solid #ccc; text-align:left; width:20%;">科目名称</th>
                    <th style="padding:4px 6px; border:1px solid #ccc; text-align:right; width:12%;">期末余额</th>
                    <th style="padding:4px 6px; border:1px solid #ccc; text-align:right; width:12%;">年初余额</th>
                    <th style="padding:4px 6px; border:1px solid #ccc; width:6%;">编码</th>
                    <th style="padding:4px 6px; border:1px solid #ccc; text-align:left; width:20%;">科目名称</th>
                    <th style="padding:4px 6px; border:1px solid #ccc; text-align:right; width:12%;">期末余额</th>
                    <th style="padding:4px 6px; border:1px solid #ccc; text-align:right; width:12%;">年初余额</th>
                </tr>
            </thead>
            <tbody>
                {html_rows}
            </tbody>
        </table>
        """, unsafe_allow_html=True)

        # --- 校验：资产 = 负债 + 所有者权益 ---
        diff_ending = asset_ending_total - right_ending_total
        diff_opening = asset_opening_total - right_opening_total

        col_chk1, col_chk2 = st.columns(2)
        with col_chk1:
            if abs(diff_ending) < 0.01:
                st.success(f"✅ 期末平衡：资产 {fmt_amt(asset_ending_total)} = 负债+权益 {fmt_amt(right_ending_total)}")
            else:
                st.error(f"❌ 期末不平衡！差额：{fmt_amt(diff_ending)}")
        with col_chk2:
            if abs(diff_opening) < 0.01:
                st.success(f"✅ 年初平衡：资产 {fmt_amt(asset_opening_total)} = 负债+权益 {fmt_amt(right_opening_total)}")
            else:
                st.error(f"❌ 年初不平衡！差额：{fmt_amt(diff_opening)}")

        # --- 导出 Excel（左右对照格式） ---
        def export_bs_excel():
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side
            import io

            wb = Workbook()
            ws = wb.active
            ws.title = "资产负债表"

            _thin = Side(style='thin')
            _medium = Side(style='medium')
            _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
            _border_top = Border(left=_thin, right=_thin, top=_medium, bottom=_thin)
            _center = Alignment(horizontal='center', vertical='center')
            _right = Alignment(horizontal='right', vertical='center')
            _left = Alignment(horizontal='left', vertical='center')
            _bold = Font(bold=True)
            _title_font = Font(bold=True, size=16)
            _info_font = Font(size=10)
            _header_font = Font(bold=True, size=10)

            # 列宽
            widths = {'A': 8, 'B': 22, 'C': 16, 'D': 16, 'E': 8, 'F': 22, 'G': 16, 'H': 16}
            for col, w in widths.items():
                ws.column_dimensions[col].width = w

            # 第1行：标题
            ws.merge_cells('A1:H1')
            ws['A1'] = '资产负债表'
            ws['A1'].font = _title_font
            ws['A1'].alignment = _center
            ws.row_dimensions[1].height = 30

            # 第2行：公司信息
            ws['A2'] = f'编制单位：{company_name or ""}'
            ws['A2'].font = _info_font
            ws.merge_cells('C2:E2')
            ws['C2'] = f'报表日期：{report_date_str}'
            ws['C2'].font = _info_font
            ws['C2'].alignment = _center
            ws.merge_cells('F2:H2')
            ws['F2'] = f'货币单位：{currency_unit}'
            ws['F2'].font = _info_font
            ws['F2'].alignment = _right
            ws.row_dimensions[2].height = 18

            # 第3行：大类标题（资产 | 负债和所有者权益）
            ws.merge_cells('A3:D3')
            ws['A3'] = '资  产'
            ws['A3'].font = _bold
            ws['A3'].alignment = _center
            ws['A3'].border = _border
            ws.merge_cells('E3:H3')
            ws['E3'] = '负债和所有者权益'
            ws['E3'].font = _bold
            ws['E3'].alignment = _center
            ws['E3'].border = _border
            ws.row_dimensions[3].height = 22

            # 第4行：列标题
            headers = ["编码", "科目名称", "期末余额", "年初余额",
                       "编码", "科目名称", "期末余额", "年初余额"]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col, value=h)
                cell.font = _header_font
                cell.alignment = _center
                cell.border = _border
            ws.row_dimensions[4].height = 20

            # 数据行
            row_idx = 5
            for i in range(max_rows):
                lr = left_rows[i]
                rr = right_rows[i]

                # 左方
                l_vals = [lr["code"], lr["name"],
                          fmt_amt(lr["ending"]) if lr["ending"] else "",
                          fmt_amt(lr["opening"]) if lr["opening"] else ""]
                l_is_total = lr.get("is_total", False)
                for col, val in enumerate(l_vals, 1):
                    cell = ws.cell(row=row_idx, column=col, value=val if val else None)
                    cell.border = _border_top if l_is_total else _border
                    cell.font = _bold if l_is_total else Font(size=10)
                    if col == 1:
                        cell.alignment = _center
                    elif col == 2:
                        cell.alignment = _left
                    else:
                        cell.alignment = _right

                # 右方
                r_vals = [rr["code"], rr["name"],
                          fmt_amt(rr["ending"]) if rr["ending"] else "",
                          fmt_amt(rr["opening"]) if rr["opening"] else ""]
                r_is_total = rr.get("is_total", False)
                r_is_section = rr.get("is_section", False)
                for col, val in enumerate(r_vals, 5):
                    cell = ws.cell(row=row_idx, column=col, value=val if val else None)
                    cell.border = _border_top if (r_is_total or r_is_section) else _border
                    if r_is_total or r_is_section:
                        cell.font = _bold if r_is_total else Font(size=10, bold=True)
                    else:
                        cell.font = Font(size=10)
                    if col == 5:
                        cell.alignment = _center
                    elif col == 6:
                        cell.alignment = _left
                    else:
                        cell.alignment = _right

                ws.row_dimensions[row_idx].height = 18
                row_idx += 1

            # 平衡校验行
            row_idx += 1
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=4)
            if abs(diff_ending) < 0.01:
                ws.cell(row=row_idx, column=1, value=f"✅ 期末平衡：资产 = 负债+权益 = {fmt_amt(asset_ending_total)}")
            else:
                ws.cell(row=row_idx, column=1, value=f"❌ 期末不平衡！差额：{fmt_amt(diff_ending)}")
            ws.cell(row=row_idx, column=1).font = Font(bold=True, size=10, color="2e7d32" if abs(diff_ending) < 0.01 else "c62828")

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return output

        st.download_button(
            label="📥 下载资产负债表（Excel）",
            data=export_bs_excel(),
            file_name=f"资产负债表_{report_date_str or '未定'}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    # ================================================================
    # 利润表
    # ================================================================
    elif report_type == "利润表":
        st.markdown("---")

        # --- 第1行：标题 ---
        st.markdown(
            '<div style="text-align:center; font-size:22px; font-weight:700; '
            'padding:12px 0;">利润表</div>',
            unsafe_allow_html=True,
        )

        # --- 第2行：公司名称 / 报表日期 / 货币单位 ---
        st.markdown(
            f'<div style="display:flex; justify-content:space-between; '
            f'font-size:13px; color:#555; padding:4px 12px 8px 12px;">'
            f'<span>编制单位：{company_name or "（未填写）"}</span>'
            f'<span>报表日期：{report_date_str or "（未选择）"}</span>'
            f'<span>货币单位：{currency_unit}</span>'
            f'</div>',
            unsafe_allow_html=True,
        )

        # --- 金额计算辅助 ---
        def fmt_amt(v):
            """格式化金额：0 显示空"""
            if abs(v) < 0.005:
                return ""
            return f"{v:,.2f}"

        def get_ending(code):
            _, _, _, _, ending = calc_account_balance(code)
            return ending

        def get_ending_list(codes):
            return sum(get_ending(c) for c in codes)

        # --- 利润表行项目定义 ---
        # is_total: 需要计算的合计行；codes: 直接取科目余额；indent: 缩进层级
        IS_ITEMS = [
            # 序号, 项目, 科目代码列表, 是否合计行, 缩进
            {"no": 1,  "name": "营业收入",                          "codes": ["6001", "6051"], "indent": 0},
            {"no": 2,  "name": "营业成本",                          "codes": ["6401", "6402"], "indent": 0},
            {"no": 3,  "name": "税金及附加",                        "codes": ["6403"],         "indent": 0},
            {"no": "", "name": "    营业税",                        "codes": [],              "indent": 1},
            {"no": "", "name": "    城市维护建设税",                "codes": [],              "indent": 1},
            {"no": "", "name": "    资源税",                        "codes": [],              "indent": 1},
            {"no": "", "name": "    土地增值税",                    "codes": [],              "indent": 1},
            {"no": "", "name": "    城镇土地使用税、房产税、车船税、印花税", "codes": [],     "indent": 1},
            {"no": "", "name": "    教育费附加、矿产资源补偿费、排污费",   "codes": [],     "indent": 1},
            {"no": 4,  "name": "销售费用",                          "codes": ["6601"],        "indent": 0},
            {"no": "", "name": "    广告费和业务宣传费",            "codes": [],              "indent": 1},
            {"no": 5,  "name": "管理费用",                          "codes": ["6602"],        "indent": 0},
            {"no": "", "name": "    业务招待费",                    "codes": [],              "indent": 1},
            {"no": "", "name": "    研究费用",                      "codes": ["6405"],        "indent": 1},
            {"no": 6,  "name": "财务费用",                          "codes": ["6603"],        "indent": 0},
            {"no": 7,  "name": "投资收益",                          "codes": ["6111"],        "indent": 0},
            {"no": 8,  "name": "营业利润",                          "is_calc": "operating_profit", "indent": 0},
            {"no": 9,  "name": "营业外收入",                       "codes": ["6301"],        "indent": 0},
            {"no": 10, "name": "营业外支出",                        "codes": ["6701"],        "indent": 0},
            {"no": "", "name": "    无法收回的长期债券投资损失",    "codes": [],              "indent": 1},
            {"no": "", "name": "    无法收回的长期股权投资损失",    "codes": [],              "indent": 1},
            {"no": "", "name": "    自然灾害等不可抗力因素造成的损失", "codes": [],            "indent": 1},
            {"no": "", "name": "    税收滞纳金",                    "codes": [],              "indent": 1},
            {"no": 11, "name": "利润总额",                          "is_calc": "total_profit", "indent": 0},
            {"no": 12, "name": "所得税费用",                        "codes": ["6711"],        "indent": 0},
            {"no": 13, "name": "净利润",                            "is_calc": "net_profit",   "indent": 0},
        ]

        # --- 计算各项金额 ---
        # 收入类科目余额在贷方，费用类科目余额在借方
        # 利润表金额 = 收入 - 费用（费用取负值或绝对值减去）
        amounts = {}
        # 营业收入
        amounts["营业收入"] = get_ending_list(["6001", "6051"])
        # 营业成本
        amounts["营业成本"] = get_ending_list(["6401", "6402"])
        # 税金及附加
        amounts["税金及附加"] = get_ending("6403")
        # 销售费用
        amounts["销售费用"] = get_ending("6601")
        # 管理费用
        amounts["管理费用"] = get_ending("6602")
        # 财务费用
        amounts["财务费用"] = get_ending("6603")
        # 投资收益
        amounts["投资收益"] = get_ending("6111")
        # 营业利润 = 营业收入 - 营业成本 - 税金及附加 - 销售费用 - 管理费用 - 财务费用 + 投资收益
        amounts["营业利润"] = (amounts["营业收入"] - amounts["营业成本"] - amounts["税金及附加"]
                              - amounts["销售费用"] - amounts["管理费用"] - amounts["财务费用"]
                              + amounts["投资收益"])
        # 营业外收入
        amounts["营业外收入"] = get_ending("6301")
        # 营业外支出
        amounts["营业外支出"] = get_ending("6701")
        # 利润总额 = 营业利润 + 营业外收入 - 营业外支出
        amounts["利润总额"] = amounts["营业利润"] + amounts["营业外收入"] - amounts["营业外支出"]
        # 所得税费用
        amounts["所得税费用"] = get_ending("6711")
        # 净利润 = 利润总额 - 所得税费用
        amounts["净利润"] = amounts["利润总额"] - amounts["所得税费用"]

        # --- 构建表格数据 ---
        rows = []
        for item in IS_ITEMS:
            name_display = item["name"]
            if item.get("is_calc"):
                val = amounts.get(item["is_calc"], 0)
            else:
                val = amounts.get(item["name"], 0) if item["codes"] else 0
            rows.append({
                "序号": item["no"],
                "项目": name_display,
                "本月金额": fmt_amt(val),
                "本年累积金额": fmt_amt(val),
            })

        is_df = pd.DataFrame(rows, columns=["序号", "项目", "本月金额", "本年累积金额"])
        st.dataframe(is_df, use_container_width=True, hide_index=True)

        # --- 关键指标高亮 ---
        col_p1, col_p2, col_p3 = st.columns(3)
        with col_p1:
            st.metric("营业利润", f"{amounts['营业利润']:,.2f}")
        with col_p2:
            st.metric("利润总额", f"{amounts['利润总额']:,.2f}")
        with col_p3:
            st.metric("净利润", f"{amounts['净利润']:,.2f}")

        # --- 导出 Excel ---
        def export_is_excel():
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side
            import io

            wb = Workbook()
            ws = wb.active
            ws.title = "利润表"

            _thin = Side(style='thin')
            _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
            _center = Alignment(horizontal='center', vertical='center')
            _right = Alignment(horizontal='right', vertical='center')
            _left = Alignment(horizontal='left', vertical='center')
            _bold = Font(bold=True)
            _title_font = Font(bold=True, size=16)
            _info_font = Font(size=11)

            # 第1行：标题
            ws.merge_cells('A1:D1')
            ws['A1'] = '利润表'
            ws['A1'].font = _title_font
            ws['A1'].alignment = _center
            ws.row_dimensions[1].height = 30

            # 第2行：公司名称 / 报表日期 / 货币单位
            ws['A2'] = f'编制单位：{company_name or ""}'
            ws['A2'].font = _info_font
            ws.merge_cells('B2:C2')
            ws['B2'] = f'报表日期：{report_date_str}'
            ws['B2'].font = _info_font
            ws['B2'].alignment = _center
            ws['D2'] = f'货币单位：{currency_unit}'
            ws['D2'].font = _info_font
            ws['D2'].alignment = _right
            ws.row_dimensions[2].height = 20

            # 第3行：表头
            headers = ["序号", "项目", "本月金额", "本年累积金额"]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col, value=h)
                cell.font = _bold
                cell.alignment = _center
                cell.border = _border
            ws.row_dimensions[3].height = 22

            # 数据行
            for ridx, row in enumerate(rows, 4):
                for col, key in enumerate(["序号", "项目", "本月金额", "本年累积金额"], 1):
                    val = row[key]
                    cell = ws.cell(row=ridx, column=col, value=val)
                    cell.border = _border
                    if col == 1:
                        cell.alignment = _center
                    elif col == 2:
                        cell.alignment = _left
                    else:
                        cell.alignment = _right
                    # 营业利润、利润总额、净利润 加粗
                    if row["项目"] in ("营业利润", "利润总额", "净利润"):
                        cell.font = _bold

            # 列宽
            ws.column_dimensions['A'].width = 6
            ws.column_dimensions['B'].width = 42
            ws.column_dimensions['C'].width = 18
            ws.column_dimensions['D'].width = 18

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return output

        st.download_button(
            label="📥 下载利润表（Excel）",
            data=export_is_excel(),
            file_name=f"利润表_{report_date_str or '未定'}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    # ================================================================
    # 现金流量表
    # ================================================================
    elif report_type == "现金流量表":
        st.markdown("---")

        # --- 第1行：标题 ---
        st.markdown(
            '<div style="text-align:center; font-size:22px; font-weight:700; '
            'padding:12px 0;">现金流量表</div>',
            unsafe_allow_html=True,
        )

        # --- 第2行：公司名称 / 报表日期 / 货币单位 ---
        st.markdown(
            f'<div style="display:flex; justify-content:space-between; '
            f'font-size:13px; color:#555; padding:4px 12px 8px 12px;">'
            f'<span>编制单位：{company_name or "（未填写）"}</span>'
            f'<span>报表日期：{report_date_str or "（未选择）"}</span>'
            f'<span>货币单位：{currency_unit}</span>'
            f'</div>',
            unsafe_allow_html=True,
        )

        # --- 金额计算辅助 ---
        def fmt_amt(v):
            """格式化金额：0 显示空"""
            if abs(v) < 0.005:
                return ""
            return f"{v:,.2f}"

        def get_ending(code):
            _, _, _, _, ending = calc_account_balance(code)
            return ending

        # --- 现金类科目编码 ---
        CASH_CODES = {"1001", "1002", "1012"}

        # 投资活动涉及的对方科目
        INVEST_CODES = {
            "1601", "1602", "1603", "1604", "1605", "1606",
            "1701", "1702", "1703", "1801",
            "1501", "1502", "1503", "1504",
            "1511", "1512", "1521", "1101",
            "1641", "1642", "1651", "1652", "1621", "1622",
        }

        # 筹资活动涉及的对方科目
        FINANCE_CODES = {
            "2001", "2501", "2502", "2701",
            "4001", "4002", "4101", "4104",
            "2231", "2232",
        }

        # --- 从凭证中分析现金流量 ---
        vdf_all = get_all_vouchers()

        # 初始化各项目金额
        # 经营活动
        op_sales_cash = 0          # 销售产成品、商品、提供劳务收到的现金
        op_other_in = 0            # 收到其他与经营活动有关的现金
        op_purchase_cash = 0       # 购买原材料、商品、接受劳务支付的现金
        op_salary_cash = 0         # 支付的职工薪酬
        op_tax_cash = 0            # 支付的税费
        op_other_out = 0           # 支付其他与经营活动有关的现金

        # 投资活动
        inv_recover = 0            # 收回短期投资、长期债券投资和长期股权投资收到的现金
        inv_income = 0             # 取得投资收益收到的现金
        inv_dispose = 0            # 处置固定资产、无形资产和其他非流动资产收回的现金净额
        inv_invest_pay = 0         # 短期投资、长期债券投资和长期股权投资支付的现金
        inv_buy_asset = 0          # 购建固定资产、无形资产和其他非流动资产支付的现金

        # 筹资活动
        fin_loan_in = 0            # 取得借款收到的现金
        fin_invest_in = 0          # 吸收投资者投资收到的现金
        fin_loan_repay = 0         # 偿还借款本金支付的现金
        fin_interest_pay = 0       # 偿还借款利息支付的现金
        fin_profit_dist = 0        # 分配利润支付的现金

        if not vdf_all.empty:
            # 找出所有涉及现金科目的凭证
            cash_voucher_nums = vdf_all[
                vdf_all["account_code"].isin(CASH_CODES)
            ]["voucher_number"].unique()

            for vnum in cash_voucher_nums:
                vdf = vdf_all[vdf_all["voucher_number"] == vnum]
                cash_rows = vdf[vdf["account_code"].isin(CASH_CODES)]
                non_cash_rows = vdf[~vdf["account_code"].isin(CASH_CODES)]

                for _, crow in cash_rows.iterrows():
                    cash_in = crow["debit_amount"] or 0   # 现金借方 = 收到现金
                    cash_out = crow["credit_amount"] or 0 # 现金贷方 = 支付现金

                    if len(non_cash_rows) > 0:
                        counterpart_code = non_cash_rows.iloc[0]["account_code"]
                    else:
                        counterpart_code = ""

                    # 经营活动
                    if counterpart_code in ("6001", "6051"):
                        op_sales_cash += cash_in
                    elif counterpart_code in ("6401", "6402", "1403", "1405", "1123"):
                        op_purchase_cash += cash_out
                    elif counterpart_code == "2211":
                        op_salary_cash += cash_out
                    elif counterpart_code in ("2221", "6403"):
                        op_tax_cash += cash_out
                    elif counterpart_code in ("6601", "6602"):
                        op_other_out += cash_out
                    elif counterpart_code == "6603":
                        fin_interest_pay += cash_out
                    elif counterpart_code in ("6301", "1221", "2202", "2203"):
                        op_other_in += cash_in
                        op_other_out += cash_out
                    # 投资活动
                    elif counterpart_code in ("1101", "1501", "1503", "1504", "1511"):
                        if cash_in > 0:
                            inv_recover += cash_in
                        if cash_out > 0:
                            inv_invest_pay += cash_out
                    elif counterpart_code == "6111":
                        inv_income += cash_in
                    elif counterpart_code in ("1601", "1602", "1701", "1702"):
                        if cash_out > 0:
                            inv_buy_asset += cash_out
                        if cash_in > 0:
                            inv_dispose += cash_in
                    # 筹资活动
                    elif counterpart_code in ("2001", "2501", "2502"):
                        if cash_in > 0:
                            fin_loan_in += cash_in
                        if cash_out > 0:
                            fin_loan_repay += cash_out
                    elif counterpart_code == "4001":
                        fin_invest_in += cash_in
                    elif counterpart_code == "4101":
                        fin_profit_dist += cash_out
                    elif counterpart_code in ("2231", "2232"):
                        if cash_out > 0:
                            fin_profit_dist += cash_out
                    else:
                        op_other_in += cash_in
                        op_other_out += cash_out

        # 净额计算
        op_net = op_sales_cash + op_other_in - op_purchase_cash - op_salary_cash - op_tax_cash - op_other_out
        inv_net = inv_recover + inv_income + inv_dispose - inv_invest_pay - inv_buy_asset
        fin_net = fin_loan_in + fin_invest_in - fin_loan_repay - fin_interest_pay - fin_profit_dist
        net_increase = op_net + inv_net + fin_net

        # 期初/期末现金余额
        opening_cash = sum(get_opening_balance(c)[0] if get_opening_balance(c)[0] else get_opening_balance(c)[1] for c in CASH_CODES)
        ending_cash = opening_cash + net_increase

        # --- 行项目定义 ---
        CF_ITEMS = [
            # 序号, 项目, 金额
            {"no": "",  "name": "一、经营活动产生的现金流量",                                            "val": None,  "bold": True},
            {"no": 1,   "name": "  销售产成品、商品、提供劳务收到的现金",                                  "val": op_sales_cash},
            {"no": 2,   "name": "  收到其他与经营活动有关的现金",                                         "val": op_other_in},
            {"no": 3,   "name": "  购买原材料、商品、接受劳务支付的现金",                                 "val": op_purchase_cash},
            {"no": 4,   "name": "  支付的职工薪酬",                                                       "val": op_salary_cash},
            {"no": 5,   "name": "  支付的税费",                                                           "val": op_tax_cash},
            {"no": 6,   "name": "  支付其他与经营活动有关的现金",                                         "val": op_other_out},
            {"no": 7,   "name": "  经营活动产生的现金流量净额",                                            "val": op_net, "bold": True},
            {"no": "",  "name": "二、投资活动产生的现金流量",                                              "val": None,  "bold": True},
            {"no": 8,   "name": "  收回短期投资、长期债券投资和长期股权投资收到的现金",                    "val": inv_recover},
            {"no": 9,   "name": "  取得投资收益收到的现金",                                                "val": inv_income},
            {"no": 10,  "name": "  处置固定资产、无形资产和其他非流动资产收回的现金净额",                  "val": inv_dispose},
            {"no": 11,  "name": "  短期投资、长期债券投资和长期股权投资支付的现金",                        "val": inv_invest_pay},
            {"no": 12,  "name": "  购建固定资产、无形资产和其他非流动资产支付的现金",                      "val": inv_buy_asset},
            {"no": 13,  "name": "  投资活动产生的现金流量净额",                                            "val": inv_net, "bold": True},
            {"no": "",  "name": "三、筹资活动产生的现金流量",                                              "val": None,  "bold": True},
            {"no": 14,  "name": "  取得借款收到的现金",                                                    "val": fin_loan_in},
            {"no": 15,  "name": "  吸收投资者投资收到的现金",                                             "val": fin_invest_in},
            {"no": 16,  "name": "  偿还借款本金支付的现金",                                                "val": fin_loan_repay},
            {"no": 17,  "name": "  偿还借款利息支付的现金",                                                "val": fin_interest_pay},
            {"no": 18,  "name": "  分配利润支付的现金",                                                    "val": fin_profit_dist},
            {"no": 19,  "name": "  筹资活动产生的现金流量净额",                                            "val": fin_net, "bold": True},
            {"no": 20,  "name": "四、现金净增加额",                                                        "val": net_increase, "bold": True},
            {"no": 21,  "name": "  期初现金余额",                                                         "val": opening_cash},
            {"no": 22,  "name": "五、期末现金余额",                                                        "val": ending_cash, "bold": True},
        ]

        # --- 构建表格数据 ---
        rows = []
        for item in CF_ITEMS:
            val = item["val"]
            val_str = fmt_amt(val) if val is not None else ""
            rows.append({
                "序号": item["no"],
                "项目": item["name"],
                "本月金额": val_str,
                "本年累积金额": val_str,
            })

        cf_df = pd.DataFrame(rows, columns=["序号", "项目", "本月金额", "本年累积金额"])
        st.dataframe(cf_df, use_container_width=True, hide_index=True)

        # --- 关键指标高亮 ---
        col_c1, col_c2, col_c3, col_c4 = st.columns(4)
        with col_c1:
            st.metric("经营活动净额", f"{op_net:,.2f}")
        with col_c2:
            st.metric("投资活动净额", f"{inv_net:,.2f}")
        with col_c3:
            st.metric("筹资活动净额", f"{fin_net:,.2f}")
        with col_c4:
            st.metric("现金净增加额", f"{net_increase:,.2f}")

        # --- 导出 Excel ---
        def export_cf_excel():
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side
            import io

            wb = Workbook()
            ws = wb.active
            ws.title = "现金流量表"

            _thin = Side(style='thin')
            _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
            _center = Alignment(horizontal='center', vertical='center')
            _right = Alignment(horizontal='right', vertical='center')
            _left = Alignment(horizontal='left', vertical='center')
            _bold = Font(bold=True)
            _title_font = Font(bold=True, size=16)
            _info_font = Font(size=11)

            # 第1行：标题
            ws.merge_cells('A1:D1')
            ws['A1'] = '现金流量表'
            ws['A1'].font = _title_font
            ws['A1'].alignment = _center
            ws.row_dimensions[1].height = 30

            # 第2行：公司名称 / 报表日期 / 货币单位
            ws['A2'] = f'编制单位：{company_name or ""}'
            ws['A2'].font = _info_font
            ws.merge_cells('B2:C2')
            ws['B2'] = f'报表日期：{report_date_str}'
            ws['B2'].font = _info_font
            ws['B2'].alignment = _center
            ws['D2'] = f'货币单位：{currency_unit}'
            ws['D2'].font = _info_font
            ws['D2'].alignment = _right
            ws.row_dimensions[2].height = 20

            # 第3行：表头
            headers = ["序号", "项目", "本月金额", "本年累积金额"]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col, value=h)
                cell.font = _bold
                cell.alignment = _center
                cell.border = _border
            ws.row_dimensions[3].height = 22

            # 数据行
            for ridx, row in enumerate(rows, 4):
                is_bold = CF_ITEMS[ridx - 4].get("bold", False)
                for col, key in enumerate(["序号", "项目", "本月金额", "本年累积金额"], 1):
                    val = row[key]
                    cell = ws.cell(row=ridx, column=col, value=val)
                    cell.border = _border
                    if col == 1:
                        cell.alignment = _center
                    elif col == 2:
                        cell.alignment = _left
                    else:
                        cell.alignment = _right
                    if is_bold:
                        cell.font = _bold

            # 列宽
            ws.column_dimensions['A'].width = 6
            ws.column_dimensions['B'].width = 52
            ws.column_dimensions['C'].width = 18
            ws.column_dimensions['D'].width = 18

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return output

        st.download_button(
            label="📥 下载现金流量表（Excel）",
            data=export_cf_excel(),
            file_name=f"现金流量表_{report_date_str or '未定'}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    # ================================================================
    # 财务可视化分析
    # ================================================================
    elif report_type == "📊 财务可视化分析":
        import plotly.express as px
        import plotly.graph_objects as go
        from plotly.subplots import make_subplots

        st.markdown("---")
        st.markdown(
            '<div style="text-align:center; font-size:22px; font-weight:700; '
            'padding:12px 0;">📊 财务可视化分析</div>',
            unsafe_allow_html=True,
        )
        st.caption("基于当前账套数据自动生成可视化图表，涵盖资产结构、损益分析、费用构成、收支趋势等。")

        # 清除缓存确保数据最新
        calc_account_balance.clear()
        get_all_opening_balances.clear()
        get_all_vouchers.clear()

        # -------- 辅助函数：获取科目余额 --------
        def _get_balance(code):
            _, _, _, _, ending = calc_account_balance(code)
            return ending

        def _get_balance_list(codes):
            return sum(_get_balance(c) for c in codes)

        # -------- 收集各类数据 --------
        # 资产类明细
        asset_data = []
        for a in ACCOUNT_CHART:
            if a["category"] != "资产":
                continue
            bal = _get_balance(a["code"])
            if abs(bal) > 0.01:
                asset_data.append({"科目": a["name"], "编码": a["code"], "余额": abs(bal)})

        # 负债类明细
        liab_data = []
        for a in ACCOUNT_CHART:
            if a["category"] != "负债":
                continue
            bal = _get_balance(a["code"])
            if abs(bal) > 0.01:
                liab_data.append({"科目": a["name"], "编码": a["code"], "余额": abs(bal)})

        # 权益类明细
        equity_data = []
        for a in ACCOUNT_CHART:
            if a["category"] != "权益":
                continue
            bal = _get_balance(a["code"])
            if abs(bal) > 0.01:
                equity_data.append({"科目": a["name"], "编码": a["code"], "余额": abs(bal)})

        # 损益类数据
        income_total = _get_balance_list(["6001", "6021", "6041", "6051", "6101", "6102", "6103",
                                          "6111", "6115", "6117", "6301"])
        expense_total = _get_balance_list(["6401", "6402", "6403", "6405",
                                           "6601", "6602", "6603", "6604", "6605", "6606",
                                           "6641", "6642", "6701", "6711"])

        # 费用明细
        expense_items = []
        for code, name in [("6401", "主营业务成本"), ("6402", "其他业务成本"),
                           ("6403", "税金及附加"), ("6601", "销售费用"),
                           ("6602", "管理费用"), ("6603", "财务费用"),
                           ("6641", "信用减值损失"), ("6642", "资产减值损失"),
                           ("6701", "营业外支出"), ("6711", "所得税费用")]:
            bal = _get_balance(code)
            if abs(bal) > 0.01:
                expense_items.append({"科目": name, "编码": code, "金额": abs(bal)})

        # 收入明细
        income_items = []
        for code, name in [("6001", "主营业务收入"), ("6051", "其他业务收入"),
                            ("6111", "投资收益"), ("6117", "其他收益"),
                            ("6301", "营业外收入"), ("6101", "公允价值变动损益")]:
            bal = _get_balance(code)
            if abs(bal) > 0.01:
                income_items.append({"科目": name, "编码": code, "金额": abs(bal)})

        # 月度凭证数据（用于趋势图）
        all_vouchers_df = get_all_vouchers()
        if not all_vouchers_df.empty and "voucher_date" in all_vouchers_df.columns:
            all_vouchers_df["voucher_date"] = pd.to_datetime(all_vouchers_df["voucher_date"], errors="coerce")
            all_vouchers_df["月份"] = all_vouchers_df["voucher_date"].dt.to_period("M").astype(str)
            monthly_data = all_vouchers_df.groupby("月份").agg(
                借方合计=("debit_amount", "sum"),
                贷方合计=("credit_amount", "sum"),
            ).reset_index()
        else:
            monthly_data = pd.DataFrame(columns=["月份", "借方合计", "贷方合计"])

        # ============================================================
        # 图表区域
        # ============================================================
        # --- 核心指标卡片 ---
        asset_total = sum(d["余额"] for d in asset_data)
        liab_total = sum(d["余额"] for d in liab_data)
        equity_total = sum(d["余额"] for d in equity_data)
        net_profit = income_total - expense_total

        m1, m2, m3, m4 = st.columns(4)
        m1.metric("总资产", fmt_money(asset_total))
        m2.metric("总负债", fmt_money(liab_total))
        m3.metric("所有者权益", fmt_money(equity_total))
        m4.metric("本期净利润", fmt_money(abs(net_profit)),
                  delta=f"{'盈利' if net_profit > 0 else '亏损'}" if net_profit != 0 else "持平")

        # 负债率
        if asset_total > 0:
            debt_ratio = liab_total / asset_total * 100
            st.info(f"📌 资产负债率：{debt_ratio:.1f}%  |  权益比率：{(equity_total / asset_total * 100) if asset_total else 0:.1f}%")

        st.markdown("---")

        # --- 子标签页 ---
        viz_sub1, viz_sub2, viz_sub3, viz_sub4 = st.tabs([
            "🏗️ 资产结构", "💰 损益分析", "📈 收支趋势", "📊 费用构成",
        ])

        # ===== 1. 资产结构 =====
        with viz_sub1:
            if asset_data:
                c1, c2 = st.columns(2)

                with c1:
                    # 资产构成饼图
                    fig_asset = px.pie(
                        asset_data, values="余额", names="科目",
                        title="资产构成",
                        color_discrete_sequence=px.colors.qualitative.Set2,
                    )
                    fig_asset.update_traces(textposition="inside", textinfo="label+percent")
                    fig_asset.update_layout(showlegend=False, height=400)
                    st.plotly_chart(fig_asset, use_container_width=True)

                with c2:
                    # 负债与权益对比
                    struct_data = [
                        {"类别": "负债", "金额": liab_total},
                        {"类别": "所有者权益", "金额": equity_total},
                    ]
                    fig_struct = px.pie(
                        struct_data, values="金额", names="类别",
                        title="负债与权益结构",
                        color="类别",
                        color_discrete_map={"负债": "#ef5350", "所有者权益": "#42a5f5"},
                    )
                    fig_struct.update_traces(textposition="inside", textinfo="label+percent")
                    fig_struct.update_layout(height=400)
                    st.plotly_chart(fig_struct, use_container_width=True)

                # 资产 TOP 10 柱状图
                if len(asset_data) > 1:
                    top_assets = sorted(asset_data, key=lambda x: x["余额"], reverse=True)[:10]
                    fig_bar = px.bar(
                        top_assets, x="余额", y="科目", orientation="h",
                        title="资产科目 TOP 10",
                        color="余额", color_continuous_scale="Blues",
                    )
                    fig_bar.update_layout(yaxis={"categoryorder": "total ascending"}, height=400)
                    st.plotly_chart(fig_bar, use_container_width=True)

                # 负债明细
                if liab_data:
                    fig_liab = px.bar(
                        liab_data, x="余额", y="科目", orientation="h",
                        title="负债明细",
                        color_discrete_sequence=["#ef5350"],
                    )
                    fig_liab.update_layout(yaxis={"categoryorder": "total ascending"}, height=350)
                    st.plotly_chart(fig_liab, use_container_width=True)
            else:
                st.info("暂无资产数据，请先录入凭证或期初余额。")

        # ===== 2. 损益分析 =====
        with viz_sub2:
            # 收入 vs 费用对比
            if income_total > 0 or expense_total > 0:
                ie_data = pd.DataFrame({
                    "类别": ["收入合计", "费用合计"],
                    "金额": [income_total, expense_total],
                })
                fig_ie = px.bar(
                    ie_data, x="类别", y="金额",
                    title="收入 vs 费用",
                    color="类别",
                    color_discrete_map={"收入合计": "#66bb6a", "费用合计": "#ef5350"},
                    text="金额",
                )
                fig_ie.update_traces(texttemplate="%{text:,.0f}", textposition="outside")
                fig_ie.update_layout(height=400, showlegend=False)
                st.plotly_chart(fig_ie, use_container_width=True)

                # 净利润瀑布图
                waterfall_data = [
                    {"项目": "收入合计", "金额": income_total, "类型": "收入"},
                    {"项目": "费用合计", "金额": -expense_total, "类型": "费用"},
                    {"项目": "净利润", "金额": net_profit, "类型": "结果"},
                ]
                fig_wf = go.Figure(go.Waterfall(
                    name="损益",
                    orientation="v",
                    measure=["relative", "relative", "total"],
                    x=[d["项目"] for d in waterfall_data],
                    y=[d["金额"] for d in waterfall_data],
                    connector={"line": {"color": "#bbb"}},
                    increasing={"marker": {"color": "#66bb6a"}},
                    decreasing={"marker": {"color": "#ef5350"}},
                    totals={"marker": {"color": "#42a5f5"}},
                ))
                fig_wf.update_layout(title="损益瀑布图", height=400, yaxis_title="金额（元）")
                st.plotly_chart(fig_wf, use_container_width=True)

                # 收入构成饼图
                if income_items:
                    c1, c2 = st.columns(2)
                    with c1:
                        fig_inc = px.pie(
                            income_items, values="金额", names="科目",
                            title="收入构成",
                            color_discrete_sequence=px.colors.qualitative.Pastel,
                        )
                        fig_inc.update_traces(textposition="inside", textinfo="label+percent")
                        fig_inc.update_layout(height=400)
                        st.plotly_chart(fig_inc, use_container_width=True)
                    with c2:
                        # 毛利率计算
                        main_revenue = _get_balance("6001") + _get_balance("6051")
                        main_cost = _get_balance("6401") + _get_balance("6402")
                        gross_profit = main_revenue - main_cost
                        gross_margin = (gross_profit / main_revenue * 100) if main_revenue > 0 else 0
                        st.metric("营业收入", fmt_money(main_revenue))
                        st.metric("营业成本", fmt_money(main_cost))
                        st.metric("毛利润", fmt_money(gross_profit))
                        st.metric("毛利率", f"{gross_margin:.1f}%")
            else:
                st.info("暂无损益数据。")

        # ===== 3. 收支趋势 =====
        with viz_sub3:
            if not monthly_data.empty:
                # 月度借贷趋势折线图
                fig_trend = go.Figure()
                fig_trend.add_trace(go.Scatter(
                    x=monthly_data["月份"], y=monthly_data["贷方合计"],
                    mode="lines+markers", name="贷方（收入方）",
                    line=dict(color="#66bb6a", width=2),
                ))
                fig_trend.add_trace(go.Scatter(
                    x=monthly_data["月份"], y=monthly_data["借方合计"],
                    mode="lines+markers", name="借方（支出方）",
                    line=dict(color="#ef5350", width=2),
                ))
                fig_trend.update_layout(
                    title="月度借贷发生额趋势",
                    xaxis_title="月份", yaxis_title="金额（元）",
                    height=400, hovermode="x unified",
                )
                st.plotly_chart(fig_trend, use_container_width=True)

                # 月度凭证数量
                monthly_count = all_vouchers_df.groupby("月份").size().reset_index(name="凭证数")
                if not monthly_count.empty:
                    fig_count = px.bar(
                        monthly_count, x="月份", y="凭证数",
                        title="月度凭证数量",
                        color_discrete_sequence=["#42a5f5"],
                    )
                    fig_count.update_layout(height=350)
                    st.plotly_chart(fig_count, use_container_width=True)

                # 月度净发生额（贷方 - 借方）
                monthly_data["净发生额"] = monthly_data["贷方合计"] - monthly_data["借方合计"]
                colors = ["#66bb6a" if v >= 0 else "#ef5350" for v in monthly_data["净发生额"]]
                fig_net = px.bar(
                    monthly_data, x="月份", y="净发生额",
                    title="月度净发生额（贷方 - 借方）",
                )
                fig_net.update_traces(marker_color=colors)
                fig_net.update_layout(height=350)
                st.plotly_chart(fig_net, use_container_width=True)
            else:
                st.info("暂无凭证数据，请先录入凭证。")

        # ===== 4. 费用构成 =====
        with viz_sub4:
            if expense_items:
                # 费用构成饼图
                fig_exp = px.pie(
                    expense_items, values="金额", names="科目",
                    title="费用构成",
                    color_discrete_sequence=px.colors.qualitative.Set3,
                )
                fig_exp.update_traces(textposition="inside", textinfo="label+percent")
                fig_exp.update_layout(height=400)
                st.plotly_chart(fig_exp, use_container_width=True)

                # 费用排序柱状图
                sorted_exp = sorted(expense_items, key=lambda x: x["金额"], reverse=True)
                fig_exp_bar = px.bar(
                    sorted_exp, x="科目", y="金额",
                    title="费用科目排序",
                    color="金额", color_continuous_scale="Reds",
                )
                fig_exp_bar.update_layout(height=400, xaxis_tickangle=-30)
                st.plotly_chart(fig_exp_bar, use_container_width=True)

                # 费用占比表
                total_exp = sum(d["金额"] for d in expense_items)
                st.markdown("#### 费用占比明细")
                for item in sorted_exp:
                    pct = item["金额"] / total_exp * 100 if total_exp > 0 else 0
                    st.markdown(
                        f"- **{item['科目']}**：{fmt_money(item['金额'])} 元（{pct:.1f}%）"
                    )
            else:
                st.info("暂无费用数据。")

            # 税务相关费用专项分析
            tax_codes = ["6601", "6602"]  # 销售/管理费用下的税务二级科目
            tax_items = []
            df_custom = get_custom_accounts()
            if not df_custom.empty and "tax_flag" in df_custom.columns:
                tax_df = df_custom[df_custom["tax_flag"].notna() & (df_custom["tax_flag"] != "")]
                if not tax_df.empty:
                    for _, row in tax_df.iterrows():
                        _, _, _, _, bal = calc_account_balance(row["full_code"])
                        if abs(bal) > 0.01:
                            tax_items.append({
                                "科目": row["full_name"],
                                "税务标记": row["tax_flag"],
                                "金额": abs(bal),
                            })

            if tax_items:
                st.markdown("---")
                st.markdown("#### 🔖 税务相关费用分析")
                fig_tax = px.bar(
                    tax_items, x="科目", y="金额",
                    title="税务标记科目金额",
                    color="税务标记",
                    color_discrete_sequence=px.colors.qualitative.Dark2,
                )
                fig_tax.update_layout(height=400, xaxis_tickangle=-30)
                st.plotly_chart(fig_tax, use_container_width=True)


st.divider()



# ============================================================
# 【💰 财务会计】模块三：AI 智能问答（需要 API）
# ============================================================
with tab3:
    st.header("🤖 AI 智能问答")
    st.caption("需要 API Key 才能使用。AI 可以查看你的财务数据，回答问题、提供建议。")

    # --- API 设置 ---
    st.subheader("⚙️ API 设置")

    with st.expander("填写 API Key 和选择模型", expanded=True):
        TEXT_MODELS = {
            "硅基流动 - DeepSeek-V3（推荐）": {
                "base_url": "https://api.siliconflow.cn/v1",
                "model": "deepseek-ai/DeepSeek-V3",
                "key_hint": "在 cloud.siliconflow.cn 获取",
            },
            "硅基流动 - Qwen2.5-72B": {
                "base_url": "https://api.siliconflow.cn/v1",
                "model": "Qwen/Qwen2.5-72B-Instruct",
                "key_hint": "在 cloud.siliconflow.cn 获取",
            },
        }

        VISION_MODELS = {
            "硅基流动 - Qwen2-VL-72B（推荐）": {
                "base_url": "https://api.siliconflow.cn/v1",
                "model": "Qwen/Qwen2-VL-72B-Instruct",
            },
            "硅基流动 - DeepSeek-VL2": {
                "base_url": "https://api.siliconflow.cn/v1",
                "model": "deepseek-ai/deepseek-vl2",
            },
        }

        col1, col2 = st.columns(2)
        with col1:
            text_choice = st.selectbox("文字模型", list(TEXT_MODELS.keys()))
            text_config = TEXT_MODELS[text_choice]
            st.caption(f"模型：`{text_config['model']}`")
        with col2:
            vision_choice = st.selectbox("视觉模型（看凭证图片用）", list(VISION_MODELS.keys()))
            vision_config = VISION_MODELS[vision_choice]
            st.caption(f"模型：`{vision_config['model']}`")

        api_key = st.text_input(
            "API Key（硅基流动注册获取）",
            type="password",
            help="在 cloud.siliconflow.cn 注册，一个 Key 同时用于文字和视觉模型",
        )
        st.caption("没有 API Key？前往 cloud.siliconflow.cn 注册，有免费额度。")

    if not api_key:
        st.info("👆 请先填写 API Key 才能使用 AI 功能。模块一和模块二不需要 API Key，可以直接使用。")
    else:
        # --- AI 工具函数 ---
        def tool_query_account(account_name: str) -> str:
            """工具：查询科目余额"""
            # 根据名称找科目编码
            matched = [a for a in ACCOUNT_CHART if account_name in a["name"]]
            if not matched:
                return f"未找到包含「{account_name}」的科目。"
            results = []
            for acc in matched:
                od, oc, p_debit, p_credit, ending = calc_account_balance(acc["code"])
                results.append(
                    f"科目【{acc['name']}】({acc['category']}类)："
                    f"期初借方={od:,.2f}，期初贷方={oc:,.2f}，"
                    f"本期借方发生额={p_debit:,.2f}，本期贷方发生额={p_credit:,.2f}，"
                    f"期末余额={ending:,.2f}"
                )
            return "\n".join(results)

        def tool_query_all_vouchers() -> str:
            """工具：查询所有凭证"""
            df = get_all_vouchers()
            if df.empty:
                return "暂无凭证记录。"
            voucher_nums = df["voucher_number"].unique()
            lines = [f"共 {len(voucher_nums)} 张凭证："]
            for vnum in voucher_nums:
                vdf = df[df["voucher_number"] == vnum]
                vdate = vdf.iloc[0]["voucher_date"]
                vsummary = vdf.iloc[0]["summary"]
                d_total = vdf["debit_amount"].sum()
                lines.append(f"  {vnum} | {vdate} | {vsummary} | 借贷合计={d_total:,.2f}")
            return "\n".join(lines)

        def tool_calc_bad_debt(account_name: str) -> str:
            """工具：按账龄分析法计算坏账准备"""
            # 简化版：直接用应收账款期末余额 × 5% 估算
            matched = [a for a in ACCOUNT_CHART if account_name in a["name"]]
            if not matched:
                return f"未找到包含「{account_name}」的科目。"
            results = []
            for acc in matched:
                _, _, _, _, ending = calc_account_balance(acc["code"])
                if acc["code"] == "1122":  # 应收账款
                    provision = ending * 0.05
                    results.append(
                        f"应收账款期末余额：{ending:,.2f} 元\n"
                        f"按5%计提坏账准备：{provision:,.2f} 元\n"
                        f"（注：实际应按账龄分析法分档计提，此处为简化估算）"
                    )
                else:
                    results.append(f"{acc['name']} 期末余额：{ending:,.2f} 元")
            return "\n".join(results)

        TOOLS = [
            {
                "type": "function",
                "function": {
                    "name": "tool_query_account",
                    "description": "查询某个财务科目的余额、发生额等数据。",
                    "parameters": {
                        "type": "object",
                        "properties": {
                            "account_name": {
                                "type": "string",
                                "description": "科目名称或关键字，例如：应收账款、银行存款、管理费用",
                            }
                        },
                        "required": ["account_name"],
                    },
                },
            },
            {
                "type": "function",
                "function": {
                    "name": "tool_query_all_vouchers",
                    "description": "查询所有已录入的凭证记录。",
                    "parameters": {"type": "object", "properties": {}},
                },
            },
            {
                "type": "function",
                "function": {
                    "name": "tool_calc_bad_debt",
                    "description": "根据应收账款余额计算坏账准备。",
                    "parameters": {
                        "type": "object",
                        "properties": {
                            "account_name": {
                                "type": "string",
                                "description": "科目名称，例如：应收账款",
                            }
                        },
                        "required": ["account_name"],
                    },
                },
            },
        ]

        TOOL_FUNCTIONS = {
            "tool_query_account": tool_query_account,
            "tool_query_all_vouchers": tool_query_all_vouchers,
            "tool_calc_bad_debt": tool_calc_bad_debt,
        }

        def chat_with_agent(messages: list) -> str:
            """和 AI 对话，AI 会自动决定要不要调用工具"""
            client = OpenAI(api_key=api_key, base_url=text_config["base_url"])
            response = client.chat.completions.create(
                model=text_config["model"],
                messages=messages,
                tools=TOOLS,
                temperature=0.3,
            )
            message = response.choices[0].message

            if not message.tool_calls:
                return message.content or "（AI 没有返回内容）"

            messages.append(message)

            for tool_call in message.tool_calls:
                tool_name = tool_call.function.name
                tool_args = json.loads(tool_call.function.arguments)
                with st.container():
                    st.info(f"🔧 智能体正在调用工具：**{tool_name}**，参数：{tool_args}")
                func = TOOL_FUNCTIONS.get(tool_name)
                result = func(**tool_args) if func else f"未知工具：{tool_name}"
                messages.append({
                    "role": "tool",
                    "tool_call_id": tool_call.id,
                    "content": result,
                })

            final_response = client.chat.completions.create(
                model=text_config["model"],
                messages=messages,
                temperature=0.3,
            )
            return final_response.choices[0].message.content

        # --- 聊天界面 ---
        st.subheader("💬 和 AI 聊天")

        if "ai_messages" not in st.session_state:
            st.session_state.ai_messages = [
                {
                    "role": "system",
                    "content": (
                        "你是一个专业的财务智能助手。你可以查询科目余额、查看所有凭证、计算坏账准备。"
                        "请用中文回答，金额用千分位格式显示。"
                        "如果用户问的数据你查不到，请告诉他可用的科目有哪些。"
                    ),
                }
            ]
            st.session_state.ai_display = []

        for msg in st.session_state.ai_display:
            with st.chat_message(msg["role"]):
                st.markdown(msg["content"])

        user_input = st.chat_input("问我一个财务问题，比如：帮我查一下应收账款余额")

        if user_input:
            with st.chat_message("user"):
                st.markdown(user_input)
            st.session_state.ai_display.append({"role": "user", "content": user_input})

            with st.chat_message("assistant"):
                with st.spinner("财务智能体正在思考..."):
                    st.session_state.ai_messages.append({"role": "user", "content": user_input})
                    try:
                        reply = chat_with_agent(st.session_state.ai_messages)
                    except Exception as e:
                        reply = f"❌ 出错了：{e}\n\n常见原因：\n- API Key 填错了\n- 网络问题\n- 账户余额不足"
                    st.markdown(reply)

            st.session_state.ai_display.append({"role": "assistant", "content": reply})
            st.session_state.ai_messages.append({"role": "assistant", "content": reply})

        # --- 凭证图片识别 ---
        st.divider()
        st.subheader("🧾 凭证图片识别")
        st.caption("上传发票/收据图片，AI 自动识别内容并建议会计分录。")

        import base64

        def encode_image(image_file):
            return base64.b64encode(image_file.getvalue()).decode("utf-8")

        def recognize_voucher(image_base64, image_type="image/jpeg"):
            client = OpenAI(api_key=api_key, base_url=vision_config["base_url"])
            response = client.chat.completions.create(
                model=vision_config["model"],
                messages=[
                    {
                        "role": "user",
                        "content": [
                            {
                                "type": "text",
                                "text": (
                                    "请仔细识别这张凭证/发票/收据上的所有信息，包括：\n"
                                    "1. 凭证类型（如：增值税专用发票、普通发票、收据、银行回单等）\n"
                                    "2. 日期\n"
                                    "3. 金额（含税/不含税）\n"
                                    "4. 摘要或商品名称\n"
                                    "5. 供应商/客户名称\n"
                                    "6. 税额（如果有）\n"
                                    "请用清晰的格式列出。如果看不清楚的地方请标注。"
                                ),
                            },
                            {
                                "type": "image_url",
                                "image_url": {"url": f"data:{image_type};base64,{image_base64}"},
                            },
                        ],
                    }
                ],
                temperature=0.1,
            )
            return response.choices[0].message.content

        def suggest_entry(recognition_result):
            client = OpenAI(api_key=api_key, base_url=text_config["base_url"])
            response = client.chat.completions.create(
                model=text_config["model"],
                messages=[
                    {
                        "role": "system",
                        "content": (
                            "你是一个专业的会计。根据用户提供的凭证识别结果，"
                            "建议对应的会计分录（借方和贷方）。\n"
                            "请按以下格式输出：\n"
                            "摘要：xxx\n"
                            "借：xxx科目  xxx元\n"
                            "贷：xxx科目  xxx元\n"
                            "并简要说明理由。"
                        ),
                    },
                    {
                        "role": "user",
                        "content": f"以下是凭证识别结果，请建议会计分录：\n\n{recognition_result}",
                    },
                ],
                temperature=0.3,
            )
            return response.choices[0].message.content

        uploaded_images = st.file_uploader(
            "上传凭证/发票图片", type=["png", "jpg", "jpeg"],
            accept_multiple_files=True,
        )

        if uploaded_images:
            for img in uploaded_images:
                st.image(img, caption=img.name, width=400)
                if st.button(f"🔍 AI 识别：{img.name}", key=f"ai_btn_{img.name}"):
                    with st.spinner("AI 正在识别凭证内容..."):
                        try:
                            img_base64 = encode_image(img)
                            img_type = f"image/{img.type.split('/')[-1]}" if img.type else "image/jpeg"
                            recognition = recognize_voucher(img_base64, img_type)
                            st.success("✅ 识别完成！")
                            st.markdown("**识别结果：**")
                            st.text(recognition)

                            with st.spinner("AI 正在建议会计分录..."):
                                entry = suggest_entry(recognition)
                            st.markdown("**建议的会计分录：**")
                            st.markdown(entry)
                        except Exception as e:
                            st.error(f"❌ 识别出错：{e}")


# ============================================================
# 【📦 供应链管理】模块四：库存管理（仓存核算）
# ============================================================
with tab4:
    st.header("📦 仓存管理")
    st.caption("产成品档案 · 出入库记录 · 自动挂钩「库存商品」科目生成凭证 ｜ 完全离线")

    sub_inv1, sub_inv2, sub_inv3, sub_inv4 = st.tabs(["产品档案", "入库管理", "出库管理", "库存查询"])

    # --- 产品档案 ---
    with sub_inv1:
        st.subheader("产品档案管理")
        st.caption("录入产成品的基本信息。每个产品默认挂钩「1405 库存商品」科目。")

        # --- 新增/编辑产品 ---
        with st.expander("➕ 新增产品", expanded=True):
            col_a, col_b = st.columns(2)
            with col_a:
                p_code = st.text_input("产品编码 *", placeholder="如：P001", key="p_code")
                p_name = st.text_input("产品名称 *", placeholder="如：矿泉水 550ml", key="p_name")
                p_spec = st.text_input("规格", placeholder="如：550ml/瓶", key="p_spec")
                p_unit = st.selectbox("单位", ["件", "箱", "瓶", "袋", "千克", "米", "个", "台", "套", "批"], key="p_unit")
            with col_b:
                p_cost = money_input("成本单价（元）", key="p_cost", min_value=0.0)
                p_sell = money_input("销售单价（元）", key="p_sell", min_value=0.0)
                # 选择关联科目（默认库存商品 1405）
                stock_accounts = [a for a in ACCOUNT_CHART if a["code"] in ("1405", "1403", "1406", "1408", "1411", "1421")]
                stock_options = [f'{a["code"]} {a["name"]}' for a in stock_accounts]
                p_acc_choice = st.selectbox("关联库存科目", stock_options, key="p_acc")
                p_acc_code = p_acc_choice.split(" ")[0]
                p_acc_name = p_acc_choice.split(" ", 1)[1]

            if st.button("💾 保存产品", key="save_product"):
                if not p_code or not p_name:
                    st.error("产品编码和名称不能为空！")
                else:
                    add_product(p_code, p_name, p_spec, p_unit, p_cost, p_sell, p_acc_code, p_acc_name)
                    st.success(f"✅ 产品「{p_name}」已保存！")
                    st.rerun()

        # --- 产品列表 ---
        st.markdown("---")
        st.subheader("产品列表")
        products_df = get_all_products()
        if products_df.empty:
            st.info("暂无产品档案，请在上方录入。")
        else:
            # 重命名列以方便显示
            display_df = products_df.rename(columns={
                "product_code": "编码",
                "product_name": "名称",
                "specification": "规格",
                "unit": "单位",
                "cost_price": "成本单价",
                "selling_price": "销售单价",
                "account_code": "科目编码",
                "account_name": "科目名称",
            })
            st.dataframe(display_df[["编码", "名称", "规格", "单位", "成本单价", "销售单价", "科目编码", "科目名称"]],
                         use_container_width=True, hide_index=True)

            # 删除产品
            with st.expander("🗑️ 删除产品"):
                del_code = st.selectbox("选择要删除的产品", products_df["product_code"].tolist(), key="del_product")
                if st.button("确认删除", key="del_btn"):
                    delete_product(del_code)
                    st.success(f"已删除产品 {del_code}")
                    st.rerun()

    # --- 入库管理 ---
    with sub_inv2:
        st.subheader("入库管理")
        st.caption("记录产成品入库，系统自动生成借贷凭证（借 库存商品，贷 对方科目）。")

        products_df = get_all_products()
        if products_df.empty:
            st.warning("请先在「产品档案」中录入产品！")
        else:
            with st.form("inbound_form"):
                col1, col2, col3 = st.columns(3)
                with col1:
                    in_date = st.date_input("入库日期", key="in_date")
                    in_product = st.selectbox(
                        "选择产品 *",
                        products_df.apply(lambda r: f'{r["product_code"]} | {r["product_name"]} | {r["specification"] or "无规格"} | {r["unit"]}', axis=1).tolist(),
                        key="in_product"
                    )
                    in_product_code = in_product.split(" | ")[0]
                    in_product_name = in_product.split(" | ")[1]
                    in_spec = in_product.split(" | ")[2] if len(in_product.split(" | ")) > 2 else ""
                    in_unit = in_product.split(" | ")[3] if len(in_product.split(" | ")) > 3 else ""

                with col2:
                    in_qty = st.number_input("入库数量 *", min_value=0.0, value=1.0, step=1.0, key="in_qty")

                    # 获取产品的成本单价作为默认值
                    sel_row = products_df[products_df["product_code"] == in_product_code]
                    default_price = float(sel_row["cost_price"].values[0]) if not sel_row.empty else 0.0

                    in_price = money_input("单位成本（元）*", default_value=default_price, key="in_price", min_value=0.0)
                    in_total = in_qty * in_price
                    st.info(f"总金额：**{in_total:,.2f} 元**")

                with col3:
                    in_summary = st.text_input("摘要", value="产成品入库", key="in_summary")

                    # 入库时贷方科目选择（钱从哪来）
                    inbound_counterparties = [
                        ("1002", "银行存款"),
                        ("1001", "库存现金"),
                        ("2202", "应付账款"),
                        ("2203", "预收账款"),
                        ("5001", "生产成本"),
                        ("5101", "制造费用"),
                    ]
                    cp_options = [f'{c} {n}' for c, n in inbound_counterparties]
                    in_cp_choice = st.selectbox("贷方科目（资金来源）*", cp_options, key="in_cp")
                    in_cp_code = in_cp_choice.split(" ")[0]
                    in_cp_name = in_cp_choice.split(" ", 1)[1]

                submitted = st.form_submit_button("📦 确认入库")
                if submitted:
                    if in_qty <= 0:
                        st.error("入库数量必须大于 0！")
                    elif in_price < 0:
                        st.error("单价不能为负！")
                    else:
                        voucher_no = add_inventory_movement(
                            str(in_date), "入库", in_product_code,
                            in_product_name, in_spec, in_unit,
                            in_qty, in_price, in_summary,
                            in_cp_code, in_cp_name
                        )
                        st.success(f"✅ 入库成功！已自动生成凭证：**{voucher_no}**")
                        st.info(f"借：库存商品（{in_product_name}） {in_total:,.2f} 元\n贷：{in_cp_name} {in_total:,.2f} 元")

        # --- 入库记录列表 ---
        st.markdown("---")
        st.subheader("入库记录")
        movements_df = get_inventory_movements()
        if movements_df.empty:
            st.info("暂无入库记录。")
        else:
            inbound_df = movements_df[movements_df["movement_type"] == "入库"].copy()
            if inbound_df.empty:
                st.info("暂无入库记录。")
            else:
                display_in = inbound_df.rename(columns={
                    "movement_date": "日期",
                    "product_code": "产品编码",
                    "product_name": "产品名称",
                    "specification": "规格",
                    "unit": "单位",
                    "quantity": "数量",
                    "unit_price": "单价",
                    "total_amount": "总金额",
                    "summary": "摘要",
                    "counterparty_account_name": "对方科目",
                    "voucher_number": "凭证编号",
                })
                display_in = fmt_money_df(display_in, ["单价", "总金额"])
                st.dataframe(display_in[["日期", "产品编码", "产品名称", "规格", "单位", "数量", "单价", "总金额", "摘要", "对方科目", "凭证编号"]],
                             use_container_width=True, hide_index=True)

    # --- 出库管理 ---
    with sub_inv3:
        st.subheader("出库管理")
        st.caption("记录产成品出库，系统自动生成借贷凭证（借 对方科目，贷 库存商品）。")

        products_df = get_all_products()
        if products_df.empty:
            st.warning("请先在「产品档案」中录入产品！")
        else:
            with st.form("outbound_form"):
                col1, col2, col3 = st.columns(3)
                with col1:
                    out_date = st.date_input("出库日期", key="out_date")
                    out_product = st.selectbox(
                        "选择产品 *",
                        products_df.apply(lambda r: f'{r["product_code"]} | {r["product_name"]} | {r["specification"] or "无规格"} | {r["unit"]}', axis=1).tolist(),
                        key="out_product"
                    )
                    out_product_code = out_product.split(" | ")[0]
                    out_product_name = out_product.split(" | ")[1]
                    out_spec = out_product.split(" | ")[2] if len(out_product.split(" | ")) > 2 else ""
                    out_unit = out_product.split(" | ")[3] if len(out_product.split(" | ")) > 3 else ""

                    # 显示当前库存
                    current_stock = get_product_stock(out_product_code)
                    st.info(f"当前库存：**{current_stock:,.2f} {out_unit}**")

                with col2:
                    out_qty = st.number_input("出库数量 *", min_value=0.0, value=1.0, step=1.0, key="out_qty")

                    sel_row = products_df[products_df["product_code"] == out_product_code]
                    default_out_price = float(sel_row["cost_price"].values[0]) if not sel_row.empty else 0.0

                    out_price = money_input("单位成本（元）*", default_value=default_out_price, key="out_price", min_value=0.0)
                    out_total = out_qty * out_price
                    st.info(f"总金额：**{out_total:,.2f} 元**")

                with col3:
                    out_summary = st.text_input("摘要", value="产成品出库", key="out_summary")

                    # 出库时借方科目选择（去向）
                    outbound_counterparties = [
                        ("6401", "主营业务成本"),
                        ("6601", "销售费用"),
                        ("6602", "管理费用"),
                        ("1122", "应收账款"),
                        ("1002", "银行存款"),
                        ("1001", "库存现金"),
                    ]
                    cp_out_options = [f'{c} {n}' for c, n in outbound_counterparties]
                    out_cp_choice = st.selectbox("借方科目（资金去向）*", cp_out_options, key="out_cp")
                    out_cp_code = out_cp_choice.split(" ")[0]
                    out_cp_name = out_cp_choice.split(" ", 1)[1]

                submitted = st.form_submit_button("📤 确认出库")
                if submitted:
                    if out_qty <= 0:
                        st.error("出库数量必须大于 0！")
                    elif out_qty > current_stock:
                        st.error(f"库存不足！当前库存仅 {current_stock:,.2f} {out_unit}")
                    else:
                        voucher_no = add_inventory_movement(
                            str(out_date), "出库", out_product_code,
                            out_product_name, out_spec, out_unit,
                            out_qty, out_price, out_summary,
                            out_cp_code, out_cp_name
                        )
                        st.success(f"✅ 出库成功！已自动生成凭证：**{voucher_no}**")
                        st.info(f"借：{out_cp_name} {out_total:,.2f} 元\n贷：库存商品（{out_product_name}） {out_total:,.2f} 元")

        # --- 出库记录列表 ---
        st.markdown("---")
        st.subheader("出库记录")
        movements_df = get_inventory_movements()
        if movements_df.empty:
            st.info("暂无出库记录。")
        else:
            outbound_df = movements_df[movements_df["movement_type"] == "出库"].copy()
            if outbound_df.empty:
                st.info("暂无出库记录。")
            else:
                display_out = outbound_df.rename(columns={
                    "movement_date": "日期",
                    "product_code": "产品编码",
                    "product_name": "产品名称",
                    "specification": "规格",
                    "unit": "单位",
                    "quantity": "数量",
                    "unit_price": "单价",
                    "total_amount": "总金额",
                    "summary": "摘要",
                    "counterparty_account_name": "对方科目",
                    "voucher_number": "凭证编号",
                })
                display_out = fmt_money_df(display_out, ["单价", "总金额"])
                st.dataframe(display_out[["日期", "产品编码", "产品名称", "规格", "单位", "数量", "单价", "总金额", "摘要", "对方科目", "凭证编号"]],
                             use_container_width=True, hide_index=True)

    # --- 库存查询 ---
    with sub_inv4:
        st.subheader("库存查询")
        st.caption("查看所有产品的当前库存数量和金额。")

        stock_df = get_all_stock()
        if stock_df.empty:
            st.info("暂无产品档案。")
        else:
            # 格式化显示
            display_stock = stock_df.copy()
            display_stock["成本单价"] = display_stock["成本单价"].apply(lambda x: f"{x:,.2f}")
            display_stock["销售单价"] = display_stock["销售单价"].apply(lambda x: f"{x:,.2f}")
            display_stock["当前库存"] = display_stock["当前库存"].apply(lambda x: f"{x:,.2f}")
            display_stock["库存金额"] = display_stock["库存金额"].apply(lambda x: f"{x:,.2f}")
            st.dataframe(display_stock, use_container_width=True, hide_index=True)

            # 汇总指标
            total_items = len(stock_df)
            total_stock_value = stock_df["库存金额"].sum(numeric_only=True)
            if total_stock_value is None or isinstance(total_stock_value, str):
                total_stock_value = 0
            # 重新获取数值用于汇总
            raw_stock_df = get_all_stock()
            total_stock_value = raw_stock_df["库存金额"].astype(float).sum()

            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("产品种类", f"{total_items} 种")
            with col2:
                st.metric("库存总金额", f"{total_stock_value:,.2f} 元")
            with col3:
                # 库存商品科目余额（从凭证系统计算）
                _, _, _, _, stock_balance = calc_account_balance("1405")
                st.metric("库存商品科目余额", f"{stock_balance:,.2f} 元")

            st.caption("💡 提示：库存总金额应与「库存商品（1405）」科目余额一致，两者通过自动生成的凭证保持同步。")

            # --- 导出 Excel ---
            st.markdown("---")
            def export_stock_excel(df):
                from io import BytesIO
                output = BytesIO()
                with pd.ExcelWriter(output, engine="openpyxl") as writer:
                    df.to_excel(writer, index=False, sheet_name="库存汇总")
                output.seek(0)
                return output

            st.download_button(
                label="📥 下载库存汇总（Excel）",
                data=export_stock_excel(raw_stock_df),
                file_name="库存汇总表.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

            # --- 单个产品出入库明细 ---
            st.markdown("---")
            st.subheader("单个产品出入库明细")
            sel_product = st.selectbox(
                "选择产品查看流水",
                stock_df.apply(lambda r: f'{r["产品编码"]} | {r["产品名称"]}', axis=1).tolist(),
                key="stock_detail_sel"
            )
            if sel_product:
                sel_code = sel_product.split(" | ")[0]
                detail_df = get_inventory_movements(sel_code)
                if detail_df.empty:
                    st.info("该产品暂无出入库记录。")
                else:
                    detail_display = detail_df.rename(columns={
                        "movement_date": "日期",
                        "movement_type": "类型",
                        "product_code": "编码",
                        "product_name": "名称",
                        "specification": "规格",
                        "unit": "单位",
                        "quantity": "数量",
                        "unit_price": "单价",
                        "total_amount": "金额",
                        "summary": "摘要",
                        "counterparty_account_name": "对方科目",
                        "voucher_number": "凭证编号",
                    })
                    st.dataframe(detail_display[["日期", "类型", "编码", "名称", "规格", "单位", "数量", "单价", "金额", "摘要", "对方科目", "凭证编号"]],
                                 use_container_width=True, hide_index=True)


# ============================================================
# 【📦 供应链管理】模块五：多平台商品管理（商品档案）
# ============================================================
with tab5:
    st.header("🏪 商品档案管理")
    st.caption(
        "SPU/SKU 主数据 · 平台店铺 · 一品多商绑定 · BOM 物料清单 · 分类标签 ｜ 支持 Excel 批量导入导出"
    )

    # ---------- 公共选项 ----------
    M5_PLATFORMS = ["淘宝", "拼多多", "抖音", "视频号", "京东", "快手"]
    M5_LISTING_STATUS = ["未上架", "在售", "下架", "审核中"]
    M5_PRODUCT_TYPES = ["外购", "自产", "组装"]
    M5_UNITS = ["件", "箱", "瓶", "袋", "千克", "克", "米", "个", "台", "套", "批"]
    _m5_acc_list = [a for a in ACCOUNT_CHART if a["category"] == "资产" and a["code"].startswith("14")]
    M5_ACC_OPTIONS = [f'{a["code"]} {a["name"]}' for a in _m5_acc_list]

    m5_sub1, m5_sub2, m5_sub3, m5_sub4, m5_sub5, m5_sub6, m5_sub7 = st.tabs([
        "标准商品库(SPU)", "SKU规格管理", "平台店铺", "一品多商绑定",
        "BOM物料清单", "分类标签", "批量导入导出"
    ])

    # ---------- 1. 标准商品库(SPU) ----------
    with m5_sub1:
        st.subheader("标准商品库（SPU）")
        st.caption("ERP 内部统一商品主数据。每个 SPU 可挂多个 SKU、绑定多个平台店铺。")

        with st.form("m5_spu_form"):
            col1, col2, col3 = st.columns(3)
            with col1:
                spu_code = st.text_input("SPU 编码 *", placeholder="如：SPU001", key="m5_spu_code")
                spu_name = st.text_input("SPU 名称 *", placeholder="如：纯棉T恤", key="m5_spu_name")
                brand = st.text_input("品牌", key="m5_spu_brand")
                specification = st.text_input("规格描述", placeholder="如：圆领短袖", key="m5_spu_spec")
            with col2:
                cat_df = get_all_categories()
                if not cat_df.empty:
                    cat_options = cat_df.apply(
                        lambda r: f'{r["category_code"]} | {r["category_name"]}', axis=1).tolist()
                    cat_choice = st.selectbox("商品分类", cat_options, key="m5_spu_cat")
                    spu_cat_code = cat_choice.split(" | ")[0]
                    spu_cat_name = cat_choice.split(" | ", 1)[1] if " | " in cat_choice else ""
                else:
                    st.info("暂无分类，可先在「分类标签」中维护。")
                    spu_cat_code = ""
                    spu_cat_name = ""
                main_unit = st.selectbox("主计量单位", M5_UNITS, key="m5_spu_unit")
                barcode = st.text_input("条形码", key="m5_spu_barcode")
                product_type = st.selectbox("商品类型", M5_PRODUCT_TYPES, key="m5_spu_type")
            with col3:
                cost_price = money_input("成本价（元）", key="m5_spu_cost", min_value=0.0)
                selling_price = money_input("售价（元）", key="m5_spu_sell", min_value=0.0)
                tax_rate = st.number_input("税率", min_value=0.0, value=0.13, step=0.01, key="m5_spu_tax")
                acc_choice = st.selectbox("关联库存科目", M5_ACC_OPTIONS, key="m5_spu_acc")
                acc_code = acc_choice.split(" ")[0]
                acc_name = acc_choice.split(" ", 1)[1]
            col4, col5 = st.columns(2)
            with col4:
                supplier = st.text_input("供应商", key="m5_spu_supplier")
            with col5:
                default_warehouse = st.text_input("默认仓库", placeholder="如：主仓库", key="m5_spu_wh")
            description = st.text_area("商品描述", key="m5_spu_desc")
            submitted = st.form_submit_button("💾 保存 SPU")
            if submitted:
                if not spu_code or not spu_name:
                    st.error("SPU 编码和名称不能为空！")
                else:
                    add_spu(spu_code, spu_name, spu_cat_code, spu_cat_name, brand, specification,
                            main_unit, barcode, product_type, cost_price, selling_price, tax_rate,
                            supplier, default_warehouse, acc_code, acc_name, description)
                    st.success(f"✅ SPU「{spu_name}」已保存！")

        st.markdown("---")
        st.subheader("SPU 列表")
        spus_df = get_all_spus()
        if spus_df.empty:
            st.info("暂无 SPU 数据。")
        else:
            spu_display = spus_df.rename(columns={
                "spu_code": "SPU编码", "spu_name": "名称", "category_name": "分类",
                "brand": "品牌", "specification": "规格", "main_unit": "单位",
                "barcode": "条码", "product_type": "类型", "cost_price": "成本价",
                "selling_price": "售价", "tax_rate": "税率", "supplier": "供应商",
                "default_warehouse": "仓库", "account_code": "科目编码",
                "account_name": "科目名称", "is_active": "启用", "created_at": "创建时间",
            })
            show_cols = ["SPU编码", "名称", "分类", "品牌", "规格", "单位", "条码", "类型",
                         "成本价", "售价", "税率", "供应商", "仓库", "科目编码", "科目名称", "创建时间"]
            spu_display = fmt_money_df(spu_display, ["成本价", "售价"])
            st.dataframe(spu_display[show_cols], use_container_width=True, hide_index=True)

            with st.expander("🗑️ 删除 SPU"):
                del_spu = st.selectbox(
                    "选择要删除的 SPU",
                    spus_df.apply(lambda r: f'{r["spu_code"]} | {r["spu_name"]}', axis=1).tolist(),
                    key="m5_spu_del")
                if st.button("确认删除", key="m5_spu_del_btn"):
                    del_spu_code = del_spu.split(" | ")[0]
                    delete_spu(del_spu_code)
                    st.success(f"已删除 SPU {del_spu_code}")
                    st.rerun()

    # ---------- 2. SKU规格管理 ----------
    with m5_sub2:
        st.subheader("SKU 规格管理")
        st.caption("为选定的 SPU 维护规格组合（SKU）。规格属性用文本输入，如：颜色:红色;尺码:M")

        spus_df = get_all_spus()
        if spus_df.empty:
            st.warning("请先在「标准商品库(SPU)」中创建 SPU！")
        else:
            spu_options = spus_df.apply(
                lambda r: f'{r["spu_code"]} | {r["spu_name"]}', axis=1).tolist()
            sel_spu = st.selectbox("选择 SPU *", spu_options, key="m5_sku_spu")
            sel_spu_code = sel_spu.split(" | ")[0]
            sel_spu_name = sel_spu.split(" | ", 1)[1] if " | " in sel_spu else ""

            with st.form("m5_sku_form"):
                col1, col2 = st.columns(2)
                with col1:
                    default_sku_code = f"{sel_spu_code}-001"
                    sku_code = st.text_input("SKU 编码 *", value=default_sku_code, key="m5_sku_code")
                    sku_name = st.text_input("SKU 名称 *", placeholder="如：纯棉T恤 红色M", key="m5_sku_name")
                    spec_attrs = st.text_input("规格属性", placeholder="颜色:红色;尺码:M", key="m5_sku_spec")
                    barcode = st.text_input("条形码", key="m5_sku_barcode")
                with col2:
                    sku_cost = money_input("成本价（元）", key="m5_sku_cost", min_value=0.0)
                    sku_sell = money_input("售价（元）", key="m5_sku_sell", min_value=0.0)
                    sku_weight = st.number_input("重量", min_value=0.0, value=0.0, step=0.01, key="m5_sku_weight")
                    weight_unit = st.selectbox("重量单位", ["克", "千克", "磅", "盎司"], key="m5_sku_wunit")
                submitted = st.form_submit_button("💾 保存 SKU")
                if submitted:
                    if not sku_code or not sku_name:
                        st.error("SKU 编码和名称不能为空！")
                    else:
                        add_sku(sku_code, sel_spu_code, sku_name, spec_attrs, barcode,
                                sku_cost, sku_sell, sku_weight, weight_unit)
                        st.success(f"✅ SKU「{sku_name}」已保存！")

            st.markdown("---")
            st.subheader(f"{sel_spu_name} 的 SKU 列表")
            skus_df = get_skus_by_spu(sel_spu_code)
            if skus_df.empty:
                st.info("该 SPU 暂无 SKU。")
            else:
                sku_display = skus_df.rename(columns={
                    "sku_code": "SKU编码", "spu_code": "SPU编码", "sku_name": "名称",
                    "spec_attrs": "规格属性", "barcode": "条码", "cost_price": "成本价",
                    "selling_price": "售价", "weight": "重量", "weight_unit": "重量单位",
                    "is_active": "启用", "created_at": "创建时间",
                })
                show_cols = ["SKU编码", "名称", "规格属性", "条码", "成本价", "售价", "重量", "重量单位", "创建时间"]
                sku_display = fmt_money_df(sku_display, ["成本价", "售价"])
                st.dataframe(sku_display[show_cols], use_container_width=True, hide_index=True)

                with st.expander("🗑️ 删除 SKU"):
                    del_sku = st.selectbox(
                        "选择要删除的 SKU",
                        skus_df.apply(lambda r: f'{r["sku_code"]} | {r["sku_name"]}', axis=1).tolist(),
                        key="m5_sku_del")
                    if st.button("确认删除", key="m5_sku_del_btn"):
                        del_sku_code = del_sku.split(" | ")[0]
                        delete_sku(del_sku_code)
                        st.success(f"已删除 SKU {del_sku_code}")
                        st.rerun()

    # ---------- 3. 平台店铺 ----------
    with m5_sub3:
        st.subheader("平台店铺管理")
        st.caption("管理各电商平台店铺信息，含 API 对接预留字段（app_key / app_secret / access_token）。")

        with st.form("m5_shop_form"):
            col1, col2, col3 = st.columns(3)
            with col1:
                shop_id = st.text_input("店铺编号 *", placeholder="如：TB001", key="m5_shop_id")
                platform = st.selectbox("平台 *", M5_PLATFORMS, key="m5_shop_platform")
                shop_name = st.text_input("店铺名称 *", placeholder="如：官方旗舰店", key="m5_shop_name")
            with col2:
                shop_url = st.text_input("店铺网址", key="m5_shop_url")
                app_key = st.text_input("API AppKey", key="m5_shop_appkey")
                app_secret = st.text_input("API AppSecret", type="password", key="m5_shop_appsecret")
            with col3:
                access_token = st.text_input("Access Token", type="password", key="m5_shop_token")
                token_expire = st.date_input("Token 到期日", key="m5_shop_expire")
                api_status = st.selectbox("对接状态", ["未对接", "已对接", "异常"], key="m5_shop_status")
            submitted = st.form_submit_button("💾 保存店铺")
            if submitted:
                if not shop_id or not shop_name:
                    st.error("店铺编号和名称不能为空！")
                else:
                    add_shop(shop_id, platform, shop_name, shop_url, app_key, app_secret,
                             access_token, token_expire.strftime("%Y-%m-%d"), api_status)
                    st.success(f"✅ 店铺「{shop_name}」已保存！")

        st.markdown("---")
        st.subheader("店铺列表")
        shops_df = get_all_shops()
        if shops_df.empty:
            st.info("暂无店铺数据。")
        else:
            shop_display = shops_df.copy()
            # 密钥脱敏显示
            shop_display["api_app_secret"] = shop_display["api_app_secret"].apply(
                lambda x: "***" if x else "")
            shop_display["api_access_token"] = shop_display["api_access_token"].apply(
                lambda x: "***" if x else "")
            shop_display = shop_display.rename(columns={
                "shop_id": "店铺编号", "platform": "平台", "shop_name": "店铺名称",
                "shop_url": "网址", "api_app_key": "AppKey", "api_app_secret": "AppSecret",
                "api_access_token": "AccessToken", "api_token_expire": "Token到期",
                "api_status": "对接状态", "is_active": "启用", "created_at": "创建时间",
            })
            show_cols = ["店铺编号", "平台", "店铺名称", "网址", "AppKey", "AppSecret",
                         "AccessToken", "Token到期", "对接状态", "创建时间"]
            st.dataframe(shop_display[show_cols], use_container_width=True, hide_index=True)

            with st.expander("🗑️ 删除店铺"):
                del_shop = st.selectbox(
                    "选择要删除的店铺",
                    shops_df.apply(lambda r: f'{r["shop_id"]} | {r["shop_name"]}', axis=1).tolist(),
                    key="m5_shop_del")
                if st.button("确认删除", key="m5_shop_del_btn"):
                    del_shop_id = del_shop.split(" | ")[0]
                    delete_shop(del_shop_id)
                    st.success(f"已删除店铺 {del_shop_id}")
                    st.rerun()

    # ---------- 4. 一品多商绑定 ----------
    with m5_sub4:
        st.subheader("一品多商绑定")
        st.caption("将 SPU/SKU 绑定到各平台店铺，记录平台商品 ID、链接、价格与上下架状态。")

        spus_df = get_all_spus()
        if spus_df.empty:
            st.warning("请先创建 SPU！")
        else:
            spu_options = spus_df.apply(
                lambda r: f'{r["spu_code"]} | {r["spu_name"]}', axis=1).tolist()
            sel_spu = st.selectbox("选择 SPU *", spu_options, key="m5_link_spu")
            sel_spu_code = sel_spu.split(" | ")[0]

            # 该 SPU 下的 SKU（可选，留空表示整 SPU 绑定）
            skus_df = get_skus_by_spu(sel_spu_code)
            sku_options = ["（整 SPU，不指定 SKU）"]
            sku_map = {"（整 SPU，不指定 SKU）": ""}
            if not skus_df.empty:
                for _, r in skus_df.iterrows():
                    label = f'{r["sku_code"]} | {r["sku_name"]}'
                    sku_options.append(label)
                    sku_map[label] = r["sku_code"]
            sel_sku = st.selectbox("选择 SKU（可选）", sku_options, key="m5_link_sku")
            sel_sku_code = sku_map.get(sel_sku, "")

            shops_df = get_all_shops()
            with st.form("m5_link_form"):
                col1, col2, col3 = st.columns(3)
                with col1:
                    if shops_df.empty:
                        st.warning("请先在「平台店铺」创建店铺！")
                        sel_platform = st.selectbox("平台 *", M5_PLATFORMS, key="m5_link_platform")
                        sel_shop_id = ""
                    else:
                        shop_opts = shops_df.apply(
                            lambda r: f'{r["shop_id"]} | {r["platform"]} | {r["shop_name"]}', axis=1).tolist()
                        sel_shop = st.selectbox("选择店铺 *", shop_opts, key="m5_link_shop")
                        parts = sel_shop.split(" | ")
                        sel_shop_id = parts[0]
                        sel_platform = parts[1] if len(parts) > 1 else ""
                    platform_product_id = st.text_input("平台商品ID", key="m5_link_pid")
                    platform_sku_id = st.text_input("平台SKU ID", key="m5_link_sid")
                with col2:
                    platform_title = st.text_input("平台商品标题", key="m5_link_title")
                    platform_item_url = st.text_input("平台商品链接", key="m5_link_url")
                with col3:
                    platform_price = money_input("平台售价（元）", key="m5_link_price", min_value=0.0)
                    platform_stock = st.number_input("平台库存", min_value=0, value=0, step=1, key="m5_link_stock")
                    listing_status = st.selectbox("上架状态", M5_LISTING_STATUS, key="m5_link_status")
                submitted = st.form_submit_button("🔗 添加绑定")
                if submitted:
                    if not sel_spu_code:
                        st.error("请选择 SPU！")
                    elif shops_df.empty:
                        st.error("请先创建店铺！")
                    else:
                        add_platform_link(sel_spu_code, sel_sku_code, sel_platform, sel_shop_id,
                                          platform_product_id, platform_sku_id, platform_title,
                                          platform_item_url, platform_price, platform_stock, listing_status)
                        st.success("✅ 平台绑定已添加！")

        st.markdown("---")
        st.subheader("平台绑定列表")
        filter_platform = st.selectbox("按平台筛选", ["全部"] + M5_PLATFORMS, key="m5_link_filter")
        links_df = get_platform_links(platform=None if filter_platform == "全部" else filter_platform)
        if links_df.empty:
            st.info("暂无平台绑定记录。")
        else:
            link_display = links_df.rename(columns={
                "id": "ID", "spu_code": "SPU编码", "spu_name": "SPU名称",
                "sku_code": "SKU编码", "sku_name": "SKU名称", "platform": "平台",
                "shop_id": "店铺编号", "shop_name": "店铺名称",
                "platform_product_id": "平台商品ID", "platform_sku_id": "平台SKU ID",
                "platform_title": "标题", "platform_item_url": "链接",
                "platform_price": "平台售价", "platform_stock": "平台库存",
                "listing_status": "上架状态", "last_sync_time": "最后同步",
                "sync_status": "同步状态", "created_at": "创建时间",
            })
            show_cols = ["ID", "SPU编码", "SPU名称", "SKU编码", "平台", "店铺名称",
                         "平台商品ID", "标题", "平台售价", "平台库存", "上架状态", "同步状态", "创建时间"]
            st.dataframe(link_display[show_cols], use_container_width=True, hide_index=True)

            st.markdown("---")
            st.subheader("更新上架状态")
            with st.form("m5_link_status_form"):
                status_link = st.selectbox(
                    "选择绑定记录",
                    [f'{r["id"]} | {r["spu_code"]} @ {r["platform"]}' for _, r in links_df.iterrows()],
                    key="m5_link_status_sel")
                new_status = st.selectbox("新上架状态", M5_LISTING_STATUS, key="m5_link_status_new")
                upd_submitted = st.form_submit_button("更新状态")
                if upd_submitted:
                    upd_id = int(status_link.split(" | ")[0])
                    update_listing_status(upd_id, new_status)
                    st.success(f"已更新绑定 {upd_id} 的状态为「{new_status}」")
                    st.rerun()

    # ---------- 5. BOM物料清单 ----------
    with m5_sub5:
        st.subheader("BOM 物料清单")
        st.caption("创建 BOM（母件 SKU），添加子件明细，自动计算总成本。总成本 = SUM(数量 × 单价 × (1 + 损耗率))")

        all_skus_df = get_all_skus()
        with st.form("m5_bom_form"):
            col1, col2, col3 = st.columns(3)
            with col1:
                bom_code = st.text_input("BOM 编号 *", placeholder="如：BOM001", key="m5_bom_code")
                bom_type = st.selectbox("BOM 类型", ["生产", "组装", "委外"], key="m5_bom_type")
                version = st.text_input("版本", value="V1.0", key="m5_bom_ver")
            with col2:
                if all_skus_df.empty:
                    st.warning("请先创建 SKU！")
                    parent_sku_code = ""
                    parent_sku_name = ""
                else:
                    sku_opts = all_skus_df.apply(
                        lambda r: f'{r["sku_code"]} | {r["sku_name"]}', axis=1).tolist()
                    parent_sel = st.selectbox("母件 SKU *", sku_opts, key="m5_bom_parent")
                    parent_sku_code = parent_sel.split(" | ")[0]
                    parent_sku_name = parent_sel.split(" | ", 1)[1] if " | " in parent_sel else ""
            with col3:
                bom_desc = st.text_area("BOM 说明", key="m5_bom_desc")
            submitted = st.form_submit_button("💾 创建/更新 BOM")
            if submitted:
                if not bom_code or not parent_sku_code:
                    st.error("BOM 编号和母件 SKU 不能为空！")
                else:
                    add_bom(bom_code, parent_sku_code, parent_sku_name, bom_type, version, bom_desc)
                    st.success(f"✅ BOM「{bom_code}」已保存！")

        st.markdown("---")
        st.subheader("BOM 列表")
        boms_df = get_all_boms()
        if boms_df.empty:
            st.info("暂无 BOM 数据。")
        else:
            bom_display = boms_df.rename(columns={
                "bom_code": "BOM编号", "parent_sku_code": "母件SKU", "parent_sku_name": "母件名称",
                "bom_type": "类型", "version": "版本", "total_cost": "总成本",
                "status": "状态", "description": "说明", "created_at": "创建时间",
            })
            show_cols = ["BOM编号", "母件SKU", "母件名称", "类型", "版本", "总成本", "状态", "创建时间"]
            bom_display = fmt_money_df(bom_display, ["总成本"])
            st.dataframe(bom_display[show_cols], use_container_width=True, hide_index=True)

            st.markdown("---")
            st.subheader("BOM 明细维护")
            bom_opts = boms_df.apply(
                lambda r: f'{r["bom_code"]} | {r["parent_sku_name"]}', axis=1).tolist()
            sel_bom = st.selectbox("选择 BOM", bom_opts, key="m5_bom_sel")
            sel_bom_code = sel_bom.split(" | ")[0]

            items_df = get_bom_items(sel_bom_code)
            with st.form("m5_bom_item_form"):
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    if all_skus_df.empty:
                        st.warning("无可用 SKU")
                        child_sku_code = ""
                        child_sku_name = ""
                    else:
                        child_opts = all_skus_df.apply(
                            lambda r: f'{r["sku_code"]} | {r["sku_name"]}', axis=1).tolist()
                        child_sel = st.selectbox("子件 SKU *", child_opts, key="m5_bom_item_child")
                        child_sku_code = child_sel.split(" | ")[0]
                        child_sku_name = child_sel.split(" | ", 1)[1] if " | " in child_sel else ""
                    seq = st.number_input("序号", min_value=1, value=1, step=1, key="m5_bom_item_seq")
                with col2:
                    item_qty = st.number_input("数量 *", min_value=0.0, value=1.0, step=0.01, key="m5_bom_item_qty")
                    item_unit = st.selectbox("单位", M5_UNITS, key="m5_bom_item_unit")
                with col3:
                    item_cost = money_input("单价（元）*", key="m5_bom_item_cost", min_value=0.0)
                    loss_rate = st.number_input("损耗率", min_value=0.0, value=0.0, step=0.01,
                                                help="如 0.05 表示 5% 损耗", key="m5_bom_item_loss")
                with col4:
                    line_total = item_qty * item_cost * (1 + loss_rate)
                    st.metric("本行成本", f"{line_total:,.2f} 元")
                submitted = st.form_submit_button("➕ 添加子件")
                if submitted:
                    if not sel_bom_code or not child_sku_code:
                        st.error("请选择 BOM 和子件 SKU！")
                    else:
                        add_bom_item(sel_bom_code, seq, child_sku_code, child_sku_name,
                                     item_qty, item_unit, item_cost, loss_rate)
                        st.success(f"✅ 子件「{child_sku_name}」已添加！")
                        st.rerun()

            st.markdown("**子件明细**")
            if items_df.empty:
                st.info("该 BOM 暂无子件明细。")
            else:
                item_display = items_df.rename(columns={
                    "id": "ID", "bom_code": "BOM编号", "seq": "序号",
                    "child_sku_code": "子件SKU", "child_sku_name": "子件名称",
                    "quantity": "数量", "unit": "单位", "unit_cost": "单价",
                    "total_cost": "行成本", "loss_rate": "损耗率", "is_active": "启用",
                })
                show_cols = ["序号", "子件SKU", "子件名称", "数量", "单位", "单价", "损耗率", "行成本"]
                item_display = fmt_money_df(item_display, ["单价", "行成本"])
                st.dataframe(item_display[show_cols], use_container_width=True, hide_index=True)
                total_cost = calc_bom_cost(sel_bom_code)
                st.success(f"🧮 BOM「{sel_bom_code}」总成本：**{total_cost:,.2f} 元**")

            with st.expander("🗑️ 删除 BOM"):
                del_bom = st.selectbox("选择要删除的 BOM", bom_opts, key="m5_bom_del")
                if st.button("确认删除", key="m5_bom_del_btn"):
                    del_bom_code = del_bom.split(" | ")[0]
                    delete_bom(del_bom_code)
                    st.success(f"已删除 BOM {del_bom_code}")
                    st.rerun()

    # ---------- 6. 分类标签 ----------
    with m5_sub6:
        st.subheader("分类标签管理")
        st.caption("维护商品分类树与标签体系，并为 SPU 打标签。")

        col_cat, col_tag = st.columns(2)

        with col_cat:
            st.markdown("#### 商品分类")
            cat_df = get_all_categories()
            with st.form("m5_cat_form"):
                cat_code = st.text_input("分类编码 *", placeholder="如：C001", key="m5_cat_code")
                cat_name = st.text_input("分类名称 *", placeholder="如：食品", key="m5_cat_name")
                parent_options = ["（顶层分类）"]
                parent_map = {"（顶层分类）": ""}
                for _, r in cat_df.iterrows():
                    label = f'{r["category_code"]} | {r["category_name"]}'
                    parent_options.append(label)
                    parent_map[label] = r["category_code"]
                parent_sel = st.selectbox("父分类", parent_options, key="m5_cat_parent")
                parent_code_val = parent_map.get(parent_sel, "")
                if parent_code_val:
                    parent_row = cat_df[cat_df["category_code"] == parent_code_val]
                    cat_level = int(parent_row["level"].values[0]) + 1 if not parent_row.empty else 1
                else:
                    cat_level = 1
                st.info(f"自动层级：第 {cat_level} 级")
                submitted = st.form_submit_button("💾 保存分类")
                if submitted:
                    if not cat_code or not cat_name:
                        st.error("分类编码和名称不能为空！")
                    else:
                        add_category(cat_code, cat_name, parent_code_val, cat_level)
                        st.success(f"✅ 分类「{cat_name}」已保存！")

            st.markdown("**分类列表**")
            if cat_df.empty:
                st.info("暂无分类。")
            else:
                cat_display = cat_df.rename(columns={
                    "category_code": "编码", "category_name": "名称",
                    "parent_code": "父编码", "level": "层级",
                    "is_active": "启用", "created_at": "创建时间",
                })
                st.dataframe(cat_display[["编码", "名称", "父编码", "层级", "创建时间"]],
                             use_container_width=True, hide_index=True)

        with col_tag:
            st.markdown("#### 商品标签")
            tags_df = get_all_tags()
            with st.form("m5_tag_form"):
                tag_name = st.text_input("标签名称 *", placeholder="如：热销", key="m5_tag_name")
                tag_color = st.color_picker("标签颜色", value="#1890ff", key="m5_tag_color")
                submitted = st.form_submit_button("💾 保存标签")
                if submitted:
                    if not tag_name:
                        st.error("标签名称不能为空！")
                    else:
                        add_tag(tag_name, tag_color)
                        st.success(f"✅ 标签「{tag_name}」已保存！")

            st.markdown("**标签列表**")
            if tags_df.empty:
                st.info("暂无标签。")
            else:
                tag_display = tags_df.rename(columns={
                    "tag_id": "ID", "tag_name": "名称",
                    "tag_color": "颜色", "created_at": "创建时间",
                })
                st.dataframe(tag_display[["ID", "名称", "颜色", "创建时间"]],
                             use_container_width=True, hide_index=True)

        st.markdown("---")
        st.markdown("#### 为 SPU 打标签")
        spus_df = get_all_spus()
        if spus_df.empty:
            st.info("请先创建 SPU。")
        elif tags_df.empty:
            st.info("请先创建标签。")
        else:
            col_t1, col_t2 = st.columns(2)
            with col_t1:
                spu_opts = spus_df.apply(
                    lambda r: f'{r["spu_code"]} | {r["spu_name"]}', axis=1).tolist()
                tag_spu = st.selectbox("选择 SPU", spu_opts, key="m5_tag_spu")
                tag_spu_code = tag_spu.split(" | ")[0]
            with col_t2:
                tag_opts = tags_df.apply(
                    lambda r: f'{r["tag_id"]} | {r["tag_name"]}', axis=1).tolist()
                tag_sel = st.selectbox("选择标签", tag_opts, key="m5_tag_sel")
                tag_sel_id = int(tag_sel.split(" | ")[0])
            if st.button("🏷️ 绑定标签", key="m5_tag_bind_btn"):
                add_product_tag(tag_spu_code, tag_sel_id)
                st.success(f"✅ 已为 SPU {tag_spu_code} 绑定标签！")
                st.rerun()

            st.markdown(f"**{tag_spu_code} 已有标签**")
            spu_tags_df = get_tags_by_spu(tag_spu_code)
            if spu_tags_df.empty:
                st.info("该 SPU 暂无标签。")
            else:
                st.dataframe(spu_tags_df.rename(columns={
                    "tag_id": "ID", "tag_name": "名称",
                    "tag_color": "颜色", "created_at": "创建时间",
                })[["ID", "名称", "颜色"]], use_container_width=True, hide_index=True)

    # ---------- 7. 批量导入导出 ----------
    with m5_sub7:
        st.subheader("批量导入导出")
        st.caption("支持 SPU 与 SKU 的 Excel 模板下载、数据导出与导入。工作表名：SPU标准商品 / SKU规格")

        col_exp, col_imp = st.columns(2)
        with col_exp:
            st.markdown("#### 导出 / 模板下载")
            st.caption("点击下方按钮导出当前所有 SPU 与 SKU 数据（空库时可作为导入模板）。")
            export_data = export_products_excel()
            st.download_button(
                label="📥 下载 Excel（SPU + SKU）",
                data=export_data,
                file_name="商品主数据.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="m5_export_btn",
            )

        with col_imp:
            st.markdown("#### 导入数据")
            uploaded = st.file_uploader("选择 Excel 文件", type=["xlsx", "xls"], key="m5_upload")
            if uploaded is not None:
                file_bytes = uploaded.read()
                try:
                    spu_n, sku_n = import_products_from_excel(file_bytes)
                    st.success(f"✅ 导入完成：SPU {spu_n} 条，SKU {sku_n} 条")
                except Exception as e:
                    st.error(f"导入失败：{e}")


# ============================================================
# 【👥 客户关系】模块六：电商CRM会员管理
# ============================================================
with tab6:
    st.header("👥 电商 CRM 会员管理")
    st.caption("会员档案 · 平台账号 · 等级体系 · 消费记录 · 积分 · 标签与黑名单管理")

    M6_PLATFORMS = ["淘宝", "拼多多", "抖音", "视频号", "京东", "快手"]
    M6_GENDERS = ["男", "女", "未知"]
    M6_TAG_TYPES = ["消费行为", "客户属性", "营销标签", "自定义"]
    M6_TAG_COLORS = ["#1890ff", "#52c41a", "#faad14", "#f5222d", "#722ed1", "#13c2c2"]
    M6_POINT_TYPES = ["手动增加", "兑换扣减", "活动奖励", "等级奖励"]
    M6_BLOCK_TYPES = ["永久", "临时"]

    # 确保预置等级存在（仅首次运行）
    if not st.session_state.get("_preset_levels_done"):
        init_preset_levels()
        st.session_state["_preset_levels_done"] = True

    m6_sub1, m6_sub2, m6_sub3, m6_sub4, m6_sub5, m6_sub6, m6_sub7, m6_sub8 = st.tabs([
        "会员档案", "平台账号聚合", "会员等级", "消费记录",
        "积分管理", "客户标签", "黑名单管理", "📊 客户画像分析",
    ])

    # ---------- 1. 会员档案 ----------
    with m6_sub1:
        st.subheader("会员档案")
        st.caption("录入和维护客户基础信息。已存在客户仅更新档案，保留消费统计。")

        with st.form("crm_customer_form"):
            col1, col2, col3 = st.columns(3)
            with col1:
                customer_id = st.text_input("客户ID *", placeholder="如：C001", key="crm_cust_id")
                customer_name = st.text_input("姓名 *", key="crm_cust_name")
                phone = st.text_input("手机", key="crm_cust_phone")
            with col2:
                email = st.text_input("邮箱", key="crm_cust_email")
                gender = st.selectbox("性别", M6_GENDERS, key="crm_cust_gender")
                birthday = st.date_input("生日", value=None, key="crm_cust_birthday")
            with col3:
                province = st.text_input("省份", key="crm_cust_province")
                city = st.text_input("城市", key="crm_cust_city")
                register_date = st.date_input("注册日期", value=datetime.now().date(), key="crm_cust_register")
            remark = st.text_area("备注", key="crm_cust_remark")
            submitted = st.form_submit_button("💾 保存客户")
            if submitted:
                if not customer_id or not customer_name:
                    st.error("客户ID和姓名不能为空！")
                else:
                    add_customer(customer_id, customer_name, phone, email, gender,
                                 str(birthday) if birthday else "", province, city,
                                 str(register_date) if register_date else "", remark)
                    st.success(f"✅ 客户「{customer_name}」已保存！")

        st.markdown("---")
        st.subheader("客户列表")
        cust_df = get_all_customers()
        if cust_df.empty:
            st.info("暂无客户数据，请先在上方录入。")
        else:
            display = cust_df.rename(columns={
                "customer_id": "客户ID", "customer_name": "姓名", "phone": "手机",
                "level_name": "等级", "total_points": "积分",
                "total_consumption": "消费总额", "total_orders": "订单数",
                "avg_order_value": "客单价", "last_order_date": "最后消费", "status": "状态",
            })
            show_cols = ["客户ID", "姓名", "手机", "等级", "积分", "消费总额", "订单数", "客单价", "最后消费", "状态"]
            view = display[show_cols].copy()
            view = fmt_money_df(view, ["消费总额", "客单价"])
            styler = view.style.map(
                lambda v: "color: #f5222d; font-weight: bold" if v == "黑名单" else "",
                subset=["状态"])
            styler.hide(axis="index")
            st.dataframe(styler, use_container_width=True)

            with st.expander("🗑️ 删除客户（级联删除所有关联数据）"):
                del_options = cust_df.apply(
                    lambda r: f'{r["customer_id"]} | {r["customer_name"]}', axis=1).tolist()
                del_choice = st.selectbox("选择要删除的客户", del_options, key="crm_cust_del")
                st.warning("删除将同时移除该客户的平台账号、消费记录、积分日志、标签关联与黑名单记录。")
                if st.button("确认删除", key="crm_cust_del_btn"):
                    del_id = del_choice.split(" | ")[0]
                    delete_customer(del_id)
                    st.success(f"已删除客户 {del_id}")
                    st.rerun()

    # ---------- 2. 平台账号聚合 ----------
    with m6_sub2:
        st.subheader("平台账号聚合")
        st.caption("一个客户可绑定多个平台账号，实现全平台身份聚合。")

        cust_df = get_all_customers()
        if cust_df.empty:
            st.info("暂无客户数据，请先在「会员档案」中录入。")
        else:
            cust_options = cust_df.apply(
                lambda r: f'{r["customer_id"]} | {r["customer_name"]}', axis=1).tolist()
            sel_cust = st.selectbox("选择客户", cust_options, key="crm_pa_cust")
            sel_cust_id = sel_cust.split(" | ")[0]
            sel_cust_name = sel_cust.split(" | ", 1)[1] if " | " in sel_cust else ""

            with st.form("crm_pa_form"):
                col1, col2 = st.columns(2)
                with col1:
                    pa_platform = st.selectbox("平台 *", M6_PLATFORMS, key="crm_pa_platform")
                    pa_user_id = st.text_input("平台用户ID", placeholder="如：淘宝会员号", key="crm_pa_uid")
                with col2:
                    pa_nick = st.text_input("平台昵称", key="crm_pa_nick")
                    shop_df = get_all_shops()
                    shop_opts_all = ["（不关联）"]
                    if not shop_df.empty:
                        shop_opts_all += shop_df.apply(
                            lambda r: f'{r["shop_id"]} | {r["shop_name"]}（{r["platform"]}）',
                            axis=1).tolist()
                    shop_choice = st.selectbox("关联店铺", shop_opts_all, key="crm_pa_shop")
                    pa_shop_id = "" if shop_choice == "（不关联）" else shop_choice.split(" | ")[0]
                submitted = st.form_submit_button("💾 绑定平台账号")
                if submitted:
                    if not pa_user_id and not pa_nick:
                        st.error("平台用户ID和昵称至少填写一项！")
                    else:
                        add_platform_account(sel_cust_id, pa_platform, pa_user_id, pa_nick, pa_shop_id)
                        st.success(f"✅ 已为「{sel_cust_name}」绑定 {pa_platform} 账号！")

            st.markdown("---")
            st.subheader(f"{sel_cust_name} 的平台账号")
            pa_df = get_customer_platform_accounts(sel_cust_id)
            if pa_df.empty:
                st.info("该客户暂未绑定平台账号。")
            else:
                pa_display = pa_df.rename(columns={
                    "id": "记录ID", "platform": "平台", "platform_user_id": "平台用户ID",
                    "platform_nick": "平台昵称", "shop_id": "关联店铺",
                    "bind_date": "绑定日期", "created_at": "创建时间",
                })
                show_cols = ["记录ID", "平台", "平台用户ID", "平台昵称", "关联店铺", "绑定日期", "创建时间"]
                st.dataframe(pa_display[show_cols], use_container_width=True, hide_index=True)

                with st.expander("🗑️ 解除绑定"):
                    del_opts = pa_df.apply(
                        lambda r: f'{int(r["id"])} | {r["platform"]} - {r["platform_nick"] or r["platform_user_id"]}',
                        axis=1).tolist()
                    del_pa = st.selectbox("选择要解除的绑定", del_opts, key="crm_pa_del")
                    if st.button("确认解除", key="crm_pa_del_btn"):
                        del_id = int(del_pa.split(" | ")[0])
                        delete_platform_account(del_id)
                        st.success("已解除绑定")
                        st.rerun()

    # ---------- 3. 会员等级 ----------
    with m6_sub3:
        st.subheader("会员等级")
        st.caption("维护等级体系。客户达到对应积分与消费额时自动升级。")

        with st.form("crm_level_form"):
            col1, col2, col3 = st.columns(3)
            with col1:
                lvl_code = st.text_input("等级编码 *", placeholder="如：L1", key="crm_lvl_code")
                lvl_name = st.text_input("等级名称 *", placeholder="如：普通会员", key="crm_lvl_name")
                sort_order = st.number_input("排序", min_value=0, value=1, step=1, key="crm_lvl_sort")
            with col2:
                min_points = st.number_input("最低积分", min_value=0, value=0, step=100, key="crm_lvl_pts")
                min_consumption = money_input("最低消费额", key="crm_lvl_cons", min_value=0.0)
            with col3:
                discount_rate = st.number_input("折扣率", min_value=0.0, max_value=1.0, value=1.0, step=0.01, key="crm_lvl_disc")
                description = st.text_input("等级说明", key="crm_lvl_desc")
            submitted = st.form_submit_button("💾 保存等级")
            if submitted:
                if not lvl_code or not lvl_name:
                    st.error("等级编码和名称不能为空！")
                else:
                    add_level(lvl_code, lvl_name, min_points, min_consumption, discount_rate, description, sort_order)
                    st.success(f"✅ 等级「{lvl_name}」已保存！")

        st.markdown("---")
        st.subheader("等级列表")
        lvl_df = get_all_levels()
        if lvl_df.empty:
            st.info("暂无等级数据。")
        else:
            lvl_display = lvl_df.rename(columns={
                "level_code": "等级编码", "level_name": "等级名称", "min_points": "最低积分",
                "min_consumption": "最低消费额", "discount_rate": "折扣率",
                "description": "说明", "sort_order": "排序", "created_at": "创建时间",
            })
            show_cols = ["等级编码", "等级名称", "最低积分", "最低消费额", "折扣率", "说明", "排序", "创建时间"]
            lvl_display = fmt_money_df(lvl_display, ["最低消费额"])
            st.dataframe(lvl_display[show_cols], use_container_width=True, hide_index=True)

            with st.expander("🗑️ 删除等级"):
                del_opts = lvl_df.apply(
                    lambda r: f'{r["level_code"]} | {r["level_name"]}', axis=1).tolist()
                del_lvl = st.selectbox("选择要删除的等级", del_opts, key="crm_lvl_del")
                if st.button("确认删除", key="crm_lvl_del_btn"):
                    del_code = del_lvl.split(" | ")[0]
                    delete_level(del_code)
                    st.success(f"已删除等级 {del_code}")
                    st.rerun()

    # ---------- 4. 消费记录 ----------
    with m6_sub4:
        st.subheader("消费记录")
        st.caption("录入消费订单，自动计算积分、更新统计并检查等级升级。")

        cust_df = get_all_customers()
        if cust_df.empty:
            st.info("暂无客户数据，请先在「会员档案」中录入。")
        else:
            cust_options = cust_df.apply(
                lambda r: f'{r["customer_id"]} | {r["customer_name"]}', axis=1).tolist()
            sel_cust = st.selectbox("选择客户", cust_options, key="crm_con_cust")
            sel_cust_id = sel_cust.split(" | ")[0]
            sel_cust_name = sel_cust.split(" | ", 1)[1] if " | " in sel_cust else ""

            con_df = get_consumption_log(sel_cust_id)
            total_con = float(con_df["order_amount"].sum()) if not con_df.empty else 0.0
            order_cnt = int(len(con_df))
            avg_con = total_con / order_cnt if order_cnt > 0 else 0.0
            max_con = float(con_df["order_amount"].max()) if not con_df.empty else 0.0
            m1, m2, m3, m4 = st.columns(4)
            m1.metric("总消费（元）", fmt_money(total_con))
            m2.metric("总订单数", order_cnt)
            m3.metric("平均客单价", fmt_money(avg_con))
            m4.metric("最大单笔", fmt_money(max_con))

            with st.form("crm_con_form"):
                col1, col2 = st.columns(2)
                with col1:
                    con_date = st.date_input("订单日期 *", value=datetime.now().date(), key="crm_con_date")
                    con_platform = st.selectbox("平台 *", M6_PLATFORMS, key="crm_con_platform")
                    con_order_id = st.text_input("订单号", placeholder="如：TB2024001", key="crm_con_oid")
                with col2:
                    con_amount = money_input("订单金额（元） *", key="crm_con_amount", min_value=0.0)
                    points_rate = st.number_input("积分倍率", min_value=0.0, value=1.0, step=0.1, key="crm_con_rate")
                    con_summary = st.text_input("商品摘要", placeholder="如：纯棉T恤 x2", key="crm_con_summary")
                submitted = st.form_submit_button("💾 录入消费")
                if submitted:
                    if con_amount <= 0:
                        st.error("订单金额必须大于0！")
                    else:
                        earned = add_consumption(sel_cust_id, str(con_date), con_platform,
                                                 con_order_id, con_amount, con_summary, points_rate)
                        st.success(f"✅ 已录入消费 {con_amount} 元，获得积分 {earned}！")
                        st.rerun()

            st.markdown("---")
            st.subheader(f"{sel_cust_name} 的消费历史")
            if con_df.empty:
                st.info("该客户暂无消费记录。")
            else:
                con_display = con_df.rename(columns={
                    "id": "记录ID", "order_date": "订单日期", "platform": "平台",
                    "order_id": "订单号", "order_amount": "订单金额",
                    "product_summary": "商品摘要", "points_earned": "获得积分",
                    "points_used": "使用积分", "created_at": "录入时间",
                })
                show_cols = ["记录ID", "订单日期", "平台", "订单号", "订单金额", "商品摘要", "获得积分", "使用积分", "录入时间"]
                con_display = fmt_money_df(con_display, ["订单金额"])
                st.dataframe(con_display[show_cols], use_container_width=True, hide_index=True)

                from io import BytesIO
                export_buf = BytesIO()
                with pd.ExcelWriter(export_buf, engine="openpyxl") as writer:
                    con_display[show_cols].to_excel(writer, index=False, sheet_name="消费记录")
                export_buf.seek(0)
                st.download_button(
                    label="📥 导出消费记录 Excel",
                    data=export_buf,
                    file_name=f"{sel_cust_id}_消费记录.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="crm_con_export",
                )

    # ---------- 5. 积分管理 ----------
    with m6_sub5:
        st.subheader("积分管理")
        st.caption("查看积分变动日志，手动调整积分（正数增加、负数扣减）。")

        cust_df = get_all_customers()
        if cust_df.empty:
            st.info("暂无客户数据，请先在「会员档案」中录入。")
        else:
            cust_options = cust_df.apply(
                lambda r: f'{r["customer_id"]} | {r["customer_name"]}', axis=1).tolist()
            sel_cust = st.selectbox("选择客户", cust_options, key="crm_pts_cust")
            sel_cust_id = sel_cust.split(" | ")[0]
            sel_cust_name = sel_cust.split(" | ", 1)[1] if " | " in sel_cust else ""

            cust_info = get_customer_by_id(sel_cust_id)
            if cust_info:
                cm1, cm2 = st.columns(2)
                cm1.metric("当前积分余额", cust_info.get("total_points") or 0)
                cm2.metric("会员等级", cust_info.get("level_name") or "普通会员")

            with st.form("crm_pts_form"):
                col1, col2 = st.columns(2)
                with col1:
                    change_type = st.selectbox("调整类型", M6_POINT_TYPES, key="crm_pts_type")
                    points_change = st.number_input("积分变动值（负数为扣减）", value=0, step=1, key="crm_pts_change")
                with col2:
                    pts_desc = st.text_input("描述", placeholder="如：生日礼积分", key="crm_pts_desc")
                    pts_order = st.text_input("关联订单号", key="crm_pts_order")
                submitted = st.form_submit_button("💾 调整积分")
                if submitted:
                    if points_change == 0:
                        st.error("积分变动值不能为0！")
                    else:
                        real_change = int(points_change) if change_type != "兑换扣减" else -abs(int(points_change))
                        adjust_points(sel_cust_id, real_change, change_type,
                                      pts_desc or change_type, pts_order)
                        sign = "增加" if real_change > 0 else "扣减"
                        st.success(f"✅ 已{sign}积分 {abs(real_change)}！")
                        st.rerun()

            st.markdown("---")
            st.subheader(f"{sel_cust_name} 的积分变动日志")
            pts_df = get_points_log(sel_cust_id)
            if pts_df.empty:
                st.info("该客户暂无积分变动记录。")
            else:
                pts_display = pts_df.rename(columns={
                    "id": "日志ID", "change_type": "变动类型", "points_change": "变动积分",
                    "balance_after": "变动后余额", "description": "描述",
                    "related_order": "关联订单", "created_at": "时间",
                })
                show_cols = ["日志ID", "变动类型", "变动积分", "变动后余额", "描述", "关联订单", "时间"]
                st.dataframe(pts_display[show_cols], use_container_width=True, hide_index=True)

    # ---------- 6. 客户标签 ----------
    with m6_sub6:
        st.subheader("客户标签")
        st.caption("管理标签库并为客户打标签，支持消费行为、客户属性、营销等分类。")

        with st.form("crm_tag_form"):
            col1, col2 = st.columns(2)
            with col1:
                tag_name = st.text_input("标签名称 *", placeholder="如：高消费客户", key="crm_tag_name")
                tag_type = st.selectbox("标签类型", M6_TAG_TYPES, key="crm_tag_type")
            with col2:
                tag_color = st.selectbox("标签颜色", M6_TAG_COLORS, key="crm_tag_color")
                auto_rule = st.text_input("自动规则", placeholder="如：消费满10000自动打标", key="crm_tag_rule")
            submitted = st.form_submit_button("💾 保存标签")
            if submitted:
                if not tag_name:
                    st.error("标签名称不能为空！")
                else:
                    add_customer_tag(tag_name, tag_type, tag_color, auto_rule)
                    st.success(f"✅ 标签「{tag_name}」已保存！")

        st.markdown("---")
        st.subheader("标签库")
        tag_df = get_all_customer_tags()
        if tag_df.empty:
            st.info("暂无标签数据。")
        else:
            tag_display = tag_df.rename(columns={
                "tag_id": "标签ID", "tag_name": "标签名称", "tag_type": "类型",
                "tag_color": "颜色", "auto_rule": "自动规则", "created_at": "创建时间",
            })
            show_cols = ["标签ID", "标签名称", "类型", "颜色", "自动规则", "创建时间"]
            st.dataframe(tag_display[show_cols], use_container_width=True, hide_index=True)

        st.markdown("---")
        st.subheader("为客户打标签")
        cust_df = get_all_customers()
        if cust_df.empty or tag_df.empty:
            st.info("需要先有客户和标签才能打标签。")
        else:
            col_a, col_b = st.columns(2)
            with col_a:
                cust_options = cust_df.apply(
                    lambda r: f'{r["customer_id"]} | {r["customer_name"]}', axis=1).tolist()
                tag_cust = st.selectbox("选择客户", cust_options, key="crm_tag_bind_cust")
                tag_cust_id = tag_cust.split(" | ")[0]
                tag_cust_name = tag_cust.split(" | ", 1)[1] if " | " in tag_cust else ""
            with col_b:
                tag_options = tag_df.apply(
                    lambda r: f'{int(r["tag_id"])} | {r["tag_name"]}', axis=1).tolist()
                tag_choice = st.selectbox("选择标签", tag_options, key="crm_tag_bind_tag")
                bind_tag_id = int(tag_choice.split(" | ")[0])
            if st.button("📌 绑定标签", key="crm_tag_bind_btn"):
                bind_customer_tag(tag_cust_id, bind_tag_id)
                st.success(f"✅ 已为「{tag_cust_name}」打上标签！")
                st.rerun()

            st.markdown("---")
            st.subheader(f"{tag_cust_name} 的标签")
            ctags_df = get_customer_tags(tag_cust_id)
            if ctags_df.empty:
                st.info("该客户暂无标签。")
            else:
                ctags_display = ctags_df.rename(columns={
                    "tag_id": "标签ID", "tag_name": "标签名称", "tag_type": "类型",
                    "tag_color": "颜色", "is_auto": "自动", "bind_time": "绑定时间",
                })
                show_cols = ["标签ID", "标签名称", "类型", "颜色", "自动", "绑定时间"]
                st.dataframe(ctags_display[show_cols], use_container_width=True, hide_index=True)

                with st.expander("🗑️ 解除标签"):
                    unbind_opts = ctags_df.apply(
                        lambda r: f'{int(r["tag_id"])} | {r["tag_name"]}', axis=1).tolist()
                    unbind_choice = st.selectbox("选择要解除的标签", unbind_opts, key="crm_tag_unbind")
                    if st.button("确认解除", key="crm_tag_unbind_btn"):
                        unbind_id = int(unbind_choice.split(" | ")[0])
                        unbind_customer_tag(tag_cust_id, unbind_id)
                        st.success("已解除标签")
                        st.rerun()

    # ---------- 7. 黑名单管理 ----------
    with m6_sub7:
        st.subheader("黑名单管理")
        st.caption("管理拉黑客户，支持永久与临时拉黑，移出后恢复客户状态。")

        bl_df = get_blacklist()
        total_bl = int(len(bl_df))
        perm_bl = int((bl_df["block_type"] == "永久").sum()) if not bl_df.empty else 0
        temp_bl = int((bl_df["block_type"] == "临时").sum()) if not bl_df.empty else 0
        bm1, bm2, bm3 = st.columns(3)
        bm1.metric("黑名单总数", total_bl)
        bm2.metric("永久拉黑", perm_bl)
        bm3.metric("临时拉黑", temp_bl)

        st.markdown("---")
        with st.form("crm_bl_form"):
            cust_df = get_all_customers()
            col1, col2 = st.columns(2)
            with col1:
                if cust_df.empty:
                    st.info("暂无客户数据。")
                    bl_cust_id = ""
                    bl_cust_name = ""
                else:
                    bl_options = cust_df.apply(
                        lambda r: f'{r["customer_id"]} | {r["customer_name"]}', axis=1).tolist()
                    bl_choice = st.selectbox("选择客户 *", bl_options, key="crm_bl_cust")
                    bl_cust_id = bl_choice.split(" | ")[0]
                    bl_cust_name = bl_choice.split(" | ", 1)[1] if " | " in bl_choice else ""
                bl_reason = st.text_input("拉黑原因 *", placeholder="如：恶意退款", key="crm_bl_reason")
                bl_type = st.selectbox("拉黑类型", M6_BLOCK_TYPES, key="crm_bl_type")
            with col2:
                bl_start = st.date_input("开始日期 *", value=datetime.now().date(), key="crm_bl_start")
                bl_end = st.date_input("结束日期", value=None, key="crm_bl_end")
                bl_operator = st.text_input("操作人", key="crm_bl_operator")
            submitted = st.form_submit_button("💾 加入黑名单")
            if submitted:
                if not bl_cust_id or not bl_reason:
                    st.error("客户和拉黑原因不能为空！")
                else:
                    end_str = str(bl_end) if bl_end else ("永久" if bl_type == "永久" else "")
                    add_to_blacklist(bl_cust_id, bl_cust_name, bl_reason, bl_type,
                                     str(bl_start), end_str, bl_operator)
                    st.success(f"✅ 已将「{bl_cust_name}」加入黑名单！")
                    st.rerun()

        st.markdown("---")
        st.subheader("黑名单列表")
        if bl_df.empty:
            st.info("暂无黑名单记录。")
        else:
            bl_display = bl_df.rename(columns={
                "id": "记录ID", "customer_id": "客户ID", "customer_name": "姓名",
                "reason": "拉黑原因", "block_type": "类型", "start_date": "开始日期",
                "end_date": "结束日期", "operator": "操作人", "status": "状态",
                "created_at": "创建时间",
            })
            show_cols = ["记录ID", "客户ID", "姓名", "拉黑原因", "类型", "开始日期", "结束日期", "操作人", "状态", "创建时间"]
            st.dataframe(bl_display[show_cols], use_container_width=True, hide_index=True)

            with st.expander("✅ 移出黑名单"):
                rm_opts = bl_df.apply(
                    lambda r: f'{int(r["id"])} | {r["customer_name"]}（{r["block_type"]}）',
                    axis=1).tolist()
                rm_choice = st.selectbox("选择要移出的记录", rm_opts, key="crm_bl_rm")
                if st.button("确认移出", key="crm_bl_rm_btn"):
                    rm_id = int(rm_choice.split(" | ")[0])
                    remove_from_blacklist(rm_id)
                    st.success("已移出黑名单，客户状态已恢复")
                    st.rerun()

    # ---------- 8. 客户画像分析 ----------
    with m6_sub8:
        st.subheader("📊 客户画像分析")
        st.caption("全维度客户数据分析看板：会员等级分布、平台来源、消费特征、RFM分层、地域分布、增长趋势。")

        import plotly.express as px
        import plotly.graph_objects as go
        from io import BytesIO as _BytesIO

        customers_df = get_all_customers()

        if customers_df.empty:
            st.warning("暂无客户数据，请先在「会员档案」中录入客户。")
        else:
            # ========== 1. 核心指标卡片 ==========
            total_customers = len(customers_df)
            active_customers = len(customers_df[customers_df["status"] == "正常"])
            blacklist_count = len(customers_df[customers_df["status"] == "黑名单"])
            total_consumption = customers_df["total_consumption"].sum()
            total_points = customers_df["total_points"].sum()
            avg_order_value = customers_df["avg_order_value"].mean()

            col1, col2, col3, col4, col5, col6 = st.columns(6)
            with col1:
                st.metric("客户总数", f"{total_customers}")
            with col2:
                st.metric("正常客户", f"{active_customers}")
            with col3:
                st.metric("黑名单", f"{blacklist_count}")
            with col4:
                st.metric("消费总额", fmt_money(total_consumption))
            with col5:
                st.metric("积分总额", f"{int(total_points):,}")
            with col6:
                st.metric("平均客单价", fmt_money(avg_order_value))

            st.markdown("---")

            # ========== 2. 会员等级分布（饼图）==========
            col_a, col_b = st.columns(2)
            with col_a:
                st.markdown("#### 会员等级分布")
                level_counts = customers_df.groupby("level_name").size().reset_index(name="人数")
                if not level_counts.empty:
                    fig_level = px.pie(
                        level_counts, values="人数", names="level_name",
                        color="level_name",
                        color_discrete_sequence=px.colors.qualitative.Set2,
                        hole=0.4,
                    )
                    fig_level.update_layout(height=320, margin=dict(l=20, r=20, t=20, b=20))
                    st.plotly_chart(fig_level, use_container_width=True)
                else:
                    st.info("暂无等级数据")

            # ========== 3. 客户状态分布（饼图）==========
            with col_b:
                st.markdown("#### 客户状态分布")
                status_counts = customers_df.groupby("status").size().reset_index(name="人数")
                if not status_counts.empty:
                    fig_status = px.pie(
                        status_counts, values="人数", names="status",
                        color="status",
                        color_discrete_map={"正常": "#52c41a", "黑名单": "#ff4d4f"},
                        hole=0.4,
                    )
                    fig_status.update_layout(height=320, margin=dict(l=20, r=20, t=20, b=20))
                    st.plotly_chart(fig_status, use_container_width=True)
                else:
                    st.info("暂无状态数据")

            st.markdown("---")

            # ========== 4. 平台来源分布（柱状图）==========
            col_c, col_d = st.columns(2)
            with col_c:
                st.markdown("#### 平台来源分布")
                try:
                    pa_df = get_all_platform_accounts_for_analysis()
                    if pa_df is not None and not pa_df.empty:
                        platform_counts = pa_df.groupby("platform").size().reset_index(name="客户数")
                        fig_platform = px.bar(
                            platform_counts, x="platform", y="客户数",
                            color="platform",
                            color_discrete_sequence=px.colors.qualitative.Pastel,
                            text="客户数",
                        )
                        fig_platform.update_layout(height=320, margin=dict(l=20, r=20, t=20, b=20), showlegend=False)
                        st.plotly_chart(fig_platform, use_container_width=True)
                    else:
                        st.info("暂无平台账号绑定数据")
                except Exception:
                    st.info("暂无平台账号绑定数据")

            # ========== 5. 性别分布 ==========
            with col_d:
                st.markdown("#### 性别分布")
                gender_counts = customers_df.groupby("gender").size().reset_index(name="人数")
                if not gender_counts.empty:
                    fig_gender = px.pie(
                        gender_counts, values="人数", names="gender",
                        color="gender",
                        color_discrete_map={"男": "#1890ff", "女": "#eb4845", "未知": "#bfbfbf"},
                        hole=0.4,
                    )
                    fig_gender.update_layout(height=320, margin=dict(l=20, r=20, t=20, b=20))
                    st.plotly_chart(fig_gender, use_container_width=True)
                else:
                    st.info("暂无性别数据")

            st.markdown("---")

            # ========== 6. 消费 TOP10 客户排行（横向柱状图）==========
            col_e, col_f = st.columns(2)
            with col_e:
                st.markdown("#### 消费 TOP10 客户")
                top10 = customers_df.nlargest(10, "total_consumption")[["customer_name", "total_consumption"]]
                if not top10.empty:
                    fig_top = px.bar(
                        top10, x="total_consumption", y="customer_name",
                        orientation="h",
                        color="total_consumption",
                        color_continuous_scale="Viridis",
                        text="total_consumption",
                    )
                    fig_top.update_layout(height=360, margin=dict(l=20, r=20, t=20, b=20), showlegend=False)
                    fig_top.update_traces(texttemplate="%{text:,.2f}", textposition="outside")
                    st.plotly_chart(fig_top, use_container_width=True)
                else:
                    st.info("暂无消费数据")

            # ========== 7. 地域分布 TOP10 ==========
            with col_f:
                st.markdown("#### 地域分布 TOP10")
                region_counts = customers_df.groupby("province").size().reset_index(name="客户数")
                region_counts = region_counts[region_counts["province"].notna() & (region_counts["province"] != "")]
                if not region_counts.empty:
                    region_counts = region_counts.nlargest(10, "客户数")
                    fig_region = px.bar(
                        region_counts, x="province", y="客户数",
                        color="客户数",
                        color_continuous_scale="Blugrn",
                    )
                    fig_region.update_layout(height=360, margin=dict(l=20, r=20, t=20, b=20), showlegend=False)
                    st.plotly_chart(fig_region, use_container_width=True)
                else:
                    st.info("暂无地域数据")

            st.markdown("---")

            # ========== 8. RFM 客户分层分析 ==========
            st.markdown("#### RFM 客户分层分析")
            st.caption("R=最近消费距今天数，F=消费频次，M=消费金额。自动分为8类客户。")

            consumption_df = get_all_consumption_for_analysis()
            if consumption_df is not None and not consumption_df.empty:
                import datetime as _dt

                today = _dt.date.today()
                rfm = consumption_df.groupby("customer_id").agg(
                    last_order_date=("order_date", "max"),
                    frequency=("order_id", "count"),
                    monetary=("order_amount", "sum"),
                ).reset_index()

                rfm["R_days"] = rfm["last_order_date"].apply(
                    lambda x: (today - _dt.datetime.strptime(x, "%Y-%m-%d").date()).days if x and len(str(x)) >= 10 else 999
                )

                # 用中位数分高低
                r_med = rfm["R_days"].median()
                f_med = rfm["frequency"].median()
                m_med = rfm["monetary"].median()

                def rfm_label(row):
                    r_high = row["R_days"] <= r_med
                    f_high = row["frequency"] > f_med
                    m_high = row["monetary"] > m_med
                    if r_high and f_high and m_high:
                        return "重要价值客户"
                    elif r_high and not f_high and m_high:
                        return "重要发展客户"
                    elif not r_high and f_high and m_high:
                        return "重要保持客户"
                    elif not r_high and not f_high and m_high:
                        return "重要挽留客户"
                    elif r_high and f_high and not m_high:
                        return "一般价值客户"
                    elif r_high and not f_high and not m_high:
                        return "一般发展客户"
                    elif not r_high and f_high and not m_high:
                        return "一般保持客户"
                    else:
                        return "一般挽留客户"

                rfm["客户类型"] = rfm.apply(rfm_label, axis=1)

                rfm_summary = rfm.groupby("客户类型").agg(
                    人数=("customer_id", "count"),
                    平均消费=("monetary", "mean"),
                    平均频次=("frequency", "mean"),
                ).reset_index()

                # 类型排序
                type_order = [
                    "重要价值客户", "重要发展客户", "重要保持客户", "重要挽留客户",
                    "一般价值客户", "一般发展客户", "一般保持客户", "一般挽留客户",
                ]
                rfm_summary["sort"] = rfm_summary["客户类型"].apply(
                    lambda x: type_order.index(x) if x in type_order else 99
                )
                rfm_summary = rfm_summary.sort_values("sort").drop(columns=["sort"])

                col_rfm1, col_rfm2 = st.columns([3, 2])
                with col_rfm1:
                    fig_rfm = px.bar(
                        rfm_summary, x="客户类型", y="人数",
                        color="客户类型",
                        color_discrete_sequence=px.colors.qualitative.Set3,
                        text="人数",
                    )
                    fig_rfm.update_layout(height=350, margin=dict(l=20, r=20, t=20, b=20), showlegend=False)
                    st.plotly_chart(fig_rfm, use_container_width=True)
                with col_rfm2:
                    rfm_summary_display = fmt_money_df(rfm_summary, ["平均消费"])
                    st.dataframe(rfm_summary_display, use_container_width=True, hide_index=True)

                st.markdown("---")
                st.markdown("#### RFM 分层解读")
                rfm_explain = [
                    ("重要价值客户", "最近消费+高频+高金额", "VIP客户，重点维护，提供专属服务"),
                    ("重要发展客户", "最近消费+低频+高金额", "高潜力客户，提升复购频率"),
                    ("重要保持客户", "较久未消费+高频+高金额", "曾忠实但可能流失，主动触达"),
                    ("重要挽留客户", "较久未消费+低频+高金额", "大额偶购客户，需唤醒"),
                    ("一般价值客户", "最近消费+高频+低金额", "活跃小客，引导提升客单价"),
                    ("一般发展客户", "最近消费+低频+低金额", "新客或偶购，培育习惯"),
                    ("一般保持客户", "较久未消费+高频+低金额", "曾经活跃小客，需激活"),
                    ("一般挽留客户", "较久未消费+低频+低金额", "低价值客户，低成本维护"),
                ]
                for label, desc, action in rfm_explain:
                    if label in rfm_summary["客户类型"].values:
                        st.markdown(f"**{label}** — {desc}")
                        st.caption(f"  策略：{action}")
            else:
                st.info("暂无消费记录数据，录入消费记录后可生成RFM分析。")

            st.markdown("---")

            # ========== 9. 会员等级 vs 消费金额（散点图）==========
            st.markdown("#### 等级与消费金额关系")
            scatter_df = customers_df[customers_df["total_consumption"] > 0].copy()
            if not scatter_df.empty:
                fig_scatter = px.scatter(
                    scatter_df,
                    x="total_orders",
                    y="total_consumption",
                    color="level_name",
                    size="total_points",
                    hover_data=["customer_name", "phone"],
                    color_discrete_sequence=px.colors.qualitative.Set2,
                )
                fig_scatter.update_layout(
                    height=380, margin=dict(l=20, r=20, t=20, b=20),
                    xaxis_title="消费频次（订单数）",
                    yaxis_title="消费总金额（元）",
                )
                st.plotly_chart(fig_scatter, use_container_width=True)
            else:
                st.info("暂无消费数据")

            st.markdown("---")

            # ========== 10. 导出分析报告 ==========
            st.markdown("#### 导出分析报告")
            def export_crm_report():
                output = _BytesIO()
                with pd.ExcelWriter(output, engine="openpyxl") as writer:
                    customers_df.to_excel(writer, sheet_name="客户总表", index=False)
                    if not level_counts.empty:
                        level_counts.to_excel(writer, sheet_name="等级分布", index=False)
                    if consumption_df is not None and not consumption_df.empty:
                        rfm.to_excel(writer, sheet_name="RFM分析", index=False)
                        rfm_summary.to_excel(writer, sheet_name="RFM汇总", index=False)
                output.seek(0)
                return output

            st.download_button(
                label="📥 导出CRM分析报告（Excel）",
                data=export_crm_report(),
                file_name="CRM客户画像分析报告.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="crm_report_export",
            )


# ============================================================
# 【📦 供应链管理】模块七：全渠道订单OMS（订单履约）
# ============================================================
with tab7:
    import plotly.express as px
    st.header("📋 订单履约 OMS")
    st.caption("订单录入 · 智能审核 · 合并拆分 · 打单发货 · 异常拦截 · 物流同步 · 统计看板")

    # 初始化预置物流公司（仅首次运行）
    if not st.session_state.get("_preset_logistics_done"):
        init_preset_logistics()
        st.session_state["_preset_logistics_done"] = True

    oms_sub1, oms_sub2, oms_sub3, oms_sub4, oms_sub5, oms_sub6, oms_sub7 = st.tabs([
        "订单看板", "订单列表与录入", "订单审核", "订单合并与拆分",
        "批量打单发货", "异常订单管理", "物流管理",
    ])

    # 订单列表中文显示映射
    OMS_ORDER_DISPLAY = {
        "order_id": "订单号", "platform": "平台", "shop_name": "店铺",
        "customer_name": "客户", "receiver_name": "收货人",
        "receiver_phone": "电话", "actual_amount": "实付金额",
        "total_amount": "商品总额", "order_status": "状态",
        "order_type": "类型", "order_time": "下单时间",
        "is_abnormal": "异常标记", "logistics_company": "物流公司",
        "logistics_number": "物流单号", "created_at": "创建时间",
        "abnormal_reason": "异常原因", "parent_order_id": "合并自",
        "split_from": "拆分自",
    }

    # ---------- 1. 订单看板 ----------
    with oms_sub1:
        st.subheader("订单看板")
        stats = get_order_stats()
        sc = stats["status_counts"]
        c1, c2, c3, c4, c5, c6 = st.columns(6)
        c1.metric("待审核", sc.get("待审核", 0))
        c2.metric("待发货", sc.get("待发货", 0))
        c3.metric("待打单", sc.get("待打单", 0))
        c4.metric("已发货", sc.get("已发货", 0))
        c5.metric("异常订单", stats["abnormal"])
        c6.metric("今日新增", stats["today_new"])

        st.markdown("---")
        col_l, col_r = st.columns(2)
        with col_l:
            st.markdown("#### 各状态订单数量")
            status_rows = [{"状态": s, "数量": sc.get(s, 0)} for s in OMS_STATUSES if s in sc]
            if not status_rows:
                st.info("暂无订单数据")
            else:
                status_df = pd.DataFrame(status_rows)
                fig_s = px.bar(status_df, x="状态", y="数量", color="状态",
                               text="数量",
                               color_discrete_sequence=px.colors.qualitative.Set3)
                fig_s.update_layout(height=380, margin=dict(l=20, r=20, t=20, b=20),
                                    showlegend=False)
                st.plotly_chart(fig_s, use_container_width=True)
        with col_r:
            st.markdown("#### 各平台订单数量")
            pc = stats["platform_counts"]
            if not pc:
                st.info("暂无平台数据")
            else:
                plat_df = pd.DataFrame([{"平台": k, "数量": v} for k, v in pc.items()])
                fig_p = px.pie(plat_df, names="平台", values="数量",
                               color_discrete_sequence=px.colors.qualitative.Pastel)
                fig_p.update_layout(height=380, margin=dict(l=20, r=20, t=20, b=20))
                st.plotly_chart(fig_p, use_container_width=True)

        st.markdown("---")
        st.markdown("#### 最近10条订单")
        recent_df = get_all_orders()
        if recent_df.empty:
            st.info("暂无订单数据")
        else:
            rview = recent_df.head(10).rename(columns=OMS_ORDER_DISPLAY)
            rview["异常标记"] = rview["异常标记"].apply(lambda x: "⚠️异常" if x == 1 else "正常")
            rview = fmt_money_df(rview, ["实付金额", "商品总额"])
            rshow = ["订单号", "平台", "客户", "收货人", "实付金额", "状态", "类型", "下单时间", "异常标记"]
            st.dataframe(rview[rshow], use_container_width=True, hide_index=True)

    # ---------- 2. 订单列表与录入 ----------
    with oms_sub2:
        st.subheader("订单列表与录入")
        col_f1, col_f2, col_f3 = st.columns([1, 1, 2])
        with col_f1:
            f_status = st.selectbox("按状态筛选", ["全部"] + OMS_STATUSES, key="oms_filter_status")
        with col_f2:
            f_platform = st.selectbox("按平台筛选", ["全部"] + OMS_PLATFORMS, key="oms_filter_platform")
        with col_f3:
            f_search = st.text_input("搜索订单号/客户", placeholder="输入关键字", key="oms_search")

        orders_df = get_all_orders(
            status_filter=None if f_status == "全部" else f_status,
            platform_filter=None if f_platform == "全部" else f_platform,
        )
        if f_search and not orders_df.empty:
            mask = orders_df.apply(
                lambda r: f_search.lower() in str(r.get("order_id", "")).lower()
                or f_search.lower() in str(r.get("customer_name", "")).lower(),
                axis=1)
            orders_df = orders_df[mask]

        col_exp, _ = st.columns([1, 3])
        with col_exp:
            if not orders_df.empty:
                from io import BytesIO
                export_buf = BytesIO()
                with pd.ExcelWriter(export_buf, engine="openpyxl") as writer:
                    orders_df.to_excel(writer, sheet_name="订单列表", index=False)
                export_buf.seek(0)
                st.download_button("📥 导出订单Excel", data=export_buf,
                                   file_name="订单列表.xlsx", key="oms_order_export",
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        if orders_df.empty:
            st.info("暂无符合条件的订单，请在下方录入。")
        else:
            display = orders_df.rename(columns=OMS_ORDER_DISPLAY)
            display["异常标记"] = display["异常标记"].apply(lambda x: "⚠️异常" if x == 1 else "正常")
            display = fmt_money_df(display, ["实付金额", "商品总额"])
            oshow = ["订单号", "平台", "客户", "收货人", "实付金额", "状态", "类型", "下单时间", "异常标记"]
            view = display[oshow].copy()

            def _hl_abnormal(row):
                if row["异常标记"] == "⚠️异常":
                    return ["background-color: #fff1f0; color: #cf1322"] * len(row)
                return [""] * len(row)
            styler = view.style.apply(_hl_abnormal, axis=1)
            styler.hide(axis="index")
            st.dataframe(styler, use_container_width=True)

        # 展开式录入表单
        with st.expander("➕ 新增订单"):
            st.markdown("**基本信息**")
            b1, b2, b3, b4 = st.columns(4)
            with b1:
                in_order_id = st.text_input("订单号 *", placeholder="如：DD2024001", key="oms_in_oid")
            with b2:
                in_platform = st.selectbox("平台 *", OMS_PLATFORMS, key="oms_in_platform")
            with b3:
                shop_df = get_all_shops()
                shop_platform_df = shop_df[shop_df["platform"] == in_platform] if not shop_df.empty else shop_df
                if not shop_platform_df.empty:
                    shop_opts = shop_platform_df.apply(
                        lambda r: f'{r["shop_id"]} | {r["shop_name"]}', axis=1).tolist()
                    shop_sel = st.selectbox("店铺", [""] + shop_opts, key="oms_in_shop")
                    in_shop_id = shop_sel.split(" | ")[0] if shop_sel and " | " in shop_sel else ""
                    in_shop_name = shop_sel.split(" | ", 1)[1] if shop_sel and " | " in shop_sel else ""
                else:
                    st.selectbox("店铺", ["（无店铺，请先在模块五录入）"], key="oms_in_shop_empty")
                    in_shop_id = ""
                    in_shop_name = ""
            with b4:
                in_order_type = st.selectbox("订单类型", OMS_ORDER_TYPES, key="oms_in_type")

            cust_df = get_all_customers()
            if cust_df.empty:
                st.warning("暂无客户数据，请先在模块六录入客户。")
                cust_opts = []
            else:
                cust_opts = cust_df.apply(
                    lambda r: f'{r["customer_id"]} | {r["customer_name"]}', axis=1).tolist()
            cust_sel = st.selectbox("客户 *", [""] + cust_opts, key="oms_in_customer")
            in_cust_id = cust_sel.split(" | ")[0] if cust_sel and " | " in cust_sel else ""
            in_cust_name = cust_sel.split(" | ", 1)[1] if cust_sel and " | " in cust_sel else ""

            st.markdown("**收货信息**")
            r1, r2, r3, r4 = st.columns(4)
            with r1:
                in_rname = st.text_input("收货人", key="oms_in_rname")
            with r2:
                in_rphone = st.text_input("电话", key="oms_in_rphone")
            with r3:
                in_rprovince = st.text_input("省份", key="oms_in_rprov")
            with r4:
                in_rcity = st.text_input("城市", key="oms_in_rcity")
            r5, r6 = st.columns(2)
            with r5:
                in_rdistrict = st.text_input("区/县", key="oms_in_rdist")
            with r6:
                in_raddr = st.text_input("详细地址", key="oms_in_raddr")

            st.markdown("**金额信息**（实付金额 = 商品总额 - 优惠金额 + 运费）")
            a1, a2, a3, a4 = st.columns(4)
            with a1:
                in_discount = money_input("优惠金额", key="oms_in_discount", min_value=0.0)
            with a2:
                in_shipping = money_input("运费", key="oms_in_shipping", min_value=0.0)
            with a3:
                in_payment = st.selectbox("支付方式", OMS_PAYMENT_METHODS, key="oms_in_payment")
            with a4:
                in_ordertime = st.date_input("下单日期", value=datetime.now().date(), key="oms_in_odate")

            st.markdown("**备注与发票**")
            n1, n2 = st.columns(2)
            with n1:
                in_seller_remark = st.text_input("卖家备注", key="oms_in_sremark")
            with n2:
                in_buyer_remark = st.text_input("买家备注", key="oms_in_bremark")
            n3, n4 = st.columns(2)
            with n3:
                in_invoice_req = st.checkbox("需要发票", key="oms_in_invreq")
            with n4:
                in_invoice_title = st.text_input("发票抬头", key="oms_in_invtitle", disabled=not in_invoice_req)

            st.markdown("**订单明细**（动态添加行）")
            if "oms_item_rows" not in st.session_state:
                st.session_state.oms_item_rows = 2
            ca, cd, ci = st.columns([1, 1, 3])
            with ca:
                if st.button("➕ 添加明细行", key="oms_item_add"):
                    st.session_state.oms_item_rows += 1
                    st.rerun()
            with cd:
                if st.session_state.oms_item_rows > 2:
                    if st.button("➖ 删除最后一行", key="oms_item_del"):
                        st.session_state.oms_item_rows -= 1
                        st.rerun()
            with ci:
                st.caption(f"当前 {st.session_state.oms_item_rows} 行")

            spu_df = get_all_spus()
            if not spu_df.empty:
                spu_opts = [""] + spu_df.apply(
                    lambda r: f'{r["spu_code"]} | {r["spu_name"]}', axis=1).tolist()
            else:
                spu_opts = [""]
            detail_lines = []
            for i in range(st.session_state.oms_item_rows):
                dc = st.columns([2.2, 2.2, 1.1, 1.3, 0.8])
                with dc[0]:
                    spu_sel = st.selectbox("SPU", spu_opts, key=f"oms_d_spu_{i}", label_visibility="collapsed")
                    d_spu_code = spu_sel.split(" | ")[0] if spu_sel and " | " in spu_sel else ""
                    d_spu_name = spu_sel.split(" | ", 1)[1] if spu_sel and " | " in spu_sel else ""
                with dc[1]:
                    if d_spu_code:
                        sku_df_i = get_skus_by_spu(d_spu_code)
                        if not sku_df_i.empty:
                            sku_opts_i = [""] + sku_df_i.apply(
                                lambda r: f'{r["sku_code"]} | {r["sku_name"]}', axis=1).tolist()
                        else:
                            sku_opts_i = [""]
                    else:
                        sku_df_i = pd.DataFrame()
                        sku_opts_i = [""]
                    sku_sel = st.selectbox("SKU", sku_opts_i, key=f"oms_d_sku_{i}", label_visibility="collapsed")
                    d_sku_code = sku_sel.split(" | ")[0] if sku_sel and " | " in sku_sel else ""
                    d_sku_name = sku_sel.split(" | ", 1)[1] if sku_sel and " | " in sku_sel else ""
                    d_spec = ""
                    if d_sku_code and not sku_df_i.empty:
                        m = sku_df_i[sku_df_i["sku_code"] == d_sku_code]
                        if not m.empty:
                            d_spec = str(m.iloc[0].get("spec_attrs", "") or "")
                with dc[2]:
                    d_qty = st.number_input("数量", min_value=0, value=0, step=1, key=f"oms_d_qty_{i}", label_visibility="collapsed")
                with dc[3]:
                    d_price = money_input("单价", key=f"oms_d_price_{i}", min_value=0.0,
                                          label_visibility="collapsed")
                with dc[4]:
                    d_gift = st.checkbox("赠品", key=f"oms_d_gift_{i}")
                if d_sku_code and d_qty > 0:
                    detail_lines.append({
                        "spu_code": d_spu_code, "spu_name": d_spu_name,
                        "sku_code": d_sku_code, "sku_name": d_sku_name,
                        "spec_attrs": d_spec, "quantity": int(d_qty),
                        "unit_price": float(d_price), "total_price": float(d_qty) * float(d_price),
                        "is_gift": d_gift,
                    })

            in_total_amount = sum(l["total_price"] for l in detail_lines if not l["is_gift"])
            in_actual = in_total_amount - float(in_discount) + float(in_shipping)
            st.markdown(f"**商品总额：** {fmt_money(in_total_amount)} 元　|　**实付金额：** {fmt_money(in_actual)} 元")

            if st.button("💾 保存订单", key="oms_save_order", type="primary"):
                if not in_order_id:
                    st.error("订单号不能为空！")
                elif not in_cust_id:
                    st.error("请选择客户！")
                elif not detail_lines:
                    st.error("请至少录入一条订单明细！")
                else:
                    add_order(in_order_id, in_platform, in_shop_id, in_shop_name,
                              in_cust_id, in_cust_name, in_rname, in_rphone,
                              in_rprovince, in_rcity, in_rdistrict, in_raddr,
                              in_order_type, in_total_amount, in_discount,
                              in_shipping, in_actual, in_payment,
                              str(in_ordertime), in_seller_remark, in_buyer_remark,
                              in_invoice_req, in_invoice_title)
                    for l in detail_lines:
                        add_order_item(in_order_id, l["spu_code"], l["spu_name"],
                                       l["sku_code"], l["sku_name"], l["spec_attrs"],
                                       l["quantity"], l["unit_price"], l["total_price"],
                                       l["is_gift"])
                    # 录入后确认付款并流转到待审核，同时做异常检查
                    update_order_status(in_order_id, "待审核", "录入员", "订单录入并确认付款")
                    check_abnormal(in_order_id)
                    st.success(f"✅ 订单「{in_order_id}」已保存，已自动进行异常检查。")
                    st.session_state.oms_item_rows = 2
                    st.rerun()

    # ---------- 3. 订单审核 ----------
    with oms_sub3:
        st.subheader("订单审核")
        review_df = get_all_orders(status_filter="待审核")
        if st.button("⚡ 一键批量审核所有待审核订单", key="oms_batch_review", type="primary"):
            res = batch_auto_review()
            st.success(f"批量审核完成：共 {res['total']} 单，通过 {res['passed']} 单，拦截 {res['failed']} 单。")
            st.rerun()
        st.markdown("---")
        if review_df.empty:
            st.info("暂无待审核订单。")
        else:
            for _, row in review_df.iterrows():
                oid = row["order_id"]
                items = get_order_items(oid)
                abn = "⚠️异常" if row.get("is_abnormal") == 1 else "正常"
                title = f"{oid} | {row.get('customer_name', '')} | 实付 {fmt_money(float(row.get('actual_amount', 0) or 0))} | 明细 {len(items)} 条 | {abn}"
                with st.expander(title):
                    rc1, rc2, rc3, rc4 = st.columns(4)
                    rc1.metric("实付金额", fmt_money(float(row.get('actual_amount', 0) or 0)))
                    rc2.metric("明细数", len(items))
                    rc3.metric("平台", row.get("platform", ""))
                    rc4.metric("异常", abn)
                    st.markdown("**收货信息**")
                    addr_str = f"{row.get('receiver_name', '')}　{row.get('receiver_phone', '')}　{row.get('receiver_province', '')}{row.get('receiver_city', '')}{row.get('receiver_district', '')}{row.get('receiver_address', '')}"
                    st.caption(addr_str)
                    st.markdown("**订单明细**")
                    if items.empty:
                        st.warning("无明细")
                    else:
                        idisp = items.rename(columns={
                            "spu_name": "商品", "sku_name": "规格", "spec_attrs": "属性",
                            "quantity": "数量", "unit_price": "单价", "total_price": "小计",
                            "is_gift": "赠品"})
                        idisp["赠品"] = idisp["赠品"].apply(lambda x: "是" if x == 1 else "否")
                        idisp = fmt_money_df(idisp, ["单价", "小计"])
                        st.dataframe(idisp[["商品", "规格", "属性", "数量", "单价", "小计", "赠品"]],
                                     use_container_width=True, hide_index=True)
                    bc1, bc2 = st.columns(2)
                    with bc1:
                        if st.button("✅ 通过审核", key=f"oms_pass_{oid}"):
                            update_order_status(oid, "待发货", "审核员", "手动审核通过")
                            st.success(f"订单 {oid} 已通过审核。")
                            st.rerun()
                    with bc2:
                        if st.button("🚫 手动拦截", key=f"oms_block_{oid}"):
                            st.session_state[f"oms_block_show_{oid}"] = True
                            st.rerun()
                    if st.session_state.get(f"oms_block_show_{oid}"):
                        with st.form(f"oms_block_form_{oid}"):
                            block_reason = st.text_input("拦截原因", key=f"oms_block_reason_{oid}")
                            bs = st.form_submit_button("确认拦截")
                            if bs:
                                if not block_reason:
                                    st.error("请填写拦截原因")
                                else:
                                    manual_intercept_order(oid, block_reason, "高", "审核员")
                                    st.success(f"订单 {oid} 已拦截并标记异常。")
                                    st.session_state[f"oms_block_show_{oid}"] = False
                                    st.rerun()

    # ---------- 4. 订单合并与拆分 ----------
    with oms_sub4:
        st.subheader("订单合并与拆分")

        st.markdown("##### 上半部分：合并订单（选择多个待发货订单）")
        merge_df = get_all_orders(status_filter="待发货")
        if merge_df.empty:
            st.info("暂无待发货订单可供合并。")
        else:
            merge_opts = merge_df.apply(
                lambda r: f'{r["order_id"]} | {r.get("customer_name", "")} | {fmt_money(float(r.get("actual_amount", 0) or 0))}',
                axis=1).tolist()
            merge_sel = st.multiselect("选择要合并的订单（至少2个）", merge_opts, key="oms_merge_sel")
            with st.form("oms_merge_form"):
                merge_reason = st.text_input("合并原因", key="oms_merge_reason")
                merge_op = st.text_input("操作人", value="运营", key="oms_merge_op")
                ms = st.form_submit_button("🔗 执行合并")
                if ms:
                    if len(merge_sel) < 2:
                        st.error("请至少选择2个订单进行合并！")
                    elif not merge_reason:
                        st.error("请填写合并原因！")
                    else:
                        m_ids = [s.split(" | ")[0] for s in merge_sel]
                        new_id = merge_orders(m_ids, merge_op, merge_reason)
                        if new_id:
                            st.success(f"✅ 合并成功，新订单号：{new_id}")
                            st.rerun()
                        else:
                            st.error("合并失败，请重试。")

        st.markdown("---")
        st.markdown("##### 下半部分：拆分订单（选择一个订单，拆出部分明细）")
        split_pool_df = get_all_orders(status_filter="待发货")
        if split_pool_df.empty:
            st.info("暂无可拆分的待发货订单。")
        else:
            split_opts = split_pool_df.apply(
                lambda r: f'{r["order_id"]} | {r.get("customer_name", "")}',
                axis=1).tolist()
            split_sel = st.selectbox("选择要拆分的订单", [""] + split_opts, key="oms_split_sel")
            split_oid = split_sel.split(" | ")[0] if split_sel and " | " in split_sel else ""
            if split_oid:
                split_items_df = get_order_items(split_oid)
                if split_items_df.empty:
                    st.warning("该订单无明细可拆分。")
                else:
                    st.markdown("**勾选要拆出的明细行并输入数量**")
                    split_picks = []
                    for _, it in split_items_df.iterrows():
                        sku_code = it.get("sku_code", "")
                        orig_qty = int(it.get("quantity", 0) or 0)
                        sc1, sc2, sc3, sc4, sc5 = st.columns([0.6, 2.5, 1.2, 1.2, 1.2])
                        with sc1:
                            chk = st.checkbox("拆", key=f"oms_split_chk_{sku_code}_{it.get('id')}")
                        with sc2:
                            st.caption(f"{it.get('sku_name', '')}（{it.get('spec_attrs', '')}）")
                        with sc3:
                            st.caption(f"原数量：{orig_qty}")
                        with sc4:
                            sq = st.number_input("拆出数量", min_value=0, max_value=orig_qty,
                                                 value=0, step=1, key=f"oms_split_qty_{sku_code}_{it.get('id')}")
                        with sc5:
                            st.caption(f"单价：{fmt_money(float(it.get('unit_price', 0) or 0))}")
                        if chk and sq > 0:
                            split_picks.append({"sku_code": sku_code, "quantity": int(sq)})
                    with st.form("oms_split_form"):
                        split_reason = st.text_input("拆分原因", key="oms_split_reason")
                        split_op = st.text_input("操作人", value="运营", key="oms_split_op")
                        ss = st.form_submit_button("✂️ 执行拆分")
                        if ss:
                            if not split_picks:
                                st.error("请勾选并填写要拆出的明细数量！")
                            elif not split_reason:
                                st.error("请填写拆分原因！")
                            else:
                                new_id = split_order(split_oid, split_picks, split_op, split_reason)
                                if new_id:
                                    st.success(f"✅ 拆分成功，新订单号：{new_id}")
                                    st.rerun()
                                else:
                                    st.error("拆分失败，请重试。")

    # ---------- 5. 批量打单发货 ----------
    with oms_sub5:
        st.subheader("批量打单发货")
        ship_pool_df = get_all_orders()
        ship_pool_df = ship_pool_df[ship_pool_df["order_status"].isin(["待发货", "待打单", "待揽收"])] if not ship_pool_df.empty else ship_pool_df
        logi_df = get_logistics_companies()
        if not logi_df.empty:
            logi_opts = logi_df["company_name"].tolist()
        else:
            logi_opts = ["（请先在物流管理中添加公司）"]

        col_exp2, _ = st.columns([1, 3])
        with col_exp2:
            if not ship_pool_df.empty:
                from io import BytesIO
                ship_exp_buf = BytesIO()
                ship_exp_view = ship_pool_df.rename(columns=OMS_ORDER_DISPLAY)
                with pd.ExcelWriter(ship_exp_buf, engine="openpyxl") as writer:
                    ship_exp_view.to_excel(writer, sheet_name="待发货清单", index=False)
                ship_exp_buf.seek(0)
                st.download_button("📥 导出待发货清单", data=ship_exp_buf,
                                   file_name="待发货清单.xlsx", key="oms_ship_export",
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        if ship_pool_df.empty:
            st.info("暂无待发货/待打单/待揽收订单。")
        else:
            st.caption("为每个订单选择物流公司并填写物流单号，然后点击底部批量发货。")
            for _, row in ship_pool_df.iterrows():
                oid = row["order_id"]
                with st.container(border=True):
                    sc1, sc2, sc3, sc4 = st.columns([2, 2, 2, 2])
                    with sc1:
                        st.caption(f"订单号：{oid}")
                        st.caption(f"客户：{row.get('customer_name', '')}")
                    with sc2:
                        st.caption(f"收货人：{row.get('receiver_name', '')} {row.get('receiver_phone', '')}")
                        st.caption(f"地址：{row.get('receiver_province', '')}{row.get('receiver_city', '')}{row.get('receiver_district', '')}{row.get('receiver_address', '')}")
                    with sc3:
                        st.selectbox("物流公司", logi_opts, key=f"oms_ship_lg_{oid}", label_visibility="collapsed")
                    with sc4:
                        st.text_input("物流单号", key=f"oms_ship_num_{oid}", label_visibility="collapsed", placeholder="输入物流单号")

            if st.button("🚚 批量发货（已填单号的订单）", key="oms_batch_ship", type="primary"):
                ship_list = []
                for _, row in ship_pool_df.iterrows():
                    oid = row["order_id"]
                    lg = st.session_state.get(f"oms_ship_lg_{oid}", "")
                    num = st.session_state.get(f"oms_ship_num_{oid}", "")
                    if num:
                        ship_list.append({
                            "order_id": oid,
                            "logistics_company": lg,
                            "logistics_number": num,
                        })
                if not ship_list:
                    st.warning("没有填写物流单号的订单，无法发货。")
                else:
                    cnt = batch_ship(ship_list, "打单员")
                    st.success(f"✅ 批量发货完成，共发货 {cnt} 单。")
                    st.rerun()

    # ---------- 6. 异常订单管理 ----------
    with oms_sub6:
        st.subheader("异常订单管理")
        abn_df = get_abnormal_orders()
        pending_df = abn_df[abn_df["handle_status"] == "待处理"] if not abn_df.empty else abn_df
        high_cnt = len(pending_df[pending_df["risk_level"] == "高"]) if not pending_df.empty else 0
        mid_cnt = len(pending_df[pending_df["risk_level"] == "中"]) if not pending_df.empty else 0
        low_cnt = len(pending_df[pending_df["risk_level"] == "低"]) if not pending_df.empty else 0
        ac1, ac2, ac3, ac4 = st.columns(4)
        ac1.metric("待处理异常", len(pending_df))
        ac2.metric("高风险", high_cnt)
        ac3.metric("中风险", mid_cnt)
        ac4.metric("低风险", low_cnt)
        st.markdown("---")
        if abn_df.empty:
            st.info("暂无异常订单记录。")
        else:
            adisp = abn_df.rename(columns={
                "id": "记录ID", "order_id": "订单号", "customer_name": "客户",
                "platform": "平台", "intercept_reason": "拦截原因",
                "intercept_type": "拦截类型", "risk_level": "风险等级",
                "handle_status": "处理状态", "handle_result": "处理结果",
                "operator": "处理人", "created_at": "拦截时间", "handled_at": "处理时间",
            })
            ashow = ["记录ID", "订单号", "客户", "平台", "拦截原因", "拦截类型", "风险等级", "处理状态", "处理结果", "处理人", "拦截时间", "处理时间"]
            view = adisp[ashow].copy()

            def _hl_risk(row):
                if row["风险等级"] == "高":
                    return ["background-color: #fff1f0; color: #cf1322"] * len(row)
                elif row["风险等级"] == "中":
                    return ["background-color: #fffbe6; color: #d48806"] * len(row)
                return [""] * len(row)
            styler = view.style.apply(_hl_risk, axis=1)
            styler.hide(axis="index")
            st.dataframe(styler, use_container_width=True)

            st.markdown("---")
            st.markdown("##### 处理异常订单")
            pending_opts = pending_df.apply(
                lambda r: f'{r["id"]} | {r["order_id"]} | {r.get("customer_name", "")} | {r.get("risk_level", "")}',
                axis=1).tolist() if not pending_df.empty else []
            if not pending_opts:
                st.info("没有待处理的异常订单。")
            else:
                with st.form("oms_handle_form"):
                    h_sel = st.selectbox("选择待处理异常记录", [""] + pending_opts, key="oms_handle_sel")
                    h_result = st.text_input("处理结果", key="oms_handle_result")
                    h_op = st.text_input("处理人", value="客服", key="oms_handle_op")
                    hs = st.form_submit_button("✅ 标记已处理")
                    if hs:
                        if not h_sel or not h_result:
                            st.error("请选择记录并填写处理结果！")
                        else:
                            abn_id = int(h_sel.split(" | ")[0])
                            handle_abnormal(abn_id, h_result, h_op)
                            st.success(f"异常记录 {abn_id} 已标记为已处理。")
                            st.rerun()

            st.markdown("---")
            st.markdown("##### 查看订单详情")
            detail_opts = abn_df.apply(
                lambda r: f'{r["order_id"]} | {r.get("customer_name", "")}', axis=1).tolist()
            d_sel = st.selectbox("选择订单查看详情", [""] + detail_opts, key="oms_abn_detail_sel")
            d_oid = d_sel.split(" | ")[0] if d_sel and " | " in d_sel else ""
            if d_oid:
                od = get_order_by_id(d_oid)
                if od:
                    st.caption(f"状态：{od.get('order_status')}　实付：{fmt_money(float(od.get('actual_amount',0) or 0))}　异常原因：{od.get('abnormal_reason','')}")
                    di = get_order_items(d_oid)
                    if not di.empty:
                        didisp = di.rename(columns={
                            "spu_name": "商品", "sku_name": "规格", "quantity": "数量",
                            "unit_price": "单价", "total_price": "小计", "is_gift": "赠品"})
                        didisp["赠品"] = didisp["赠品"].apply(lambda x: "是" if x == 1 else "否")
                        didisp = fmt_money_df(didisp, ["单价", "小计"])
                        st.dataframe(didisp[["商品", "规格", "数量", "单价", "小计", "赠品"]],
                                     use_container_width=True, hide_index=True)

    # ---------- 7. 物流管理 ----------
    with oms_sub7:
        st.subheader("物流管理")

        st.markdown("##### 物流公司列表（可添加/删除）")
        logi_list_df = get_logistics_companies()
        if not logi_list_df.empty:
            ldisp = logi_list_df.rename(columns={
                "company_code": "公司编码", "company_name": "公司名称",
                "api_code": "API编码", "is_active": "启用", "sort_order": "排序"})
            ldisp["启用"] = ldisp["启用"].apply(lambda x: "是" if x == 1 else "否")
            st.dataframe(ldisp[["公司编码", "公司名称", "API编码", "启用", "排序"]],
                         use_container_width=True, hide_index=True)
        else:
            st.info("暂无物流公司。")

        with st.form("oms_logi_form"):
            lc1, lc2, lc3 = st.columns(3)
            with lc1:
                lc_code = st.text_input("公司编码 *", placeholder="如：SF", key="oms_logi_code")
            with lc2:
                lc_name = st.text_input("公司名称 *", placeholder="如：顺丰速运", key="oms_logi_name")
            with lc3:
                lc_api = st.text_input("API编码", placeholder="如：shunfeng", key="oms_logi_api")
            ls = st.form_submit_button("💾 添加/更新物流公司")
            if ls:
                if not lc_code or not lc_name:
                    st.error("公司编码和名称不能为空！")
                else:
                    add_logistics_company(lc_code, lc_name, lc_api)
                    st.success(f"✅ 物流公司「{lc_name}」已保存。")
                    st.rerun()

        with st.expander("🗑️ 删除物流公司"):
            if not logi_list_df.empty:
                del_opts = logi_list_df.apply(
                    lambda r: f'{r["company_code"]} | {r["company_name"]}', axis=1).tolist()
                del_sel = st.selectbox("选择要删除的物流公司", [""] + del_opts, key="oms_logi_del")
                if st.button("确认删除", key="oms_logi_del_btn"):
                    if del_sel and " | " in del_sel:
                        delete_logistics_company(del_sel.split(" | ")[0])
                        st.success("已删除。")
                        st.rerun()
            else:
                st.info("暂无物流公司可删除。")

        st.markdown("---")
        st.markdown("##### 物流同步状态查询")
        shipped_df = get_all_orders(status_filter="已发货")
        if shipped_df.empty:
            st.info("暂无已发货订单。")
        else:
            sh_opts = shipped_df.apply(
                lambda r: f'{r["order_id"]} | {r.get("customer_name", "")} | {r.get("logistics_company", "")} {r.get("logistics_number", "")}',
                axis=1).tolist()
            sh_sel = st.selectbox("选择已发货订单", [""] + sh_opts, key="oms_logi_query")
            sh_oid = sh_sel.split(" | ")[0] if sh_sel and " | " in sh_sel else ""
            if sh_oid:
                sod = get_order_by_id(sh_oid)
                if sod:
                    qc1, qc2, qc3, qc4 = st.columns(4)
                    qc1.metric("物流公司", sod.get("logistics_company", "") or "—")
                    qc2.metric("物流单号", sod.get("logistics_number", "") or "—")
                    qc3.metric("发货时间", sod.get("ship_time", "") or "—")
                    qc4.metric("物流状态", sod.get("logistics_status", "") or "运输中")
                    st.caption("提示：当前为离线模拟状态，正式对接各平台物流API后将自动同步实时轨迹。")

            st.markdown("---")
            st.markdown("##### 导出物流清单")
            from io import BytesIO
            logi_exp_buf = BytesIO()
            logi_exp_view = shipped_df.rename(columns=OMS_ORDER_DISPLAY)
            with pd.ExcelWriter(logi_exp_buf, engine="openpyxl") as writer:
                logi_exp_view.to_excel(writer, sheet_name="物流清单", index=False)
            logi_exp_buf.seek(0)
            st.download_button("📥 导出物流清单Excel", data=logi_exp_buf,
                               file_name="物流清单.xlsx", key="oms_logi_export",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ============================================================
# 【📈 数据分析】模块八：BI 数据报表（商业智能）
# ============================================================
with tab8:
    import plotly.express as px
    import plotly.graph_objects as go
    from io import BytesIO
    from datetime import date, timedelta

    st.header("📈 BI 智能数据报表")
    st.caption("经营总览 · 店铺分析 · 商品销售 · 库存周转 · 售后分析 · 综合报表导出 ｜ 完全离线")

    def _bi_date_filter(prefix):
        """渲染日期范围筛选器，返回 (date_from_str, date_to_str)"""
        today = date.today()
        key_from = f"{prefix}_from"
        key_to = f"{prefix}_to"
        if key_from not in st.session_state:
            st.session_state[key_from] = today - timedelta(days=30)
        if key_to not in st.session_state:
            st.session_state[key_to] = today

        fc, tc, qc = st.columns([1.2, 1.2, 2.6])
        with fc:
            df_val = st.date_input("开始日期", key=key_from)
        with tc:
            dt_val = st.date_input("结束日期", key=key_to)
        with qc:
            st.markdown("**快捷选择**")
            q1, q2, q3, q4 = st.columns(4)
            if q1.button("近7天", key=f"{prefix}_q7", use_container_width=True):
                st.session_state[key_from] = today - timedelta(days=7)
                st.session_state[key_to] = today
                st.rerun()
            if q2.button("近30天", key=f"{prefix}_q30", use_container_width=True):
                st.session_state[key_from] = today - timedelta(days=30)
                st.session_state[key_to] = today
                st.rerun()
            if q3.button("本月", key=f"{prefix}_qm", use_container_width=True):
                st.session_state[key_from] = today.replace(day=1)
                st.session_state[key_to] = today
                st.rerun()
            if q4.button("本季度", key=f"{prefix}_qq", use_container_width=True):
                q_month = (today.month - 1) // 3 * 3 + 1
                st.session_state[key_from] = date(today.year, q_month, 1)
                st.session_state[key_to] = today
                st.rerun()

        df_str = df_val.strftime("%Y-%m-%d") if df_val else None
        dt_str = dt_val.strftime("%Y-%m-%d") if dt_val else None
        return df_str, dt_str

    bi_sub1, bi_sub2, bi_sub3, bi_sub4, bi_sub5, bi_sub6 = st.tabs([
        "经营总览", "店铺经营分析", "商品销售分析",
        "库存周转分析", "售后分析", "综合报表导出",
    ])

    # ---------- 1. 经营总览 ----------
    with bi_sub1:
        st.subheader("经营总览")
        df_str, dt_str = _bi_date_filter("bi_ov")
        ov = bi_overview(df_str, dt_str)

        m1, m2, m3, m4, m5, m6 = st.columns(6)
        m1.metric("总订单数", f"{ov['total_orders']}")
        m2.metric("已完成订单", f"{ov['completed_orders']}")
        m3.metric("总营收", f"¥{fmt_money(ov['total_revenue'])}")
        m4.metric("平均客单价", f"¥{fmt_money(ov['avg_order_value'])}")
        m5.metric("退款总额", f"¥{fmt_money(ov['total_refund'])}")
        m6.metric("退款率", f"{ov['refund_rate']:.2%}")

        n1, n2, n3, n4, n5, n6 = st.columns(6)
        n1.metric("总客户数", f"{ov['total_customers']}")
        n2.metric("新客数", f"{ov['new_customers']}")
        n3.metric("商品销售件数", f"{int(ov['total_products_sold'])}")
        n4.metric("异常订单数", f"{ov['abnormal_count']}")
        n5.metric("售后单数", f"{ov['after_sales_count']}")
        n6.metric("售后率", f"{ov['after_sales_rate']:.2%}")

        st.markdown("---")

        col_l, col_r = st.columns(2)
        with col_l:
            st.markdown("#### 每日销售趋势")
            daily_df = bi_daily_trend(df_str, dt_str)
            if daily_df.empty:
                st.info("暂无销售趋势数据。")
            else:
                fig = go.Figure()
                fig.add_trace(go.Scatter(
                    x=daily_df["date"], y=daily_df["order_count"],
                    name="订单数", mode="lines+markers",
                    line=dict(color="#1890ff")))
                fig.add_trace(go.Scatter(
                    x=daily_df["date"], y=daily_df["revenue"],
                    name="营收", mode="lines+markers",
                    yaxis="y2", line=dict(color="#ff7f0e")))
                fig.update_layout(
                    yaxis=dict(title="订单数"),
                    yaxis2=dict(title="营收(¥)", overlaying="y", side="right"),
                    legend=dict(orientation="h", y=1.12),
                    height=400, margin=dict(l=20, r=20, t=20, b=20))
                st.plotly_chart(fig, use_container_width=True)
        with col_r:
            st.markdown("#### 平台销售分布")
            plat_df = bi_platform_distribution(df_str, dt_str)
            if plat_df.empty:
                st.info("暂无平台数据。")
            else:
                fig_p = px.pie(plat_df, values="revenue", names="platform",
                               color="platform",
                               color_discrete_sequence=px.colors.qualitative.Set2,
                               hole=0.4)
                fig_p.update_layout(height=400, margin=dict(l=20, r=20, t=20, b=20))
                st.plotly_chart(fig_p, use_container_width=True)

        col_l2, col_r2 = st.columns(2)
        with col_l2:
            st.markdown("#### 下单时段分布")
            hour_df = bi_hourly_distribution(df_str, dt_str)
            fig_h = px.bar(hour_df, x="hour", y="order_count",
                           color_discrete_sequence=["#1890ff"])
            fig_h.update_layout(xaxis=dict(title="小时", dtick=1),
                                height=380, margin=dict(l=20, r=20, t=20, b=20))
            st.plotly_chart(fig_h, use_container_width=True)
        with col_r2:
            st.markdown("#### 订单状态分布")
            conn = sqlite3.connect(DB_PATH)
            where_s, params_s = _bi_date_where("order_time", df_str, dt_str)
            cond_s = f" AND {where_s}" if where_s else ""
            status_df = pd.read_sql_query(
                f"SELECT order_status, COUNT(*) AS count FROM orders WHERE 1=1{cond_s} GROUP BY order_status ORDER BY count DESC",
                conn, params=params_s)
            conn.close()
            if status_df.empty:
                st.info("暂无订单状态数据。")
            else:
                fig_st = px.bar(status_df, x="order_status", y="count",
                                color="order_status",
                                color_discrete_sequence=px.colors.qualitative.Set3)
                fig_st.update_layout(height=380, margin=dict(l=20, r=20, t=20, b=20),
                                     showlegend=False)
                st.plotly_chart(fig_st, use_container_width=True)

    # ---------- 2. 店铺经营分析 ----------
    with bi_sub2:
        st.subheader("店铺经营分析")
        df_str, dt_str = _bi_date_filter("bi_shop")
        shop_df = bi_shop_analysis(df_str, dt_str)
        if shop_df.empty:
            st.info("暂无店铺经营数据。")
        else:
            display_shop = shop_df.copy()
            display_shop["revenue"] = display_shop["revenue"].round(2)
            display_shop["avg_order_value"] = display_shop["avg_order_value"].round(2)
            display_shop["refund_amount"] = display_shop["refund_amount"].round(2)
            display_shop["completion_rate"] = display_shop["completion_rate"].apply(lambda x: f"{x:.2%}")
            display_shop.columns = ["店铺名", "平台", "订单数", "营收", "客单价", "退款数", "退款金额", "完成率"]
            display_shop = fmt_money_df(display_shop, ["营收", "客单价", "退款金额"])
            st.dataframe(display_shop, use_container_width=True, hide_index=True)

            st.markdown("---")
            col_l, col_r = st.columns(2)
            with col_l:
                st.markdown("#### 各店铺营收对比")
                fig1 = px.bar(shop_df, x="shop_name", y="revenue", color="shop_name",
                              color_discrete_sequence=px.colors.qualitative.Set2)
                fig1.update_layout(height=380, margin=dict(l=20, r=20, t=20, b=20), showlegend=False)
                st.plotly_chart(fig1, use_container_width=True)
            with col_r:
                st.markdown("#### 各店铺完成率")
                fig2 = px.bar(shop_df, x="completion_rate", y="shop_name", orientation="h",
                              color="shop_name", color_discrete_sequence=px.colors.qualitative.Set3)
                fig2.update_layout(height=380, margin=dict(l=20, r=20, t=20, b=20), showlegend=False)
                st.plotly_chart(fig2, use_container_width=True)

            st.markdown("---")
            st.markdown("##### 导出Excel")
            buf = BytesIO()
            with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                display_shop.to_excel(writer, sheet_name="店铺经营分析", index=False)
            buf.seek(0)
            st.download_button("📥 导出店铺经营分析", data=buf,
                               file_name="店铺经营分析.xlsx", key="bi_shop_export",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    # ---------- 3. 商品销售分析 ----------
    with bi_sub3:
        st.subheader("商品销售分析")
        df_str, dt_str = _bi_date_filter("bi_prod")
        prod_df = bi_product_sales(df_str, dt_str)
        if prod_df.empty:
            st.info("暂无商品销售数据。")
        else:
            display_prod = prod_df.copy()
            display_prod["total_revenue"] = display_prod["total_revenue"].round(2)
            display_prod["avg_price"] = display_prod["avg_price"].round(2)
            display_prod.columns = ["SPU编码", "名称", "销量", "销售额", "平均售价", "订单数", "退款件数"]
            display_prod = fmt_money_df(display_prod, ["销售额", "平均售价"])
            st.dataframe(display_prod, use_container_width=True, hide_index=True)

            st.markdown("---")
            col_l, col_r = st.columns(2)
            with col_l:
                st.markdown("#### 商品销量 TOP10")
                top_qty = prod_df.nlargest(10, "total_qty")
                fig1 = px.bar(top_qty, x="total_qty", y="spu_name", orientation="h",
                              color="spu_name", color_discrete_sequence=px.colors.qualitative.Set2)
                fig1.update_layout(height=400, margin=dict(l=20, r=20, t=20, b=20), showlegend=False)
                st.plotly_chart(fig1, use_container_width=True)
            with col_r:
                st.markdown("#### 商品销售额 TOP10")
                top_rev = prod_df.nlargest(10, "total_revenue")
                fig2 = px.bar(top_rev, x="total_revenue", y="spu_name", orientation="h",
                              color="spu_name", color_discrete_sequence=px.colors.qualitative.Set3)
                fig2.update_layout(height=400, margin=dict(l=20, r=20, t=20, b=20), showlegend=False)
                st.plotly_chart(fig2, use_container_width=True)

            st.markdown("---")
            st.markdown("##### 导出Excel")
            buf = BytesIO()
            with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                display_prod.to_excel(writer, sheet_name="商品销售分析", index=False)
            buf.seek(0)
            st.download_button("📥 导出商品销售分析", data=buf,
                               file_name="商品销售分析.xlsx", key="bi_prod_export",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    # ---------- 4. 库存周转分析 ----------
    with bi_sub4:
        st.subheader("库存周转分析")
        df_str, dt_str = _bi_date_filter("bi_inv")
        inv_df = bi_inventory_turnover(df_str, dt_str)
        if inv_df.empty:
            st.info("暂无库存数据。")
        else:
            display_inv = inv_df.copy()
            display_inv["stock_value"] = display_inv["stock_value"].round(2)
            display_inv["turnover_rate"] = display_inv["turnover_rate"].apply(lambda x: f"{x:.2%}")
            display_inv.columns = ["产品编码", "名称", "入库量", "出库量", "当前库存", "库存金额", "周转率"]
            display_inv = fmt_money_df(display_inv, ["库存金额"])
            st.dataframe(display_inv, use_container_width=True, hide_index=True)

            low_stock = inv_df[inv_df["current_stock"] < 10]
            if not low_stock.empty:
                st.warning(f"⚠️ 库存预警：共 {len(low_stock)} 个产品库存不足10件，请及时补货。")
                warn_display = low_stock[["product_code", "product_name", "current_stock"]].copy()
                warn_display.columns = ["产品编码", "名称", "当前库存"]
                st.dataframe(warn_display, use_container_width=True, hide_index=True)

            st.markdown("---")
            col_l, col_r = st.columns(2)
            with col_l:
                st.markdown("#### 入库 vs 出库对比")
                chart_inv = inv_df.head(15)
                fig1 = px.bar(chart_inv, x="product_name", y=["inbound_qty", "outbound_qty"],
                              barmode="group",
                              color_discrete_sequence=["#52c41a", "#ff4d4f"])
                fig1.update_layout(height=400, margin=dict(l=20, r=20, t=20, b=20))
                st.plotly_chart(fig1, use_container_width=True)
            with col_r:
                st.markdown("#### 库存金额 TOP10")
                top_val = inv_df.nlargest(10, "stock_value")
                fig2 = px.bar(top_val, x="stock_value", y="product_name", orientation="h",
                              color="product_name", color_discrete_sequence=px.colors.qualitative.Set3)
                fig2.update_layout(height=400, margin=dict(l=20, r=20, t=20, b=20), showlegend=False)
                st.plotly_chart(fig2, use_container_width=True)

            st.markdown("---")
            st.markdown("##### 导出Excel")
            buf = BytesIO()
            with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                display_inv.to_excel(writer, sheet_name="库存周转分析", index=False)
            buf.seek(0)
            st.download_button("📥 导出库存周转分析", data=buf,
                               file_name="库存周转分析.xlsx", key="bi_inv_export",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    # ---------- 5. 售后分析 ----------
    with bi_sub5:
        st.subheader("售后分析")
        df_str, dt_str = _bi_date_filter("bi_as")
        as_data = bi_after_sales_analysis(df_str, dt_str)

        a1, a2, a3, a4, a5 = st.columns(5)
        a1.metric("售后总数", f"{as_data['total_count']}")
        a2.metric("退货数", f"{as_data['refund_count']}")
        a3.metric("换货数", f"{as_data['exchange_count']}")
        a4.metric("退款总额", f"¥{fmt_money(as_data['total_refund_amount'])}")
        a5.metric("退款率", f"{as_data['refund_rate']:.2%}")

        st.markdown("---")
        col_l, col_r = st.columns(2)
        with col_l:
            st.markdown("#### 售后类型分布")
            type_data = pd.DataFrame({
                "type": ["退货", "换货"],
                "count": [as_data["refund_count"], as_data["exchange_count"]],
            })
            if type_data["count"].sum() == 0:
                st.info("暂无售后类型数据。")
            else:
                fig1 = px.pie(type_data, values="count", names="type", hole=0.4,
                              color="type", color_discrete_sequence=px.colors.qualitative.Set2)
                fig1.update_layout(height=380, margin=dict(l=20, r=20, t=20, b=20))
                st.plotly_chart(fig1, use_container_width=True)
        with col_r:
            st.markdown("#### 售后原因 TOP10")
            top_r = as_data["top_reasons"]
            if top_r.empty:
                st.info("暂无售后原因数据。")
            else:
                fig2 = px.bar(top_r, x="reason", y="count", color="reason",
                              color_discrete_sequence=px.colors.qualitative.Set3)
                fig2.update_layout(height=380, margin=dict(l=20, r=20, t=20, b=20), showlegend=False)
                st.plotly_chart(fig2, use_container_width=True)

        col_l2, col_r2 = st.columns(2)
        with col_l2:
            st.markdown("#### 各平台售后分布")
            plat = as_data["platform_distribution"]
            if plat.empty:
                st.info("暂无平台售后数据。")
            else:
                fig3 = px.pie(plat, values="count", names="platform", hole=0.4,
                              color="platform", color_discrete_sequence=px.colors.qualitative.Set3)
                fig3.update_layout(height=380, margin=dict(l=20, r=20, t=20, b=20))
                st.plotly_chart(fig3, use_container_width=True)
        with col_r2:
            st.markdown("#### 每日售后趋势")
            daily_as = bi_daily_trend(df_str, dt_str)
            if daily_as.empty:
                st.info("暂无售后趋势数据。")
            else:
                fig4 = px.line(daily_as, x="date", y="refund_amount",
                               color_discrete_sequence=["#ff4d4f"])
                fig4.update_layout(height=380, margin=dict(l=20, r=20, t=20, b=20))
                st.plotly_chart(fig4, use_container_width=True)

        st.markdown("---")
        st.markdown("##### 售后单管理")

        as_list = get_all_after_sales()
        if not as_list.empty:
            as_display = as_list.copy()
            as_display["refund_amount"] = as_display["refund_amount"].round(2)
            as_display = fmt_money_df(as_display, ["refund_amount"])
            show_cols = ["after_sales_id", "order_id", "platform", "customer_name",
                         "type", "reason", "spu_name", "quantity", "refund_amount",
                         "status", "apply_date"]
            show_cols = [c for c in show_cols if c in as_display.columns]
            st.dataframe(as_display[show_cols], use_container_width=True, hide_index=True)
        else:
            st.info("暂无售后单。")

        with st.expander("➕ 录入售后单"):
            with st.form("bi_as_add_form"):
                fc1, fc2 = st.columns(2)
                with fc1:
                    as_id = st.text_input("售后单号 *", key="bi_as_id",
                                          placeholder="如：AS20260807001")
                    order_id = st.text_input("关联订单号 *", key="bi_as_order",
                                             placeholder="留空则手动填写下方信息")
                with fc2:
                    as_type = st.selectbox("售后类型", ["退货", "换货"], key="bi_as_type")
                    as_reason = st.text_input("售后原因", key="bi_as_reason",
                                              placeholder="如：质量问题")
                fc3, fc4, fc5 = st.columns(3)
                with fc3:
                    as_platform = st.text_input("平台", key="bi_as_plat",
                                                placeholder="留空自动获取")
                with fc4:
                    as_shop = st.text_input("店铺", key="bi_as_shop",
                                            placeholder="留空自动获取")
                with fc5:
                    as_customer = st.text_input("客户", key="bi_as_cust",
                                                placeholder="留空自动获取")
                fc6, fc7 = st.columns(2)
                with fc6:
                    as_spu_code = st.text_input("SPU编码", key="bi_as_spucode")
                    as_sku_code = st.text_input("SKU编码", key="bi_as_skucode")
                with fc7:
                    as_spu_name = st.text_input("商品名称", key="bi_as_spuname")
                    as_qty = st.number_input("数量", min_value=1, value=1, step=1, key="bi_as_qty")
                fc8, fc9 = st.columns(2)
                with fc8:
                    as_refund = money_input("退款金额", key="bi_as_refund", min_value=0.0)
                with fc9:
                    as_apply = st.date_input("申请日期", key="bi_as_apply")
                if st.form_submit_button("💾 提交售后单", use_container_width=True):
                    if not as_id or not order_id:
                        st.error("售后单号和关联订单号不能为空！")
                    else:
                        if not as_platform or not as_shop or not as_customer:
                            od = get_order_by_id(order_id)
                            if od:
                                as_platform = as_platform or od.get("platform", "") or ""
                                as_shop = as_shop or od.get("shop_name", "") or ""
                                as_customer = as_customer or od.get("customer_name", "") or ""
                        add_after_sales(as_id, order_id, as_platform, as_shop, "",
                                        as_customer, as_type, as_reason, as_spu_code,
                                        as_spu_name, as_sku_code, "", as_qty,
                                        as_refund, as_apply.strftime("%Y-%m-%d"))
                        st.success(f"✅ 售后单「{as_id}」已提交。")
                        st.rerun()

        with st.expander("🔄 更新售后状态"):
            if not as_list.empty:
                as_opts = as_list.apply(
                    lambda r: f'{r["after_sales_id"]} | {r.get("customer_name","")} | {r.get("status","")}',
                    axis=1).tolist()
                sel = st.selectbox("选择售后单", [""] + as_opts, key="bi_as_sel")
                if sel and " | " in sel:
                    sel_id = sel.split(" | ")[0]
                    uc1, uc2 = st.columns(2)
                    with uc1:
                        new_status = st.selectbox(
                            "新状态", ["待处理", "处理中", "已退款", "已完成", "已拒绝"],
                            key="bi_as_nstatus")
                    with uc2:
                        operator = st.text_input("处理人", key="bi_as_op", placeholder="如：张三")
                    handle_result = st.text_input("处理结果", key="bi_as_result")
                    if st.button("确认更新", key="bi_as_upd_btn", use_container_width=True):
                        update_after_sales_status(sel_id, new_status, handle_result, operator)
                        st.success(f"✅ 售后单「{sel_id}」状态已更新为「{new_status}」。")
                        st.rerun()
            else:
                st.info("暂无售后单可处理。")

        with st.expander("🗑️ 删除售后单"):
            if not as_list.empty:
                del_opts = as_list.apply(
                    lambda r: f'{r["after_sales_id"]} | {r.get("customer_name","")}',
                    axis=1).tolist()
                del_sel = st.selectbox("选择要删除的售后单", [""] + del_opts, key="bi_as_del")
                if st.button("确认删除", key="bi_as_del_btn"):
                    if del_sel and " | " in del_sel:
                        delete_after_sales(del_sel.split(" | ")[0])
                        st.success("已删除。")
                        st.rerun()
            else:
                st.info("暂无售后单可删除。")

        st.markdown("---")
        st.markdown("##### 导出Excel")
        if not as_list.empty:
            buf = BytesIO()
            with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                as_list.to_excel(writer, sheet_name="售后单列表", index=False)
            buf.seek(0)
            st.download_button("📥 导出售后单Excel", data=buf,
                               file_name="售后分析.xlsx", key="bi_as_export",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    # ---------- 6. 综合报表导出 ----------
    with bi_sub6:
        st.subheader("综合报表导出")
        df_str, dt_str = _bi_date_filter("bi_exp")
        st.caption("一键导出完整 BI 报告，包含经营总览、店铺分析、商品销售、库存周转、售后分析、平台分布、日趋势等多个 Sheet。")

        ov = bi_overview(df_str, dt_str)
        shop_df = bi_shop_analysis(df_str, dt_str)
        prod_df = bi_product_sales(df_str, dt_str)
        inv_df = bi_inventory_turnover(df_str, dt_str)
        as_data = bi_after_sales_analysis(df_str, dt_str)
        plat_df = bi_platform_distribution(df_str, dt_str)
        daily_df = bi_daily_trend(df_str, dt_str)

        ov_rows = [
            {"指标": "总订单数", "数值": ov["total_orders"]},
            {"指标": "已完成订单", "数值": ov["completed_orders"]},
            {"指标": "总营收", "数值": round(ov["total_revenue"], 2)},
            {"指标": "平均客单价", "数值": round(ov["avg_order_value"], 2)},
            {"指标": "退款总额", "数值": round(ov["total_refund"], 2)},
            {"指标": "退款率", "数值": f"{ov['refund_rate']:.2%}"},
            {"指标": "总客户数", "数值": ov["total_customers"]},
            {"指标": "新客数", "数值": ov["new_customers"]},
            {"指标": "商品销售件数", "数值": int(ov["total_products_sold"])},
            {"指标": "异常订单数", "数值": ov["abnormal_count"]},
            {"指标": "售后单数", "数值": ov["after_sales_count"]},
            {"指标": "售后率", "数值": f"{ov['after_sales_rate']:.2%}"},
        ]
        ov_df = pd.DataFrame(ov_rows)

        as_summary = pd.DataFrame([
            {"指标": "售后总数", "数值": as_data["total_count"]},
            {"指标": "退货数", "数值": as_data["refund_count"]},
            {"指标": "换货数", "数值": as_data["exchange_count"]},
            {"指标": "退款总额", "数值": round(as_data["total_refund_amount"], 2)},
            {"指标": "退款率", "数值": f"{as_data['refund_rate']:.2%}"},
        ])

        buf = BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            ov_df.to_excel(writer, sheet_name="经营总览", index=False)
            if not shop_df.empty:
                shop_df.to_excel(writer, sheet_name="店铺分析", index=False)
            if not prod_df.empty:
                prod_df.to_excel(writer, sheet_name="商品销售", index=False)
            if not inv_df.empty:
                inv_df.to_excel(writer, sheet_name="库存周转", index=False)
            as_summary.to_excel(writer, sheet_name="售后分析", index=False)
            if not as_data["top_reasons"].empty:
                as_data["top_reasons"].to_excel(writer, sheet_name="售后原因TOP", index=False)
            if not plat_df.empty:
                plat_df.to_excel(writer, sheet_name="平台分布", index=False)
            if not daily_df.empty:
                daily_df.to_excel(writer, sheet_name="日趋势", index=False)
        buf.seek(0)
        st.download_button(
            "📊 一键导出综合BI报告", data=buf,
            file_name=f"BI综合报告_{df_str}_{dt_str}.xlsx", key="bi_exp_all",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        st.markdown("---")
        st.markdown("##### 各模块独立下载")
        ec1, ec2, ec3 = st.columns(3)
        with ec1:
            if not shop_df.empty:
                buf1 = BytesIO()
                with pd.ExcelWriter(buf1, engine="openpyxl") as writer:
                    shop_df.to_excel(writer, sheet_name="店铺分析", index=False)
                buf1.seek(0)
                st.download_button("📋 店铺分析", data=buf1, file_name="店铺分析.xlsx",
                                   key="bi_exp_shop_dl", use_container_width=True,
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        with ec2:
            if not prod_df.empty:
                buf2 = BytesIO()
                with pd.ExcelWriter(buf2, engine="openpyxl") as writer:
                    prod_df.to_excel(writer, sheet_name="商品销售", index=False)
                buf2.seek(0)
                st.download_button("📋 商品销售", data=buf2, file_name="商品销售.xlsx",
                                   key="bi_exp_prod_dl", use_container_width=True,
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        with ec3:
            if not inv_df.empty:
                buf3 = BytesIO()
                with pd.ExcelWriter(buf3, engine="openpyxl") as writer:
                    inv_df.to_excel(writer, sheet_name="库存周转", index=False)
                buf3.seek(0)
                st.download_button("📋 库存周转", data=buf3, file_name="库存周转.xlsx",
                                   key="bi_exp_inv_dl", use_container_width=True,
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ============================================================
# 底部说明
# ============================================================
st.divider()
st.caption(
    "💡 **系统说明：**\n"
    "- 📝 模块一：离线记账，无需 API，免费使用\n"
    "- 📊 模块二：三大报表（正在重新设计中）\n"
    "- 🤖 模块三：AI 智能问答 + 凭证识别，需要 API Key\n"
    "- 📦 模块四：库存管理，产成品出入库自动生成凭证，免费使用\n"
    "\n"
    "**安装依赖：** `pip install streamlit openai pandas openpyxl`"
)
