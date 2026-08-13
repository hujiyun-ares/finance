# -*- coding: utf-8 -*-
"""
财务会计系统 - 独立财务模块
==========================

从 financial_agent.py 提取的财务部分，可独立运行的 Streamlit 应用。

模块一：离线记账（免费，不需要 API）
模块二：离线报表（免费，不需要 API）
模块三：AI 智能问答（需要 API Key）

运行方法：
  pip install streamlit openai pandas openpyxl plotly
  streamlit run financial_module.py
"""

import json
import os
import re
import hashlib
import sqlite3
from datetime import datetime
import streamlit as st
import pandas as pd
from openai import OpenAI

# set_page_config 必须是第一个 Streamlit 命令
st.set_page_config(page_title="财务会计系统", page_icon="💰", layout="wide")

# ============================================================
# 数据目录：固定路径，确保数据不会因工作目录变化而丢失
# ============================================================
import pathlib
_DATA_DIR = pathlib.Path.home() / ".finance_erp_data"
_DATA_DIR.mkdir(parents=True, exist_ok=True)

# ============================================================
# 数据库兼容层：SQLite / PostgreSQL 自动切换
# ----------------------------------------------------------
# - 本地沙盒（无 SUPABASE_DB_URL）→ 使用 SQLite（原有逻辑）
# - Streamlit Cloud（有 SUPABASE_DB_URL）→ 使用 Supabase PostgreSQL
#   数据永久保存，不受应用重启影响
# ============================================================

# 尝试获取 Supabase 连接字符串
def _get_supabase_url():
    """从环境变量或 Streamlit Secrets 获取 Supabase 连接字符串"""
    url = os.environ.get("SUPABASE_DB_URL", "")
    if not url:
        try:
            url = st.secrets.get("SUPABASE_DB_URL", "")
        except Exception:
            pass
    return url

_SUPABASE_URL = _get_supabase_url()
_USE_POSTGRES = bool(_SUPABASE_URL)

# 表主键映射（用于 INSERT OR REPLACE -> ON CONFLICT 转换）—— 仅保留财务相关表
_TABLE_PK = {
    'opening_balances': 'account_code',
    'custom_accounts': 'full_code',
}




def _db_path_to_schema(db_path):
    """将数据库文件路径转换为 PostgreSQL schema 名称"""
    name = db_path.replace(".db", "")
    name = re.sub(r'[^a-zA-Z0-9_]', '_', name)
    return name


def _adapt_sql(sql):
    """将 SQLite SQL 语法转换为 PostgreSQL 兼容语法"""
    # ? 占位符 → %s
    sql = sql.replace("?", "%s")
    # AUTOINCREMENT → SERIAL
    sql = sql.replace("INTEGER PRIMARY KEY AUTOINCREMENT", "SERIAL PRIMARY KEY")
    # datetime('now') → CURRENT_TIMESTAMP
    sql = sql.replace("datetime('now')", "CURRENT_TIMESTAMP")
    # INSERT OR IGNORE → INSERT ... ON CONFLICT DO NOTHING
    sql = _convert_insert_or_ignore(sql)
    # INSERT OR REPLACE → INSERT ... ON CONFLICT ... DO UPDATE
    sql = _convert_insert_or_replace(sql)
    return sql


def _convert_insert_or_ignore(sql):
    """将 SQLite 的 INSERT OR IGNORE 转换为 PostgreSQL 的 ON CONFLICT DO NOTHING"""
    pattern = r'INSERT\s+OR\s+IGNORE\s+INTO\s+(\w+)'
    match = re.search(pattern, sql, re.IGNORECASE)
    if not match:
        return sql
    table = match.group(1)
    pk = _TABLE_PK.get(table)
    # 先把 INSERT OR IGNORE 替换为普通 INSERT
    sql = re.sub(r'INSERT\s+OR\s+IGNORE\s+INTO', 'INSERT INTO', sql, flags=re.IGNORECASE)
    if pk:
        sql = sql.rstrip().rstrip(';')
        sql += f"\n        ON CONFLICT ({pk}) DO NOTHING"
    return sql


def _convert_insert_or_replace(sql):
    """将 SQLite 的 INSERT OR REPLACE 转换为 PostgreSQL 的 ON CONFLICT"""
    pattern = r'INSERT\s+OR\s+REPLACE\s+INTO\s+(\w+)\s*\(([^)]+)\)\s*VALUES'
    match = re.search(pattern, sql, re.IGNORECASE | re.DOTALL)

    if not match:
        return sql

    table = match.group(1)
    cols_str = match.group(2)
    cols = [c.strip() for c in cols_str.split(',')]

    pk = _TABLE_PK.get(table)
    if not pk:
        return sql.replace("INSERT OR REPLACE", "INSERT")

    non_pk_cols = [c for c in cols if c != pk]

    old_part = match.group(0)
    new_part = f"INSERT INTO {table} ({cols_str}) VALUES"
    sql = sql.replace(old_part, new_part, 1)

    if non_pk_cols:
        set_clause = ", ".join(f"{c} = EXCLUDED.{c}" for c in non_pk_cols)
        sql = sql.rstrip().rstrip(';')
        sql += f"\n        ON CONFLICT ({pk}) DO UPDATE SET {set_clause}"
    else:
        sql = sql.rstrip().rstrip(';')
        sql += f"\n        ON CONFLICT ({pk}) DO NOTHING"

    return sql


if _USE_POSTGRES:
    try:
        import psycopg2
        from psycopg2.extras import RealDictCursor
    except ImportError:
        # psycopg2 未安装时自动回退到 SQLite
        _USE_POSTGRES = False

if _USE_POSTGRES:
    class _PgCursor:
        """PostgreSQL 游标包装器，自动转换 SQLite SQL"""

        def __init__(self, cursor):
            self._cursor = cursor

        def execute(self, sql, params=None):
            pg_sql = _adapt_sql(sql)
            if params is not None:
                self._cursor.execute(pg_sql, params)
            else:
                self._cursor.execute(pg_sql)

        def executemany(self, sql, params_seq):
            pg_sql = _adapt_sql(sql)
            self._cursor.executemany(pg_sql, params_seq)

        def fetchall(self):
            return self._cursor.fetchall()

        def fetchone(self):
            return self._cursor.fetchone()

        def fetchmany(self, size=None):
            return self._cursor.fetchmany(size) if size else self._cursor.fetchmany()

        @property
        def lastrowid(self):
            return None

        @property
        def rowcount(self):
            return self._cursor.rowcount

        @property
        def description(self):
            return self._cursor.description

        def close(self):
            self._cursor.close()

    class _PgConnection:
        """PostgreSQL 连接包装器，模拟 sqlite3.Connection 接口"""

        def __init__(self, db_path):
            self._schema = _db_path_to_schema(db_path)
            self._conn = psycopg2.connect(_SUPABASE_URL)
            self._conn.autocommit = True
            self._row_factory = None

            cur = self._conn.cursor()
            # 防止并发 CREATE SCHEMA 导致 UniqueViolation
            try:
                cur.execute(f"CREATE SCHEMA IF NOT EXISTS {self._schema}")
            except Exception:
                pass  # schema 已存在或并发冲突，忽略
            cur.execute(f"SET search_path TO {self._schema}")
            cur.close()

        @property
        def row_factory(self):
            return self._row_factory

        @row_factory.setter
        def row_factory(self, value):
            self._row_factory = value

        def cursor(self):
            if self._row_factory is not None:
                return _PgCursor(self._conn.cursor(cursor_factory=RealDictCursor))
            return _PgCursor(self._conn.cursor())

        def commit(self):
            self._conn.commit()

        def rollback(self):
            self._conn.rollback()

        def close(self):
            self._conn.close()

        def execute(self, sql, params=None):
            cur = self.cursor()
            cur.execute(sql, params)
            return cur

    class _SQLiteCompat:
        """模拟 sqlite3 模块，底层使用 PostgreSQL"""
        Row = dict

        def connect(self, db_path, **kwargs):
            return _PgConnection(db_path)

    sqlite3 = _SQLiteCompat()
else:
    import sqlite3


# ============================================================
# 全局辅助：会计格式化函数
# ============================================================
def fmt_money(v):
    """将数值格式化为会计专用格式：千分位逗号 + 两位小数，如 300,000.00"""
    try:
        return f"{float(v):,.2f}"
    except (TypeError, ValueError):
        return "0.00"


def fmt_money_df(df, columns):
    """对 DataFrame 的指定列应用会计格式化（返回新 DataFrame）"""
    result = df.copy()
    for col in columns:
        if col in result.columns:
            result[col] = result[col].apply(lambda x: fmt_money(x) if pd.notna(x) else x)
    return result


def _format_money_cb(key):
    """on_change 回调：将金额输入框的值自动格式化为千分位逗号"""
    val = st.session_state.get(key, "")
    if val and str(val).strip():
        cleaned = str(val).replace(",", "").replace("，", "").replace(" ", "").strip()
        try:
            num = float(cleaned)
            st.session_state[key] = f"{num:,.2f}"
        except ValueError:
            pass  # 非数字，保持原样


def money_input(label, default_value=0.0, key=None, min_value=0.0,
                label_visibility="visible", placeholder="0.00",
                on_change_cb=None, on_change_args=None):
    """
    会计专用金额输入框（替代 st.number_input）。
    - 初始值显示千分位逗号，如 300,000.00
    - 用户可以输入带逗号或不带逗号的数字
    - 通过 on_change 回调自动格式化千分位逗号
    - 返回 float 类型的数值
    - on_change_cb: 自定义 on_change 回调（替代默认的 _format_money_cb）
    - on_change_args: 传给回调的参数元组
    """
    init_flag = f"_money_init_{key}"
    if init_flag not in st.session_state:
        if default_value and float(default_value) > 0:
            st.session_state[key] = f"{float(default_value):,.2f}"
        else:
            st.session_state[key] = ""
        st.session_state[init_flag] = True

    _cb = on_change_cb if on_change_cb else _format_money_cb
    _cb_args = tuple(on_change_args) if on_change_args else (key,)

    try:
        raw = st.text_input(
            label, key=key, label_visibility=label_visibility,
            placeholder=placeholder, on_change=_cb, args=_cb_args
        )
    except Exception as e:
        # 表单内不允许 on_change，退回普通 text_input
        st.session_state[f"_dbg_except_{key}"] = str(e)
        raw = st.text_input(
            label, key=key, label_visibility=label_visibility,
            placeholder=placeholder,
        )

    if raw is None or str(raw).strip() == "":
        return 0.0

    cleaned = str(raw).replace(",", "").replace("，", "").replace(" ", "").strip()
    try:
        num = float(cleaned)
        if min_value is not None and num < min_value:
            return float(min_value)
        return num
    except ValueError:
        return 0.0


# ============================================================
# 第 0.5 部分：用户认证系统
# ============================================================

AUTH_DB = str(_DATA_DIR / "erp_users.db")


def init_auth_db():
    """创建用户认证数据库"""
    conn = sqlite3.connect(AUTH_DB)
    c = conn.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS users (
            username TEXT PRIMARY KEY,
            password_hash TEXT NOT NULL,
            display_name TEXT,
            company TEXT,
            created_at TEXT
        )
    """)
    # 多公司管理表：一个用户下可以管理多家公司
    c.execute("""
        CREATE TABLE IF NOT EXISTS user_companies (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL,
            company_code TEXT NOT NULL,
            company_name TEXT NOT NULL,
            tax_id TEXT,
            address TEXT,
            phone TEXT,
            is_default INTEGER DEFAULT 0,
            created_at TEXT,
            UNIQUE(username, company_code)
        )
    """)
    conn.commit()
    conn.close()


def hash_password(password):
    """密码哈希"""
    return hashlib.sha256(password.encode()).hexdigest()


def register_user(username, password, display_name, company):
    """注册新用户"""
    conn = sqlite3.connect(AUTH_DB)
    c = conn.cursor()
    c.execute("SELECT username FROM users WHERE username = ?", (username,))
    if c.fetchone():
        conn.close()
        return False, "用户名已存在"
    from datetime import datetime
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    c.execute("""
        INSERT INTO users (username, password_hash, display_name, company, created_at)
        VALUES (?, ?, ?, ?, ?)
    """, (username, hash_password(password), display_name, company, now_str))
    # 注册时同时创建第一家公司
    if company:
        company_code = "CO001"
    else:
        company = "默认公司"
        company_code = "CO001"
    c.execute("""
        INSERT INTO user_companies (username, company_code, company_name, is_default, created_at)
        VALUES (?, ?, ?, 1, ?)
    """, (username, company_code, company, now_str))
    conn.commit()
    conn.close()
    return True, "注册成功"


def verify_user(username, password):
    """验证用户登录"""
    conn = sqlite3.connect(AUTH_DB)
    c = conn.cursor()
    c.execute("SELECT username, password_hash, display_name, company FROM users WHERE username = ?",
              (username,))
    row = c.fetchone()
    if not row:
        conn.close()
        return False, "用户名不存在"
    if row[1] != hash_password(password):
        conn.close()
        return False, "密码错误"
    # 获取该用户的所有公司
    c.execute("""
        SELECT company_code, company_name, is_default
        FROM user_companies WHERE username = ?
        ORDER BY is_default DESC, id ASC
    """, (username,))
    companies = [{"code": r[0], "name": r[1], "is_default": bool(r[2])} for r in c.fetchall()]
    conn.close()
    return True, {"username": row[0], "display_name": row[2], "company": row[3], "companies": companies}


def create_company(username, company_code, company_name, tax_id="", address="", phone=""):
    """为用户新建一家公司"""
    conn = sqlite3.connect(AUTH_DB)
    c = conn.cursor()
    from datetime import datetime
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    try:
        c.execute("""
            INSERT INTO user_companies (username, company_code, company_name, tax_id, address, phone, is_default, created_at)
            VALUES (?, ?, ?, ?, ?, ?, 0, ?)
        """, (username, company_code, company_name, tax_id, address, phone, now_str))
        conn.commit()
        conn.close()
        return True, "公司创建成功"
    except Exception as e:
        conn.close()
        if "UNIQUE" in str(e):
            return False, "公司编码已存在，请换一个"
        return False, str(e)


def get_user_companies(username):
    """获取用户的所有公司列表"""
    conn = sqlite3.connect(AUTH_DB)
    c = conn.cursor()
    c.execute("""
        SELECT company_code, company_name, tax_id, address, phone, is_default, created_at
        FROM user_companies WHERE username = ?
        ORDER BY is_default DESC, id ASC
    """, (username,))
    rows = c.fetchall()
    conn.close()
    return [
        {"code": r[0], "name": r[1], "tax_id": r[2] or "", "address": r[3] or "",
         "phone": r[4] or "", "is_default": bool(r[5]), "created_at": r[6] or ""}
        for r in rows
    ]


def delete_company(username, company_code):
    """删除一家公司（不能删除默认公司或最后一家公司）"""
    conn = sqlite3.connect(AUTH_DB)
    c = conn.cursor()
    # 检查是否是最后一家公司
    c.execute("SELECT COUNT(*) FROM user_companies WHERE username = ?", (username,))
    count = c.fetchone()[0]
    if count <= 1:
        conn.close()
        return False, "至少保留一家公司，无法删除"
    # 检查是否是默认公司
    c.execute("SELECT is_default FROM user_companies WHERE username = ? AND company_code = ?",
              (username, company_code))
    row = c.fetchone()
    if not row:
        conn.close()
        return False, "公司不存在"
    if row[0]:
        conn.close()
        return False, "不能删除默认公司，请先切换到其他公司"
    c.execute("DELETE FROM user_companies WHERE username = ? AND company_code = ?",
              (username, company_code))
    conn.commit()
    conn.close()
    return True, "公司已删除"


def update_company(username, company_code, company_name=None, tax_id=None, address=None, phone=None):
    """更新公司信息"""
    conn = sqlite3.connect(AUTH_DB)
    c = conn.cursor()
    updates = []
    params = []
    if company_name is not None:
        updates.append("company_name = ?")
        params.append(company_name)
    if tax_id is not None:
        updates.append("tax_id = ?")
        params.append(tax_id)
    if address is not None:
        updates.append("address = ?")
        params.append(address)
    if phone is not None:
        updates.append("phone = ?")
        params.append(phone)
    if not updates:
        conn.close()
        return False, "没有需要更新的字段"
    params.extend([username, company_code])
    c.execute(f"""
        UPDATE user_companies SET {', '.join(updates)}
        WHERE username = ? AND company_code = ?
    """, params)
    conn.commit()
    conn.close()
    return True, "公司信息已更新"


def get_user_db_path(username, company_code=None):
    """获取用户专属数据库路径
    如果提供了 company_code，则每家公司有独立数据库
    否则回退到旧版单数据库路径（兼容）
    """
    if company_code:
        return str(_DATA_DIR / f"erp_data_{username}_{company_code}.db")
    return str(_DATA_DIR / f"erp_data_{username}.db")


def login_page():
    """登录/注册页面"""
    st.markdown("## 💰 财务会计系统")
    st.caption("离线记账 · 自动报表 · AI 智能问答 ｜ 财务会计专用版")

    tab_login, tab_register = st.tabs(["🔐 登录", "📝 注册"])

    with tab_login:
        # 不使用 st.form，避免在 st.tabs 内密码框无法输入的问题
        username = st.text_input("用户名", key="login_user")
        password = st.text_input("密码", type="password", key="login_pwd")
        if st.button("登录", key="login_btn", use_container_width=True):
            if not username or not password:
                st.error("请输入用户名和密码")
            else:
                success, result = verify_user(username, password)
                if success:
                    st.session_state["logged_in"] = True
                    st.session_state["username"] = username
                    st.session_state["display_name"] = result["display_name"]
                    st.session_state["company"] = result["company"]
                    st.session_state["user_companies"] = result.get("companies", [])
                    # 自动选择默认公司
                    if result.get("companies"):
                        default_co = result["companies"][0]
                        st.session_state["current_company_code"] = default_co["code"]
                        st.session_state["current_company_name"] = default_co["name"]
                    st.rerun()
                else:
                    st.error(result)

    with tab_register:
        reg_user = st.text_input("设置用户名 *", key="reg_user", placeholder="字母或数字，如：admin")
        reg_pwd = st.text_input("设置密码 *", type="password", key="reg_pwd")
        reg_pwd2 = st.text_input("确认密码 *", type="password", key="reg_pwd2")
        reg_name = st.text_input("显示名称", key="reg_name", placeholder="如：张三")
        reg_company = st.text_input("公司名称", key="reg_company", placeholder="如：XX电子商务有限公司")
        if st.button("注册", key="reg_btn", use_container_width=True):
            if not reg_user or not reg_pwd:
                st.error("用户名和密码不能为空")
            elif reg_pwd != reg_pwd2:
                st.error("两次输入的密码不一致")
            elif len(reg_pwd) < 4:
                st.error("密码至少4位")
            else:
                success, msg = register_user(reg_user, reg_pwd, reg_name or reg_user, reg_company)
                if success:
                    st.success("注册成功！请切换到「登录」标签登录。")
                else:
                    st.error(msg)

    st.markdown("---")
    st.caption("一个账号可管理多家公司，每家公司拥有独立的财务数据库。")


# ============================================================
# 第 0.8 部分：登录状态检查
# ============================================================

# 自动迁移旧数据库文件到固定目录
def _migrate_old_dbs():
    """将工作目录下已有的 .db 文件迁移到固定数据目录"""
    import shutil
    cwd = pathlib.Path.cwd()
    for db_file in cwd.glob("erp_*.db"):
        target = _DATA_DIR / db_file.name
        if not target.exists():
            try:
                shutil.copy2(str(db_file), str(target))
            except Exception:
                pass

_migrate_old_dbs()

init_auth_db()

if "logged_in" not in st.session_state:
    st.session_state["logged_in"] = False

if not st.session_state["logged_in"]:
    login_page()
    st.stop()

# 已登录 → 设置用户专属数据库
_username = st.session_state.get("username", "default")

# === 多公司管理：旧用户迁移 ===
# 如果用户登录后没有公司信息（旧用户首次登录），自动创建默认公司
if not st.session_state.get("user_companies"):
    _existing_companies = get_user_companies(_username)
    if _existing_companies:
        # 数据库中已有公司记录但 session_state 中没有
        st.session_state["user_companies"] = _existing_companies
        _default_co = _existing_companies[0]
        st.session_state["current_company_code"] = _default_co["code"]
        st.session_state["current_company_name"] = _default_co["name"]
    else:
        # 旧用户没有公司记录，自动创建一家默认公司
        _old_company_name = st.session_state.get("company", "") or "默认公司"
        create_company(_username, "CO001", _old_company_name)
        _existing_companies = get_user_companies(_username)
        st.session_state["user_companies"] = _existing_companies
        if _existing_companies:
            _default_co = _existing_companies[0]
            st.session_state["current_company_code"] = _default_co["code"]
            st.session_state["current_company_name"] = _default_co["name"]

# 根据当前选中的公司设置 DB_PATH
_current_co_code = st.session_state.get("current_company_code")
_current_co_name = st.session_state.get("current_company_name", "")

# 如果没有当前公司但有旧数据库文件，使用旧路径（兼容）
if _current_co_code:
    DB_PATH = get_user_db_path(_username, _current_co_code)
    # 检查是否有旧数据库需要迁移
    _old_db = get_user_db_path(_username)
    import os as _os_mod
    if _os_mod.path.exists(_old_db) and not _os_mod.path.exists(DB_PATH):
        import shutil as _shutil_mod
        try:
            _shutil_mod.copy2(_old_db, DB_PATH)
        except Exception:
            pass
else:
    DB_PATH = get_user_db_path(_username)


# ============================================================
# 第 1 部分：会计科目表
# ----------------------------------------------------------
# 这是整个系统的基础。
# 每个科目都有：
#   - code: 科目编码（国家标准编码）
#   - name: 科目名称
#   - category: 大类（资产/负债/权益/成本/损益-收入/损益-费用）
#   - direction: 余额方向（借/贷）
#     · 资产、成本类科目正常余额在借方
#     · 负债、权益类科目正常余额在贷方
#     · 收入类科目余额方向是贷方
#     · 费用类科目余额方向是借方
# ============================================================

ACCOUNT_CHART = [
    # ================================================================
    # 一、资产类（余额在借方，抵减科目余额在贷方）
    # ================================================================
    # --- 流动资产 ---
    {"code": "1001", "name": "库存现金",               "category": "资产", "direction": "借"},
    {"code": "1002", "name": "银行存款",               "category": "资产", "direction": "借"},
    {"code": "1012", "name": "其他货币资金",           "category": "资产", "direction": "借"},
    {"code": "1101", "name": "交易性金融资产",         "category": "资产", "direction": "借"},
    {"code": "1102", "name": "交易性金融资产公允价值变动", "category": "资产", "direction": "借"},
    {"code": "1121", "name": "应收票据",               "category": "资产", "direction": "借"},
    {"code": "1122", "name": "应收账款",               "category": "资产", "direction": "借"},
    {"code": "1123", "name": "预付账款",               "category": "资产", "direction": "借"},
    {"code": "1131", "name": "应收股利",               "category": "资产", "direction": "借"},
    {"code": "1132", "name": "应收利息",               "category": "资产", "direction": "借"},
    {"code": "1221", "name": "其他应收款",             "category": "资产", "direction": "借"},
    {"code": "1231", "name": "坏账准备",               "category": "资产", "direction": "贷"},   # 抵减
    {"code": "1241", "name": "应收代位追偿款",         "category": "资产", "direction": "借"},
    {"code": "1251", "name": "应收分保账款",           "category": "资产", "direction": "借"},
    {"code": "1252", "name": "应收分保合同准备金",     "category": "资产", "direction": "借"},
    {"code": "1401", "name": "在途物资",               "category": "资产", "direction": "借"},
    {"code": "1403", "name": "原材料",                 "category": "资产", "direction": "借"},
    {"code": "1405", "name": "库存商品",               "category": "资产", "direction": "借"},
    {"code": "1406", "name": "发出商品",               "category": "资产", "direction": "借"},
    {"code": "1408", "name": "委托加工物资",           "category": "资产", "direction": "借"},
    {"code": "1411", "name": "周转材料",               "category": "资产", "direction": "借"},
    {"code": "1421", "name": "消耗性生物资产",         "category": "资产", "direction": "借"},
    {"code": "1471", "name": "存货跌价准备",           "category": "资产", "direction": "贷"},   # 抵减
    {"code": "1481", "name": "持有待售资产",           "category": "资产", "direction": "借"},
    {"code": "1482", "name": "持有待售资产减值准备",   "category": "资产", "direction": "贷"},   # 抵减
    {"code": "1491", "name": "合同资产",               "category": "资产", "direction": "借"},
    {"code": "1492", "name": "合同资产减值准备",       "category": "资产", "direction": "贷"},   # 抵减
    # --- 非流动资产 ---
    {"code": "1501", "name": "债权投资",               "category": "资产", "direction": "借"},
    {"code": "1502", "name": "债权投资减值准备",       "category": "资产", "direction": "贷"},   # 抵减
    {"code": "1503", "name": "其他债权投资",           "category": "资产", "direction": "借"},
    {"code": "1504", "name": "其他权益工具投资",       "category": "资产", "direction": "借"},
    {"code": "1511", "name": "长期股权投资",           "category": "资产", "direction": "借"},
    {"code": "1512", "name": "长期股权投资减值准备",   "category": "资产", "direction": "贷"},   # 抵减
    {"code": "1521", "name": "投资性房地产",           "category": "资产", "direction": "借"},
    {"code": "1531", "name": "长期应收款",             "category": "资产", "direction": "借"},
    {"code": "1541", "name": "未实现融资收益",         "category": "资产", "direction": "贷"},
    {"code": "1601", "name": "固定资产",               "category": "资产", "direction": "借"},
    {"code": "1602", "name": "累计折旧",               "category": "资产", "direction": "贷"},   # 抵减
    {"code": "1603", "name": "固定资产减值准备",       "category": "资产", "direction": "贷"},   # 抵减
    {"code": "1604", "name": "在建工程",               "category": "资产", "direction": "借"},
    {"code": "1605", "name": "工程物资",               "category": "资产", "direction": "借"},
    {"code": "1606", "name": "固定资产清理",           "category": "资产", "direction": "借"},
    {"code": "1621", "name": "生产性生物资产",         "category": "资产", "direction": "借"},
    {"code": "1622", "name": "生产性生物资产累计折旧", "category": "资产", "direction": "贷"},   # 抵减
    {"code": "1623", "name": "公益性生物资产",         "category": "资产", "direction": "借"},
    {"code": "1641", "name": "使用权资产",             "category": "资产", "direction": "借"},
    {"code": "1642", "name": "使用权资产累计折旧",     "category": "资产", "direction": "贷"},   # 抵减
    {"code": "1651", "name": "油气资产",               "category": "资产", "direction": "借"},
    {"code": "1652", "name": "油气资产折耗",           "category": "资产", "direction": "贷"},   # 抵减
    {"code": "1701", "name": "无形资产",               "category": "资产", "direction": "借"},
    {"code": "1702", "name": "累计摊销",               "category": "资产", "direction": "贷"},   # 抵减
    {"code": "1703", "name": "无形资产减值准备",       "category": "资产", "direction": "贷"},   # 抵减
    {"code": "1801", "name": "长期待摊费用",           "category": "资产", "direction": "借"},
    {"code": "1811", "name": "递延所得税资产",         "category": "资产", "direction": "借"},
    {"code": "1901", "name": "待处理财产损溢",         "category": "资产", "direction": "借"},
    # --- 金融行业专用 ---
    {"code": "1021", "name": "结算备付金",             "category": "资产", "direction": "借"},
    {"code": "1031", "name": "存出保证金",             "category": "资产", "direction": "借"},
    {"code": "1111", "name": "买入返售金融资产",       "category": "资产", "direction": "借"},
    {"code": "1302", "name": "贷款损失准备",           "category": "资产", "direction": "贷"},   # 抵减
    {"code": "1303", "name": "贴现资产",               "category": "资产", "direction": "借"},
    {"code": "1304", "name": "拆出资金",               "category": "资产", "direction": "借"},
    {"code": "1431", "name": "贵金属",                 "category": "资产", "direction": "借"},
    {"code": "1441", "name": "抵债资产",               "category": "资产", "direction": "借"},
    {"code": "1451", "name": "损余物资",               "category": "资产", "direction": "借"},
    {"code": "1461", "name": "融资租赁资产",           "category": "资产", "direction": "借"},
    {"code": "1661", "name": "勘探费用",               "category": "资产", "direction": "借"},
    {"code": "1671", "name": "独立账户资产",           "category": "资产", "direction": "借"},

    # ================================================================
    # 二、负债类（余额在贷方）
    # ================================================================
    {"code": "2001", "name": "短期借款",               "category": "负债", "direction": "贷"},
    {"code": "2201", "name": "应付票据",               "category": "负债", "direction": "贷"},
    {"code": "2202", "name": "应付账款",               "category": "负债", "direction": "贷"},
    {"code": "2203", "name": "合同负债",               "category": "负债", "direction": "贷"},
    {"code": "2211", "name": "应付职工薪酬",           "category": "负债", "direction": "贷"},
    {"code": "2221", "name": "应交税费",               "category": "负债", "direction": "贷"},
    {"code": "2231", "name": "应付利息",               "category": "负债", "direction": "贷"},
    {"code": "2232", "name": "应付股利",               "category": "负债", "direction": "贷"},
    {"code": "2241", "name": "其他应付款",             "category": "负债", "direction": "贷"},
    {"code": "2501", "name": "长期借款",               "category": "负债", "direction": "贷"},
    {"code": "2502", "name": "应付债券",               "category": "负债", "direction": "贷"},
    {"code": "2701", "name": "长期应付款",             "category": "负债", "direction": "贷"},
    {"code": "2711", "name": "递延收益",               "category": "负债", "direction": "贷"},
    {"code": "2801", "name": "预计负债",               "category": "负债", "direction": "贷"},

    # ================================================================
    # 三、所有者权益类（余额在贷方，库存股为借方抵减）
    # ================================================================
    {"code": "4001", "name": "实收资本",               "category": "权益", "direction": "贷"},
    {"code": "4002", "name": "资本公积",               "category": "权益", "direction": "贷"},
    {"code": "4101", "name": "盈余公积",               "category": "权益", "direction": "贷"},
    {"code": "4103", "name": "本年利润",               "category": "权益", "direction": "贷"},
    {"code": "4104", "name": "利润分配",               "category": "权益", "direction": "贷"},
    {"code": "4201", "name": "库存股",                 "category": "权益", "direction": "借"},   # 抵减
    {"code": "4301", "name": "其他综合收益",           "category": "权益", "direction": "贷"},

    # ================================================================
    # 四、成本类（余额在借方）
    # ================================================================
    {"code": "5001", "name": "生产成本",               "category": "成本", "direction": "借"},
    {"code": "5101", "name": "制造费用",               "category": "成本", "direction": "借"},
    {"code": "5201", "name": "劳务成本",               "category": "成本", "direction": "借"},
    {"code": "5301", "name": "研发支出",               "category": "成本", "direction": "借"},
    {"code": "5401", "name": "合同履约",               "category": "成本", "direction": "借"},
    {"code": "5501", "name": "生产成本辅助核算",       "category": "成本", "direction": "借"},

    # ================================================================
    # 五、损益类 - 收入（余额在贷方）
    # ================================================================
    {"code": "6001", "name": "主营业务收入",           "category": "损益-收入", "direction": "贷"},
    {"code": "6021", "name": "手续费及佣金收入",       "category": "损益-收入", "direction": "贷"},
    {"code": "6041", "name": "租赁收入",               "category": "损益-收入", "direction": "贷"},
    {"code": "6051", "name": "其他业务收入",           "category": "损益-收入", "direction": "贷"},
    {"code": "6101", "name": "公允价值变动损益",       "category": "损益-收入", "direction": "贷"},
    {"code": "6102", "name": "套期损益",               "category": "损益-收入", "direction": "贷"},
    {"code": "6103", "name": "净敞口套期损益",         "category": "损益-收入", "direction": "贷"},
    {"code": "6111", "name": "投资收益",               "category": "损益-收入", "direction": "贷"},
    {"code": "6115", "name": "资产处置损益",           "category": "损益-收入", "direction": "贷"},
    {"code": "6117", "name": "其他收益",               "category": "损益-收入", "direction": "贷"},
    {"code": "6301", "name": "营业外收入",             "category": "损益-收入", "direction": "贷"},
    {"code": "6391", "name": "报废收入（保险）",       "category": "损益-收入", "direction": "贷"},

    # ================================================================
    # 五、损益类 - 费用（余额在借方）
    # ================================================================
    {"code": "6401", "name": "主营业务成本",           "category": "损益-费用", "direction": "借"},
    {"code": "6402", "name": "其他业务成本",           "category": "损益-费用", "direction": "借"},
    {"code": "6403", "name": "税金及附加",             "category": "损益-费用", "direction": "借"},
    {"code": "6405", "name": "研发费用",               "category": "损益-费用", "direction": "借"},
    {"code": "6601", "name": "销售费用",               "category": "损益-费用", "direction": "借"},
    {"code": "6602", "name": "管理费用",               "category": "损益-费用", "direction": "借"},
    {"code": "6603", "name": "财务费用",               "category": "损益-费用", "direction": "借"},
    {"code": "6604", "name": "勘探费用",               "category": "损益-费用", "direction": "借"},
    {"code": "6605", "name": "租赁费用",               "category": "损益-费用", "direction": "借"},
    {"code": "6606", "name": "汇兑损益",               "category": "损益-费用", "direction": "借"},
    {"code": "6607", "name": "退保金（保险）",         "category": "损益-费用", "direction": "借"},
    {"code": "6608", "name": "赔付支出（保险）",       "category": "损益-费用", "direction": "借"},
    {"code": "6611", "name": "保单红利支出（保险）",   "category": "损益-费用", "direction": "借"},
    {"code": "6621", "name": "分出保费（保险）",       "category": "损益-费用", "direction": "借"},
    {"code": "6622", "name": "分保费用（保险）",       "category": "损益-费用", "direction": "借"},
    {"code": "6631", "name": "手续及佣金支出（金融）", "category": "损益-费用", "direction": "借"},
    {"code": "6641", "name": "信用减值损失",           "category": "损益-费用", "direction": "借"},
    {"code": "6642", "name": "资产减值损失",           "category": "损益-费用", "direction": "借"},
    {"code": "6701", "name": "营业外支出",             "category": "损益-费用", "direction": "借"},
    {"code": "6711", "name": "所得税费用",             "category": "损益-费用", "direction": "借"},
    {"code": "6801", "name": "以前年度损益调整",       "category": "损益-费用", "direction": "借"},
    {"code": "6901", "name": "其他资产损失",           "category": "损益-费用", "direction": "借"},
    {"code": "6902", "name": "财务担保合同",           "category": "损益-费用", "direction": "借"},
]

# 科目编码 → 科目信息 的映射表，方便快速查找
ACCOUNT_MAP = {a["code"]: a for a in ACCOUNT_CHART}

# 科目名称列表，用于下拉选择框
ACCOUNT_NAMES = [f"{a['code']} {a['name']}" for a in ACCOUNT_CHART]




# ============================================================
# 第 2 部分：数据库初始化
# ============================================================
# DB_PATH 已在第 0.8 部分根据登录用户动态设置


def init_database():
    """创建数据库表（如果不存在的话）"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()

    # 科目表：存科目编码、名称、分类、余额方向
    c.execute("""
        CREATE TABLE IF NOT EXISTS accounts (
            code TEXT PRIMARY KEY,
            name TEXT,
            category TEXT,
            direction TEXT
        )
    """)

    # 期初余额表：记录每个科目的初始余额
    c.execute("""
        CREATE TABLE IF NOT EXISTS opening_balances (
            account_code TEXT PRIMARY KEY,
            opening_debit REAL DEFAULT 0,
            opening_credit REAL DEFAULT 0
        )
    """)

    # 凭证表：每一行是一条分录（借方或贷方）
    c.execute("""
        CREATE TABLE IF NOT EXISTS vouchers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            voucher_number TEXT,
            voucher_date TEXT,
            summary TEXT,
            account_code TEXT,
            account_name TEXT,
            debit_amount REAL DEFAULT 0,
            credit_amount REAL DEFAULT 0
        )
    """)
    # 凭证审核相关字段（ALTER TABLE 兼容旧数据库）
    for _col, _type in [
        ("audit_status", "TEXT DEFAULT '未审核'"),
        ("auditor", "TEXT DEFAULT ''"),
        ("audit_date", "TEXT DEFAULT ''"),
    ]:
        try:
            c.execute(f"ALTER TABLE vouchers ADD COLUMN {_col} {_type}")
        except Exception:
            pass  # 列已存在

    c.execute("""
        CREATE TABLE IF NOT EXISTS custom_accounts (
            full_code TEXT PRIMARY KEY,
            parent_code TEXT NOT NULL,
            parent_name TEXT NOT NULL,
            sub_code TEXT NOT NULL,
            sub_name TEXT NOT NULL,
            full_name TEXT NOT NULL,
            category TEXT NOT NULL,
            direction TEXT NOT NULL,
            tax_flag TEXT DEFAULT '',
            created_at TEXT
        )
    """)

    # 如果科目表是空的，把预置科目写进去
    c.execute("SELECT COUNT(*) FROM accounts")
    if c.fetchone()[0] == 0:
        for a in ACCOUNT_CHART:
            c.execute(
                "INSERT INTO accounts VALUES (?, ?, ?, ?)",
                (a["code"], a["name"], a["category"], a["direction"]),
            )

    # ============================================================
    # 预置常用二级科目（参照企业会计准则应用指南 + 常见财务软件模板）
    # 如果 custom_accounts 表为空，自动创建以下二级科目：
    #   一、资产类    1001/1012/1221/1601/1602/1701/1702/1801
    #   二、负债类    2211/2221/2241
    #   三、所有者权益  4001/4002/4101/4104
    #   四、成本类    5001/5101
    #   五、损益类    6601/6602/6603/6301/6701
    # 注：1002银行存款、1122应收账款、1123预付账款、2202应付账款等
    #     通常启用辅助核算而非预置二级科目，此处不预置。
    # ============================================================
    c.execute("SELECT COUNT(*) FROM custom_accounts")
    if c.fetchone()[0] == 0:
        _preset_subs = [
            # (parent_code, parent_name, sub_name, category, direction, tax_flag)

            # ===== 一、资产类 =====
            # 1001 库存现金
            ("1001", "库存现金", "人民币",           "资产", "借", ""),
            ("1001", "库存现金", "外币",             "资产", "借", ""),
            # 1012 其他货币资金
            ("1012", "其他货币资金", "外埠存款",       "资产", "借", ""),
            ("1012", "其他货币资金", "银行汇票存款",   "资产", "借", ""),
            ("1012", "其他货币资金", "银行本票存款",   "资产", "借", ""),
            ("1012", "其他货币资金", "信用证保证金",   "资产", "借", ""),
            ("1012", "其他货币资金", "存出投资款",     "资产", "借", ""),
            # 1221 其他应收款
            ("1221", "其他应收款", "员工借款",         "资产", "借", ""),
            ("1221", "其他应收款", "备用金",           "资产", "借", ""),
            ("1221", "其他应收款", "押金保证金",       "资产", "借", ""),
            ("1221", "其他应收款", "其他应收",         "资产", "借", ""),
            # 1601 固定资产
            ("1601", "固定资产", "房屋建筑物",         "资产", "借", ""),
            ("1601", "固定资产", "机器设备",           "资产", "借", ""),
            ("1601", "固定资产", "运输设备",           "资产", "借", ""),
            ("1601", "固定资产", "电子设备",           "资产", "借", ""),
            ("1601", "固定资产", "办公设备",           "资产", "借", ""),
            # 1602 累计折旧（与固定资产一一对应，贷方抵减）
            ("1602", "累计折旧", "房屋建筑物折旧",     "资产", "贷", ""),
            ("1602", "累计折旧", "机器设备折旧",       "资产", "贷", ""),
            ("1602", "累计折旧", "运输设备折旧",       "资产", "贷", ""),
            ("1602", "累计折旧", "电子设备折旧",       "资产", "贷", ""),
            ("1602", "累计折旧", "办公设备折旧",       "资产", "贷", ""),
            # 1701 无形资产
            ("1701", "无形资产", "土地使用权",         "资产", "借", ""),
            ("1701", "无形资产", "软件及专利权",       "资产", "借", ""),
            # 1702 累计摊销（与无形资产对应，贷方抵减）
            ("1702", "累计摊销", "土地使用权摊销",     "资产", "贷", ""),
            ("1702", "累计摊销", "软件及专利权摊销",   "资产", "贷", ""),
            # 1801 长期待摊费用
            ("1801", "长期待摊费用", "开办费",                   "资产", "借", ""),
            ("1801", "长期待摊费用", "租入固定资产改良支出",     "资产", "借", ""),

            # ===== 二、负债类 =====
            # 2211 应付职工薪酬（几乎所有软件标准预置）
            ("2211", "应付职工薪酬", "工资薪金",         "负债", "贷", "工资薪金"),
            ("2211", "应付职工薪酬", "职工福利费",       "负债", "贷", "职工福利费"),
            ("2211", "应付职工薪酬", "社会保险费",       "负债", "贷", "社保公积金"),
            ("2211", "应付职工薪酬", "住房公积金",       "负债", "贷", "社保公积金"),
            ("2211", "应付职工薪酬", "工会经费",         "负债", "贷", "工会经费"),
            ("2211", "应付职工薪酬", "职工教育经费",     "负债", "贷", "职工教育经费"),
            # 2221 应交税费（强制标准二级，全部软件预置）
            #   应交增值税的三级明细（进项/销项等）展开为二级科目
            ("2221", "应交税费", "应交增值税-进项税额",         "负债", "贷", ""),
            ("2221", "应交税费", "应交增值税-销项税额",         "负债", "贷", ""),
            ("2221", "应交税费", "应交增值税-进项税额转出",     "负债", "贷", ""),
            ("2221", "应交税费", "应交增值税-已交税金",         "负债", "贷", ""),
            ("2221", "应交税费", "未交增值税",                   "负债", "贷", ""),
            ("2221", "应交税费", "城建税",                       "负债", "贷", ""),
            ("2221", "应交税费", "教育费附加",                   "负债", "贷", ""),
            ("2221", "应交税费", "地方教育附加",                 "负债", "贷", ""),
            ("2221", "应交税费", "企业所得税",                   "负债", "贷", ""),
            ("2221", "应交税费", "个人所得税",                   "负债", "贷", ""),
            ("2221", "应交税费", "印花税",                       "负债", "贷", ""),
            ("2221", "应交税费", "房产税",                       "负债", "贷", ""),
            ("2221", "应交税费", "土地使用税",                   "负债", "贷", ""),
            # 2241 其他应付款
            ("2241", "其他应付款", "押金保证金",       "负债", "贷", ""),
            ("2241", "其他应付款", "代扣款项",         "负债", "贷", ""),
            ("2241", "其他应付款", "其他",             "负债", "贷", ""),

            # ===== 三、所有者权益类 =====
            # 4001 实收资本
            ("4001", "实收资本", "法人资本",           "权益", "贷", ""),
            ("4001", "实收资本", "个人资本",           "权益", "贷", ""),
            # 4002 资本公积
            ("4002", "资本公积", "资本溢价",           "权益", "贷", ""),
            # 4101 盈余公积
            ("4101", "盈余公积", "法定盈余公积",       "权益", "贷", ""),
            ("4101", "盈余公积", "任意盈余公积",       "权益", "贷", ""),
            # 4104 利润分配
            ("4104", "利润分配", "提取法定盈余公积",   "权益", "贷", ""),
            ("4104", "利润分配", "应付利润",           "权益", "贷", ""),
            ("4104", "利润分配", "未分配利润",         "权益", "贷", ""),

            # ===== 四、成本类（生产型企业模板预置）=====
            # 5001 生产成本
            ("5001", "生产成本", "直接材料",           "成本", "借", ""),
            ("5001", "生产成本", "直接人工",           "成本", "借", ""),
            ("5001", "生产成本", "制造费用转入",       "成本", "借", ""),
            # 5101 制造费用
            ("5101", "制造费用", "车间人工",           "成本", "借", ""),
            ("5101", "制造费用", "车间折旧",           "成本", "借", ""),
            ("5101", "制造费用", "车间水电费",         "成本", "借", ""),
            ("5101", "制造费用", "机物料消耗",         "成本", "借", ""),

            # ===== 五、损益类 =====
            # 6601 销售费用
            ("6601", "销售费用", "工资",               "损益-费用", "借", "工资薪金"),
            ("6601", "销售费用", "广告费及业务宣传费", "损益-费用", "借", "广告费"),
            ("6601", "销售费用", "运输费",             "损益-费用", "借", ""),
            ("6601", "销售费用", "差旅费",             "损益-费用", "借", ""),
            ("6601", "销售费用", "业务招待费",         "损益-费用", "借", "业务招待费"),
            ("6601", "销售费用", "折旧费",             "损益-费用", "借", ""),
            ("6601", "销售费用", "职工福利费",         "损益-费用", "借", "职工福利费"),
            ("6601", "销售费用", "销售佣金",           "损益-费用", "借", ""),
            # 6602 管理费用（通用模板标配，预置最完整）
            ("6602", "管理费用", "工资薪金",           "损益-费用", "借", "工资薪金"),
            ("6602", "管理费用", "办公费",             "损益-费用", "借", ""),
            ("6602", "管理费用", "差旅费",             "损益-费用", "借", ""),
            ("6602", "管理费用", "业务招待费",         "损益-费用", "借", "业务招待费"),
            ("6602", "管理费用", "折旧费",             "损益-费用", "借", ""),
            ("6602", "管理费用", "水电费",             "损益-费用", "借", ""),
            ("6602", "管理费用", "租赁费",             "损益-费用", "借", ""),
            ("6602", "管理费用", "社保公积金",         "损益-费用", "借", "社保公积金"),
            ("6602", "管理费用", "咨询审计费",         "损益-费用", "借", ""),
            ("6602", "管理费用", "印花税及税费",       "损益-费用", "借", ""),
            ("6602", "管理费用", "职工福利费",         "损益-费用", "借", "职工福利费"),
            ("6602", "管理费用", "研发费用",           "损益-费用", "借", "研发费用"),
            ("6602", "管理费用", "工会经费",           "损益-费用", "借", "工会经费"),
            ("6602", "管理费用", "职工教育经费",       "损益-费用", "借", "职工教育经费"),
            ("6602", "管理费用", "邮电通讯费",         "损益-费用", "借", ""),
            ("6602", "管理费用", "车辆使用费",         "损益-费用", "借", ""),
            ("6602", "管理费用", "会议费",             "损益-费用", "借", ""),
            # 6603 财务费用
            ("6603", "财务费用", "利息支出",           "损益-费用", "借", ""),
            ("6603", "财务费用", "利息收入",           "损益-费用", "贷", ""),
            ("6603", "财务费用", "银行手续费",         "损益-费用", "借", ""),
            ("6603", "财务费用", "汇兑损益",           "损益-费用", "借", ""),
            # 6301 营业外收入
            ("6301", "营业外收入", "政府补助",         "损益-收入", "贷", ""),
            ("6301", "营业外收入", "罚款收入",         "损益-收入", "贷", ""),
            ("6301", "营业外收入", "其他",             "损益-收入", "贷", ""),
            # 6701 营业外支出
            ("6701", "营业外支出", "罚款支出",         "损益-费用", "借", ""),
            ("6701", "营业外支出", "公益性捐赠",       "损益-费用", "借", "公益性捐赠"),
            ("6701", "营业外支出", "资产盘亏损失",     "损益-费用", "借", ""),
        ]
        # 按父科目分别编号（01、02、03...），避免跨科目编号混乱
        _sub_counters = {}
        for p_code, p_name, s_name, cat, direction, tax_flag in _preset_subs:
            _sub_counters[p_code] = _sub_counters.get(p_code, 0) + 1
            sub_code = f"{_sub_counters[p_code]:02d}"
            full_code = f"{p_code}{sub_code}"
            full_name = f"{p_name}-{s_name}"
            c.execute("""
                INSERT OR IGNORE INTO custom_accounts
                (full_code, parent_code, parent_name, sub_code, sub_name, full_name,
                 category, direction, tax_flag, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (full_code, p_code, p_name, sub_code, s_name, full_name,
                  cat, direction, tax_flag,
                  datetime.now().strftime("%Y-%m-%d %H:%M:%S")))

    conn.commit()
    conn.close()


def get_next_voucher_number():
    """获取下一个凭证编号，如：记字第001号"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT COUNT(DISTINCT voucher_number) FROM vouchers")
    count = c.fetchone()[0]
    conn.close()
    return f"记字第{count + 1:03d}号"


def save_voucher(voucher_number, voucher_date, summary, lines):
    """
    保存一张凭证到数据库
    lines 是一个列表，每个元素：{account_code, account_name, debit, credit}
    """
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    for line in lines:
        c.execute("""
            INSERT INTO vouchers
            (voucher_number, voucher_date, summary, account_code, account_name, debit_amount, credit_amount)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (
            voucher_number, voucher_date, summary,
            line["account_code"], line["account_name"],
            float(line["debit"]), float(line["credit"]),
        ))
    conn.commit()
    conn.close()
    get_all_vouchers.clear()


def delete_voucher(voucher_number):
    """删除一张凭证（所有分录行）"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("DELETE FROM vouchers WHERE voucher_number = ?", (voucher_number,))
    conn.commit()
    conn.close()
    get_all_vouchers.clear()
    get_all_opening_balances.clear()
    calc_account_balance.clear()


def update_voucher(voucher_number, new_date, new_summary, new_lines):
    """
    修改一张凭证：先删除旧分录，再写入新分录。
    new_lines: [{account_code, account_name, debit, credit}, ...]
    """
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    # 删除旧分录
    c.execute("DELETE FROM vouchers WHERE voucher_number = ?", (voucher_number,))
    # 写入新分录
    for line in new_lines:
        c.execute("""
            INSERT INTO vouchers
            (voucher_number, voucher_date, summary, account_code, account_name, debit_amount, credit_amount)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (
            voucher_number,
            new_date,
            new_summary,
            line["account_code"],
            line["account_name"],
            float(line["debit"]),
            float(line["credit"]),
        ))
    conn.commit()
    conn.close()
    get_all_vouchers.clear()
    get_all_opening_balances.clear()
    calc_account_balance.clear()


def get_voucher_by_number(voucher_number):
    """获取一张凭证的所有分录，返回列表 [{account_code, account_name, debit, credit}, ...]"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("""
        SELECT account_code, account_name, debit_amount, credit_amount
        FROM vouchers WHERE voucher_number = ?
        ORDER BY id
    """, (voucher_number,))
    rows = c.fetchall()
    conn.close()
    return [
        {"account_code": r[0], "account_name": r[1], "debit": r[2] or 0, "credit": r[3] or 0}
        for r in rows
    ]


def get_voucher_info(voucher_number):
    """获取凭证的基本信息（日期、摘要）"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("""
        SELECT voucher_date, summary FROM vouchers
        WHERE voucher_number = ? LIMIT 1
    """, (voucher_number,))
    row = c.fetchone()
    conn.close()
    if row:
        return {"date": row[0], "summary": row[1]}
    return None


@st.cache_data(ttl=60, show_spinner=False)
def get_all_vouchers():
    """获取所有凭证，返回 DataFrame，缓存 1 分钟"""
    conn = sqlite3.connect(DB_PATH)
    df = pd.read_sql_query("SELECT * FROM vouchers ORDER BY id", conn)
    conn.close()
    return df


# ============================================================
# 凭证审核功能（对标用友U8/金蝶K3 制单→审核→记账 三段式内控）
# ============================================================

def get_voucher_audit_status(voucher_number):
    """获取单张凭证的审核状态"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("""
        SELECT audit_status, auditor, audit_date
        FROM vouchers WHERE voucher_number = ? LIMIT 1
    """, (voucher_number,))
    row = c.fetchone()
    conn.close()
    if row:
        return {"status": row[0] or "未审核", "auditor": row[1] or "", "audit_date": row[2] or ""}
    return {"status": "未审核", "auditor": "", "audit_date": ""}


def audit_voucher(voucher_number, auditor_name):
    """审核一张凭证（未审核 → 已审核）
    审核后凭证锁定，不能修改/删除
    """
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    # 检查凭证是否存在且未审核
    c.execute("SELECT audit_status FROM vouchers WHERE voucher_number = ? LIMIT 1", (voucher_number,))
    row = c.fetchone()
    if not row:
        conn.close()
        return False, "凭证不存在"
    current_status = row[0] or "未审核"
    if current_status == "已审核":
        conn.close()
        return False, "该凭证已审核，无需重复审核"
    from datetime import datetime
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    c.execute("""
        UPDATE vouchers SET audit_status = '已审核', auditor = ?, audit_date = ?
        WHERE voucher_number = ?
    """, (auditor_name, now_str, voucher_number))
    conn.commit()
    conn.close()
    get_all_vouchers.clear()
    return True, f"凭证 {voucher_number} 审核通过"


def unaudit_voucher(voucher_number):
    """反审核一张凭证（已审核 → 未审核）
    取消审核后凭证可修改/删除
    """
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT audit_status FROM vouchers WHERE voucher_number = ? LIMIT 1", (voucher_number,))
    row = c.fetchone()
    if not row:
        conn.close()
        return False, "凭证不存在"
    current_status = row[0] or "未审核"
    if current_status != "已审核":
        conn.close()
        return False, "该凭证未审核，无需反审核"
    c.execute("""
        UPDATE vouchers SET audit_status = '未审核', auditor = '', audit_date = ''
        WHERE voucher_number = ?
    """, (voucher_number,))
    conn.commit()
    conn.close()
    get_all_vouchers.clear()
    return True, f"凭证 {voucher_number} 已取消审核"


def batch_audit_vouchers(voucher_numbers, auditor_name):
    """批量审核多张凭证"""
    success_count = 0
    fail_count = 0
    messages = []
    for vnum in voucher_numbers:
        ok, msg = audit_voucher(vnum, auditor_name)
        if ok:
            success_count += 1
        else:
            fail_count += 1
            messages.append(f"{vnum}: {msg}")
    summary = f"批量审核完成：成功 {success_count} 张"
    if fail_count:
        summary += f"，失败 {fail_count} 张"
    return True, summary


def get_pending_audit_vouchers():
    """获取所有待审核凭证列表"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("""
        SELECT DISTINCT voucher_number, voucher_date, summary
        FROM vouchers
        WHERE audit_status IS NULL OR audit_status = '' OR audit_status = '未审核'
        ORDER BY voucher_date, voucher_number
    """)
    rows = c.fetchall()
    conn.close()
    return [{"number": r[0], "date": r[1], "summary": r[2]} for r in rows]


def get_audited_vouchers():
    """获取所有已审核凭证列表"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("""
        SELECT DISTINCT voucher_number, voucher_date, summary, auditor, audit_date
        FROM vouchers
        WHERE audit_status = '已审核'
        ORDER BY audit_date DESC, voucher_number
    """)
    rows = c.fetchall()
    conn.close()
    return [{"number": r[0], "date": r[1], "summary": r[2],
             "auditor": r[3] or "", "audit_date": r[4] or ""} for r in rows]


@st.cache_data(ttl=30, show_spinner=False)
def get_all_opening_balances():
    """一次查询获取所有期初余额，返回 {account_code: (debit, credit)} 字典"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT account_code, opening_debit, opening_credit FROM opening_balances")
    rows = c.fetchall()
    conn.close()
    return {row[0]: (row[1] or 0, row[2] or 0) for row in rows}


def get_opening_balance(account_code):
    """获取某个科目的期初余额（借方和贷方）"""
    _all = get_all_opening_balances()
    if account_code in _all:
        return _all[account_code]
    return 0, 0


def save_opening_balance(account_code, debit, credit):
    """保存某个科目的期初余额"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("""
        INSERT OR REPLACE INTO opening_balances
        (account_code, opening_debit, opening_credit)
        VALUES (?, ?, ?)
    """, (account_code, float(debit), float(credit)))
    conn.commit()
    conn.close()
    get_all_opening_balances.clear()


# ============================================================
# 第 2.2 部分：自定义二级科目管理
# ============================================================

# 税务特殊科目标记
TAX_SPECIAL_FLAGS = {
    "": "无特殊税务标记",
    "业务招待费": "业务招待费：按实际发生额60%扣除，且不超过当年销售收入的5‰",
    "广告费": "广告费：不超过当年销售（营业）收入15%的部分准予扣除",
    "公益性捐赠": "公益性捐赠：不超过年度利润总额12%的部分准予扣除",
    "职工福利费": "职工福利费：不超过工资薪金总额14%的部分准予扣除",
    "工会经费": "工会经费：不超过工资薪金总额2%的部分准予扣除",
    "职工教育经费": "职工教育经费：不超过工资薪金总额8%的部分准予扣除",
    "研发费用": "研发费用：按规定加计扣除（一般企业75%，特定领域100%）",
}


def get_custom_accounts(parent_code=None):
    """获取自定义二级科目列表，可按一级科目筛选"""
    conn = sqlite3.connect(DB_PATH)
    if parent_code:
        df = pd.read_sql_query(
            "SELECT * FROM custom_accounts WHERE parent_code = ? ORDER BY full_code",
            conn, params=(parent_code,)
        )
    else:
        df = pd.read_sql_query(
            "SELECT * FROM custom_accounts ORDER BY parent_code, sub_code", conn
        )
    conn.close()
    return df


def get_next_sub_code(parent_code):
    """获取下一个二级科目序号，如 '01', '02'..."""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute(
        "SELECT sub_code FROM custom_accounts WHERE parent_code = ? ORDER BY sub_code DESC LIMIT 1",
        (parent_code,)
    )
    row = c.fetchone()
    conn.close()
    if row:
        next_num = int(row[0]) + 1
    else:
        next_num = 1
    return f"{next_num:02d}"


def add_custom_account(parent_code, parent_name, sub_name, category, direction, tax_flag=""):
    """添加一个自定义二级科目"""
    sub_code = get_next_sub_code(parent_code)
    full_code = f"{parent_code}{sub_code}"
    full_name = f"{parent_name}-{sub_name}"
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("""
        INSERT INTO custom_accounts
        (full_code, parent_code, parent_name, sub_code, sub_name, full_name, category, direction, tax_flag, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (full_code, parent_code, parent_name, sub_code, sub_name,
          full_name, category, direction, tax_flag,
          datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
    conn.commit()
    conn.close()
    return full_code, full_name


def delete_custom_account(full_code):
    """删除一个自定义二级科目（同时清理期初余额）"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("DELETE FROM custom_accounts WHERE full_code = ?", (full_code,))
    c.execute("DELETE FROM opening_balances WHERE account_code = ?", (full_code,))
    conn.commit()
    conn.close()


@st.cache_data(ttl=300, show_spinner=False)
def get_all_accounts_for_display():
    """
    合并系统科目 + 自定义二级科目，返回统一格式的列表。
    缓存 5 分钟，减少数据库查询次数。
    每个元素: {"code", "name", "display_name", "category", "direction",
               "is_custom", "parent_code", "tax_flag", "level"}
    level: 1=一级科目, 2=二级科目
    """
    result = []
    # 系统一级科目
    for a in ACCOUNT_CHART:
        result.append({
            "code": a["code"],
            "name": a["name"],
            "display_name": f"{a['code']} {a['name']}",
            "category": a["category"],
            "direction": a["direction"],
            "is_custom": False,
            "parent_code": None,
            "tax_flag": "",
            "level": 1,
        })
    # 自定义二级科目
    df = get_custom_accounts()
    for _, row in df.iterrows():
        result.append({
            "code": row["full_code"],
            "name": row["full_name"],
            "display_name": f"{row['full_code']} {row['full_name']}  ← {row['parent_name']}",
            "category": row["category"],
            "direction": row["direction"],
            "is_custom": True,
            "parent_code": row["parent_code"],
            "tax_flag": row.get("tax_flag", ""),
            "level": 2,
        })
    return result


@st.cache_data(ttl=300, show_spinner=False)
def get_account_names_dynamic():
    """获取动态科目名称列表（用于下拉选择框），缓存 5 分钟"""
    accounts = get_all_accounts_for_display()
    return [a["display_name"] for a in accounts]


def get_account_info_dynamic(code):
    """根据科目编码获取科目信息（先查系统科目，再查自定义科目）"""
    # 先查系统一级科目
    acc = ACCOUNT_MAP.get(code)
    if acc:
        return {
            "code": acc["code"],
            "name": acc["name"],
            "display_name": f"{acc['code']} {acc['name']}",
            "category": acc["category"],
            "direction": acc["direction"],
            "is_custom": False,
            "parent_code": None,
            "tax_flag": "",
            "level": 1,
        }
    # 再查自定义二级科目
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT * FROM custom_accounts WHERE full_code = ?", (code,))
    row = c.fetchone()
    conn.close()
    if row:
        return {
            "code": row[0],
            "name": row[5],
            "display_name": f"{row[0]} {row[5]}",
            "category": row[6],
            "direction": row[7],
            "is_custom": True,
            "parent_code": row[1],
            "tax_flag": row[8] or "",
            "level": 2,
        }
    return None


@st.cache_data(ttl=300, show_spinner=False)
def get_sub_account_codes(parent_code):
    """获取某个一级科目下所有二级科目的编码列表"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute(
        "SELECT full_code FROM custom_accounts WHERE parent_code = ?",
        (parent_code,)
    )
    codes = [row[0] for row in c.fetchall()]
    conn.close()
    return codes

def _now():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


# ============================================================
# 第 3 部分：核心会计计算函数
# ============================================================

@st.cache_data(ttl=30, show_spinner=False)
def calc_account_balance(account_code):
    """
    计算某个科目的期末余额
    返回：(期初借方, 期初贷方, 本期借方发生额, 本期贷方发生额, 期末余额)

    计算规则：
      借方科目（资产/成本）：期末 = 期初借方 + 本期借方 - 本期贷方
      贷方科目（负债/权益）：期末 = 期初贷方 + 本期贷方 - 本期借方

    如果是一级科目，自动汇总其名下所有二级科目的余额。
    """
    # 先判断是系统科目还是自定义二级科目
    account = ACCOUNT_MAP.get(account_code)
    if account:
        direction = account["direction"]
        is_primary = True
    else:
        # 查自定义二级科目
        info = get_account_info_dynamic(account_code)
        if info:
            direction = info["direction"]
            is_primary = False
        else:
            return 0, 0, 0, 0, 0

    # 本科目自身的期初余额
    opening_debit, opening_credit = get_opening_balance(account_code)

    # 本科目自身的本期发生额
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("""
        SELECT COALESCE(SUM(debit_amount), 0), COALESCE(SUM(credit_amount), 0)
        FROM vouchers WHERE account_code = ?
    """, (account_code,))
    period_debit, period_credit = c.fetchone()

    # 如果是一级科目，汇总所有二级科目的余额
    if is_primary:
        sub_codes = get_sub_account_codes(account_code)
        for sub_code in sub_codes:
            sub_od, sub_oc = get_opening_balance(sub_code)
            opening_debit += sub_od
            opening_credit += sub_oc
            c.execute("""
                SELECT COALESCE(SUM(debit_amount), 0), COALESCE(SUM(credit_amount), 0)
                FROM vouchers WHERE account_code = ?
            """, (sub_code,))
            sub_pd, sub_pc = c.fetchone()
            period_debit += sub_pd or 0
            period_credit += sub_pc or 0

    conn.close()

    period_debit = period_debit or 0
    period_credit = period_credit or 0

    # 计算期末余额
    if direction == "借":
        ending_balance = opening_debit + period_debit - period_credit
    else:
        ending_balance = opening_credit + period_credit - period_debit

    return opening_debit, opening_credit, period_debit, period_credit, ending_balance


# ============================================================
# 期末自动结转功能
# ----------------------------------------------------------
# 月末：将所有损益类科目余额结转至"本年利润"（4103）
#   收入类（贷方余额）→ 借记收入科目，贷记 4103
#   费用类（借方余额）→ 借记 4103，贷记费用科目
# 年末：再将 4103 余额转入"利润分配-未分配利润"（410403）
#   并可选提取法定盈余公积（净利润的 10%）
# ============================================================

# 需要结转的损益类科目（按类别分组）
_INCOME_ACCOUNT_CODES = [
    "6001", "6021", "6041", "6051",       # 主营/其他业务收入、租赁收入
    "6101", "6102", "6103",               # 公允价值变动损益等
    "6111", "6115", "6117",               # 投资收益、资产处置损益、其他收益
    "6301",                                # 营业外收入
]

_EXPENSE_ACCOUNT_CODES = [
    "6401", "6402", "6403",               # 主营/其他业务成本、税金及附加
    "6405",                                # 研发费用
    "6601", "6602", "6603",               # 销售/管理/财务费用
    "6604", "6605", "6606",               # 勘探/租赁费用、汇兑损益
    "6641", "6642",                        # 信用/资产减值损失
    "6701",                                # 营业外支出
    "6711",                                # 所得税费用
]

# 本年利润科目
_PROFIT_ACCOUNT_CODE = "4103"
# 利润分配-未分配利润科目编码（自定义二级科目）
_UNDISTRIBUTED_PROFIT_CODE = "410403"
# 盈余公积-法定盈余公积
_LEGAL_RESERVE_CODE = "410101"
# 利润分配-提取法定盈余公积
_EXTRACT_LEGAL_RESERVE_CODE = "410401"


def get_carryforward_preview():
    """
    计算期末结转预览数据。
    返回 dict:
      - income_items: [{code, name, balance, direction}]  收入类科目（贷方余额）
      - expense_items: [{code, name, balance, direction}] 费用类科目（借方余额）
      - total_income: 收入合计
      - total_expense: 费用合计
      - net_profit: 净利润（收入 - 费用）
      - profit_balance: 本年利润科目当前余额
    """
    # 清除缓存以确保数据最新
    calc_account_balance.clear()
    get_all_opening_balances.clear()
    get_all_vouchers.clear()

    income_items = []
    expense_items = []

    # 收入类科目（正常余额在贷方）
    for code in _INCOME_ACCOUNT_CODES:
        acc = ACCOUNT_MAP.get(code)
        if not acc:
            continue
        od, oc, pd, pc, ending = calc_account_balance(code)
        # 收入类科目贷方余额 = 期初贷方 + 本期贷方 - 本期借方
        balance = ending
        if abs(balance) > 0.01:
            income_items.append({
                "code": code,
                "name": acc["name"],
                "balance": balance,
                "direction": "贷" if balance > 0 else "借",
            })

    # 费用类科目（正常余额在借方）
    for code in _EXPENSE_ACCOUNT_CODES:
        acc = ACCOUNT_MAP.get(code)
        if not acc:
            continue
        od, oc, pd, pc, ending = calc_account_balance(code)
        balance = ending
        if abs(balance) > 0.01:
            expense_items.append({
                "code": code,
                "name": acc["name"],
                "balance": balance,
                "direction": "借" if balance > 0 else "贷",
            })

    total_income = sum(item["balance"] for item in income_items if item["balance"] > 0)
    total_expense = sum(item["balance"] for item in expense_items if item["balance"] > 0)
    # 负余额的收入类科目视为费用
    total_income -= sum(abs(item["balance"]) for item in income_items if item["balance"] < 0)
    # 负余额的费用类科目视为收入
    total_expense -= sum(abs(item["balance"]) for item in expense_items if item["balance"] < 0)

    # 本年利润当前余额
    _, _, _, _, profit_balance = calc_account_balance(_PROFIT_ACCOUNT_CODE)

    return {
        "income_items": income_items,
        "expense_items": expense_items,
        "total_income": total_income,
        "total_expense": total_expense,
        "net_profit": total_income - total_expense,
        "profit_balance": profit_balance,
    }


def execute_monthly_carryforward(carryforward_date):
    """
    执行月末损益结转：将所有损益类科目余额结转至"本年利润"。
    生成两张凭证：
      1. 结转收入：借各收入科目，贷 4103
      2. 结转费用：借 4103，贷各费用科目
    返回：生成的凭证编号列表
    """
    preview = get_carryforward_preview()
    date_str = carryforward_date.strftime("%Y-%m-%d")
    voucher_numbers = []

    # —— 凭证 1：结转本期收入 ——
    income_lines = []
    for item in preview["income_items"]:
        balance = item["balance"]
        if balance > 0.01:
            # 贷方余额 → 借记结转
            income_lines.append({
                "account_code": item["code"],
                "account_name": item["name"],
                "debit": balance,
                "credit": 0,
            })
        elif balance < -0.01:
            # 借方余额（净损失）→ 贷记结转
            income_lines.append({
                "account_code": item["code"],
                "account_name": item["name"],
                "debit": 0,
                "credit": abs(balance),
            })

    total_income_credit = sum(l["debit"] for l in income_lines)
    total_income_debit = sum(l["credit"] for l in income_lines)
    income_net = total_income_credit - total_income_debit

    if income_lines:
        # 添加本年利润抵消行
        if income_net > 0:
            income_lines.append({
                "account_code": _PROFIT_ACCOUNT_CODE,
                "account_name": ACCOUNT_MAP[_PROFIT_ACCOUNT_CODE]["name"],
                "debit": 0,
                "credit": income_net,
            })
        elif income_net < 0:
            income_lines.append({
                "account_code": _PROFIT_ACCOUNT_CODE,
                "account_name": ACCOUNT_MAP[_PROFIT_ACCOUNT_CODE]["name"],
                "debit": abs(income_net),
                "credit": 0,
            })

        v_num = get_next_voucher_number()
        save_voucher(v_num, date_str, "月末结转：结转本期收入至本年利润", income_lines)
        voucher_numbers.append(v_num)

    # —— 凭证 2：结转本期费用 ——
    expense_lines = []
    for item in preview["expense_items"]:
        balance = item["balance"]
        if balance > 0.01:
            # 借方余额 → 贷记结转
            expense_lines.append({
                "account_code": item["code"],
                "account_name": item["name"],
                "debit": 0,
                "credit": balance,
            })
        elif balance < -0.01:
            # 贷方余额 → 借记结转
            expense_lines.append({
                "account_code": item["code"],
                "account_name": item["name"],
                "debit": abs(balance),
                "credit": 0,
            })

    total_expense_debit = sum(l["debit"] for l in expense_lines)
    total_expense_credit = sum(l["credit"] for l in expense_lines)
    expense_net = total_expense_credit - total_expense_debit

    if expense_lines:
        if expense_net > 0:
            expense_lines.insert(0, {
                "account_code": _PROFIT_ACCOUNT_CODE,
                "account_name": ACCOUNT_MAP[_PROFIT_ACCOUNT_CODE]["name"],
                "debit": expense_net,
                "credit": 0,
            })
        elif expense_net < 0:
            expense_lines.insert(0, {
                "account_code": _PROFIT_ACCOUNT_CODE,
                "account_name": ACCOUNT_MAP[_PROFIT_ACCOUNT_CODE]["name"],
                "debit": 0,
                "credit": abs(expense_net),
            })

        v_num = get_next_voucher_number()
        save_voucher(v_num, date_str, "月末结转：结转本期费用至本年利润", expense_lines)
        voucher_numbers.append(v_num)

    # 清除缓存
    calc_account_balance.clear()
    get_all_vouchers.clear()
    get_all_opening_balances.clear()

    return voucher_numbers


def execute_yearly_carryforward(carryforward_date, extract_legal_reserve=True, reserve_rate=0.10):
    """
    执行年末结转：
      1. 先执行月末损益结转（确保所有损益科目清零）
      2. 将"本年利润"余额转入"利润分配-未分配利润"
      3. （可选）提取法定盈余公积
    返回：{voucher_numbers, net_profit, reserve_amount}
    """
    date_str = carryforward_date.strftime("%Y-%m-%d")
    all_voucher_numbers = []

    # 步骤 1：先执行月末损益结转
    monthly_vouchers = execute_monthly_carryforward(carryforward_date)
    all_voucher_numbers.extend(monthly_vouchers)

    # 步骤 2：本年利润 → 利润分配-未分配利润
    _, _, _, _, profit_balance = calc_account_balance(_PROFIT_ACCOUNT_CODE)
    voucher_numbers = []

    if abs(profit_balance) > 0.01:
        lines = []
        if profit_balance > 0:
            # 净利润：借本年利润，贷未分配利润
            lines.append({
                "account_code": _PROFIT_ACCOUNT_CODE,
                "account_name": ACCOUNT_MAP[_PROFIT_ACCOUNT_CODE]["name"],
                "debit": profit_balance,
                "credit": 0,
            })
            lines.append({
                "account_code": _UNDISTRIBUTED_PROFIT_CODE,
                "account_name": "利润分配-未分配利润",
                "debit": 0,
                "credit": profit_balance,
            })
        else:
            # 净亏损：借未分配利润，贷本年利润
            abs_balance = abs(profit_balance)
            lines.append({
                "account_code": _UNDISTRIBUTED_PROFIT_CODE,
                "account_name": "利润分配-未分配利润",
                "debit": abs_balance,
                "credit": 0,
            })
            lines.append({
                "account_code": _PROFIT_ACCOUNT_CODE,
                "account_name": ACCOUNT_MAP[_PROFIT_ACCOUNT_CODE]["name"],
                "debit": 0,
                "credit": abs_balance,
            })

        v_num = get_next_voucher_number()
        save_voucher(v_num, date_str, "年末结转：本年利润转入利润分配-未分配利润", lines)
        all_voucher_numbers.append(v_num)

    # 步骤 3：提取法定盈余公积
    reserve_amount = 0
    if extract_legal_reserve and profit_balance > 0:
        reserve_amount = round(profit_balance * reserve_rate, 2)
        if reserve_amount > 0.01:
            lines = [
                {
                    "account_code": _EXTRACT_LEGAL_RESERVE_CODE,
                    "account_name": "利润分配-提取法定盈余公积",
                    "debit": reserve_amount,
                    "credit": 0,
                },
                {
                    "account_code": _LEGAL_RESERVE_CODE,
                    "account_name": "盈余公积-法定盈余公积",
                    "debit": 0,
                    "credit": reserve_amount,
                },
            ]
            v_num = get_next_voucher_number()
            save_voucher(v_num, date_str, "年末结转：提取法定盈余公积", lines)
            all_voucher_numbers.append(v_num)

    # 清除缓存
    calc_account_balance.clear()
    get_all_vouchers.clear()
    get_all_opening_balances.clear()

    return {
        "voucher_numbers": all_voucher_numbers,
        "net_profit": profit_balance,
        "reserve_amount": reserve_amount,
    }


def get_account_ledger(account_code):
    """
    获取某个科目的明细账（所有凭证流水 + 运行余额）。
    返回 dict：
      - account_code, account_name, direction
      - opening_balance: 期初余额
      - entries: [{date, voucher_number, summary, debit, credit, balance, direction_label}]
      - total_debit, total_credit: 本期借贷合计
      - ending_balance: 期末余额
    """
    account = ACCOUNT_MAP.get(account_code)
    if account:
        direction = account["direction"]
        acc_name = account["name"]
    else:
        info = get_account_info_dynamic(account_code)
        if info:
            direction = info["direction"]
            acc_name = info["name"]
        else:
            return None

    opening_debit, opening_credit = get_opening_balance(account_code)
    if direction == "借":
        running = (opening_debit or 0) - (opening_credit or 0)
    else:
        running = (opening_credit or 0) - (opening_debit or 0)

    conn = sqlite3.connect(DB_PATH)
    df = pd.read_sql_query(
        "SELECT * FROM vouchers WHERE account_code = ? ORDER BY voucher_date, id",
        conn, params=(account_code,))
    conn.close()

    entries = []
    total_debit = 0.0
    total_credit = 0.0

    for _, row in df.iterrows():
        debit = float(row["debit_amount"] or 0)
        credit = float(row["credit_amount"] or 0)
        total_debit += debit
        total_credit += credit

        if direction == "借":
            running += debit - credit
        else:
            running += credit - debit

        direction_label = "借" if running >= 0 else "贷"
        entries.append({
            "date": row["voucher_date"],
            "voucher_number": row["voucher_number"],
            "summary": row["summary"],
            "debit": debit,
            "credit": credit,
            "balance": running,
            "direction_label": direction_label,
        })

    return {
        "account_code": account_code,
        "account_name": acc_name,
        "direction": direction,
        "opening_balance": running - (total_debit - total_credit) if direction == "借"
                         else running - (total_credit - total_debit),
        "entries": entries,
        "total_debit": total_debit,
        "total_credit": total_credit,
        "ending_balance": running,
    }


# ============================================================
# 第 4 部分：主界面 —— 八个 Tab
# ============================================================

# 只在首次加载时初始化数据库（避免每次 rerun 都创建 31 张表）
if not st.session_state.get("_db_initialized"):
    try:
        init_database()
        st.session_state["_db_initialized"] = True
    except Exception as _init_err:
        st.error(f"⚠️ 数据库初始化失败: {_init_err}")
        st.info(f"数据库模式: {'PostgreSQL' if _USE_POSTGRES else 'SQLite'} | DB_PATH: {DB_PATH}")
        st.stop()

# --- 侧边栏：系统导航 ---
with st.sidebar:
    st.markdown("## 💰 财务会计系统")
    _disp = st.session_state.get("display_name", "")
    _comp_name = st.session_state.get("current_company_name", "")
    if _disp:
        st.markdown(f"👤 **{_disp}**")

    st.markdown("---")

    # 公司切换
    _user_companies = st.session_state.get("user_companies", [])
    _current_co_code = st.session_state.get("current_company_code", "")
    if _user_companies:
        _co_options = [f"{c['code']} | {c['name']}" for c in _user_companies]
        _current_idx = 0
        for i, c in enumerate(_user_companies):
            if c["code"] == _current_co_code:
                _current_idx = i
                break
        _selected_co = st.selectbox(
            "🏢 当前公司",
            _co_options,
            index=_current_idx,
            key="company_switcher",
        )
        _selected_co_code = _selected_co.split(" | ")[0] if " | " in _selected_co else _selected_co
        if _selected_co_code != _current_co_code:
            st.session_state["current_company_code"] = _selected_co_code
            st.session_state["current_company_name"] = _selected_co.split(" | ", 1)[1] if " | " in _selected_co else ""
            st.session_state["_db_initialized"] = False
            st.rerun()

    # 公司管理 & 退出
    _sb_col1, _sb_col2 = st.columns(2)
    with _sb_col1:
        if st.button("🏢 公司管理", key="company_mgmt_btn", use_container_width=True):
            st.session_state["_show_company_mgmt"] = not st.session_state.get("_show_company_mgmt", False)
            st.rerun()
    with _sb_col2:
        if st.button("🚪 退出登录", key="logout_btn", use_container_width=True):
            for k in list(st.session_state.keys()):
                del st.session_state[k]
            st.rerun()

    st.markdown("---")

    # 功能导航树（参照用友/金蝶业务域分组）
    st.markdown("### 📋 功能导航")
    st.markdown("""
    **💰 财务会计**
    > 📝 记账 · 📊 报表 · 🤖 AI问答
    """)
    st.caption("💡 点击上方对应标签页切换业务域")

# --- 主界面标题 ---
st.title("💰 财务会计系统")
st.caption("离线记账 · 自动报表 · AI 智能问答 ｜ 财务会计专用版")

# --- 公司管理弹窗 ---
if st.session_state.get("_show_company_mgmt"):
    _mg_username = st.session_state.get("username", "default")
    st.markdown("### 🏢 公司管理")

    _mg_tab1, _mg_tab2, _mg_tab3 = st.tabs(["📋 公司列表", "➕ 新建公司", "✏️ 编辑公司"])

    with _mg_tab1:
        _companies_list = get_user_companies(_mg_username)
        if _companies_list:
            _list_data = []
            for c in _companies_list:
                _list_data.append({
                    "公司编码": c["code"],
                    "公司名称": c["name"],
                    "税号": c["tax_id"],
                    "联系电话": c["phone"],
                    "是否默认": "✅ 是" if c["is_default"] else "—",
                    "创建时间": c["created_at"],
                })
            st.dataframe(_list_data, use_container_width=True, hide_index=True)

            # 删除公司
            st.markdown("---")
            st.markdown("#### 🗑️ 删除公司")
            _del_options = [f"{c['code']} | {c['name']}" for c in _companies_list]
            _del_selected = st.selectbox("选择要删除的公司", _del_options, key="del_company_sel")
            if st.button("确认删除", key="del_company_btn", type="secondary"):
                _del_code = _del_selected.split(" | ")[0] if " | " in _del_selected else _del_selected
                _ok, _msg = delete_company(_mg_username, _del_code)
                if _ok:
                    st.success(_msg)
                    st.session_state["user_companies"] = get_user_companies(_mg_username)
                    if _del_code == st.session_state.get("current_company_code"):
                        _remaining = st.session_state["user_companies"]
                        if _remaining:
                            st.session_state["current_company_code"] = _remaining[0]["code"]
                            st.session_state["current_company_name"] = _remaining[0]["name"]
                        st.session_state["_db_initialized"] = False
                    st.rerun()
                else:
                    st.error(_msg)
        else:
            st.info("暂无公司记录")

    with _mg_tab2:
        with st.form("create_company_form"):
            _new_code = st.text_input("公司编码 *", placeholder="如：CO002", key="new_co_code")
            _new_name = st.text_input("公司名称 *", placeholder="如：XX科技有限公司", key="new_co_name")
            _new_tax = st.text_input("统一社会信用代码（税号）", placeholder="选填", key="new_co_tax")
            _new_addr = st.text_input("公司地址", placeholder="选填", key="new_co_addr")
            _new_phone = st.text_input("联系电话", placeholder="选填", key="new_co_phone")
            _new_submit = st.form_submit_button("➕ 创建公司")

            if _new_submit:
                if not _new_code or not _new_name:
                    st.error("公司编码和名称不能为空")
                else:
                    _ok, _msg = create_company(_mg_username, _new_code.strip(), _new_name.strip(),
                                               _new_tax.strip(), _new_addr.strip(), _new_phone.strip())
                    if _ok:
                        st.success(_msg)
                        st.session_state["user_companies"] = get_user_companies(_mg_username)
                        st.rerun()
                    else:
                        st.error(_msg)

    with _mg_tab3:
        _edit_companies = get_user_companies(_mg_username)
        if _edit_companies:
            _edit_options = [f"{c['code']} | {c['name']}" for c in _edit_companies]
            _edit_selected = st.selectbox("选择公司", _edit_options, key="edit_company_sel")
            _edit_code = _edit_selected.split(" | ")[0] if " | " in _edit_selected else _edit_selected

            _edit_co = None
            for c in _edit_companies:
                if c["code"] == _edit_code:
                    _edit_co = c
                    break

            if _edit_co:
                with st.form("edit_company_form"):
                    _ed_name = st.text_input("公司名称", value=_edit_co["name"], key="ed_co_name")
                    _ed_tax = st.text_input("税号", value=_edit_co["tax_id"], key="ed_co_tax")
                    _ed_addr = st.text_input("公司地址", value=_edit_co["address"], key="ed_co_addr")
                    _ed_phone = st.text_input("联系电话", value=_edit_co["phone"], key="ed_co_phone")
                    _ed_submit = st.form_submit_button("💾 保存修改")

                    if _ed_submit:
                        _ok, _msg = update_company(_mg_username, _edit_code,
                                                   company_name=_ed_name.strip(),
                                                   tax_id=_ed_tax.strip(),
                                                   address=_ed_addr.strip(),
                                                   phone=_ed_phone.strip())
                        if _ok:
                            st.success(_msg)
                            if _edit_code == st.session_state.get("current_company_code"):
                                st.session_state["current_company_name"] = _ed_name.strip()
                            st.session_state["user_companies"] = get_user_companies(_mg_username)
                            st.rerun()
                        else:
                            st.error(_msg)
        else:
            st.info("暂无可编辑的公司")

    st.markdown("---")


# === 财务会计导航 ===
tab1, tab2, tab3 = st.tabs(["📝 记账", "📊 报表", "🤖 AI 问答"])

# ============================================================
# 【💰 财务会计】模块一：记账（总账核算）
# ============================================================
with tab1:
    st.header("📝 总账记账")
    st.caption("期初余额 · 录入凭证 · 凭证查询 · ✅凭证审核 · 明细账 · 科目管理 · 🔄 期末结转 ｜ 完全离线，无需 API")

    sub1, sub2, sub3, sub3a, sub4, sub5, sub6 = st.tabs(["期初余额", "录入凭证", "凭证查询", "✅ 凭证审核", "明细账", "科目管理", "🔄 期末结转"])

    # --- 期初余额 ---
    with sub1:
        st.subheader("期初余额录入")

        # 动态获取科目列表（系统科目 + 自定义二级科目）
        _dynamic_names = get_account_names_dynamic()

        # 录入区域：科目 + 金额
        cols_ob_in = st.columns([3, 2, 1])
        with cols_ob_in[0]:
            selected_ob = st.selectbox(
                "科目", _dynamic_names,
                key="ob_input_acc", label_visibility="collapsed",
                placeholder="搜索科目名称或编号...",
            )
        with cols_ob_in[1]:
            # 保存后清空输入框的标志位（必须在 widget 创建前处理）
            if st.session_state.pop("_clear_ob_amt", False):
                st.session_state["ob_input_amt"] = ""
                st.session_state["_money_init_ob_input_amt"] = True
            ob_amt_input = money_input(
                "金额", key="ob_input_amt", min_value=0.0,
                label_visibility="collapsed", placeholder="0.00",
            )
        with cols_ob_in[2]:
            save_clicked = st.button("💾 保存", key="ob_save_btn", use_container_width=True)

        code_ob = selected_ob.split(" ")[0]
        # 从显示名中提取科目名称（处理自定义科目带 ← 的情况）
        _name_part = selected_ob.split(" ", 1)[1] if " " in selected_ob else selected_ob
        name_ob = _name_part.split("  ←")[0].strip() if "  ←" in _name_part else _name_part.strip()
        acc_info_ob = get_account_info_dynamic(code_ob) or {}
        direction_ob = acc_info_ob.get("direction", "借")
        color_ob = "#1976d2" if direction_ob == "借" else "#e53935"
        st.caption(f"已选：**{code_ob} {name_ob}** ｜ 方向：<span style='color:{color_ob};font-weight:600'>{direction_ob}方</span>", unsafe_allow_html=True)

        if save_clicked:
            if ob_amt_input > 0:
                if direction_ob == "借":
                    save_opening_balance(code_ob, ob_amt_input, 0)
                else:
                    save_opening_balance(code_ob, 0, ob_amt_input)
                st.success(f"✅ 已保存：{name_ob} = {ob_amt_input:,.2f} 元（{direction_ob}方）")
                # 设置清空标志，在下一次 rerun 时清空输入框
                st.session_state["_clear_ob_amt"] = True
                st.rerun()
            else:
                st.warning("请输入金额")

        # 已保存的期初余额列表（包含一级科目和二级科目）
        all_opening = []
        # 系统一级科目
        for a in ACCOUNT_CHART:
            od_v, oc_v = get_opening_balance(a["code"])
            if od_v or oc_v:
                amt_v = od_v if a["direction"] == "借" else oc_v
                all_opening.append({
                    "code": a["code"],
                    "name": a["name"],
                    "direction": a["direction"],
                    "amount": amt_v,
                    "raw_debit": od_v,
                    "raw_credit": oc_v,
                    "level": 1,
                })
        # 自定义二级科目
        df_custom_ob = get_custom_accounts()
        for _, row in df_custom_ob.iterrows():
            od_v, oc_v = get_opening_balance(row["full_code"])
            if od_v or oc_v:
                amt_v = od_v if row["direction"] == "借" else oc_v
                all_opening.append({
                    "code": row["full_code"],
                    "name": row["full_name"],
                    "direction": row["direction"],
                    "amount": amt_v,
                    "raw_debit": od_v,
                    "raw_credit": oc_v,
                    "level": 2,
                })

        st.markdown("---")

        if all_opening:
            # 表头
            st.markdown(
                '<div style="display:flex; font-weight:600; font-size:13px; '
                'padding:6px 0; border-bottom:2px solid #333; margin-bottom:4px;">'
                '<span style="flex:1; text-align:center;">科目</span>'
                '<span style="flex:0.5; text-align:center;">方向</span>'
                '<span style="flex:1; text-align:center;">期初余额</span>'
                '<span style="flex:0.5; text-align:center;">删除</span>'
                '</div>',
                unsafe_allow_html=True,
            )
            for o in all_opening:
                cols_ob = st.columns([1, 0.5, 1, 0.5])
                with cols_ob[0]:
                    st.markdown(
                        f'<div style="padding-top:5px; font-size:14px;">'
                        f'{o["code"]} {o["name"]}</div>',
                        unsafe_allow_html=True,
                    )
                with cols_ob[1]:
                    c = "#1976d2" if o["direction"] == "借" else "#e53935"
                    st.markdown(
                        f'<div style="text-align:center; padding-top:5px; '
                        f'font-size:13px; color:{c}; font-weight:600;">{o["direction"]}方</div>',
                        unsafe_allow_html=True,
                    )
                with cols_ob[2]:
                    st.markdown(
                        f'<div style="text-align:center; padding-top:5px; '
                        f'font-size:14px; font-weight:600;">{o["amount"]:,.2f}</div>',
                        unsafe_allow_html=True,
                    )
                with cols_ob[3]:
                    if st.button("🗑", key=f"ob_del_{o['code']}", help=f"删除{o['name']}的期初余额"):
                        save_opening_balance(o["code"], 0, 0)
                        st.rerun()

            # 合计行
            total_debit_ob = sum(o["amount"] for o in all_opening if o["direction"] == "借")
            total_credit_ob = sum(o["amount"] for o in all_opening if o["direction"] == "贷")
            st.markdown(
                f'<div style="display:flex; font-weight:600; font-size:14px; '
                f'padding:8px 0; border-top:2px solid #333; border-bottom:2px solid #333;">'
                f'<span style="flex:1;">合计</span>'
                f'<span style="flex:0.5;"></span>'
                f'<span style="flex:1; text-align:center;">'
                f'借: {total_debit_ob:,.2f} ｜ 贷: {total_credit_ob:,.2f}</span>'
                f'<span style="flex:0.5;"></span>'
                f'</div>',
                unsafe_allow_html=True,
            )
            if abs(total_debit_ob - total_credit_ob) > 0.01:
                st.warning(f"⚠️ 借贷不平衡！差额：{abs(total_debit_ob - total_credit_ob):,.2f} 元")
            else:
                st.success("✅ 借贷平衡")
        else:
            st.info("💡 尚未录入任何期初余额。在上方选择科目并输入金额后点击「保存」。")

    # --- 录入凭证 ---
    with sub2:
        st.subheader("录入记账凭证")

        # 凭证金额专用回调：Enter 格式化 + 最后一行自动新增行
        def _voucher_money_cb(key):
            val = st.session_state.get(key, "")
            if not val or not str(val).strip():
                return
            cleaned = str(val).replace(",", "").replace("，", "").replace(" ", "").strip()
            try:
                num = float(cleaned)
                formatted = f"{num:,.2f}"
                current = str(val).strip()
                if current != formatted:
                    # Enter 格式化金额
                    st.session_state[key] = formatted
                    # 如果是最后一行的金额，标记需要新增行
                    parts = key.split("_")
                    if len(parts) >= 3:
                        _row_idx = int(parts[-1])
                        if _row_idx == st.session_state.get("voucher_row_count", 2) - 1:
                            st.session_state["_vch_add_row"] = True
            except ValueError:
                pass

        # 中文大写金额转换函数
        def to_chinese_amount(amount):
            """把数字金额转成中文大写，如：壹仟贰佰元整"""
            if abs(amount) < 0.005:
                return "零元整"
            digits = "零壹贰叁肆伍陆柒捌玖"
            units = ["", "拾", "佰", "仟", "万", "拾", "佰", "仟", "亿", "拾", "佰", "仟"]
            # 分离整数和小数部分
            integer_part = int(abs(amount))
            decimal_part = round(abs(amount) - integer_part, 2)

            # 整数部分转换
            if integer_part == 0:
                int_str = ""
            else:
                int_str = ""
                s = str(integer_part)
                length = len(s)
                for i, ch in enumerate(s):
                    n = int(ch)
                    pos = length - 1 - i
                    if n == 0:
                        if int_str and not int_str.endswith("零") and not int_str.endswith("亿") and not int_str.endswith("万"):
                            int_str += "零"
                    else:
                        int_str += digits[n] + units[pos]
                # 清理末尾多余的零
                int_str = int_str.rstrip("零")
                int_str += "元"

            # 小数部分
            if decimal_part == 0:
                dec_str = "整"
            else:
                jiao = int(decimal_part * 10)
                fen = int(round((decimal_part * 100) % 10))
                dec_str = ""
                if jiao > 0:
                    dec_str += digits[jiao] + "角"
                if fen > 0:
                    dec_str += digits[fen] + "分"
                if not dec_str:
                    dec_str = "整"

            result = int_str + dec_str
            if amount < 0:
                result = "负" + result
            return result

        # === 凭证头部 ===
        st.markdown(
            '<div style="text-align:center; padding:8px 0; font-size:20px; font-weight:600; '
            'border-bottom: 2px solid #333; margin-bottom:16px;">记 账 凭 证</div>',
            unsafe_allow_html=True,
        )

        # 凭证字、号、日期、附单据
        VOUCHER_TYPES = {"记": "记账凭证", "收": "收款凭证", "付": "付款凭证", "转": "转账凭证"}

        col_vtype, col_vnum, col_vdate, col_attach = st.columns([1, 1, 1.5, 1])
        with col_vtype:
            voucher_type = st.selectbox("凭证字", list(VOUCHER_TYPES.keys()), key="vtype")
        with col_vnum:
            next_num = get_next_voucher_number()
            # 提取数字部分
            base_num = int(next_num.replace("记字第", "").replace("号", "")) if "记字第" in next_num else 1
            vnum = st.number_input("字第 ___ 号", min_value=1, value=base_num, step=1, key="vnum")
        with col_vdate:
            voucher_date = st.date_input("凭证日期")
        with col_attach:
            attach_count = st.number_input("附单据", min_value=0, value=0, step=1, key="attach")

        full_voucher_number = f"{voucher_type}字第{vnum:03d}号"

        # === 分录明细表 ===
        st.markdown("**分录明细**（每行填一条，至少一行借、一行贷）")

        # 动态行数：默认2行，可增减
        if "voucher_row_count" not in st.session_state:
            st.session_state.voucher_row_count = 2

        # 增减行按钮
        col_add, col_del, col_info = st.columns([1, 1, 3])
        with col_add:
            if st.button("➕ 添加一行", key="add_row"):
                st.session_state.voucher_row_count += 1
                st.rerun()
        with col_del:
            if st.session_state.voucher_row_count > 2:
                if st.button("➖ 删除最后一行", key="del_row"):
                    # 清空最后一行的所有数据
                    last = st.session_state.voucher_row_count - 1
                    for suffix in ["sum", "acc", "debit", "credit"]:
                        k = f"row_{suffix}_{last}"
                        if k in st.session_state:
                            del st.session_state[k]
                        # 清除 money_input 初始化标记
                        ik = f"_money_init_{k}"
                        if ik in st.session_state:
                            del st.session_state[ik]
                    st.session_state.voucher_row_count -= 1
                    st.rerun()
        with col_info:
            st.caption(f"当前 {st.session_state.voucher_row_count} 行（填入金额按 Enter 格式化，最后一行自动新增）")

        row_count = st.session_state.voucher_row_count
        lines = []

        # 用 HTML 画表头
        st.markdown(
            '<div style="display:flex; font-weight:600; font-size:13px; '
            'padding:6px 0; border-bottom:2px solid #333; margin-bottom:4px;">'
            '<span style="width:40px; text-align:center;">行次</span>'
            '<span style="flex:1; text-align:center;">摘要</span>'
            '<span style="flex:1.2; text-align:center;">会计科目</span>'
            '<span style="width:120px; text-align:center;">借方金额</span>'
            '<span style="width:120px; text-align:center;">贷方金额</span>'
            '</div>',
            unsafe_allow_html=True,
        )

        for i in range(row_count):
            cols = st.columns([0.5, 2, 2.5, 1.5, 1.5])
            with cols[0]:
                st.markdown(
                    f'<div style="text-align:center; padding-top:8px; font-size:14px; '
                    f'color:#666;">{i+1}</div>',
                    unsafe_allow_html=True,
                )
            with cols[1]:
                row_summary = st.text_input(
                    f"摘要{i+1}", label_visibility="collapsed",
                    key=f"row_sum_{i}", placeholder="摘要"
                )
            with cols[2]:
                selected = st.selectbox(
                    f"科目{i+1}", get_account_names_dynamic(),
                    key=f"row_acc_{i}", label_visibility="collapsed"
                )
            with cols[3]:
                debit = money_input(
                    f"借{i+1}", key=f"row_debit_{i}", min_value=0.0,
                    label_visibility="collapsed",
                    on_change_cb=_voucher_money_cb, on_change_args=(f"row_debit_{i}",),
                )
            with cols[4]:
                credit = money_input(
                    f"贷{i+1}", key=f"row_credit_{i}", min_value=0.0,
                    label_visibility="collapsed",
                    on_change_cb=_voucher_money_cb, on_change_args=(f"row_credit_{i}",),
                )

            if debit > 0 or credit > 0:
                code = selected.split(" ")[0]
                # 从显示名中提取科目名称（处理自定义科目带 ← 的情况）
                _n_part = selected.split(" ", 1)[1] if " " in selected else selected
                name = _n_part.split("  ←")[0].strip() if "  ←" in _n_part else _n_part.strip()
                lines.append({
                    "account_code": code,
                    "account_name": name,
                    "debit": debit,
                    "credit": credit,
                    "summary": row_summary if row_summary else "",
                })

        # === 最后一行金额格式化后自动新增行 ===
        if st.session_state.pop("_vch_add_row", False):
            st.session_state.voucher_row_count += 1
            st.rerun()

        # === 合计行 ===
        total_debit = sum(l["debit"] for l in lines)
        total_credit = sum(l["credit"] for l in lines)
        is_balanced = abs(total_debit - total_credit) < 0.01
        chinese_total = to_chinese_amount(max(total_debit, total_credit))

        st.markdown("---")
        st.markdown(
            f'<div style="display:flex; font-weight:600; font-size:14px; '
            f'padding:8px 0; border-top:2px solid #333; border-bottom:2px solid #333;">'
            f'<span style="width:40px;"></span>'
            f'<span style="flex:1;">合计：{chinese_total}</span>'
            f'<span style="flex:1.2;"></span>'
            f'<span style="width:120px; text-align:center;">{total_debit:,.2f}</span>'
            f'<span style="width:120px; text-align:center;">{total_credit:,.2f}</span>'
            f'</div>',
            unsafe_allow_html=True,
        )

        # 借贷平衡校验
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("借方合计", f"{total_debit:,.2f} 元")
        with col2:
            st.metric("贷方合计", f"{total_credit:,.2f} 元")
        with col3:
            if is_balanced:
                st.metric("校验结果", "✅ 平衡")
            else:
                st.metric("校验结果", "❌ 不平衡")

        if not is_balanced and (total_debit > 0 or total_credit > 0):
            st.error(f"借贷不平衡！差额：{abs(total_debit - total_credit):,.2f} 元。请调整金额使借贷相等。")

        # 保存 + 清空按钮
        col_save, col_clear = st.columns([1, 1])
        with col_save:
            if st.button("💾 保存凭证", type="primary", use_container_width=True):
                if not lines:
                    st.warning("请至少填写一行分录")
                elif not is_balanced:
                    st.warning("借贷不平衡，无法保存。请调整金额。")
                elif len(lines) < 2:
                    st.warning("至少需要两行分录（一行借、一行贷）")
                else:
                    # 用第一行的摘要作为主摘要
                    main_summary = lines[0]["summary"] if lines[0]["summary"] else "综合凭证"
                    save_voucher(
                        full_voucher_number,
                        voucher_date.strftime("%Y-%m-%d"),
                        main_summary,
                        lines,
                    )
                    st.success(f"✅ 凭证已保存！编号：{full_voucher_number}")
                    # 重置行数为2，清除所有行数据
                    _max_rows = st.session_state.voucher_row_count
                    for _r in range(_max_rows):
                        for _suf in ["sum", "acc", "debit", "credit"]:
                            _rk = f"row_{_suf}_{_r}"
                            if _rk in st.session_state:
                                del st.session_state[_rk]
                            _ik = f"_money_init_{_rk}"
                            if _ik in st.session_state:
                                del st.session_state[_ik]
                    st.session_state.voucher_row_count = 2
                    st.rerun()
        with col_clear:
            if st.button("🔄 清空重填", use_container_width=True):
                _max_rows = st.session_state.voucher_row_count
                for _r in range(_max_rows):
                    for _suf in ["sum", "acc", "debit", "credit"]:
                        _rk = f"row_{_suf}_{_r}"
                        if _rk in st.session_state:
                            del st.session_state[_rk]
                        _ik = f"_money_init_{_rk}"
                        if _ik in st.session_state:
                            del st.session_state[_ik]
                st.session_state.voucher_row_count = 2
                st.rerun()

    # --- 凭证查询 ---
    with sub3:
        st.subheader("凭证查询")

        df = get_all_vouchers()
        if df.empty:
            st.info("暂无凭证记录。请先在「录入凭证」中添加。")
        else:
            # 按凭证编号分组显示
            voucher_numbers = df["voucher_number"].unique()

            st.write(f"共 {len(voucher_numbers)} 张凭证")

            # --- 导出凭证为格式化 Excel ---
            def _write_voucher_ws(ws, vdf_single, vnum, vdate, vsummary):
                """将一张记账凭证写入指定 worksheet（带专业格式）"""
                from openpyxl.styles import Font, Alignment, Border, Side
                import math

                _thin = Side(style='thin')
                _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
                _title_font = Font(name='Arial', size=18, bold=True)
                _hdr_font = Font(name='Arial', size=10, bold=True)
                _data_font = Font(name='Arial', size=10)
                _center = Alignment(horizontal='center', vertical='center')
                _left = Alignment(horizontal='left', vertical='center')
                _right = Alignment(horizontal='right', vertical='center')

                for col, w in [('A', 25), ('B', 30), ('C', 18), ('D', 18)]:
                    ws.column_dimensions[col].width = w

                # 第1行：标题
                ws.merge_cells('A1:D1')
                ws['A1'] = '记 账 凭 证'
                ws['A1'].font = _title_font
                ws['A1'].alignment = _center
                ws.row_dimensions[1].height = 30

                # 第2行：日期 / 凭证编号 / 附件
                ws['A2'] = f'日期：{vdate}'
                ws['A2'].font = _data_font
                ws['A2'].alignment = _left
                ws.merge_cells('B2:C2')
                ws['B2'] = f'凭证编号：{vnum}'
                ws['B2'].font = _data_font
                ws['B2'].alignment = _center
                ws['D2'] = '附件：    张'
                ws['D2'].font = _data_font
                ws['D2'].alignment = _right
                ws.row_dimensions[2].height = 20

                # 第3行：摘要
                ws.merge_cells('A3:D3')
                ws['A3'] = f'摘要：{vsummary}'
                ws['A3'].font = _data_font
                ws['A3'].alignment = _left
                ws.row_dimensions[3].height = 20

                # 第4行：表头
                for i, h in enumerate(['摘要', '科目', '借方金额', '贷方金额'], 1):
                    c = ws.cell(row=4, column=i, value=h)
                    c.font = _hdr_font
                    c.alignment = _center
                    c.border = _border
                ws.row_dimensions[4].height = 20

                # 数据行
                d_total = 0.0
                c_total = 0.0
                r = 5
                for _, row in vdf_single.iterrows():
                    ws.row_dimensions[r].height = 22
                    debit = float(row["debit_amount"] or 0)
                    credit = float(row["credit_amount"] or 0)
                    d_total += debit
                    c_total += credit
                    vals = [
                        vsummary,
                        row["account_name"],
                        f"{debit:,.2f}" if debit else "",
                        f"{credit:,.2f}" if credit else "",
                    ]
                    for ci, val in enumerate(vals, 1):
                        c = ws.cell(row=r, column=ci, value=val if val else None)
                        c.font = _data_font
                        c.border = _border
                        c.number_format = '@'
                        c.alignment = _left if ci <= 2 else _right
                    r += 1

                # 合计行
                ws.row_dimensions[r].height = 22
                for ci, val in enumerate(['合计', '', f"{d_total:,.2f}", f"{c_total:,.2f}"], 1):
                    c = ws.cell(row=r, column=ci, value=val if val else None)
                    c.font = _hdr_font
                    c.border = _border
                    c.number_format = '@'
                    c.alignment = _center if ci <= 2 else _right

                # 大写金额行
                r += 2
                ws.merge_cells(f'A{r}:D{r}')

                def _to_chinese(amount):
                    if abs(amount) < 0.005:
                        return "零元整"
                    digits = "零壹贰叁肆伍陆柒捌玖"
                    units = ["", "拾", "佰", "仟", "万", "拾", "佰", "仟", "亿", "拾", "佰", "仟"]
                    integer_part = int(abs(amount))
                    decimal_part = round(abs(amount) - integer_part, 2)
                    if integer_part == 0:
                        int_str = ""
                    else:
                        int_str = ""
                        s = str(integer_part)
                        length = len(s)
                        for i, ch in enumerate(s):
                            n = int(ch)
                            pos = length - 1 - i
                            if n == 0:
                                if int_str and not int_str.endswith("零") and not int_str.endswith("亿") and not int_str.endswith("万"):
                                    int_str += "零"
                            else:
                                int_str += digits[n] + units[pos]
                        int_str = int_str.rstrip("零") + "元"
                    if decimal_part == 0:
                        dec_str = "整"
                    else:
                        jiao = int(decimal_part * 10)
                        fen = int(round((decimal_part * 100) % 10))
                        dec_str = ""
                        if jiao > 0:
                            dec_str += digits[jiao] + "角"
                        if fen > 0:
                            dec_str += digits[fen] + "分"
                        if not dec_str:
                            dec_str = "整"
                    result = int_str + dec_str
                    return ("负" + result) if amount < 0 else result

                ws.cell(row=r, column=1).value = f'合计金额（大写）：{_to_chinese(d_total)}'
                ws.cell(row=r, column=1).font = _data_font
                ws.cell(row=r, column=1).alignment = _left

                # 制单 / 审核
                r += 2
                ws.cell(row=r, column=1).value = '制单：'
                ws.cell(row=r, column=1).font = _data_font
                ws.cell(row=r, column=3).value = '审核：'
                ws.cell(row=r, column=3).font = _data_font

            def export_single_voucher_excel(vdf_single, vnum, vdate, vsummary):
                """导出单张凭证为 Excel"""
                from openpyxl import Workbook
                import io
                wb = Workbook()
                ws = wb.active
                ws.title = vnum[:31]
                _write_voucher_ws(ws, vdf_single, vnum, vdate, vsummary)
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                return output

            def export_all_vouchers_excel():
                """导出所有凭证到一个 Excel（每张一个 sheet）"""
                from openpyxl import Workbook
                import io
                wb = Workbook()
                wb.remove(wb.active)
                for vnum in voucher_numbers:
                    vdf_single = df[df["voucher_number"] == vnum]
                    vdate = vdf_single.iloc[0]["voucher_date"]
                    vsummary = vdf_single.iloc[0]["summary"]
                    ws = wb.create_sheet(title=vnum[:31])
                    _write_voucher_ws(ws, vdf_single, vnum, vdate, vsummary)
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                return output

            st.download_button(
                label="📥 下载所有凭证（Excel）",
                data=export_all_vouchers_excel(),
                file_name="所有凭证.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

            for vnum in voucher_numbers:
                vdf = df[df["voucher_number"] == vnum]
                vdate = vdf.iloc[0]["voucher_date"]
                vsummary = vdf.iloc[0]["summary"]

                # 获取审核状态
                _v_audit_status = vdf.iloc[0].get("audit_status", "未审核") or "未审核"
                _v_auditor = vdf.iloc[0].get("auditor", "") or ""
                _v_audit_date = vdf.iloc[0].get("audit_date", "") or ""
                _audit_icon = "🟢" if _v_audit_status == "已审核" else "🟡"
                _audit_label = f"{_audit_icon} {_v_audit_status}" if _v_audit_status == "已审核" else "🟡 未审核"
                _is_audited = (_v_audit_status == "已审核")

                d_total = vdf["debit_amount"].sum()
                c_total = vdf["credit_amount"].sum()
                balanced = "✅" if abs(d_total - c_total) < 0.01 else "❌"

                with st.expander(f"{_audit_label} | {vnum} | {vdate} | {vsummary} | 借 {d_total:,.2f} = 贷 {c_total:,.2f} {balanced}"):
                    display_df = vdf[["account_name", "debit_amount", "credit_amount"]].copy()
                    display_df.columns = ["科目", "借方金额", "贷方金额"]
                    # 加合计行
                    display_df.loc[len(display_df)] = ["合计", d_total, c_total]
                    display_df = fmt_money_df(display_df, ["借方金额", "贷方金额"])
                    st.dataframe(display_df, use_container_width=True, hide_index=True)

                    # 显示审核信息
                    if _is_audited:
                        st.info(f"✅ 已审核 ｜ 审核人：{_v_auditor}  ｜  审核时间：{_v_audit_date}")
                    else:
                        st.caption("🟡 未审核")

                    st.download_button(
                        label=f"📥 下载此凭证（{vnum}）",
                        data=export_single_voucher_excel(vdf, vnum, vdate, vsummary),
                        file_name=f"凭证_{vnum}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"dl_{vnum}",
                    )

                    # --- 删除凭证 ---
                    st.markdown("---")
                    del_cols = st.columns([1, 3])
                    with del_cols[0]:
                        _del_disabled = _is_audited
                        if st.button("🗑️ 删除此凭证", key=f"del_{vnum}",
                                     type="secondary", use_container_width=True,
                                     disabled=_del_disabled,
                                     help="已审核凭证不可删除，请先反审核" if _del_disabled else None):
                            st.session_state[f"_confirm_del_{vnum}"] = True
                    with del_cols[1]:
                        if _is_audited:
                            st.caption("🔒 已审核凭证不可删除，请先在「凭证审核」中反审核")
                        if st.session_state.get(f"_confirm_del_{vnum}"):
                            st.warning("⚠️ 确认要删除这张凭证吗？此操作不可撤销！")
                            confirm_cols = st.columns(2)
                            with confirm_cols[0]:
                                if st.button("✅ 确认删除", key=f"confirm_del_{vnum}",
                                             type="primary", use_container_width=True):
                                    delete_voucher(vnum)
                                    st.session_state.pop(f"_confirm_del_{vnum}", None)
                                    st.success(f"凭证 {vnum} 已删除！")
                                    st.rerun()
                            with confirm_cols[1]:
                                if st.button("❌ 取消", key=f"cancel_del_{vnum}",
                                             use_container_width=True):
                                    st.session_state.pop(f"_confirm_del_{vnum}", None)
                                    st.rerun()

                    # --- 修改凭证 ---
                    st.markdown("---")
                    edit_cols = st.columns([1, 3])
                    with edit_cols[0]:
                        _edit_disabled = _is_audited
                        if st.button("✏️ 修改此凭证", key=f"edit_{vnum}",
                                     use_container_width=True,
                                     disabled=_edit_disabled,
                                     help="已审核凭证不可修改，请先反审核" if _edit_disabled else None):
                            # 加载凭证数据到编辑模式
                            lines = get_voucher_by_number(vnum)
                            info = get_voucher_info(vnum)
                            st.session_state[f"_edit_voucher_{vnum}"] = {
                                "lines": lines,
                                "date": info["date"] if info else vdate,
                                "summary": info["summary"] if info else vsummary,
                            }
                    with edit_cols[1]:
                        if _is_audited:
                            st.caption("🔒 已审核凭证不可修改，请先在「凭证审核」中反审核")

                    # 编辑界面
                    edit_data = st.session_state.get(f"_edit_voucher_{vnum}")
                    if edit_data:
                        st.markdown("#### ✏️ 修改凭证")
                        _dynamic_names_edit = get_account_names_dynamic()

                        # 修改日期和摘要
                        edit_c1, edit_c2 = st.columns([1, 2])
                        with edit_c1:
                            try:
                                _edit_date_obj = datetime.strptime(edit_data["date"], "%Y-%m-%d")
                            except (ValueError, TypeError):
                                _edit_date_obj = datetime.now()
                            new_date = st.date_input(
                                "凭证日期", value=_edit_date_obj,
                                key=f"edit_date_{vnum}",
                            )
                        with edit_c2:
                            new_summary = st.text_input(
                                "摘要", value=edit_data["summary"],
                                key=f"edit_summary_{vnum}",
                            )

                        # 修改分录行
                        st.markdown("##### 分录明细")
                        edit_lines = edit_data["lines"]

                        # 动态行数控制
                        if f"_edit_row_count_{vnum}" not in st.session_state:
                            st.session_state[f"_edit_row_count_{vnum}"] = len(edit_lines)

                        row_count = st.session_state[f"_edit_row_count_{vnum}"]
                        rc1, rc2 = st.columns([3, 1])
                        with rc2:
                            add_rows = st.number_input(
                                "行数", min_value=2, max_value=20,
                                value=max(row_count, 2), step=1,
                                key=f"edit_rowcnt_{vnum}",
                                label_visibility="collapsed",
                            )
                            if add_rows != row_count:
                                st.session_state[f"_edit_row_count_{vnum}"] = int(add_rows)
                                # 扩展或截断行
                                while len(edit_lines) < int(add_rows):
                                    edit_lines.append({
                                        "account_code": "",
                                        "account_name": "",
                                        "debit": 0,
                                        "credit": 0,
                                })
                                edit_lines = edit_lines[:int(add_rows)]
                                edit_data["lines"] = edit_lines
                                st.rerun()

                        # 渲染分录编辑表格
                        new_lines_data = []
                        total_debit = 0
                        total_credit = 0
                        for i in range(len(edit_lines)):
                            ec1, ec2, ec3, ec4, ec5 = st.columns([3, 2, 2, 1, 1])
                            with ec1:
                                # 构建当前选中值
                                cur_val = None
                                if edit_lines[i]["account_code"]:
                                    cur_val = f'{edit_lines[i]["account_code"]} {edit_lines[i]["account_name"]}'
                                idx_default = 0
                                if cur_val and cur_val in _dynamic_names_edit:
                                    idx_default = _dynamic_names_edit.index(cur_val)
                                sel = st.selectbox(
                                    f"科目{i+1}", _dynamic_names_edit,
                                    index=idx_default, key=f"edit_acc_{vnum}_{i}",
                                    label_visibility="collapsed",
                                )
                                if sel:
                                    parts = sel.split(" ", 1)
                                    acc_code = parts[0]
                                    acc_name = parts[1] if len(parts) > 1 else ""
                                else:
                                    acc_code = ""
                                    acc_name = ""
                            with ec2:
                                d_val = st.text_input(
                                    "借方", value=f"{edit_lines[i]['debit']:,.2f}" if edit_lines[i]['debit'] else "",
                                    key=f"edit_d_{vnum}_{i}",
                                    placeholder="0.00",
                                    label_visibility="collapsed",
                                )
                                d_num = float(str(d_val).replace(",", "").strip() or 0)
                            with ec3:
                                c_val = st.text_input(
                                    "贷方", value=f"{edit_lines[i]['credit']:,.2f}" if edit_lines[i]['credit'] else "",
                                    key=f"edit_c_{vnum}_{i}",
                                    placeholder="0.00",
                                    label_visibility="collapsed",
                                )
                                c_num = float(str(c_val).replace(",", "").strip() or 0)
                            with ec4:
                                if st.button("🗑", key=f"edit_del_row_{vnum}_{i}",
                                             help="删除此行"):
                                    if len(edit_lines) > 2:
                                        edit_lines.pop(i)
                                        st.session_state[f"_edit_row_count_{vnum}"] = len(edit_lines)
                                        st.rerun()
                            with ec5:
                                st.write("")

                            new_lines_data.append({
                                "account_code": acc_code,
                                "account_name": acc_name,
                                "debit": d_num,
                                "credit": c_num,
                            })
                            total_debit += d_num
                            total_credit += c_num

                        # 借贷平衡校验
                        balance_ok = abs(total_debit - total_credit) < 0.01
                        bal_color = "#2e7d32" if balance_ok else "#c62828"
                        st.markdown(
                            f'<div style="text-align:right; font-size:14px; '
                            f'color:{bal_color}; font-weight:600; padding:4px 0;">'
                            f'借方合计：{fmt_money(total_debit)} &nbsp;|&nbsp; '
                            f'贷方合计：{fmt_money(total_credit)} &nbsp;|&nbsp; '
                            f'{"✅ 平衡" if balance_ok else "❌ 不平衡"}'
                            f'</div>',
                            unsafe_allow_html=True,
                        )

                        # 保存 / 取消
                        save_cols = st.columns(2)
                        with save_cols[0]:
                            if st.button("💾 保存修改", key=f"save_edit_{vnum}",
                                         type="primary", use_container_width=True,
                                         disabled=not balance_ok):
                                new_date_str = new_date.strftime("%Y-%m-%d") if new_date else edit_data["date"]
                                # 过滤掉空行
                                filtered_lines = [l for l in new_lines_data if l["account_code"]]
                                if len(filtered_lines) >= 2 and abs(
                                    sum(l["debit"] for l in filtered_lines) -
                                    sum(l["credit"] for l in filtered_lines)
                                ) < 0.01:
                                    update_voucher(vnum, new_date_str, new_summary, filtered_lines)
                                    st.session_state.pop(f"_edit_voucher_{vnum}", None)
                                    st.session_state.pop(f"_edit_row_count_{vnum}", None)
                                    st.success(f"✅ 凭证 {vnum} 已修改成功！")
                                    st.rerun()
                                else:
                                    st.error("保存失败：请确保至少 2 行分录且借贷平衡。")
                        with save_cols[1]:
                            if st.button("❌ 取消修改", key=f"cancel_edit_{vnum}",
                                         use_container_width=True):
                                st.session_state.pop(f"_edit_voucher_{vnum}", None)
                                st.session_state.pop(f"_edit_row_count_{vnum}", None)
                                st.rerun()

    # --- 凭证审核 ---
    with sub3a:
        st.subheader("✅ 凭证审核")
        st.caption("制单 → 审核 → 记账 ｜ 审核后凭证锁定不可修改，需反审核后方可编辑")

        _audit_df = get_all_vouchers()
        if _audit_df.empty:
            st.info("暂无凭证记录。请先在「录入凭证」中添加。")
        else:
            # 获取当前用户名作为审核人
            _auditor = st.session_state.get("display_name", "") or st.session_state.get("username", "")

            # 统计
            _audit_voucher_numbers = _audit_df["voucher_number"].unique()
            _pending_list = []
            _audited_list = []
            for _vn in _audit_voucher_numbers:
                _vdf = _audit_df[_audit_df["voucher_number"] == _vn]
                _status = _vdf.iloc[0].get("audit_status", "未审核") or "未审核"
                if _status == "已审核":
                    _audited_list.append(_vn)
                else:
                    _pending_list.append(_vn)

            _stat_c1, _stat_c2, _stat_c3 = st.columns(3)
            with _stat_c1:
                st.metric("凭证总数", len(_audit_voucher_numbers))
            with _stat_c2:
                st.metric("待审核", len(_pending_list))
            with _stat_c3:
                st.metric("已审核", len(_audited_list))

            st.markdown("---")

            # 审核操作标签页
            _audit_tab1, _audit_tab2, _audit_tab3 = st.tabs(["📋 待审核凭证", "✅ 已审核凭证", "⚡ 批量审核"])

            # --- 待审核凭证列表 ---
            with _audit_tab1:
                if not _pending_list:
                    st.success("🎉 所有凭证已审核完毕！")
                else:
                    for _vn in _pending_list:
                        _vdf = _audit_df[_audit_df["voucher_number"] == _vn]
                        _vdate = _vdf.iloc[0]["voucher_date"]
                        _vsummary = _vdf.iloc[0]["summary"]
                        _d_total = _vdf["debit_amount"].sum()
                        _c_total = _vdf["credit_amount"].sum()
                        _balanced = "✅" if abs(_d_total - _c_total) < 0.01 else "❌"

                        with st.expander(f"🟡 {_vn} | {_vdate} | {_vsummary} | 借 {_d_total:,.2f} = 贷 {_c_total:,.2f} {_balanced}"):
                            _disp_df = _vdf[["account_name", "debit_amount", "credit_amount"]].copy()
                            _disp_df.columns = ["科目", "借方金额", "贷方金额"]
                            _disp_df.loc[len(_disp_df)] = ["合计", _d_total, _c_total]
                            _disp_df = fmt_money_df(_disp_df, ["借方金额", "贷方金额"])
                            st.dataframe(_disp_df, use_container_width=True, hide_index=True)

                            if not _balanced:
                                st.error("❌ 凭证借贷不平衡，无法审核！")
                            else:
                                _ac1, _ac2 = st.columns([1, 2])
                                with _ac1:
                                    if st.button("✅ 审核通过", key=f"audit_{_vn}",
                                                 type="primary", use_container_width=True):
                                        _ok, _msg = audit_voucher(_vn, _auditor)
                                        if _ok:
                                            st.success(_msg)
                                            st.rerun()
                                        else:
                                            st.error(_msg)
                                with _ac2:
                                    st.caption(f"审核人：{_auditor}")

            # --- 已审核凭证列表 ---
            with _audit_tab2:
                if not _audited_list:
                    st.info("暂无已审核凭证。")
                else:
                    for _vn in _audited_list:
                        _vdf = _audit_df[_audit_df["voucher_number"] == _vn]
                        _vdate = _vdf.iloc[0]["voucher_date"]
                        _vsummary = _vdf.iloc[0]["summary"]
                        _auditor_name = _vdf.iloc[0].get("auditor", "") or ""
                        _audit_dt = _vdf.iloc[0].get("audit_date", "") or ""
                        _d_total = _vdf["debit_amount"].sum()
                        _c_total = _vdf["credit_amount"].sum()
                        _balanced = "✅" if abs(_d_total - _c_total) < 0.01 else "❌"

                        with st.expander(f"🟢 {_vn} | {_vdate} | {_vsummary} | 借 {_d_total:,.2f} = 贷 {_c_total:,.2f} {_balanced} | 审核人：{_auditor_name} {_audit_dt}"):
                            _disp_df = _vdf[["account_name", "debit_amount", "credit_amount"]].copy()
                            _disp_df.columns = ["科目", "借方金额", "贷方金额"]
                            _disp_df.loc[len(_disp_df)] = ["合计", _d_total, _c_total]
                            _disp_df = fmt_money_df(_disp_df, ["借方金额", "贷方金额"])
                            st.dataframe(_disp_df, use_container_width=True, hide_index=True)

                            st.markdown(f"**审核人**：{_auditor_name}  ｜  **审核时间**：{_audit_dt}")

                            _uc1, _uc2 = st.columns([1, 2])
                            with _uc1:
                                if st.button("🔄 反审核", key=f"unaudit_{_vn}",
                                             type="secondary", use_container_width=True):
                                    _ok, _msg = unaudit_voucher(_vn)
                                    if _ok:
                                        st.success(_msg)
                                        st.rerun()
                                    else:
                                        st.error(_msg)
                            with _uc2:
                                st.caption("反审核后凭证恢复为未审核状态，可修改/删除")

            # --- 批量审核 ---
            with _audit_tab3:
                if not _pending_list:
                    st.success("🎉 没有待审核的凭证！")
                else:
                    st.markdown("#### ⚡ 批量审核")
                    st.caption(f"当前待审核凭证 {len(_pending_list)} 张")

                    # 多选凭证
                    _batch_options = []
                    for _vn in _pending_list:
                        _vdf = _audit_df[_audit_df["voucher_number"] == _vn]
                        _vdate = _vdf.iloc[0]["voucher_date"]
                        _vsummary = _vdf.iloc[0]["summary"]
                        _batch_options.append(f"{_vn} | {_vdate} | {_vsummary}")

                    _selected_batch = st.multiselect(
                        "选择要审核的凭证（可多选）",
                        _batch_options,
                        key="batch_audit_select",
                    )

                    if _selected_batch:
                        st.info(f"已选择 {len(_selected_batch)} 张凭证待审核")
                        if st.button("✅ 批量审核选中凭证", type="primary",
                                     key="batch_audit_btn", use_container_width=True):
                            _batch_nums = []
                            for _sel in _selected_batch:
                                _vn = _sel.split(" | ")[0] if " | " in _sel else _sel
                                _batch_nums.append(_vn)
                            _ok, _msg = batch_audit_vouchers(_batch_nums, _auditor)
                            st.success(_msg)
                            st.rerun()

                    # 全选快捷按钮
                    st.markdown("---")
                    _bc1, _bc2 = st.columns(2)
                    with _bc1:
                        if st.button("📋 全选待审核", key="batch_select_all"):
                            st.session_state["batch_audit_select"] = _batch_options
                            st.rerun()
                    with _bc2:
                        if st.button("🧹 清空选择", key="batch_clear_all"):
                            st.session_state["batch_audit_select"] = []
                            st.rerun()

                    # 一键审核全部
                    st.markdown("---")
                    if st.button("🔥 一键审核全部待审核凭证", type="primary",
                                 key="audit_all_btn", use_container_width=True):
                        _ok, _msg = batch_audit_vouchers(_pending_list, _auditor)
                        st.success(_msg)
                        st.rerun()

    # --- 明细账 ---
    with sub4:
        st.subheader("科目明细账")
        st.caption("每个科目就像一个「碗」，凭证就是「瓢」，碗里水位随每张凭证的借贷变动而变化。选择一个科目，查看它的全部流水。")

        # 科目选择
        selected_account = st.selectbox(
            "选择科目",
            options=get_account_names_dynamic(),
            index=0,
            key="ledger_account_select",
        )

        if selected_account:
            # 解析科目编码
            sel_code = selected_account.split(" ")[0]
            ledger = get_account_ledger(sel_code)

            if not ledger:
                st.warning("未找到该科目信息")
            else:
                acc_name = f"{ledger['account_code']} {ledger['account_name']}"
                direction_text = "借方" if ledger["direction"] == "借" else "贷方"
                st.markdown(f"### {acc_name}（{direction_text}科目）")

                # 期初余额
                opening = ledger["opening_balance"]
                opening_dir = "借" if opening >= 0 else "贷"
                st.info(
                    f"期初余额：**{abs(opening):,.2f}** 元（{opening_dir}方）"
                )

                # 构建明细表
                if not ledger["entries"]:
                    st.warning("该科目暂无凭证记录。")
                else:
                    rows = []
                    for e in ledger["entries"]:
                        rows.append({
                            "日期": e["date"],
                            "凭证编号": e["voucher_number"],
                            "摘要": e["summary"],
                            "借方": f"{e['debit']:,.2f}" if e["debit"] else "",
                            "贷方": f"{e['credit']:,.2f}" if e["credit"] else "",
                            "方向": e["direction_label"],
                            "余额": f"{abs(e['balance']):,.2f}",
                        })

                    ledger_df = pd.DataFrame(rows)
                    st.dataframe(ledger_df, use_container_width=True, hide_index=True)

                    # 本期合计
                    td = ledger["total_debit"]
                    tc = ledger["total_credit"]
                    ending = ledger["ending_balance"]
                    ending_dir = "借" if ending >= 0 else "贷"

                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.metric("本期借方合计", f"{td:,.2f} 元")
                    with col2:
                        st.metric("本期贷方合计", f"{tc:,.2f} 元")
                    with col3:
                        st.metric("期末余额", f"{abs(ending):,.2f} 元（{ending_dir}方）")

                    # --- 导出 Excel ---
                    def export_ledger_excel(ledger_data):
                        """导出明细账为格式化 Excel"""
                        from openpyxl import Workbook
                        from openpyxl.styles import Font, Alignment, Border, Side
                        import io

                        wb = Workbook()
                        ws = wb.active
                        ws.title = ledger_data["account_name"][:31]

                        _thin = Side(style='thin')
                        _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
                        _title_font = Font(name='Arial', size=14, bold=True)
                        _hdr_font = Font(name='Arial', size=10, bold=True)
                        _data_font = Font(name='Arial', size=10)
                        _center = Alignment(horizontal='center', vertical='center')
                        _left = Alignment(horizontal='left', vertical='center')
                        _right = Alignment(horizontal='right', vertical='center')

                        for col, w in [('A', 12), ('B', 15), ('C', 25),
                                       ('D', 14), ('E', 14), ('F', 6), ('G', 16)]:
                            ws.column_dimensions[col].width = w

                        # 标题
                        ws.merge_cells('A1:G1')
                        ws['A1'] = f"{ledger_data['account_code']} {ledger_data['account_name']} 明细账"
                        ws['A1'].font = _title_font
                        ws['A1'].alignment = _center
                        ws.row_dimensions[1].height = 28

                        # 期初余额行
                        op = ledger_data["opening_balance"]
                        op_dir = "借" if op >= 0 else "贷"
                        ws.merge_cells('A2:G2')
                        ws['A2'] = f"期初余额：{abs(op):,.2f} 元（{op_dir}方）"
                        ws['A2'].font = _data_font
                        ws['A2'].alignment = _left
                        ws.row_dimensions[2].height = 20

                        # 表头
                        hdrs = ['日期', '凭证编号', '摘要', '借方金额', '贷方金额', '方向', '余额']
                        for i, h in enumerate(hdrs, 1):
                            c = ws.cell(row=3, column=i, value=h)
                            c.font = _hdr_font
                            c.alignment = _center
                            c.border = _border
                        ws.row_dimensions[3].height = 20

                        # 数据行
                        r = 4
                        for e in ledger_data["entries"]:
                            ws.row_dimensions[r].height = 20
                            vals = [
                                e["date"],
                                e["voucher_number"],
                                e["summary"],
                                f"{e['debit']:,.2f}" if e["debit"] else "",
                                f"{e['credit']:,.2f}" if e["credit"] else "",
                                e["direction_label"],
                                f"{abs(e['balance']):,.2f}",
                            ]
                            for ci, val in enumerate(vals, 1):
                                c = ws.cell(row=r, column=ci, value=val if val else None)
                                c.font = _data_font
                                c.border = _border
                                c.number_format = '@'
                                if ci == 3:
                                    c.alignment = _left
                                elif ci in (1, 2, 6):
                                    c.alignment = _center
                                else:
                                    c.alignment = _right
                            r += 1

                        # 合计行
                        ws.row_dimensions[r].height = 22
                        end_dir = "借" if ledger_data["ending_balance"] >= 0 else "贷"
                        totals = [
                            '本期合计及余额',
                            '',
                            '',
                            f"{ledger_data['total_debit']:,.2f}",
                            f"{ledger_data['total_credit']:,.2f}",
                            end_dir,
                            f"{abs(ledger_data['ending_balance']):,.2f}",
                        ]
                        for ci, val in enumerate(totals, 1):
                            c = ws.cell(row=r, column=ci, value=val if val else None)
                            c.font = _hdr_font
                            c.border = _border
                            c.number_format = '@'
                            if ci == 3:
                                c.alignment = _left
                            elif ci in (1, 2, 6):
                                c.alignment = _center
                            else:
                                c.alignment = _right

                        output = io.BytesIO()
                        wb.save(output)
                        output.seek(0)
                        return output

                    st.download_button(
                        label="📥 下载明细账（Excel）",
                        data=export_ledger_excel(ledger),
                        file_name=f"明细账_{ledger['account_code']}_{ledger['account_name']}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )

    # --- 科目管理 ---
    with sub5:
        st.subheader("科目管理")
        st.caption("自定义二级科目，例如：银行存款-工商银行、销售费用-业务招待费等。二级科目的编码自动在一级科目后追加两位。")

        # --- 添加二级科目 ---
        st.markdown("#### 添加二级科目")
        cols_add = st.columns([3, 3, 2, 2])

        with cols_add[0]:
            parent_options = [f"{a['code']} {a['name']}" for a in ACCOUNT_CHART]
            selected_parent = st.selectbox(
                "选择一级科目", parent_options,
                key="custom_acc_parent", label_visibility="collapsed",
                placeholder="搜索一级科目...",
            )

        parent_code = selected_parent.split(" ")[0]
        parent_name = selected_parent.split(" ", 1)[1]
        parent_info = ACCOUNT_MAP.get(parent_code, {})
        parent_direction = parent_info.get("direction", "借")
        parent_category = parent_info.get("category", "")
        next_sc = get_next_sub_code(parent_code)
        preview_full_code = f"{parent_code}{next_sc}"

        with cols_add[1]:
            sub_name_input = st.text_input(
                "二级科目名称", key="custom_acc_subname",
                label_visibility="collapsed", placeholder="如：工商银行、业务招待费",
            )

        with cols_add[2]:
            tax_options = list(TAX_SPECIAL_FLAGS.keys())
            tax_labels = [TAX_SPECIAL_FLAGS[k] for k in tax_options]
            selected_tax_idx = st.selectbox(
                "税务标记", range(len(tax_options)),
                format_func=lambda i: tax_labels[i],
                key="custom_acc_tax", label_visibility="collapsed",
            )
            selected_tax_flag = tax_options[selected_tax_idx]

        with cols_add[3]:
            add_clicked = st.button("➕ 添加", key="custom_acc_add_btn", use_container_width=True)

        # 预览信息
        color_dir = "#1976d2" if parent_direction == "借" else "#e53935"
        st.caption(
            f"预览：编码 **{preview_full_code}** ｜ "
            f"名称：{parent_name}-{sub_name_input or '（待输入）'} ｜ "
            f"方向：<span style='color:{color_dir};font-weight:600'>{parent_direction}方</span> ｜ "
            f"分类：{parent_category}"
            + (f" ｜ <span style='color:#e65100;font-weight:600'>⚠️ {selected_tax_flag}</span>" if selected_tax_flag else ""),
            unsafe_allow_html=True,
        )

        if add_clicked:
            if sub_name_input.strip():
                full_code, full_name = add_custom_account(
                    parent_code, parent_name, sub_name_input.strip(),
                    parent_category, parent_direction, selected_tax_flag,
                )
                st.success(f"✅ 已添加：{full_code} {full_name}")
                st.rerun()
            else:
                st.warning("请输入二级科目名称")

        st.markdown("---")

        # --- 已添加的二级科目列表 ---
        st.markdown("#### 已有二级科目")
        df_custom = get_custom_accounts()

        if df_custom.empty:
            st.info("💡 尚未添加任何自定义二级科目。在上方选择一级科目并输入名称后点击「添加」。")
        else:
            # 按一级科目分组显示
            for p_code in df_custom["parent_code"].unique():
                sub_df = df_custom[df_custom["parent_code"] == p_code].sort_values("sub_code")
                p_name = sub_df.iloc[0]["parent_name"]
                direction = sub_df.iloc[0]["direction"]
                color = "#1976d2" if direction == "借" else "#e53935"

                st.markdown(
                    f'<div style="font-weight:600; font-size:14px; margin-top:12px; margin-bottom:4px;">'
                    f'{p_code} {p_name}'
                    f' <span style="color:{color}; font-size:12px;">（{direction}方）</span>'
                    f'</div>',
                    unsafe_allow_html=True,
                )

                for _, row in sub_df.iterrows():
                    cols_row = st.columns([1, 2, 2, 1])
                    with cols_row[0]:
                        st.code(row["full_code"])
                    with cols_row[1]:
                        st.write(row["full_name"])
                    with cols_row[2]:
                        if row.get("tax_flag"):
                            st.markdown(
                                f'<span style="color:#e65100; font-size:12px;">⚠️ {row["tax_flag"]}</span>',
                                unsafe_allow_html=True,
                            )
                        else:
                            st.write("—")
                    with cols_row[3]:
                        if st.button("🗑", key=f"del_custom_{row['full_code']}",
                                     help=f"删除 {row['full_name']}"):
                            delete_custom_account(row["full_code"])
                            st.rerun()

        # --- 税务说明 ---
        st.markdown("---")
        st.markdown("#### 税务扣除规则说明")
        for flag, desc in TAX_SPECIAL_FLAGS.items():
            if flag:
                st.markdown(f"- **{flag}**：{desc}")

    # --- 期末结转 ---
    with sub6:
        st.subheader("🔄 期末自动结转")
        st.caption(
            "月末将所有损益类科目（收入、费用）余额结转至「本年利润」，结转后损益类科目余额为零。\n\n"
            "年末额外将「本年利润」余额转入「利润分配-未分配利润」，并可提取法定盈余公积。"
        )

        # 结转日期 & 类型选择
        cf_cols = st.columns([2, 2, 2])
        with cf_cols[0]:
            cf_date = st.date_input(
                "结转日期",
                value=datetime.now(),
                help="月末结转通常选当月最后一天，年末结转选12月31日",
            )
        with cf_cols[1]:
            cf_type = st.radio(
                "结转类型",
                ["月末结转（损益→本年利润）", "年末结转（含利润分配）"],
                label_visibility="collapsed",
            )
        with cf_cols[2]:
            is_year_end = "年末" in cf_type
            extract_reserve = False
            reserve_rate = 0.10
            if is_year_end:
                extract_reserve = st.checkbox(
                    "提取法定盈余公积（10%）",
                    value=True,
                    help="按净利润的 10% 提取，累计达注册资本 50% 后可不再提取",
                )
                if extract_reserve:
                    reserve_rate = st.number_input(
                        "计提比例", min_value=0.0, max_value=1.0,
                        value=0.10, step=0.01, format="%.2f",
                    )

        st.markdown("---")

        # 预览按钮
        if st.button("🔍 生成结转预览", type="secondary", use_container_width=True):
            st.session_state["_cf_preview_data"] = get_carryforward_preview()

        # 显示预览数据
        preview_data = st.session_state.get("_cf_preview_data")
        if preview_data:
            st.markdown("### 📋 结转预览")

            # 收入类科目
            if preview_data["income_items"]:
                st.markdown("#### 一、收入类科目（结转至本年利润·贷方）")
                inc_rows = []
                for item in preview_data["income_items"]:
                    inc_rows.append({
                        "科目编码": item["code"],
                        "科目名称": item["name"],
                        "期末余额": fmt_money(abs(item["balance"])),
                        "余额方向": item["direction"],
                        "结转方向": "借方" if item["balance"] > 0 else "贷方",
                        "结转金额": fmt_money(abs(item["balance"])),
                    })
                st.dataframe(pd.DataFrame(inc_rows), use_container_width=True, hide_index=True)
            else:
                st.info("收入类科目余额均为零，无需结转。")

            # 费用类科目
            if preview_data["expense_items"]:
                st.markdown("#### 二、费用类科目（结转至本年利润·借方）")
                exp_rows = []
                for item in preview_data["expense_items"]:
                    exp_rows.append({
                        "科目编码": item["code"],
                        "科目名称": item["name"],
                        "期末余额": fmt_money(abs(item["balance"])),
                        "余额方向": item["direction"],
                        "结转方向": "贷方" if item["balance"] > 0 else "借方",
                        "结转金额": fmt_money(abs(item["balance"])),
                    })
                st.dataframe(pd.DataFrame(exp_rows), use_container_width=True, hide_index=True)
            else:
                st.info("费用类科目余额均为零，无需结转。")

            # 汇总
            st.markdown("---")
            sum_cols = st.columns(3)
            with sum_cols[0]:
                st.metric("收入合计", fmt_money(preview_data["total_income"]))
            with sum_cols[1]:
                st.metric("费用合计", fmt_money(preview_data["total_expense"]))
            with sum_cols[2]:
                net = preview_data["net_profit"]
                st.metric(
                    "本期净利润（估）",
                    fmt_money(abs(net)),
                    delta=f"{'盈利' if net > 0 else '亏损'}" if net != 0 else "持平",
                )

            # 本年利润当前余额
            pb = preview_data["profit_balance"]
            if abs(pb) > 0.01:
                st.info(
                    f"📌 「本年利润（4103）」当前余额：{fmt_money(abs(pb))} 元"
                    f"（{'贷方' if pb > 0 else '借方'}）"
                )

            # 年末结转额外信息
            if is_year_end and net > 0 and extract_reserve:
                reserve_amt = round(net * reserve_rate, 2)
                st.warning(
                    f"📦 年末将额外执行：\n"
                    f"- 将本年利润余额转入「利润分配-未分配利润」\n"
                    f"- 提取法定盈余公积：{fmt_money(reserve_amt)} 元"
                    f"（净利润 {fmt_money(net)} × {reserve_rate:.0%}）"
                )

            # 执行按钮
            st.markdown("---")
            btn_label = "年末结转（含利润分配）" if is_year_end else "执行月末结转"
            if st.button(f"✅ 确认执行{btn_label}", type="primary", use_container_width=True):
                with st.spinner("正在执行结转，请稍候..."):
                    if is_year_end:
                        result = execute_yearly_carryforward(
                            cf_date,
                            extract_legal_reserve=extract_reserve,
                            reserve_rate=reserve_rate,
                        )
                        v_nums = result["voucher_numbers"]
                        net_profit = result["net_profit"]
                        reserve = result["reserve_amount"]
                    else:
                        v_nums = execute_monthly_carryforward(cf_date)
                        net_profit = preview_data["net_profit"]
                        reserve = 0

                    st.session_state["_cf_result"] = {
                        "vouchers": v_nums,
                        "net_profit": net_profit,
                        "reserve": reserve,
                        "is_year_end": is_year_end,
                    }
                    # 清除预览
                    st.session_state.pop("_cf_preview_data", None)
                    st.rerun()

        # 显示执行结果
        cf_result = st.session_state.get("_cf_result")
        if cf_result:
            st.success("✅ 期末结转执行成功！")
            st.markdown("### 📝 生成的结转凭证")

            for vn in cf_result["vouchers"]:
                v_df = get_all_vouchers()
                v_rows = v_df[v_df["voucher_number"] == vn]
                if not v_rows.empty:
                    st.markdown(f"**凭证编号：{vn}** — {v_rows.iloc[0]['summary']}")
                    display_df = v_rows[[
                        "account_code", "account_name", "debit_amount", "credit_amount"
                    ]].copy()
                    display_df.columns = ["科目编码", "科目名称", "借方金额", "贷方金额"]
                    display_df["借方金额"] = display_df["借方金额"].apply(fmt_money)
                    display_df["贷方金额"] = display_df["贷方金额"].apply(fmt_money)
                    st.dataframe(display_df, use_container_width=True, hide_index=True)

            # 净利润汇总
            net = cf_result["net_profit"]
            st.markdown("---")
            res_cols = st.columns(2)
            with res_cols[0]:
                st.metric("结转后净利润", fmt_money(abs(net)),
                          delta=f"{'盈利' if net > 0 else '亏损'}")
            with res_cols[1]:
                if cf_result["is_year_end"] and cf_result["reserve"] > 0:
                    st.metric("提取法定盈余公积", fmt_money(cf_result["reserve"]))
                else:
                    st.metric("损益类科目余额", "0.00 元", delta="已全部清零")

            st.info("💡 结转后可在「凭证查询」中查看结转凭证，在「报表」中查看更新后的利润表和资产负债表。")

            if st.button("清除结果"):
                st.session_state.pop("_cf_result", None)
                st.rerun()
        elif not preview_data:
            st.markdown("---")
            st.info("👆 点击「生成结转预览」查看需要结转的科目及金额。")

            with st.expander("📖 期末结转说明"):
                st.markdown("""
#### 月末结转流程
1. **结转收入**：将主营业务收入、其他业务收入、投资收益等收入类科目余额结转至「本年利润」
2. **结转费用**：将主营业务成本、销售费用、管理费用、财务费用等费用类科目余额结转至「本年利润」
3. 结转后所有损益类科目余额为零，「本年利润」反映当期净利润/亏损

#### 年末结转流程（在月末结转基础上）
4. **结转本年利润**：将「本年利润」全年累计余额转入「利润分配-未分配利润」
5. **提取法定盈余公积**：按净利润的 10% 提取（累计达注册资本 50% 后可不再提取）

#### 注意事项
- 结转会自动生成正式凭证，可在「凭证查询」中查看
- 月末结转每月执行一次，年末结转仅在 12 月末执行
- 执行前请确保当月所有日常凭证已录入完毕
""")


st.divider()


# ============================================================
# 【💰 财务会计】模块二：报表（财务报表）
# ============================================================
with tab2:
    st.header("📊 财务报表")
    st.caption("资产负债表 · 利润表 · 现金流量表 · 财务可视化分析 ｜ 自动生成，标准格式")

    # --- 公司信息输入 ---
    col_c1, col_c2, col_c3 = st.columns(3)
    with col_c1:
        company_name = st.text_input("编制单位", value=st.session_state.get("current_company_name", ""), placeholder="填写公司名称")
    with col_c2:
        report_date = st.date_input("报表出具日期", value=None)
    with col_c3:
        currency_unit = st.text_input("货币单位", value="元")

    report_date_str = report_date.strftime("%Y-%m-%d") if report_date else ""

    # --- 报表选择 ---
    report_type = st.radio("选择报表", ["资产负债表", "利润表", "现金流量表", "📊 财务可视化分析"], horizontal=True)

    # ================================================================
    # 资产负债表
    # ================================================================
    if report_type == "资产负债表":
        st.markdown("---")

        # --- 第1行：标题 ---
        st.markdown(
            '<div style="text-align:center; font-size:22px; font-weight:700; '
            'padding:12px 0;">资产负债表</div>',
            unsafe_allow_html=True,
        )

        # --- 第2行：公司名称 / 报表日期 / 货币单位 ---
        st.markdown(
            f'<div style="display:flex; justify-content:space-between; '
            f'font-size:13px; color:#555; padding:4px 12px 8px 12px;">'
            f'<span>编制单位：{company_name or "（未填写）"}</span>'
            f'<span>报表日期：{report_date_str or "（未选择）"}</span>'
            f'<span>货币单位：{currency_unit}</span>'
            f'</div>',
            unsafe_allow_html=True,
        )

        # --- 表格主体 ---
        # 获取所有科目余额
        asset_accounts = [a for a in ACCOUNT_CHART if a["category"] == "资产"]
        liab_accounts = [a for a in ACCOUNT_CHART if a["category"] == "负债"]
        equity_accounts = [a for a in ACCOUNT_CHART if a["category"] == "权益"]

        # --- 标准会计格式：左方资产，右方负债+所有者权益，底部合计相等 ---

        def fmt_amt(v):
            """格式化金额：0 显示空"""
            if abs(v) < 0.005:
                return ""
            return f"{v:,.2f}"

        def _build_bs_rows(accounts):
            """构建资产负债表数据行"""
            rows = []
            for idx, acc in enumerate(accounts, 1):
                od, oc, pd, pc, ending = calc_account_balance(acc["code"])
                opening = od if acc["direction"] == "借" else oc
                rows.append({
                    "code": acc["code"],
                    "name": acc["name"],
                    "ending": ending,
                    "opening": opening,
                })
            return rows

        asset_rows = _build_bs_rows(asset_accounts)
        liab_rows = _build_bs_rows(liab_accounts)
        equity_rows = _build_bs_rows(equity_accounts)

        asset_ending_total = sum(r["ending"] for r in asset_rows)
        asset_opening_total = sum(r["opening"] for r in asset_rows)
        liab_ending_total = sum(r["ending"] for r in liab_rows)
        liab_opening_total = sum(r["opening"] for r in liab_rows)
        equity_ending_total = sum(r["ending"] for r in equity_rows)
        equity_opening_total = sum(r["opening"] for r in equity_rows)

        # 右方 = 负债 + 所有者权益
        right_ending_total = liab_ending_total + equity_ending_total
        right_opening_total = liab_opening_total + equity_opening_total

        # 构建左方行列表（资产 + 合计行）
        left_rows = []
        for r in asset_rows:
            left_rows.append({"code": r["code"], "name": r["name"],
                              "ending": r["ending"], "opening": r["opening"],
                              "is_total": False})
        left_rows.append({"code": "", "name": "资产总计",
                          "ending": asset_ending_total, "opening": asset_opening_total,
                          "is_total": True})

        # 构建右方行列表（负债 + 负债合计 + 权益 + 权益合计 + 总合计）
        right_rows = []
        for r in liab_rows:
            right_rows.append({"code": r["code"], "name": r["name"],
                               "ending": r["ending"], "opening": r["opening"],
                               "is_total": False, "is_section": False})
        right_rows.append({"code": "", "name": "负债合计",
                           "ending": liab_ending_total, "opening": liab_opening_total,
                           "is_total": False, "is_section": True})
        for r in equity_rows:
            right_rows.append({"code": r["code"], "name": r["name"],
                               "ending": r["ending"], "opening": r["opening"],
                               "is_total": False, "is_section": False})
        right_rows.append({"code": "", "name": "所有者权益合计",
                           "ending": equity_ending_total, "opening": equity_opening_total,
                           "is_total": False, "is_section": True})
        right_rows.append({"code": "", "name": "负债和所有者权益总计",
                           "ending": right_ending_total, "opening": right_opening_total,
                           "is_total": True, "is_section": False})

        # 补齐行数（左右对齐）
        max_rows = max(len(left_rows), len(right_rows))
        while len(left_rows) < max_rows:
            left_rows.append({"code": "", "name": "", "ending": 0, "opening": 0, "is_total": False})
        while len(right_rows) < max_rows:
            right_rows.append({"code": "", "name": "", "ending": 0, "opening": 0,
                               "is_total": False, "is_section": False})

        # --- 渲染 HTML 表格（标准左右对照格式） ---
        html_rows = ""
        for i in range(max_rows):
            lr = left_rows[i]
            rr = right_rows[i]

            # 左方行样式
            l_style = 'font-weight:700; border-top:2px solid #333;' if lr.get("is_total") else ''
            l_ending = fmt_amt(lr["ending"]) if lr["ending"] else ""
            l_opening = fmt_amt(lr["opening"]) if lr["opening"] else ""

            # 右方行样式
            r_style = ""
            if rr.get("is_total"):
                r_style = 'font-weight:700; border-top:2px solid #333;'
            elif rr.get("is_section"):
                r_style = 'font-weight:600; border-top:1px solid #999;'
            r_ending = fmt_amt(rr["ending"]) if rr["ending"] else ""
            r_opening = fmt_amt(rr["opening"]) if rr["opening"] else ""

            html_rows += f"""
            <tr>
                <td style="text-align:center; {l_style}">{lr['code']}</td>
                <td style="text-align:left; {l_style}">{lr['name']}</td>
                <td style="text-align:right; {l_style}">{l_ending}</td>
                <td style="text-align:right; {l_style}">{l_opening}</td>
                <td style="text-align:center; {r_style}">{rr['code']}</td>
                <td style="text-align:left; {r_style}">{rr['name']}</td>
                <td style="text-align:right; {r_style}">{r_ending}</td>
                <td style="text-align:right; {r_style}">{r_opening}</td>
            </tr>"""

        st.markdown(f"""
        <table style="width:100%; border-collapse:collapse; font-size:13px; margin-top:8px;">
            <thead>
                <tr style="background:#f0f4f8;">
                    <th colspan="4" style="text-align:center; padding:8px; border:1px solid #ccc; font-size:15px;">资 产</th>
                    <th colspan="4" style="text-align:center; padding:8px; border:1px solid #ccc; font-size:15px;">负债和所有者权益</th>
                </tr>
                <tr style="background:#e8eef5;">
                    <th style="padding:4px 6px; border:1px solid #ccc; width:6%;">编码</th>
                    <th style="padding:4px 6px; border:1px solid #ccc; text-align:left; width:20%;">科目名称</th>
                    <th style="padding:4px 6px; border:1px solid #ccc; text-align:right; width:12%;">期末余额</th>
                    <th style="padding:4px 6px; border:1px solid #ccc; text-align:right; width:12%;">年初余额</th>
                    <th style="padding:4px 6px; border:1px solid #ccc; width:6%;">编码</th>
                    <th style="padding:4px 6px; border:1px solid #ccc; text-align:left; width:20%;">科目名称</th>
                    <th style="padding:4px 6px; border:1px solid #ccc; text-align:right; width:12%;">期末余额</th>
                    <th style="padding:4px 6px; border:1px solid #ccc; text-align:right; width:12%;">年初余额</th>
                </tr>
            </thead>
            <tbody>
                {html_rows}
            </tbody>
        </table>
        """, unsafe_allow_html=True)

        # --- 校验：资产 = 负债 + 所有者权益 ---
        diff_ending = asset_ending_total - right_ending_total
        diff_opening = asset_opening_total - right_opening_total

        col_chk1, col_chk2 = st.columns(2)
        with col_chk1:
            if abs(diff_ending) < 0.01:
                st.success(f"✅ 期末平衡：资产 {fmt_amt(asset_ending_total)} = 负债+权益 {fmt_amt(right_ending_total)}")
            else:
                st.error(f"❌ 期末不平衡！差额：{fmt_amt(diff_ending)}")
        with col_chk2:
            if abs(diff_opening) < 0.01:
                st.success(f"✅ 年初平衡：资产 {fmt_amt(asset_opening_total)} = 负债+权益 {fmt_amt(right_opening_total)}")
            else:
                st.error(f"❌ 年初不平衡！差额：{fmt_amt(diff_opening)}")

        # --- 导出 Excel（左右对照格式） ---
        def export_bs_excel():
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side
            import io

            wb = Workbook()
            ws = wb.active
            ws.title = "资产负债表"

            _thin = Side(style='thin')
            _medium = Side(style='medium')
            _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
            _border_top = Border(left=_thin, right=_thin, top=_medium, bottom=_thin)
            _center = Alignment(horizontal='center', vertical='center')
            _right = Alignment(horizontal='right', vertical='center')
            _left = Alignment(horizontal='left', vertical='center')
            _bold = Font(bold=True)
            _title_font = Font(bold=True, size=16)
            _info_font = Font(size=10)
            _header_font = Font(bold=True, size=10)

            # 列宽
            widths = {'A': 8, 'B': 22, 'C': 16, 'D': 16, 'E': 8, 'F': 22, 'G': 16, 'H': 16}
            for col, w in widths.items():
                ws.column_dimensions[col].width = w

            # 第1行：标题
            ws.merge_cells('A1:H1')
            ws['A1'] = '资产负债表'
            ws['A1'].font = _title_font
            ws['A1'].alignment = _center
            ws.row_dimensions[1].height = 30

            # 第2行：公司信息
            ws['A2'] = f'编制单位：{company_name or ""}'
            ws['A2'].font = _info_font
            ws.merge_cells('C2:E2')
            ws['C2'] = f'报表日期：{report_date_str}'
            ws['C2'].font = _info_font
            ws['C2'].alignment = _center
            ws.merge_cells('F2:H2')
            ws['F2'] = f'货币单位：{currency_unit}'
            ws['F2'].font = _info_font
            ws['F2'].alignment = _right
            ws.row_dimensions[2].height = 18

            # 第3行：大类标题（资产 | 负债和所有者权益）
            ws.merge_cells('A3:D3')
            ws['A3'] = '资  产'
            ws['A3'].font = _bold
            ws['A3'].alignment = _center
            ws['A3'].border = _border
            ws.merge_cells('E3:H3')
            ws['E3'] = '负债和所有者权益'
            ws['E3'].font = _bold
            ws['E3'].alignment = _center
            ws['E3'].border = _border
            ws.row_dimensions[3].height = 22

            # 第4行：列标题
            headers = ["编码", "科目名称", "期末余额", "年初余额",
                       "编码", "科目名称", "期末余额", "年初余额"]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col, value=h)
                cell.font = _header_font
                cell.alignment = _center
                cell.border = _border
            ws.row_dimensions[4].height = 20

            # 数据行
            row_idx = 5
            for i in range(max_rows):
                lr = left_rows[i]
                rr = right_rows[i]

                # 左方
                l_vals = [lr["code"], lr["name"],
                          fmt_amt(lr["ending"]) if lr["ending"] else "",
                          fmt_amt(lr["opening"]) if lr["opening"] else ""]
                l_is_total = lr.get("is_total", False)
                for col, val in enumerate(l_vals, 1):
                    cell = ws.cell(row=row_idx, column=col, value=val if val else None)
                    cell.border = _border_top if l_is_total else _border
                    cell.font = _bold if l_is_total else Font(size=10)
                    if col == 1:
                        cell.alignment = _center
                    elif col == 2:
                        cell.alignment = _left
                    else:
                        cell.alignment = _right

                # 右方
                r_vals = [rr["code"], rr["name"],
                          fmt_amt(rr["ending"]) if rr["ending"] else "",
                          fmt_amt(rr["opening"]) if rr["opening"] else ""]
                r_is_total = rr.get("is_total", False)
                r_is_section = rr.get("is_section", False)
                for col, val in enumerate(r_vals, 5):
                    cell = ws.cell(row=row_idx, column=col, value=val if val else None)
                    cell.border = _border_top if (r_is_total or r_is_section) else _border
                    if r_is_total or r_is_section:
                        cell.font = _bold if r_is_total else Font(size=10, bold=True)
                    else:
                        cell.font = Font(size=10)
                    if col == 5:
                        cell.alignment = _center
                    elif col == 6:
                        cell.alignment = _left
                    else:
                        cell.alignment = _right

                ws.row_dimensions[row_idx].height = 18
                row_idx += 1

            # 平衡校验行
            row_idx += 1
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=4)
            if abs(diff_ending) < 0.01:
                ws.cell(row=row_idx, column=1, value=f"✅ 期末平衡：资产 = 负债+权益 = {fmt_amt(asset_ending_total)}")
            else:
                ws.cell(row=row_idx, column=1, value=f"❌ 期末不平衡！差额：{fmt_amt(diff_ending)}")
            ws.cell(row=row_idx, column=1).font = Font(bold=True, size=10, color="2e7d32" if abs(diff_ending) < 0.01 else "c62828")

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return output

        st.download_button(
            label="📥 下载资产负债表（Excel）",
            data=export_bs_excel(),
            file_name=f"资产负债表_{report_date_str or '未定'}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    # ================================================================
    # 利润表
    # ================================================================
    elif report_type == "利润表":
        st.markdown("---")

        # --- 第1行：标题 ---
        st.markdown(
            '<div style="text-align:center; font-size:22px; font-weight:700; '
            'padding:12px 0;">利润表</div>',
            unsafe_allow_html=True,
        )

        # --- 第2行：公司名称 / 报表日期 / 货币单位 ---
        st.markdown(
            f'<div style="display:flex; justify-content:space-between; '
            f'font-size:13px; color:#555; padding:4px 12px 8px 12px;">'
            f'<span>编制单位：{company_name or "（未填写）"}</span>'
            f'<span>报表日期：{report_date_str or "（未选择）"}</span>'
            f'<span>货币单位：{currency_unit}</span>'
            f'</div>',
            unsafe_allow_html=True,
        )

        # --- 金额计算辅助 ---
        def fmt_amt(v):
            """格式化金额：0 显示空"""
            if abs(v) < 0.005:
                return ""
            return f"{v:,.2f}"

        def get_ending(code):
            _, _, _, _, ending = calc_account_balance(code)
            return ending

        def get_ending_list(codes):
            return sum(get_ending(c) for c in codes)

        # --- 利润表行项目定义 ---
        # is_total: 需要计算的合计行；codes: 直接取科目余额；indent: 缩进层级
        IS_ITEMS = [
            # 序号, 项目, 科目代码列表, 是否合计行, 缩进
            {"no": 1,  "name": "营业收入",                          "codes": ["6001", "6051"], "indent": 0},
            {"no": 2,  "name": "营业成本",                          "codes": ["6401", "6402"], "indent": 0},
            {"no": 3,  "name": "税金及附加",                        "codes": ["6403"],         "indent": 0},
            {"no": "", "name": "    营业税",                        "codes": [],              "indent": 1},
            {"no": "", "name": "    城市维护建设税",                "codes": [],              "indent": 1},
            {"no": "", "name": "    资源税",                        "codes": [],              "indent": 1},
            {"no": "", "name": "    土地增值税",                    "codes": [],              "indent": 1},
            {"no": "", "name": "    城镇土地使用税、房产税、车船税、印花税", "codes": [],     "indent": 1},
            {"no": "", "name": "    教育费附加、矿产资源补偿费、排污费",   "codes": [],     "indent": 1},
            {"no": 4,  "name": "销售费用",                          "codes": ["6601"],        "indent": 0},
            {"no": "", "name": "    广告费和业务宣传费",            "codes": [],              "indent": 1},
            {"no": 5,  "name": "管理费用",                          "codes": ["6602"],        "indent": 0},
            {"no": "", "name": "    业务招待费",                    "codes": [],              "indent": 1},
            {"no": "", "name": "    研究费用",                      "codes": ["6405"],        "indent": 1},
            {"no": 6,  "name": "财务费用",                          "codes": ["6603"],        "indent": 0},
            {"no": 7,  "name": "投资收益",                          "codes": ["6111"],        "indent": 0},
            {"no": 8,  "name": "营业利润",                          "is_calc": "operating_profit", "indent": 0},
            {"no": 9,  "name": "营业外收入",                       "codes": ["6301"],        "indent": 0},
            {"no": 10, "name": "营业外支出",                        "codes": ["6701"],        "indent": 0},
            {"no": "", "name": "    无法收回的长期债券投资损失",    "codes": [],              "indent": 1},
            {"no": "", "name": "    无法收回的长期股权投资损失",    "codes": [],              "indent": 1},
            {"no": "", "name": "    自然灾害等不可抗力因素造成的损失", "codes": [],            "indent": 1},
            {"no": "", "name": "    税收滞纳金",                    "codes": [],              "indent": 1},
            {"no": 11, "name": "利润总额",                          "is_calc": "total_profit", "indent": 0},
            {"no": 12, "name": "所得税费用",                        "codes": ["6711"],        "indent": 0},
            {"no": 13, "name": "净利润",                            "is_calc": "net_profit",   "indent": 0},
        ]

        # --- 计算各项金额 ---
        # 收入类科目余额在贷方，费用类科目余额在借方
        # 利润表金额 = 收入 - 费用（费用取负值或绝对值减去）
        amounts = {}
        # 营业收入
        amounts["营业收入"] = get_ending_list(["6001", "6051"])
        # 营业成本
        amounts["营业成本"] = get_ending_list(["6401", "6402"])
        # 税金及附加
        amounts["税金及附加"] = get_ending("6403")
        # 销售费用
        amounts["销售费用"] = get_ending("6601")
        # 管理费用
        amounts["管理费用"] = get_ending("6602")
        # 财务费用
        amounts["财务费用"] = get_ending("6603")
        # 投资收益
        amounts["投资收益"] = get_ending("6111")
        # 营业利润 = 营业收入 - 营业成本 - 税金及附加 - 销售费用 - 管理费用 - 财务费用 + 投资收益
        amounts["营业利润"] = (amounts["营业收入"] - amounts["营业成本"] - amounts["税金及附加"]
                              - amounts["销售费用"] - amounts["管理费用"] - amounts["财务费用"]
                              + amounts["投资收益"])
        # 营业外收入
        amounts["营业外收入"] = get_ending("6301")
        # 营业外支出
        amounts["营业外支出"] = get_ending("6701")
        # 利润总额 = 营业利润 + 营业外收入 - 营业外支出
        amounts["利润总额"] = amounts["营业利润"] + amounts["营业外收入"] - amounts["营业外支出"]
        # 所得税费用
        amounts["所得税费用"] = get_ending("6711")
        # 净利润 = 利润总额 - 所得税费用
        amounts["净利润"] = amounts["利润总额"] - amounts["所得税费用"]

        # --- 构建表格数据 ---
        rows = []
        for item in IS_ITEMS:
            name_display = item["name"]
            if item.get("is_calc"):
                val = amounts.get(item["is_calc"], 0)
            else:
                val = amounts.get(item["name"], 0) if item["codes"] else 0
            rows.append({
                "序号": item["no"],
                "项目": name_display,
                "本月金额": fmt_amt(val),
                "本年累积金额": fmt_amt(val),
            })

        is_df = pd.DataFrame(rows, columns=["序号", "项目", "本月金额", "本年累积金额"])
        st.dataframe(is_df, use_container_width=True, hide_index=True)

        # --- 关键指标高亮 ---
        col_p1, col_p2, col_p3 = st.columns(3)
        with col_p1:
            st.metric("营业利润", f"{amounts['营业利润']:,.2f}")
        with col_p2:
            st.metric("利润总额", f"{amounts['利润总额']:,.2f}")
        with col_p3:
            st.metric("净利润", f"{amounts['净利润']:,.2f}")

        # --- 导出 Excel ---
        def export_is_excel():
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side
            import io

            wb = Workbook()
            ws = wb.active
            ws.title = "利润表"

            _thin = Side(style='thin')
            _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
            _center = Alignment(horizontal='center', vertical='center')
            _right = Alignment(horizontal='right', vertical='center')
            _left = Alignment(horizontal='left', vertical='center')
            _bold = Font(bold=True)
            _title_font = Font(bold=True, size=16)
            _info_font = Font(size=11)

            # 第1行：标题
            ws.merge_cells('A1:D1')
            ws['A1'] = '利润表'
            ws['A1'].font = _title_font
            ws['A1'].alignment = _center
            ws.row_dimensions[1].height = 30

            # 第2行：公司名称 / 报表日期 / 货币单位
            ws['A2'] = f'编制单位：{company_name or ""}'
            ws['A2'].font = _info_font
            ws.merge_cells('B2:C2')
            ws['B2'] = f'报表日期：{report_date_str}'
            ws['B2'].font = _info_font
            ws['B2'].alignment = _center
            ws['D2'] = f'货币单位：{currency_unit}'
            ws['D2'].font = _info_font
            ws['D2'].alignment = _right
            ws.row_dimensions[2].height = 20

            # 第3行：表头
            headers = ["序号", "项目", "本月金额", "本年累积金额"]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col, value=h)
                cell.font = _bold
                cell.alignment = _center
                cell.border = _border
            ws.row_dimensions[3].height = 22

            # 数据行
            for ridx, row in enumerate(rows, 4):
                for col, key in enumerate(["序号", "项目", "本月金额", "本年累积金额"], 1):
                    val = row[key]
                    cell = ws.cell(row=ridx, column=col, value=val)
                    cell.border = _border
                    if col == 1:
                        cell.alignment = _center
                    elif col == 2:
                        cell.alignment = _left
                    else:
                        cell.alignment = _right
                    # 营业利润、利润总额、净利润 加粗
                    if row["项目"] in ("营业利润", "利润总额", "净利润"):
                        cell.font = _bold

            # 列宽
            ws.column_dimensions['A'].width = 6
            ws.column_dimensions['B'].width = 42
            ws.column_dimensions['C'].width = 18
            ws.column_dimensions['D'].width = 18

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return output

        st.download_button(
            label="📥 下载利润表（Excel）",
            data=export_is_excel(),
            file_name=f"利润表_{report_date_str or '未定'}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    # ================================================================
    # 现金流量表
    # ================================================================
    elif report_type == "现金流量表":
        st.markdown("---")

        # --- 第1行：标题 ---
        st.markdown(
            '<div style="text-align:center; font-size:22px; font-weight:700; '
            'padding:12px 0;">现金流量表</div>',
            unsafe_allow_html=True,
        )

        # --- 第2行：公司名称 / 报表日期 / 货币单位 ---
        st.markdown(
            f'<div style="display:flex; justify-content:space-between; '
            f'font-size:13px; color:#555; padding:4px 12px 8px 12px;">'
            f'<span>编制单位：{company_name or "（未填写）"}</span>'
            f'<span>报表日期：{report_date_str or "（未选择）"}</span>'
            f'<span>货币单位：{currency_unit}</span>'
            f'</div>',
            unsafe_allow_html=True,
        )

        # --- 金额计算辅助 ---
        def fmt_amt(v):
            """格式化金额：0 显示空"""
            if abs(v) < 0.005:
                return ""
            return f"{v:,.2f}"

        def get_ending(code):
            _, _, _, _, ending = calc_account_balance(code)
            return ending

        # --- 现金类科目编码 ---
        CASH_CODES = {"1001", "1002", "1012"}

        # 投资活动涉及的对方科目
        INVEST_CODES = {
            "1601", "1602", "1603", "1604", "1605", "1606",
            "1701", "1702", "1703", "1801",
            "1501", "1502", "1503", "1504",
            "1511", "1512", "1521", "1101",
            "1641", "1642", "1651", "1652", "1621", "1622",
        }

        # 筹资活动涉及的对方科目
        FINANCE_CODES = {
            "2001", "2501", "2502", "2701",
            "4001", "4002", "4101", "4104",
            "2231", "2232",
        }

        # --- 从凭证中分析现金流量 ---
        vdf_all = get_all_vouchers()

        # 初始化各项目金额
        # 经营活动
        op_sales_cash = 0          # 销售产成品、商品、提供劳务收到的现金
        op_other_in = 0            # 收到其他与经营活动有关的现金
        op_purchase_cash = 0       # 购买原材料、商品、接受劳务支付的现金
        op_salary_cash = 0         # 支付的职工薪酬
        op_tax_cash = 0            # 支付的税费
        op_other_out = 0           # 支付其他与经营活动有关的现金

        # 投资活动
        inv_recover = 0            # 收回短期投资、长期债券投资和长期股权投资收到的现金
        inv_income = 0             # 取得投资收益收到的现金
        inv_dispose = 0            # 处置固定资产、无形资产和其他非流动资产收回的现金净额
        inv_invest_pay = 0         # 短期投资、长期债券投资和长期股权投资支付的现金
        inv_buy_asset = 0          # 购建固定资产、无形资产和其他非流动资产支付的现金

        # 筹资活动
        fin_loan_in = 0            # 取得借款收到的现金
        fin_invest_in = 0          # 吸收投资者投资收到的现金
        fin_loan_repay = 0         # 偿还借款本金支付的现金
        fin_interest_pay = 0       # 偿还借款利息支付的现金
        fin_profit_dist = 0        # 分配利润支付的现金

        if not vdf_all.empty:
            # 找出所有涉及现金科目的凭证
            cash_voucher_nums = vdf_all[
                vdf_all["account_code"].isin(CASH_CODES)
            ]["voucher_number"].unique()

            for vnum in cash_voucher_nums:
                vdf = vdf_all[vdf_all["voucher_number"] == vnum]
                cash_rows = vdf[vdf["account_code"].isin(CASH_CODES)]
                non_cash_rows = vdf[~vdf["account_code"].isin(CASH_CODES)]

                for _, crow in cash_rows.iterrows():
                    cash_in = crow["debit_amount"] or 0   # 现金借方 = 收到现金
                    cash_out = crow["credit_amount"] or 0 # 现金贷方 = 支付现金

                    if len(non_cash_rows) > 0:
                        counterpart_code = non_cash_rows.iloc[0]["account_code"]
                    else:
                        counterpart_code = ""

                    # 经营活动
                    if counterpart_code in ("6001", "6051"):
                        op_sales_cash += cash_in
                    elif counterpart_code in ("6401", "6402", "1403", "1405", "1123"):
                        op_purchase_cash += cash_out
                    elif counterpart_code == "2211":
                        op_salary_cash += cash_out
                    elif counterpart_code in ("2221", "6403"):
                        op_tax_cash += cash_out
                    elif counterpart_code in ("6601", "6602"):
                        op_other_out += cash_out
                    elif counterpart_code == "6603":
                        fin_interest_pay += cash_out
                    elif counterpart_code in ("6301", "1221", "2202", "2203"):
                        op_other_in += cash_in
                        op_other_out += cash_out
                    # 投资活动
                    elif counterpart_code in ("1101", "1501", "1503", "1504", "1511"):
                        if cash_in > 0:
                            inv_recover += cash_in
                        if cash_out > 0:
                            inv_invest_pay += cash_out
                    elif counterpart_code == "6111":
                        inv_income += cash_in
                    elif counterpart_code in ("1601", "1602", "1701", "1702"):
                        if cash_out > 0:
                            inv_buy_asset += cash_out
                        if cash_in > 0:
                            inv_dispose += cash_in
                    # 筹资活动
                    elif counterpart_code in ("2001", "2501", "2502"):
                        if cash_in > 0:
                            fin_loan_in += cash_in
                        if cash_out > 0:
                            fin_loan_repay += cash_out
                    elif counterpart_code == "4001":
                        fin_invest_in += cash_in
                    elif counterpart_code == "4101":
                        fin_profit_dist += cash_out
                    elif counterpart_code in ("2231", "2232"):
                        if cash_out > 0:
                            fin_profit_dist += cash_out
                    else:
                        op_other_in += cash_in
                        op_other_out += cash_out

        # 净额计算
        op_net = op_sales_cash + op_other_in - op_purchase_cash - op_salary_cash - op_tax_cash - op_other_out
        inv_net = inv_recover + inv_income + inv_dispose - inv_invest_pay - inv_buy_asset
        fin_net = fin_loan_in + fin_invest_in - fin_loan_repay - fin_interest_pay - fin_profit_dist
        net_increase = op_net + inv_net + fin_net

        # 期初/期末现金余额
        opening_cash = sum(get_opening_balance(c)[0] if get_opening_balance(c)[0] else get_opening_balance(c)[1] for c in CASH_CODES)
        ending_cash = opening_cash + net_increase

        # --- 行项目定义 ---
        CF_ITEMS = [
            # 序号, 项目, 金额
            {"no": "",  "name": "一、经营活动产生的现金流量",                                            "val": None,  "bold": True},
            {"no": 1,   "name": "  销售产成品、商品、提供劳务收到的现金",                                  "val": op_sales_cash},
            {"no": 2,   "name": "  收到其他与经营活动有关的现金",                                         "val": op_other_in},
            {"no": 3,   "name": "  购买原材料、商品、接受劳务支付的现金",                                 "val": op_purchase_cash},
            {"no": 4,   "name": "  支付的职工薪酬",                                                       "val": op_salary_cash},
            {"no": 5,   "name": "  支付的税费",                                                           "val": op_tax_cash},
            {"no": 6,   "name": "  支付其他与经营活动有关的现金",                                         "val": op_other_out},
            {"no": 7,   "name": "  经营活动产生的现金流量净额",                                            "val": op_net, "bold": True},
            {"no": "",  "name": "二、投资活动产生的现金流量",                                              "val": None,  "bold": True},
            {"no": 8,   "name": "  收回短期投资、长期债券投资和长期股权投资收到的现金",                    "val": inv_recover},
            {"no": 9,   "name": "  取得投资收益收到的现金",                                                "val": inv_income},
            {"no": 10,  "name": "  处置固定资产、无形资产和其他非流动资产收回的现金净额",                  "val": inv_dispose},
            {"no": 11,  "name": "  短期投资、长期债券投资和长期股权投资支付的现金",                        "val": inv_invest_pay},
            {"no": 12,  "name": "  购建固定资产、无形资产和其他非流动资产支付的现金",                      "val": inv_buy_asset},
            {"no": 13,  "name": "  投资活动产生的现金流量净额",                                            "val": inv_net, "bold": True},
            {"no": "",  "name": "三、筹资活动产生的现金流量",                                              "val": None,  "bold": True},
            {"no": 14,  "name": "  取得借款收到的现金",                                                    "val": fin_loan_in},
            {"no": 15,  "name": "  吸收投资者投资收到的现金",                                             "val": fin_invest_in},
            {"no": 16,  "name": "  偿还借款本金支付的现金",                                                "val": fin_loan_repay},
            {"no": 17,  "name": "  偿还借款利息支付的现金",                                                "val": fin_interest_pay},
            {"no": 18,  "name": "  分配利润支付的现金",                                                    "val": fin_profit_dist},
            {"no": 19,  "name": "  筹资活动产生的现金流量净额",                                            "val": fin_net, "bold": True},
            {"no": 20,  "name": "四、现金净增加额",                                                        "val": net_increase, "bold": True},
            {"no": 21,  "name": "  期初现金余额",                                                         "val": opening_cash},
            {"no": 22,  "name": "五、期末现金余额",                                                        "val": ending_cash, "bold": True},
        ]

        # --- 构建表格数据 ---
        rows = []
        for item in CF_ITEMS:
            val = item["val"]
            val_str = fmt_amt(val) if val is not None else ""
            rows.append({
                "序号": item["no"],
                "项目": item["name"],
                "本月金额": val_str,
                "本年累积金额": val_str,
            })

        cf_df = pd.DataFrame(rows, columns=["序号", "项目", "本月金额", "本年累积金额"])
        st.dataframe(cf_df, use_container_width=True, hide_index=True)

        # --- 关键指标高亮 ---
        col_c1, col_c2, col_c3, col_c4 = st.columns(4)
        with col_c1:
            st.metric("经营活动净额", f"{op_net:,.2f}")
        with col_c2:
            st.metric("投资活动净额", f"{inv_net:,.2f}")
        with col_c3:
            st.metric("筹资活动净额", f"{fin_net:,.2f}")
        with col_c4:
            st.metric("现金净增加额", f"{net_increase:,.2f}")

        # --- 导出 Excel ---
        def export_cf_excel():
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side
            import io

            wb = Workbook()
            ws = wb.active
            ws.title = "现金流量表"

            _thin = Side(style='thin')
            _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
            _center = Alignment(horizontal='center', vertical='center')
            _right = Alignment(horizontal='right', vertical='center')
            _left = Alignment(horizontal='left', vertical='center')
            _bold = Font(bold=True)
            _title_font = Font(bold=True, size=16)
            _info_font = Font(size=11)

            # 第1行：标题
            ws.merge_cells('A1:D1')
            ws['A1'] = '现金流量表'
            ws['A1'].font = _title_font
            ws['A1'].alignment = _center
            ws.row_dimensions[1].height = 30

            # 第2行：公司名称 / 报表日期 / 货币单位
            ws['A2'] = f'编制单位：{company_name or ""}'
            ws['A2'].font = _info_font
            ws.merge_cells('B2:C2')
            ws['B2'] = f'报表日期：{report_date_str}'
            ws['B2'].font = _info_font
            ws['B2'].alignment = _center
            ws['D2'] = f'货币单位：{currency_unit}'
            ws['D2'].font = _info_font
            ws['D2'].alignment = _right
            ws.row_dimensions[2].height = 20

            # 第3行：表头
            headers = ["序号", "项目", "本月金额", "本年累积金额"]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col, value=h)
                cell.font = _bold
                cell.alignment = _center
                cell.border = _border
            ws.row_dimensions[3].height = 22

            # 数据行
            for ridx, row in enumerate(rows, 4):
                is_bold = CF_ITEMS[ridx - 4].get("bold", False)
                for col, key in enumerate(["序号", "项目", "本月金额", "本年累积金额"], 1):
                    val = row[key]
                    cell = ws.cell(row=ridx, column=col, value=val)
                    cell.border = _border
                    if col == 1:
                        cell.alignment = _center
                    elif col == 2:
                        cell.alignment = _left
                    else:
                        cell.alignment = _right
                    if is_bold:
                        cell.font = _bold

            # 列宽
            ws.column_dimensions['A'].width = 6
            ws.column_dimensions['B'].width = 52
            ws.column_dimensions['C'].width = 18
            ws.column_dimensions['D'].width = 18

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return output

        st.download_button(
            label="📥 下载现金流量表（Excel）",
            data=export_cf_excel(),
            file_name=f"现金流量表_{report_date_str or '未定'}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    # ================================================================
    # 财务可视化分析
    # ================================================================
    elif report_type == "📊 财务可视化分析":
        import plotly.express as px
        import plotly.graph_objects as go
        from plotly.subplots import make_subplots

        st.markdown("---")
        st.markdown(
            '<div style="text-align:center; font-size:22px; font-weight:700; '
            'padding:12px 0;">📊 财务可视化分析</div>',
            unsafe_allow_html=True,
        )
        st.caption("基于当前账套数据自动生成可视化图表，涵盖资产结构、损益分析、费用构成、收支趋势等。")

        # 清除缓存确保数据最新
        calc_account_balance.clear()
        get_all_opening_balances.clear()
        get_all_vouchers.clear()

        # -------- 辅助函数：获取科目余额 --------
        def _get_balance(code):
            _, _, _, _, ending = calc_account_balance(code)
            return ending

        def _get_balance_list(codes):
            return sum(_get_balance(c) for c in codes)

        # -------- 收集各类数据 --------
        # 资产类明细
        asset_data = []
        for a in ACCOUNT_CHART:
            if a["category"] != "资产":
                continue
            bal = _get_balance(a["code"])
            if abs(bal) > 0.01:
                asset_data.append({"科目": a["name"], "编码": a["code"], "余额": abs(bal)})

        # 负债类明细
        liab_data = []
        for a in ACCOUNT_CHART:
            if a["category"] != "负债":
                continue
            bal = _get_balance(a["code"])
            if abs(bal) > 0.01:
                liab_data.append({"科目": a["name"], "编码": a["code"], "余额": abs(bal)})

        # 权益类明细
        equity_data = []
        for a in ACCOUNT_CHART:
            if a["category"] != "权益":
                continue
            bal = _get_balance(a["code"])
            if abs(bal) > 0.01:
                equity_data.append({"科目": a["name"], "编码": a["code"], "余额": abs(bal)})

        # 损益类数据
        income_total = _get_balance_list(["6001", "6021", "6041", "6051", "6101", "6102", "6103",
                                          "6111", "6115", "6117", "6301"])
        expense_total = _get_balance_list(["6401", "6402", "6403", "6405",
                                           "6601", "6602", "6603", "6604", "6605", "6606",
                                           "6641", "6642", "6701", "6711"])

        # 费用明细
        expense_items = []
        for code, name in [("6401", "主营业务成本"), ("6402", "其他业务成本"),
                           ("6403", "税金及附加"), ("6601", "销售费用"),
                           ("6602", "管理费用"), ("6603", "财务费用"),
                           ("6641", "信用减值损失"), ("6642", "资产减值损失"),
                           ("6701", "营业外支出"), ("6711", "所得税费用")]:
            bal = _get_balance(code)
            if abs(bal) > 0.01:
                expense_items.append({"科目": name, "编码": code, "金额": abs(bal)})

        # 收入明细
        income_items = []
        for code, name in [("6001", "主营业务收入"), ("6051", "其他业务收入"),
                            ("6111", "投资收益"), ("6117", "其他收益"),
                            ("6301", "营业外收入"), ("6101", "公允价值变动损益")]:
            bal = _get_balance(code)
            if abs(bal) > 0.01:
                income_items.append({"科目": name, "编码": code, "金额": abs(bal)})

        # 月度凭证数据（用于趋势图）
        all_vouchers_df = get_all_vouchers()
        if not all_vouchers_df.empty and "voucher_date" in all_vouchers_df.columns:
            all_vouchers_df["voucher_date"] = pd.to_datetime(all_vouchers_df["voucher_date"], errors="coerce")
            all_vouchers_df["月份"] = all_vouchers_df["voucher_date"].dt.to_period("M").astype(str)
            monthly_data = all_vouchers_df.groupby("月份").agg(
                借方合计=("debit_amount", "sum"),
                贷方合计=("credit_amount", "sum"),
            ).reset_index()
        else:
            monthly_data = pd.DataFrame(columns=["月份", "借方合计", "贷方合计"])

        # ============================================================
        # 图表区域
        # ============================================================
        # --- 核心指标卡片 ---
        asset_total = sum(d["余额"] for d in asset_data)
        liab_total = sum(d["余额"] for d in liab_data)
        equity_total = sum(d["余额"] for d in equity_data)
        net_profit = income_total - expense_total

        m1, m2, m3, m4 = st.columns(4)
        m1.metric("总资产", fmt_money(asset_total))
        m2.metric("总负债", fmt_money(liab_total))
        m3.metric("所有者权益", fmt_money(equity_total))
        m4.metric("本期净利润", fmt_money(abs(net_profit)),
                  delta=f"{'盈利' if net_profit > 0 else '亏损'}" if net_profit != 0 else "持平")

        # 负债率
        if asset_total > 0:
            debt_ratio = liab_total / asset_total * 100
            st.info(f"📌 资产负债率：{debt_ratio:.1f}%  |  权益比率：{(equity_total / asset_total * 100) if asset_total else 0:.1f}%")

        st.markdown("---")

        # --- 子标签页 ---
        viz_sub1, viz_sub2, viz_sub3, viz_sub4 = st.tabs([
            "🏗️ 资产结构", "💰 损益分析", "📈 收支趋势", "📊 费用构成",
        ])

        # ===== 1. 资产结构 =====
        with viz_sub1:
            if asset_data:
                c1, c2 = st.columns(2)

                with c1:
                    # 资产构成饼图
                    fig_asset = px.pie(
                        asset_data, values="余额", names="科目",
                        title="资产构成",
                        color_discrete_sequence=px.colors.qualitative.Set2,
                    )
                    fig_asset.update_traces(textposition="inside", textinfo="label+percent")
                    fig_asset.update_layout(showlegend=False, height=400)
                    st.plotly_chart(fig_asset, use_container_width=True)

                with c2:
                    # 负债与权益对比
                    struct_data = [
                        {"类别": "负债", "金额": liab_total},
                        {"类别": "所有者权益", "金额": equity_total},
                    ]
                    fig_struct = px.pie(
                        struct_data, values="金额", names="类别",
                        title="负债与权益结构",
                        color="类别",
                        color_discrete_map={"负债": "#ef5350", "所有者权益": "#42a5f5"},
                    )
                    fig_struct.update_traces(textposition="inside", textinfo="label+percent")
                    fig_struct.update_layout(height=400)
                    st.plotly_chart(fig_struct, use_container_width=True)

                # 资产 TOP 10 柱状图
                if len(asset_data) > 1:
                    top_assets = sorted(asset_data, key=lambda x: x["余额"], reverse=True)[:10]
                    fig_bar = px.bar(
                        top_assets, x="余额", y="科目", orientation="h",
                        title="资产科目 TOP 10",
                        color="余额", color_continuous_scale="Blues",
                    )
                    fig_bar.update_layout(yaxis={"categoryorder": "total ascending"}, height=400)
                    st.plotly_chart(fig_bar, use_container_width=True)

                # 负债明细
                if liab_data:
                    fig_liab = px.bar(
                        liab_data, x="余额", y="科目", orientation="h",
                        title="负债明细",
                        color_discrete_sequence=["#ef5350"],
                    )
                    fig_liab.update_layout(yaxis={"categoryorder": "total ascending"}, height=350)
                    st.plotly_chart(fig_liab, use_container_width=True)
            else:
                st.info("暂无资产数据，请先录入凭证或期初余额。")

        # ===== 2. 损益分析 =====
        with viz_sub2:
            # 收入 vs 费用对比
            if income_total > 0 or expense_total > 0:
                ie_data = pd.DataFrame({
                    "类别": ["收入合计", "费用合计"],
                    "金额": [income_total, expense_total],
                })
                fig_ie = px.bar(
                    ie_data, x="类别", y="金额",
                    title="收入 vs 费用",
                    color="类别",
                    color_discrete_map={"收入合计": "#66bb6a", "费用合计": "#ef5350"},
                    text="金额",
                )
                fig_ie.update_traces(texttemplate="%{text:,.0f}", textposition="outside")
                fig_ie.update_layout(height=400, showlegend=False)
                st.plotly_chart(fig_ie, use_container_width=True)

                # 净利润瀑布图
                waterfall_data = [
                    {"项目": "收入合计", "金额": income_total, "类型": "收入"},
                    {"项目": "费用合计", "金额": -expense_total, "类型": "费用"},
                    {"项目": "净利润", "金额": net_profit, "类型": "结果"},
                ]
                fig_wf = go.Figure(go.Waterfall(
                    name="损益",
                    orientation="v",
                    measure=["relative", "relative", "total"],
                    x=[d["项目"] for d in waterfall_data],
                    y=[d["金额"] for d in waterfall_data],
                    connector={"line": {"color": "#bbb"}},
                    increasing={"marker": {"color": "#66bb6a"}},
                    decreasing={"marker": {"color": "#ef5350"}},
                    totals={"marker": {"color": "#42a5f5"}},
                ))
                fig_wf.update_layout(title="损益瀑布图", height=400, yaxis_title="金额（元）")
                st.plotly_chart(fig_wf, use_container_width=True)

                # 收入构成饼图
                if income_items:
                    c1, c2 = st.columns(2)
                    with c1:
                        fig_inc = px.pie(
                            income_items, values="金额", names="科目",
                            title="收入构成",
                            color_discrete_sequence=px.colors.qualitative.Pastel,
                        )
                        fig_inc.update_traces(textposition="inside", textinfo="label+percent")
                        fig_inc.update_layout(height=400)
                        st.plotly_chart(fig_inc, use_container_width=True)
                    with c2:
                        # 毛利率计算
                        main_revenue = _get_balance("6001") + _get_balance("6051")
                        main_cost = _get_balance("6401") + _get_balance("6402")
                        gross_profit = main_revenue - main_cost
                        gross_margin = (gross_profit / main_revenue * 100) if main_revenue > 0 else 0
                        st.metric("营业收入", fmt_money(main_revenue))
                        st.metric("营业成本", fmt_money(main_cost))
                        st.metric("毛利润", fmt_money(gross_profit))
                        st.metric("毛利率", f"{gross_margin:.1f}%")
            else:
                st.info("暂无损益数据。")

        # ===== 3. 收支趋势 =====
        with viz_sub3:
            if not monthly_data.empty:
                # 月度借贷趋势折线图
                fig_trend = go.Figure()
                fig_trend.add_trace(go.Scatter(
                    x=monthly_data["月份"], y=monthly_data["贷方合计"],
                    mode="lines+markers", name="贷方（收入方）",
                    line=dict(color="#66bb6a", width=2),
                ))
                fig_trend.add_trace(go.Scatter(
                    x=monthly_data["月份"], y=monthly_data["借方合计"],
                    mode="lines+markers", name="借方（支出方）",
                    line=dict(color="#ef5350", width=2),
                ))
                fig_trend.update_layout(
                    title="月度借贷发生额趋势",
                    xaxis_title="月份", yaxis_title="金额（元）",
                    height=400, hovermode="x unified",
                )
                st.plotly_chart(fig_trend, use_container_width=True)

                # 月度凭证数量
                monthly_count = all_vouchers_df.groupby("月份").size().reset_index(name="凭证数")
                if not monthly_count.empty:
                    fig_count = px.bar(
                        monthly_count, x="月份", y="凭证数",
                        title="月度凭证数量",
                        color_discrete_sequence=["#42a5f5"],
                    )
                    fig_count.update_layout(height=350)
                    st.plotly_chart(fig_count, use_container_width=True)

                # 月度净发生额（贷方 - 借方）
                monthly_data["净发生额"] = monthly_data["贷方合计"] - monthly_data["借方合计"]
                colors = ["#66bb6a" if v >= 0 else "#ef5350" for v in monthly_data["净发生额"]]
                fig_net = px.bar(
                    monthly_data, x="月份", y="净发生额",
                    title="月度净发生额（贷方 - 借方）",
                )
                fig_net.update_traces(marker_color=colors)
                fig_net.update_layout(height=350)
                st.plotly_chart(fig_net, use_container_width=True)
            else:
                st.info("暂无凭证数据，请先录入凭证。")

        # ===== 4. 费用构成 =====
        with viz_sub4:
            if expense_items:
                # 费用构成饼图
                fig_exp = px.pie(
                    expense_items, values="金额", names="科目",
                    title="费用构成",
                    color_discrete_sequence=px.colors.qualitative.Set3,
                )
                fig_exp.update_traces(textposition="inside", textinfo="label+percent")
                fig_exp.update_layout(height=400)
                st.plotly_chart(fig_exp, use_container_width=True)

                # 费用排序柱状图
                sorted_exp = sorted(expense_items, key=lambda x: x["金额"], reverse=True)
                fig_exp_bar = px.bar(
                    sorted_exp, x="科目", y="金额",
                    title="费用科目排序",
                    color="金额", color_continuous_scale="Reds",
                )
                fig_exp_bar.update_layout(height=400, xaxis_tickangle=-30)
                st.plotly_chart(fig_exp_bar, use_container_width=True)

                # 费用占比表
                total_exp = sum(d["金额"] for d in expense_items)
                st.markdown("#### 费用占比明细")
                for item in sorted_exp:
                    pct = item["金额"] / total_exp * 100 if total_exp > 0 else 0
                    st.markdown(
                        f"- **{item['科目']}**：{fmt_money(item['金额'])} 元（{pct:.1f}%）"
                    )
            else:
                st.info("暂无费用数据。")

            # 税务相关费用专项分析
            tax_codes = ["6601", "6602"]  # 销售/管理费用下的税务二级科目
            tax_items = []
            df_custom = get_custom_accounts()
            if not df_custom.empty and "tax_flag" in df_custom.columns:
                tax_df = df_custom[df_custom["tax_flag"].notna() & (df_custom["tax_flag"] != "")]
                if not tax_df.empty:
                    for _, row in tax_df.iterrows():
                        _, _, _, _, bal = calc_account_balance(row["full_code"])
                        if abs(bal) > 0.01:
                            tax_items.append({
                                "科目": row["full_name"],
                                "税务标记": row["tax_flag"],
                                "金额": abs(bal),
                            })

            if tax_items:
                st.markdown("---")
                st.markdown("#### 🔖 税务相关费用分析")
                fig_tax = px.bar(
                    tax_items, x="科目", y="金额",
                    title="税务标记科目金额",
                    color="税务标记",
                    color_discrete_sequence=px.colors.qualitative.Dark2,
                )
                fig_tax.update_layout(height=400, xaxis_tickangle=-30)
                st.plotly_chart(fig_tax, use_container_width=True)


st.divider()



# ============================================================
# 【💰 财务会计】模块三：AI 智能问答（需要 API）
# ============================================================
with tab3:
    st.header("🤖 AI 智能问答")
    st.caption("需要 API Key 才能使用。AI 可以查看你的财务数据，回答问题、提供建议。")

    # --- API 设置 ---
    st.subheader("⚙️ API 设置")

    with st.expander("填写 API Key 和选择模型", expanded=True):
        TEXT_MODELS = {
            "硅基流动 - DeepSeek-V3（推荐）": {
                "base_url": "https://api.siliconflow.cn/v1",
                "model": "deepseek-ai/DeepSeek-V3",
                "key_hint": "在 cloud.siliconflow.cn 获取",
            },
            "硅基流动 - Qwen2.5-72B": {
                "base_url": "https://api.siliconflow.cn/v1",
                "model": "Qwen/Qwen2.5-72B-Instruct",
                "key_hint": "在 cloud.siliconflow.cn 获取",
            },
        }

        VISION_MODELS = {
            "硅基流动 - Qwen2-VL-72B（推荐）": {
                "base_url": "https://api.siliconflow.cn/v1",
                "model": "Qwen/Qwen2-VL-72B-Instruct",
            },
            "硅基流动 - DeepSeek-VL2": {
                "base_url": "https://api.siliconflow.cn/v1",
                "model": "deepseek-ai/deepseek-vl2",
            },
        }

        col1, col2 = st.columns(2)
        with col1:
            text_choice = st.selectbox("文字模型", list(TEXT_MODELS.keys()))
            text_config = TEXT_MODELS[text_choice]
            st.caption(f"模型：`{text_config['model']}`")
        with col2:
            vision_choice = st.selectbox("视觉模型（看凭证图片用）", list(VISION_MODELS.keys()))
            vision_config = VISION_MODELS[vision_choice]
            st.caption(f"模型：`{vision_config['model']}`")

        api_key = st.text_input(
            "API Key（硅基流动注册获取）",
            type="password",
            help="在 cloud.siliconflow.cn 注册，一个 Key 同时用于文字和视觉模型",
        )
        st.caption("没有 API Key？前往 cloud.siliconflow.cn 注册，有免费额度。")

    if not api_key:
        st.info("👆 请先填写 API Key 才能使用 AI 功能。模块一和模块二不需要 API Key，可以直接使用。")
    else:
        # --- AI 工具函数 ---
        def tool_query_account(account_name: str) -> str:
            """工具：查询科目余额"""
            # 根据名称找科目编码
            matched = [a for a in ACCOUNT_CHART if account_name in a["name"]]
            if not matched:
                return f"未找到包含「{account_name}」的科目。"
            results = []
            for acc in matched:
                od, oc, p_debit, p_credit, ending = calc_account_balance(acc["code"])
                results.append(
                    f"科目【{acc['name']}】({acc['category']}类)："
                    f"期初借方={od:,.2f}，期初贷方={oc:,.2f}，"
                    f"本期借方发生额={p_debit:,.2f}，本期贷方发生额={p_credit:,.2f}，"
                    f"期末余额={ending:,.2f}"
                )
            return "\n".join(results)

        def tool_query_all_vouchers() -> str:
            """工具：查询所有凭证"""
            df = get_all_vouchers()
            if df.empty:
                return "暂无凭证记录。"
            voucher_nums = df["voucher_number"].unique()
            lines = [f"共 {len(voucher_nums)} 张凭证："]
            for vnum in voucher_nums:
                vdf = df[df["voucher_number"] == vnum]
                vdate = vdf.iloc[0]["voucher_date"]
                vsummary = vdf.iloc[0]["summary"]
                d_total = vdf["debit_amount"].sum()
                lines.append(f"  {vnum} | {vdate} | {vsummary} | 借贷合计={d_total:,.2f}")
            return "\n".join(lines)

        def tool_calc_bad_debt(account_name: str) -> str:
            """工具：按账龄分析法计算坏账准备"""
            # 简化版：直接用应收账款期末余额 × 5% 估算
            matched = [a for a in ACCOUNT_CHART if account_name in a["name"]]
            if not matched:
                return f"未找到包含「{account_name}」的科目。"
            results = []
            for acc in matched:
                _, _, _, _, ending = calc_account_balance(acc["code"])
                if acc["code"] == "1122":  # 应收账款
                    provision = ending * 0.05
                    results.append(
                        f"应收账款期末余额：{ending:,.2f} 元\n"
                        f"按5%计提坏账准备：{provision:,.2f} 元\n"
                        f"（注：实际应按账龄分析法分档计提，此处为简化估算）"
                    )
                else:
                    results.append(f"{acc['name']} 期末余额：{ending:,.2f} 元")
            return "\n".join(results)

        TOOLS = [
            {
                "type": "function",
                "function": {
                    "name": "tool_query_account",
                    "description": "查询某个财务科目的余额、发生额等数据。",
                    "parameters": {
                        "type": "object",
                        "properties": {
                            "account_name": {
                                "type": "string",
                                "description": "科目名称或关键字，例如：应收账款、银行存款、管理费用",
                            }
                        },
                        "required": ["account_name"],
                    },
                },
            },
            {
                "type": "function",
                "function": {
                    "name": "tool_query_all_vouchers",
                    "description": "查询所有已录入的凭证记录。",
                    "parameters": {"type": "object", "properties": {}},
                },
            },
            {
                "type": "function",
                "function": {
                    "name": "tool_calc_bad_debt",
                    "description": "根据应收账款余额计算坏账准备。",
                    "parameters": {
                        "type": "object",
                        "properties": {
                            "account_name": {
                                "type": "string",
                                "description": "科目名称，例如：应收账款",
                            }
                        },
                        "required": ["account_name"],
                    },
                },
            },
        ]

        TOOL_FUNCTIONS = {
            "tool_query_account": tool_query_account,
            "tool_query_all_vouchers": tool_query_all_vouchers,
            "tool_calc_bad_debt": tool_calc_bad_debt,
        }

        def chat_with_agent(messages: list) -> str:
            """和 AI 对话，AI 会自动决定要不要调用工具"""
            client = OpenAI(api_key=api_key, base_url=text_config["base_url"])
            response = client.chat.completions.create(
                model=text_config["model"],
                messages=messages,
                tools=TOOLS,
                temperature=0.3,
            )
            message = response.choices[0].message

            if not message.tool_calls:
                return message.content or "（AI 没有返回内容）"

            messages.append(message)

            for tool_call in message.tool_calls:
                tool_name = tool_call.function.name
                tool_args = json.loads(tool_call.function.arguments)
                with st.container():
                    st.info(f"🔧 智能体正在调用工具：**{tool_name}**，参数：{tool_args}")
                func = TOOL_FUNCTIONS.get(tool_name)
                result = func(**tool_args) if func else f"未知工具：{tool_name}"
                messages.append({
                    "role": "tool",
                    "tool_call_id": tool_call.id,
                    "content": result,
                })

            final_response = client.chat.completions.create(
                model=text_config["model"],
                messages=messages,
                temperature=0.3,
            )
            return final_response.choices[0].message.content

        # --- 聊天界面 ---
        st.subheader("💬 和 AI 聊天")

        if "ai_messages" not in st.session_state:
            st.session_state.ai_messages = [
                {
                    "role": "system",
                    "content": (
                        "你是一个专业的财务智能助手。你可以查询科目余额、查看所有凭证、计算坏账准备。"
                        "请用中文回答，金额用千分位格式显示。"
                        "如果用户问的数据你查不到，请告诉他可用的科目有哪些。"
                    ),
                }
            ]
            st.session_state.ai_display = []

        for msg in st.session_state.ai_display:
            with st.chat_message(msg["role"]):
                st.markdown(msg["content"])

        user_input = st.chat_input("问我一个财务问题，比如：帮我查一下应收账款余额")

        if user_input:
            with st.chat_message("user"):
                st.markdown(user_input)
            st.session_state.ai_display.append({"role": "user", "content": user_input})

            with st.chat_message("assistant"):
                with st.spinner("财务智能体正在思考..."):
                    st.session_state.ai_messages.append({"role": "user", "content": user_input})
                    try:
                        reply = chat_with_agent(st.session_state.ai_messages)
                    except Exception as e:
                        reply = f"❌ 出错了：{e}\n\n常见原因：\n- API Key 填错了\n- 网络问题\n- 账户余额不足"
                    st.markdown(reply)

            st.session_state.ai_display.append({"role": "assistant", "content": reply})
            st.session_state.ai_messages.append({"role": "assistant", "content": reply})

        # --- 凭证图片识别 ---
        st.divider()
        st.subheader("🧾 凭证图片识别")
        st.caption("上传发票/收据图片，AI 自动识别内容并建议会计分录。")

        import base64

        def encode_image(image_file):
            return base64.b64encode(image_file.getvalue()).decode("utf-8")

        def recognize_voucher(image_base64, image_type="image/jpeg"):
            client = OpenAI(api_key=api_key, base_url=vision_config["base_url"])
            response = client.chat.completions.create(
                model=vision_config["model"],
                messages=[
                    {
                        "role": "user",
                        "content": [
                            {
                                "type": "text",
                                "text": (
                                    "请仔细识别这张凭证/发票/收据上的所有信息，包括：\n"
                                    "1. 凭证类型（如：增值税专用发票、普通发票、收据、银行回单等）\n"
                                    "2. 日期\n"
                                    "3. 金额（含税/不含税）\n"
                                    "4. 摘要或商品名称\n"
                                    "5. 供应商/客户名称\n"
                                    "6. 税额（如果有）\n"
                                    "请用清晰的格式列出。如果看不清楚的地方请标注。"
                                ),
                            },
                            {
                                "type": "image_url",
                                "image_url": {"url": f"data:{image_type};base64,{image_base64}"},
                            },
                        ],
                    }
                ],
                temperature=0.1,
            )
            return response.choices[0].message.content

        def suggest_entry(recognition_result):
            client = OpenAI(api_key=api_key, base_url=text_config["base_url"])
            response = client.chat.completions.create(
                model=text_config["model"],
                messages=[
                    {
                        "role": "system",
                        "content": (
                            "你是一个专业的会计。根据用户提供的凭证识别结果，"
                            "建议对应的会计分录（借方和贷方）。\n"
                            "请按以下格式输出：\n"
                            "摘要：xxx\n"
                            "借：xxx科目  xxx元\n"
                            "贷：xxx科目  xxx元\n"
                            "并简要说明理由。"
                        ),
                    },
                    {
                        "role": "user",
                        "content": f"以下是凭证识别结果，请建议会计分录：\n\n{recognition_result}",
                    },
                ],
                temperature=0.3,
            )
            return response.choices[0].message.content

        uploaded_images = st.file_uploader(
            "上传凭证/发票图片", type=["png", "jpg", "jpeg"],
            accept_multiple_files=True,
        )

        if uploaded_images:
            for img in uploaded_images:
                st.image(img, caption=img.name, width=400)
                if st.button(f"🔍 AI 识别：{img.name}", key=f"ai_btn_{img.name}"):
                    with st.spinner("AI 正在识别凭证内容..."):
                        try:
                            img_base64 = encode_image(img)
                            img_type = f"image/{img.type.split('/')[-1]}" if img.type else "image/jpeg"
                            recognition = recognize_voucher(img_base64, img_type)
                            st.success("✅ 识别完成！")
                            st.markdown("**识别结果：**")
                            st.text(recognition)

                            with st.spinner("AI 正在建议会计分录..."):
                                entry = suggest_entry(recognition)
                            st.markdown("**建议的会计分录：**")
                            st.markdown(entry)
                        except Exception as e:
                            st.error(f"❌ 识别出错：{e}")



# ============================================================
# 底部说明
# ============================================================
st.divider()
st.caption(
    "💡 **系统说明：**\n"
    "- 📝 模块一：离线记账，无需 API，免费使用\n"
    "- 📊 模块二：三大报表（资产负债表、利润表、现金流量表）+ 财务可视化\n"
    "- 🤖 模块三：AI 智能问答 + 凭证识别，需要 API Key\n"
    "\n"
    "**安装依赖：** `pip install streamlit openai pandas openpyxl plotly`"
)
